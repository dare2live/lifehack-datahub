"""Parser for Liaoning undergraduate projection score spreadsheets."""
from __future__ import annotations

import io
import json
from pathlib import Path
from typing import Any
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook


TIE_BREAKER_FIELDS = [
    "语数成绩",
    "语数最高成绩",
    "外语成绩",
    "首选科目成绩",
    "再选科目最高成绩",
    "再选科目次高成绩",
    "志愿号",
]


def parse_ln_projection_score_files(
    paths: list[Path],
    *,
    score_year: int,
    batch: str,
    source_date: str,
    password_candidates: list[str] | None = None,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for path in paths:
        rows.extend(parse_ln_projection_score_file(
            path,
            score_year=score_year,
            batch=batch,
            source_date=source_date,
            password_candidates=password_candidates or [],
        ))
    return rows


def parse_ln_projection_score_file(
    path: Path,
    *,
    score_year: int,
    batch: str,
    source_date: str,
    password_candidates: list[str],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for wb in _load_workbooks(path, password_candidates):
        rows.extend(_rows_from_workbook(wb, score_year=score_year, batch=batch, source_date=source_date))
    return rows


def _rows_from_workbook(wb, *, score_year: int, batch: str, source_date: str) -> list[dict[str, Any]]:
    ws = wb.active
    subject_cat = _subject_from_worksheet(ws)
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        school_code, school_name, major_code, major_full, min_score = row[:5]
        if not school_code or not school_name or not major_code or not major_full:
            continue
        rows.append({
            "school_code": _clean_text(school_code),
            "school_name": _clean_text(school_name),
            "major_code": _clean_text(major_code),
            "major_full": _clean_text(major_full),
            "batch": batch,
            "subject_cat": subject_cat,
            "score_year": score_year,
            "min_score": _coerce_int(min_score),
            "tie_breaker_json": _tie_breaker_json(row[5:12]),
            "source_date": source_date,
        })
    return rows


def _load_workbooks(path: Path, password_candidates: list[str]):
    if path.suffix.lower() == ".zip":
        with ZipFile(path) as z:
            workbooks = []
            for name in z.namelist():
                if name.startswith("__MACOSX/") or not name.lower().endswith((".xlsx", ".xlsm")):
                    continue
                workbooks.append(_load_workbook_bytes(z.read(name), path, password_candidates))
            if not workbooks:
                raise ValueError(f"zip contains no xlsx workbook: {path}")
            return workbooks
    return [_load_workbook_path(path, password_candidates)]


def _load_workbook_path(path: Path, password_candidates: list[str]):
    try:
        return load_workbook(path, read_only=True, data_only=True)
    except BadZipFile:
        pass

    with path.open("rb") as f:
        msoffcrypto = _msoffcrypto()
        office = msoffcrypto.OfficeFile(f)
        if not office.is_encrypted():
            raise ValueError(f"workbook is not readable as xlsx and is not encrypted: {path}")
        for password in password_candidates:
            out = io.BytesIO()
            try:
                office.load_key(password=password)
                office.decrypt(out)
                out.seek(0)
                return load_workbook(out, read_only=True, data_only=True)
            except Exception:
                f.seek(0)
                office = msoffcrypto.OfficeFile(f)
        raise ValueError(f"encrypted workbook cannot be decrypted: {path}")


def _load_workbook_bytes(data: bytes, source_path: Path, password_candidates: list[str]):
    try:
        return load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except BadZipFile:
        pass

    msoffcrypto = _msoffcrypto()
    office = msoffcrypto.OfficeFile(io.BytesIO(data))
    if not office.is_encrypted():
        raise ValueError(f"workbook in zip is not readable as xlsx and is not encrypted: {source_path}")
    for password in password_candidates:
        out = io.BytesIO()
        try:
            office.load_key(password=password)
            office.decrypt(out)
            out.seek(0)
            return load_workbook(out, read_only=True, data_only=True)
        except Exception:
            office = msoffcrypto.OfficeFile(io.BytesIO(data))
    raise ValueError(f"encrypted workbook in zip cannot be decrypted: {source_path}")


def _msoffcrypto():
    try:
        import msoffcrypto
    except ImportError as exc:
        raise RuntimeError(
            "msoffcrypto is required only for encrypted Liaoning projection score workbooks"
        ) from exc
    return msoffcrypto


def _subject_from_worksheet(ws) -> str:
    title = str(ws.cell(row=1, column=1).value or ws.title)
    if "物理" in title or "物理" in ws.title:
        return "物理类"
    if "历史" in title or "历史" in ws.title:
        return "历史类"
    return str(ws.title).strip()


def _tie_breaker_json(values: tuple[Any, ...]) -> str:
    payload = {
        field: _clean_text(value)
        for field, value in zip(TIE_BREAKER_FIELDS, values)
        if value not in (None, "")
    }
    return json.dumps(payload, ensure_ascii=False)


def _clean_text(value: Any) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value or "").strip()


def _coerce_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).strip()
    return int(float(text)) if text else None
