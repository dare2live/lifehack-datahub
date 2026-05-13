"""Parse local Liaoning application workbooks into standard cleaned tables."""
from __future__ import annotations

import csv
import hashlib
import json
import re
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from datahub.config import CONFIG_DIR, get_table_schema


PLAN_TABLE = "fa_dim_ln_admission_plan"
SCORE_TABLE = "fa_fact_ln_score_history"


def parse_ln_application_workbooks(
    paths: list[Path],
    *,
    config_path: Path | None = None,
    profile: str = "default",
) -> dict[str, Any]:
    config = _load_profile(config_path, profile)
    config["built_at"] = datetime.utcnow().replace(microsecond=0).isoformat()
    plan_rows: list[dict[str, Any]] = []
    score_rows: list[dict[str, Any]] = []
    sheet_reports = []

    for path in paths:
        workbook_plan_rows, workbook_score_rows, workbook_report = _parse_workbook(path, config)
        plan_rows.extend(workbook_plan_rows)
        score_rows.extend(workbook_score_rows)
        sheet_reports.extend(workbook_report)

    plan_rows, plan_duplicates = _deduplicate(
        plan_rows,
        get_table_schema(PLAN_TABLE)["primary_key"],
        config["duplicate_policy"],
        PLAN_TABLE,
    )
    score_rows, score_duplicates = _deduplicate(
        score_rows,
        get_table_schema(SCORE_TABLE)["primary_key"],
        config["duplicate_policy"],
        SCORE_TABLE,
    )

    return {
        "tables": {
            PLAN_TABLE: plan_rows,
            SCORE_TABLE: score_rows,
        },
        "report": {
            "profile": profile,
            "input_files": [str(path) for path in paths],
            "matched_sheets": [item for item in sheet_reports if item["matched"]],
            "ignored_sheets": [item for item in sheet_reports if not item["matched"]],
            "row_counts": {
                PLAN_TABLE: len(plan_rows),
                SCORE_TABLE: len(score_rows),
            },
            "duplicate_counts": {
                PLAN_TABLE: plan_duplicates,
                SCORE_TABLE: score_duplicates,
            },
            "notes": "Parser output is cleaned CSV only. Build data packages with build-local; do not commit source Excel or cleaned CSV.",
        },
    }


def write_application_workbook_outputs(
    result: dict[str, Any],
    *,
    plan_output: Path,
    score_output: Path,
    report_output: Path | None = None,
) -> dict[str, Any]:
    plan_schema = get_table_schema(PLAN_TABLE)
    score_schema = get_table_schema(SCORE_TABLE)
    _write_csv(plan_output, result["tables"][PLAN_TABLE], plan_schema["columns"])
    _write_csv(score_output, result["tables"][SCORE_TABLE], score_schema["columns"])
    report = {
        **result["report"],
        "outputs": {
            PLAN_TABLE: str(plan_output),
            SCORE_TABLE: str(score_output),
        },
    }
    if report_output:
        report_output.parent.mkdir(parents=True, exist_ok=True)
        report_output.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return report


def _parse_workbook(path: Path, config: dict[str, Any]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    plan_rows: list[dict[str, Any]] = []
    score_rows: list[dict[str, Any]] = []
    reports = []
    for ws in wb.worksheets:
        rule = _match_sheet(ws.title, config["sheet_rules"])
        if not rule:
            reports.append({
                "file": str(path),
                "sheet": ws.title,
                "matched": False,
                "reason": "no sheet rule matched",
            })
            continue
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            reports.append({
                "file": str(path),
                "sheet": ws.title,
                "matched": False,
                "reason": "empty sheet",
            })
            continue
        headers = [_clean_header(cell) for cell in header]
        sheet_plan_count = 0
        sheet_score_count = 0
        skipped_missing_key = 0
        skipped_score_values = Counter()
        for row in rows:
            if not any(value not in (None, "") for value in row):
                continue
            base = _base_row(row, headers, rule, config)
            if not all(base.get(column) for column in ["school_code", "school_name", "major_code", "major_full"]):
                skipped_missing_key += 1
                continue
            plan_rows.append(_plan_row(row, headers, base, config, source_file=path.name))
            sheet_plan_count += 1
            built_score_rows, skip_counts = _score_rows(row, headers, base, config)
            score_rows.extend(built_score_rows)
            sheet_score_count += len(built_score_rows)
            skipped_score_values.update(skip_counts)
        reports.append({
            "file": str(path),
            "sheet": ws.title,
            "matched": True,
            "batch": rule["batch"],
            "subject_cat": rule["subject_cat"],
            "plan_rows": sheet_plan_count,
            "score_rows": sheet_score_count,
            "skipped_missing_key_rows": skipped_missing_key,
            "skipped_score_values": dict(sorted(skipped_score_values.items())),
        })
    return plan_rows, score_rows, reports


def _base_row(
    row: tuple[Any, ...],
    headers: list[str],
    rule: dict[str, Any],
    config: dict[str, Any],
) -> dict[str, Any]:
    aliases = config["field_aliases"]
    return {
        "school_code": _clean_code(_value(row, headers, aliases["school_code"])),
        "school_name": _clean_text(_value(row, headers, aliases["school_name"])),
        "major_code": _clean_code(_value(row, headers, aliases["major_code"])),
        "major_full": _clean_text(_value(row, headers, aliases["major_full"])),
        "batch": rule["batch"],
        "subject_cat": rule["subject_cat"],
    }


def _plan_row(
    row: tuple[Any, ...],
    headers: list[str],
    base: dict[str, Any],
    config: dict[str, Any],
    *,
    source_file: str,
) -> dict[str, Any]:
    aliases = config["field_aliases"]
    plan_count_year = int(config["plan_count_year"])
    plan_year = int(config.get("plan_year") or plan_count_year)
    score_year = _score_year_config(config, plan_count_year)
    return {
        "id": _make_id(base["school_code"], base["major_code"], base["batch"], base["subject_cat"], plan_year),
        **base,
        "year": plan_year,
        "major_short": _clean_text(_value(row, headers, aliases.get("major_short", []))),
        "department": _clean_text(_value(row, headers, aliases.get("department", []))),
        "school_tier": _clean_text(_value(row, headers, aliases.get("school_tier", []))),
        "region": _clean_text(_value(row, headers, aliases.get("region", []))),
        "plan_count": _coerce_number(_value(row, headers, score_year.get("plan_count_aliases", []))),
        "tuition": _coerce_number(_value(row, headers, aliases.get("tuition", []))),
        "school_type": _clean_text(_value(row, headers, aliases.get("school_type", []))),
        "school_nature": _clean_text(_value(row, headers, aliases.get("school_nature", []))),
        "city": _clean_text(_value(row, headers, aliases.get("city", []))),
        "city_level_tag": _clean_text(_value(row, headers, aliases.get("city_level_tag", []))),
        "school_rank": _clean_text(_value(row, headers, aliases.get("school_rank", []))),
        "subject_eval": _clean_text(_value(row, headers, aliases.get("subject_eval", []))),
        "top_discipline": _flag_value(_value(row, headers, aliases.get("top_discipline", []))),
        "grad_school_type": _clean_text(_value(row, headers, aliases.get("grad_school_type", []))),
        "postgrad_rate": _coerce_number(_value(row, headers, aliases.get("postgrad_rate", []))),
        "keep_research_rate": _coerce_number(_value(row, headers, aliases.get("keep_research_rate", []))),
        "transfer_policy": _clean_text(_value(row, headers, aliases.get("transfer_policy", []))),
        "study_years": _coerce_number(_value(row, headers, aliases.get("study_years", []))),
        "is_new": _flag_value(_value(row, headers, aliases.get("is_new", []))),
        "postgrad_field": _clean_text(_value(row, headers, aliases.get("postgrad_field", []))),
        "core_courses": _clean_text(_value(row, headers, aliases.get("core_courses", []))),
        "employment_directions": _clean_text(_value(row, headers, aliases.get("employment_directions", []))),
        "notes": _clean_text(_value(row, headers, aliases.get("notes", []))),
        "subject_requirement": _clean_text(_value(row, headers, aliases.get("subject_requirement", []))),
        "source_file": source_file,
        "source_date": config.get("source_date"),
        "availability_date": config.get("availability_date") or config.get("source_date"),
        "built_at": config["built_at"],
    }


def _score_rows(
    row: tuple[Any, ...],
    headers: list[str],
    base: dict[str, Any],
    config: dict[str, Any],
) -> tuple[list[dict[str, Any]], Counter[str]]:
    empty_values = {str(item).strip() for item in config["empty_values"]}
    rows = []
    skip_counts: Counter[str] = Counter()
    for score_year in config["score_years"]:
        year = int(score_year["score_year"])
        min_score = _value(row, headers, score_year["min_score_aliases"])
        min_rank = _value(row, headers, score_year["min_rank_aliases"])
        if _is_empty_metric(min_score, empty_values) or _is_empty_metric(min_rank, empty_values):
            skip_counts[str(year)] += 1
            continue
        rows.append({
            "id": _make_id(base["school_code"], base["major_code"], base["batch"], base["subject_cat"], year),
            "school_code": base["school_code"],
            "major_code": base["major_code"],
            "batch": base["batch"],
            "subject_cat": base["subject_cat"],
            "score_year": year,
            "min_score": _coerce_number(min_score),
            "min_rank": _coerce_number(min_rank),
            "plan_count": _coerce_number(_value(row, headers, score_year.get("plan_count_aliases", []))),
            "score_type": config.get("score_type", "最低分"),
            "source_date": config.get("source_date"),
            "availability_date": config.get("availability_date") or config.get("source_date"),
            "built_at": config["built_at"],
        })
    return rows, skip_counts


def _load_profile(config_path: Path | None, profile: str) -> dict[str, Any]:
    path = config_path or CONFIG_DIR / "ln_application_workbook.json"
    data = json.loads(path.read_text(encoding="utf-8"))
    profiles = data.get("profiles", {})
    if profile not in profiles:
        raise KeyError(f"unknown ln application workbook profile: {profile}")
    config = profiles[profile]
    required = [
        "sheet_rules",
        "field_aliases",
        "score_years",
        "duplicate_policy",
        "plan_count_year",
        "source_date",
        "availability_date",
        "empty_values",
    ]
    missing = [key for key in required if key not in config]
    if missing:
        raise ValueError(f"ln application workbook profile missing config: {', '.join(missing)}")
    return config


def _match_sheet(title: str, rules: list[dict[str, Any]]) -> dict[str, Any] | None:
    for rule in rules:
        exact_names = {str(item) for item in rule.get("sheet_names", [])}
        contains = [str(item) for item in rule.get("sheet_name_contains", [])]
        if title in exact_names or any(item in title for item in contains):
            return rule
    return None


def _score_year_config(config: dict[str, Any], year: int) -> dict[str, Any]:
    for item in config["score_years"]:
        if int(item["score_year"]) == year:
            return item
    raise ValueError(f"plan_count_year is not registered in score_years: {year}")


def _value(row: tuple[Any, ...], headers: list[str], aliases: list[str]) -> Any:
    clean_aliases = {_clean_header(alias) for alias in aliases}
    for index, header in enumerate(headers):
        if header in clean_aliases and index < len(row):
            return row[index]
    return None


def _deduplicate(
    rows: list[dict[str, Any]],
    primary_key: list[str],
    duplicate_policy: str,
    table_name: str,
) -> tuple[list[dict[str, Any]], int]:
    seen: dict[tuple[Any, ...], dict[str, Any]] = {}
    duplicates = 0
    for row in rows:
        key = tuple(row.get(column) for column in primary_key)
        if key in seen:
            duplicates += 1
            if duplicate_policy == "error":
                raise ValueError(f"{table_name} duplicate primary key: {key}")
            if duplicate_policy == "replace":
                seen[key] = row
            continue
        seen[key] = row
    if duplicate_policy not in {"keep_first", "replace", "error"}:
        raise ValueError(f"unknown duplicate_policy: {duplicate_policy}")
    return list(seen.values()), duplicates


def _is_empty_metric(value: Any, empty_values: set[str]) -> bool:
    if value is None:
        return True
    if isinstance(value, (int, float)) and float(value) == 0:
        return True
    return str(value).strip() in empty_values


def _clean_header(value: Any) -> str:
    return str(value or "").strip().replace("\n", "").replace(" ", "").lower()


def _clean_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return text if text else None


def _clean_code(value: Any) -> str | None:
    text = _clean_text(value)
    return text.upper() if text else None


def _coerce_number(value: Any) -> int | float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        number = float(value)
    else:
        text = str(value).strip().replace(",", "")
        percent = text.endswith("%")
        match = re.search(r"-?\d+(?:\.\d+)?", text)
        if not match:
            return None
        number = float(match.group(0))
        if percent:
            number = number / 100
    return int(number) if number.is_integer() else number


def _flag_value(value: Any) -> int:
    text = str(value or "").strip()
    return 0 if text in ("", "None", "nan", "0", "否", "-", "/") else 1


def _make_id(*parts: Any) -> str:
    raw = "||".join(str(part) for part in parts)
    return hashlib.md5(raw.encode()).hexdigest()[:16]


def _write_csv(path: Path, rows: list[dict[str, Any]], columns: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
