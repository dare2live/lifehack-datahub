"""Parse local tabular source files at the DataHub boundary."""
from __future__ import annotations

import csv
from pathlib import Path


def parse_tabular(path: Path, sheet: str | None = None) -> list[dict[str, object]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _parse_delimited(path, ",")
    if suffix == ".tsv":
        return _parse_delimited(path, "\t")
    if suffix in {".xlsx", ".xlsm"}:
        return _parse_xlsx(path, sheet)
    raise ValueError(f"unsupported tabular file type: {path.suffix}")


def _parse_delimited(path: Path, delimiter: str) -> list[dict[str, object]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f, delimiter=delimiter)
        return [dict(row) for row in reader]


def _parse_xlsx(path: Path, sheet: str | None) -> list[dict[str, object]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to parse .xlsx files") from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        return []
    columns = [str(cell).strip() if cell is not None else "" for cell in header]
    result: list[dict[str, object]] = []
    for row in rows:
        item = {columns[i]: row[i] if i < len(row) else None for i in range(len(columns)) if columns[i]}
        if any(value not in (None, "") for value in item.values()):
            result.append(item)
    return result
