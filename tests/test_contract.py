from pathlib import Path
import base64
import csv
import hashlib
import io
import json
from email.message import Message
from urllib.error import HTTPError
from zipfile import ZipFile

import duckdb
import pytest
from openpyxl import Workbook

from datahub.builders.admission_plan_snapshot import build_admission_plan_snapshot_package
from datahub.builders.admission_plan_package_audit import audit_admission_plan_package_against_core
from datahub.builders.admission_plan_reconciliation_plan import build_admission_plan_reconciliation_plan
from datahub.builders.admission_plan_reconciliation_audit import audit_admission_plan_reconciliation_plan
from datahub.builders.admission_plan_reconciliation_batch import (
    build_admission_plan_reconciliation_review_batch,
    merge_admission_plan_reconciliation_review_batch,
)
from datahub.builders.admission_plan_reconciliation_delete_plan import build_admission_plan_delete_plan_from_reconciliation_plan
from datahub.builders.career_score import build_career_score_package
from datahub.builders.career_civil_service_signal_plan import build_civil_service_signal_plan
from datahub.builders.career_source_audit import audit_career_source_plan
from datahub.builders.career_source_coverage import audit_career_source_coverage
from datahub.builders.career_source_batch import (
    build_career_source_review_batch,
    merge_career_source_review_batch,
)
from datahub.builders.career_source_package import build_career_signal_package_from_source_plan
from datahub.builders.career_source_plan import PLAN_COLUMNS as CAREER_PLAN_COLUMNS, build_career_source_plan
from datahub.builders.career_source_seed_merge import (
    apply_career_source_review_seeds,
    audit_career_source_review_seeds,
)
from datahub.builders.career_shortage_page import (
    apply_career_shortage_page_to_plan,
    parse_shortage_ranking,
)
from datahub.builders.city_context_collection_audit import audit_city_context_collection_plan
from datahub.builders.city_context_collection_batch import (
    build_city_context_review_batch,
    merge_city_context_review_batch,
)
from datahub.builders.city_context_collection_package import build_city_context_packages_from_collection_plan
from datahub.builders.city_context_collection_plan import build_city_context_collection_plan
from datahub.builders.city_context_target_cities import build_city_context_target_cities
from datahub.builders.campus_living_score import audit_campus_living_score_inputs, build_campus_living_score_package
from datahub.builders.city_development_score import build_city_development_score_package
from datahub.builders.city_listed_company_signal import build_city_listed_company_signal_package
from datahub.builders.school_city_industry_fit import (
    audit_school_city_industry_fit_inputs,
    build_school_city_industry_fit_package,
)
from datahub.builders.data_update_policy_audit import audit_data_update_policy
from datahub.builders.data_update_batch_plan import build_data_update_batch_plan
from datahub.builders.data_update_plan import build_data_update_plan
from datahub.builders.data_update_readiness_plan import build_data_update_readiness_plan
from datahub.builders.entity_normalization_registry import build_entity_normalization_registry_package
from datahub.builders.major_city_employment_fit import build_major_city_employment_fit_package
from datahub.builders.major_outcome_civil_service import build_major_outcome_from_civil_service_package
from datahub.connectors.page_images import download_page_images
from datahub.builders.outcome_collection_audit import audit_outcome_collection_plan
from datahub.builders.outcome_collection_core_coverage_audit import audit_outcome_collection_core_coverage
from datahub.builders.outcome_collection_batch import (
    build_outcome_collection_batch,
    merge_outcome_collection_batch,
)
from datahub.builders.outcome_collection_seed_merge import (
    apply_outcome_collection_review_seeds,
    audit_outcome_collection_review_seeds,
)
from datahub.builders.outcome_candidate_merge import merge_outcome_report_candidates
from datahub.builders.outcome_collection_package import build_outcome_packages_from_collection_plan
from datahub.builders.outcome_collection_verified_inherit import inherit_verified_outcome_collection_rows
from datahub.builders.major_mapping_review import build_major_mapping_review_package
from datahub.builders.local_package import build_local_package
from datahub.builders.operational_gap_report import build_operational_gap_report
from datahub.builders.release_bundle import build_release_bundle
from datahub.builders.outcome_collection_plan import PLAN_COLUMNS as OUTCOME_PLAN_COLUMNS, build_outcome_collection_plan
from datahub.builders.outcome_report_intake_merge import merge_outcome_report_intake_results
from datahub.builders.outcome_report_intake_plan import build_outcome_report_intake_plan
from datahub.builders.outcome_report_extraction_plan import build_outcome_report_extraction_plan
from datahub.builders.outcome_report_extraction_runner import run_outcome_report_extraction_plan
from datahub.builders.outcome_report_source_audit import audit_outcome_report_source_plan
from datahub.builders.outcome_report_source_batch import (
    build_outcome_report_source_review_batch,
    merge_outcome_report_source_review_batch,
)
from datahub.builders.outcome_report_source_plan import build_outcome_report_source_plan
from datahub.builders.outcome_report_source_seed_merge import (
    apply_outcome_report_source_seeds,
    audit_outcome_report_source_seeds,
)
from datahub.builders.outcome_scoped_stock_review import build_scoped_outcome_stock_review
from datahub.builders.outcome_scoped_stock_review_batch import build_scoped_outcome_stock_review_batch
from datahub.builders.outcome_scoped_stock_review_export import export_approved_scoped_stock_review_candidates
from datahub.builders.outcome_scoped_stock_review_workspace import build_scoped_outcome_stock_review_workspace
from datahub.builders.outcome_scoped_stock_review_workspace_audit import audit_scoped_outcome_stock_review_workspace
from datahub.builders.policy_tables import (
    build_policy_industry_map_package,
    build_policy_plan_history_package,
)
from datahub.builders.region_profile_from_amap import build_region_profile_package_from_amap_district
from datahub.builders.score_history_from_projection import build_score_history_from_projection_package
from datahub.builders.score_history_major_name_reference import (
    apply_score_history_major_name_reference_decisions,
    apply_score_history_pair_name_reference_decisions,
)
from datahub.builders.score_history_package_audit import audit_score_history_package_against_core
from datahub.builders.score_history_reconciliation_audit import audit_score_history_reconciliation_plan
from datahub.builders.score_history_reconciliation_auto_decision import apply_score_history_reconciliation_auto_decisions
from datahub.builders.score_history_reconciliation_batch import (
    build_score_history_reconciliation_review_batch,
    merge_score_history_reconciliation_review_batch,
)
from datahub.builders.score_history_reconciliation_delete_plan import build_score_history_delete_plan_from_reconciliation_plan
from datahub.builders.score_history_reconciliation_package import build_score_history_package_from_reconciliation_plan
from datahub.builders.score_history_reconciliation_plan import PLAN_COLUMNS, build_score_history_reconciliation_plan
from datahub.builders.score_history_snapshot import build_score_history_snapshot_package
from datahub.builders.score_source_coverage import audit_score_source_coverage
from datahub.builders.score_distribution_csv_audit import audit_score_distribution_csvs
from datahub.builders import score_distribution_image_groups as image_group_parser
from datahub.builders.score_distribution_review_workspace import (
    build_score_distribution_review_workspace,
    merge_score_distribution_review_workspace,
)
from datahub.builders.score_distribution_readiness import audit_score_distribution_readiness
from datahub.config import (
    get_table_schema,
    load_career_data_sources,
    load_data_update_policy,
    load_entity_normalization,
    load_json_config,
    load_outcome_collection,
    load_outcome_collection_review_seeds,
    load_outcome_metrics,
    load_outcome_report_sources,
    load_pipeline_error_policy,
    load_source_schemas,
)
from datahub.builders.school_identity import build_school_identity_package
from datahub.builders.school_profile_merge import build_merged_school_profile_package
from datahub.builders.school_identity_review_audit import audit_school_identity_review_plan
from datahub.builders.school_identity_review_batch import (
    build_school_identity_review_batch,
    merge_school_identity_review_batch,
)
from datahub.builders.school_identity_review_plan import build_school_identity_review_plan
from datahub.builders.school_identity_review_seed_merge import (
    apply_school_identity_review_seeds,
    audit_school_identity_review_seeds,
)
from datahub.builders.school_location_geocode_audit import audit_school_location_geocode_input
from datahub.builders.school_location_from_amap import build_school_location_package_from_amap_geocode
from datahub.builders.school_location_geocode_plan import build_school_location_geocode_input_plan
from datahub.connectors.amap_web_api import fetch_amap_web_api
from datahub.connectors.amap_web_api_readiness import audit_amap_web_api_readiness
from datahub.connectors.manual_files import intake_manual_assets
from datahub.connectors.macos_vision_ocr import ocr_page_images
from datahub.connectors.outcome_report_download import (
    build_outcome_report_manual_intake_queue,
    download_outcome_report_intake_assets,
)
from datahub.connectors.remote_files import download_remote_assets
from datahub.connectors.scs_resources import download_scs_resources
from datahub.connectors.registry import discover_assets
from datahub.connectors.source_candidates import probe_source_candidates
from datahub.parsers.ln_projection_score import parse_ln_projection_score_file
from datahub.parsers.ln_application_workbook import (
    parse_ln_application_workbooks,
    write_application_workbook_outputs,
)
from datahub.parsers.ln_score_distribution_ocr import (
    apply_score_distribution_review,
    build_score_distribution_review_tasks,
    parse_ln_score_distribution_ocr_jsonl,
    prefill_score_distribution_review_suggestions,
    write_candidate_csv,
    write_cleaned_score_distribution_csv,
    write_review_task_csv,
)
from datahub.parsers import ln_score_distribution_grid_images as grid_parser
from datahub.parsers.ln_score_distribution import parse_ln_score_distribution_lines
from datahub.parsers.moe_major_catalog import parse_moe_major_catalog_lines
from datahub.parsers.moe_school_profile import parse_moe_school_profile_rows
from datahub.parsers.scs_position_workbook import (
    parse_scs_position_workbook,
    write_scs_position_csv,
)
from datahub.parsers.digital_occupation_catalog import (
    build_broad_occupation_catalog_seed_rows,
    merge_occupation_catalog_csvs,
    parse_digital_occupation_catalog_html,
    write_digital_occupation_catalog_csv,
)
from datahub.parsers.outcome_report import (
    CANDIDATE_COLUMNS,
    extract_outcome_metric_candidates_from_lines,
    extract_outcome_metric_candidates_from_ofd,
    extract_outcome_metric_candidates_from_report,
    write_outcome_metric_candidate_csv,
)
from datahub.source_audit import audit_sources
from datahub.validators.package_validator import validate_manifest


def test_missing_manifest_fails(tmp_path: Path):
    report = validate_manifest(tmp_path / "manifest.json")
    assert report["errors"]


def test_manifest_requires_fa_prefix(tmp_path: Path):
    path = tmp_path / "manifest.json"
    path.write_text(
        '{"package_id":"p","built_at":"now","tables":[{"name":"bad_table"}],"files":[],"hashes":{},"quality_report":"quality_report.json"}',
        encoding="utf-8",
    )
    report = validate_manifest(path)
    assert any("fa_" in err for err in report["errors"])


def test_manifest_rejects_duplicate_json_keys(tmp_path: Path):
    path = tmp_path / "manifest.json"
    path.write_text(
        '{"package_id":"p1","package_id":"p2","built_at":"now","tables":[],"files":[],"hashes":{},"quality_report":"quality_report.json"}',
        encoding="utf-8",
    )
    report = validate_manifest(path)
    assert any("duplicate JSON key" in err for err in report["errors"])


def test_manifest_rejects_non_object(tmp_path: Path):
    path = tmp_path / "manifest.json"
    path.write_text('["not", "an", "object"]\n', encoding="utf-8")
    report = validate_manifest(path)
    assert report["errors"] == ["manifest must be an object"]


def test_manifest_requires_declared_quality_report_file(tmp_path: Path):
    path = tmp_path / "manifest.json"
    path.write_text(
        '{"package_id":"p","built_at":"now","tables":[],"files":[],"hashes":{},"quality_report":"missing_quality_report.json"}',
        encoding="utf-8",
    )
    report = validate_manifest(path)
    assert any("declared quality report not found" in err for err in report["errors"])


def test_manifest_rejects_quality_report_errors(tmp_path: Path):
    path = tmp_path / "manifest.json"
    (tmp_path / "quality_report.json").write_text('{"errors":["manual review is not complete"]}\n', encoding="utf-8")
    path.write_text(
        '{"package_id":"p","built_at":"now","tables":[],"files":[],"hashes":{},"quality_report":"quality_report.json"}',
        encoding="utf-8",
    )
    report = validate_manifest(path)
    assert any("quality_report has errors" in err for err in report["errors"])


def test_manifest_rejects_malformed_quality_report_errors(tmp_path: Path):
    path = tmp_path / "manifest.json"
    (tmp_path / "quality_report.json").write_text('{"errors":{}}\n', encoding="utf-8")
    path.write_text(
        '{"package_id":"p","built_at":"now","tables":[],"files":[],"hashes":{},"quality_report":"quality_report.json"}',
        encoding="utf-8",
    )
    report = validate_manifest(path)
    assert any("quality_report.errors must be a list" in err for err in report["errors"])


def test_manifest_rejects_pending_quality_readiness(tmp_path: Path):
    path = tmp_path / "manifest.json"
    (tmp_path / "quality_report.json").write_text(json.dumps({
        "errors": [],
        "readiness": {
            "progress": {
                "pending_rows": 1,
                "blocked_rows": 0,
                "blocking_decision_rows": 0,
                "unknown_status_rows": 0,
            }
        },
    }), encoding="utf-8")
    path.write_text(
        '{"package_id":"p","built_at":"now","tables":[],"files":[],"hashes":{},"quality_report":"quality_report.json"}',
        encoding="utf-8",
    )
    report = validate_manifest(path)
    assert any("quality_report readiness is not clear" in err for err in report["errors"])


def test_manifest_rejects_reviewed_reconciliation_quality_without_readiness(tmp_path: Path):
    path = tmp_path / "manifest.json"
    (tmp_path / "quality_report.json").write_text(
        '{"errors":[],"warnings":[],"decision_counts":{"use_package_row":1}}\n',
        encoding="utf-8",
    )
    path.write_text(json.dumps({
        "package_id": "p",
        "built_at": "now",
        "tables": [],
        "files": [],
        "hashes": {},
        "quality_report": "quality_report.json",
        "source_lineage": {"source_kind": "reviewed_reconciliation_plan"},
    }), encoding="utf-8")
    report = validate_manifest(path)
    assert any("quality_report missing readiness" in err for err in report["errors"])


def test_manifest_rejects_reviewed_reconciliation_malformed_decision_counts(tmp_path: Path):
    path = tmp_path / "manifest.json"
    (tmp_path / "quality_report.json").write_text(json.dumps({
        "errors": [],
        "warnings": [],
        "decision_counts": [],
        "readiness": {
            "progress": {
                "pending_rows": 0,
                "blocked_rows": 0,
                "blocking_decision_rows": 0,
                "unknown_status_rows": 0,
            }
        },
    }), encoding="utf-8")
    path.write_text(json.dumps({
        "package_id": "p",
        "built_at": "now",
        "tables": [],
        "files": [],
        "hashes": {},
        "quality_report": "quality_report.json",
        "source_lineage": {"source_kind": "reviewed_reconciliation_plan"},
    }), encoding="utf-8")
    report = validate_manifest(path)
    assert any("quality_report.decision_counts must be an object" in err for err in report["errors"])


def test_manifest_rejects_declared_hash_mismatch(tmp_path: Path):
    path = tmp_path / "manifest.json"
    (tmp_path / "quality_report.json").write_text('{"errors":[]}\n', encoding="utf-8")
    (tmp_path / "fa_test.csv").write_text("id\n1\n", encoding="utf-8")
    path.write_text(
        (
            '{"package_id":"p","built_at":"now","tables":[{"name":"fa_test","file":"fa_test.csv"}],'
            '"files":["fa_test.csv"],"hashes":{"fa_test.csv":"bad-hash"},'
            '"quality_report":"quality_report.json"}'
        ),
        encoding="utf-8",
    )
    report = validate_manifest(path)
    assert any("hash mismatch" in err for err in report["errors"])


def test_manifest_rejects_malformed_collection_fields(tmp_path: Path):
    path = tmp_path / "manifest.json"
    (tmp_path / "quality_report.json").write_text('{"errors":[]}\n', encoding="utf-8")
    path.write_text(
        (
            '{"package_id":"p","built_at":"now","tables":{},'
            '"files":"fa_test.csv","hashes":[],'
            '"quality_report":"quality_report.json"}'
        ),
        encoding="utf-8",
    )
    report = validate_manifest(path)
    assert any("manifest tables must be a list" in err for err in report["errors"])
    assert any("manifest files must be a list" in err for err in report["errors"])
    assert any("manifest hashes must be an object" in err for err in report["errors"])


def test_manifest_rejects_invalid_table_entries(tmp_path: Path):
    path = tmp_path / "manifest.json"
    (tmp_path / "quality_report.json").write_text('{"errors":[]}\n', encoding="utf-8")
    (tmp_path / "fa_test.csv").write_text("id\n1\n", encoding="utf-8")
    path.write_text(
        (
            '{"package_id":"p","built_at":"now",'
            '"tables":["fa_not_an_object",{"name":"fa_test","file":"missing.csv"},{"name":"bad-table","file":"fa_test.csv"}],'
            '"files":["fa_test.csv"],"hashes":{},'
            '"quality_report":"quality_report.json"}'
        ),
        encoding="utf-8",
    )
    report = validate_manifest(path)
    assert any("invalid manifest table entry" in err for err in report["errors"])
    assert any("file is not listed in manifest.files" in err for err in report["errors"])
    assert any("invalid table name" in err for err in report["errors"])


def test_build_release_bundle_summarizes_core_handoff_fields(tmp_path: Path):
    package_dir = tmp_path / "pkg-release"
    package_dir.mkdir()
    table_path = package_dir / "fa_test.csv"
    table_path.write_text("id\n1\n", encoding="utf-8")
    quality = {
        "row_counts": {"fa_test": 1},
        "warnings": [],
        "errors": [],
        "decision_counts": {"use_package_row": 1},
        "readiness": {
            "progress": {
                "pending_rows": 0,
                "blocked_rows": 0,
                "blocking_decision_rows": 0,
                "unknown_status_rows": 0,
            }
        },
    }
    (package_dir / "quality_report.json").write_text(json.dumps(quality), encoding="utf-8")
    manifest = {
        "package_id": "pkg-release",
        "built_at": "now",
        "tables": [{"name": "fa_test", "file": "fa_test.csv"}],
        "files": ["fa_test.csv"],
        "hashes": {"fa_test.csv": hashlib.sha256(table_path.read_bytes()).hexdigest()},
        "quality_report": "quality_report.json",
        "source_lineage": {"source_kind": "reviewed_reconciliation_plan", "plan_csv": "review.csv"},
    }
    manifest_path = package_dir / "manifest.json"
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")
    dry_run_report = tmp_path / "pkg-release-dry-run.json"
    dry_run_report.write_text('{"status":"passed","dry_run":true,"errors":[]}\n', encoding="utf-8")

    result = build_release_bundle(
        package_dirs=[package_dir],
        output=tmp_path / "release_bundle.json",
        bundle_id="release-fixture",
        load_modes={"pkg-release": "upsert_or_replace_package"},
        dry_run_reports={"pkg-release": dry_run_report},
    )

    bundle = json.loads((tmp_path / "release_bundle.json").read_text(encoding="utf-8"))
    package = bundle["packages"][0]
    assert result["ready_for_core_import"] is True
    assert bundle["bundle_id"] == "release-fixture"
    assert package["package_id"] == "pkg-release"
    assert package["target_tables"] == [
        {"name": "fa_test", "file": "fa_test.csv", "load_mode": "upsert_or_replace_package"}
    ]
    assert package["manifest"]["path"].endswith("manifest.json")
    assert package["manifest"]["sha256"] == hashlib.sha256(manifest_path.read_bytes()).hexdigest()
    assert package["quality_report"]["error_count"] == 0
    assert package["source_lineage"]["source_kind"] == "reviewed_reconciliation_plan"
    assert package["readiness"]["status"] == "passed"
    assert package["review_reconciliation"]["status"] == "passed"
    assert package["core_importer_dry_run"]["status"] == "passed"


def test_build_release_bundle_blocks_missing_formal_handoff_evidence(tmp_path: Path):
    package_dir = tmp_path / "pkg-missing-evidence"
    package_dir.mkdir()
    table_path = package_dir / "fa_test.csv"
    table_path.write_text("id\n1\n", encoding="utf-8")
    (package_dir / "quality_report.json").write_text('{"errors":[],"warnings":[]}\n', encoding="utf-8")
    (package_dir / "manifest.json").write_text(json.dumps({
        "package_id": "pkg-missing-evidence",
        "built_at": "now",
        "tables": [{"name": "fa_test", "file": "fa_test.csv"}],
        "files": ["fa_test.csv"],
        "hashes": {"fa_test.csv": hashlib.sha256(table_path.read_bytes()).hexdigest()},
        "quality_report": "quality_report.json",
    }), encoding="utf-8")

    result = build_release_bundle(
        package_dirs=[package_dir],
        output=tmp_path / "release_bundle.json",
    )

    blocker_codes = {blocker["code"] for blocker in result["blockers"]}
    assert result["ready_for_core_import"] is False
    assert "load_mode_missing" in blocker_codes
    assert "readiness_not_passed" in blocker_codes
    assert "review_reconciliation_not_passed" in blocker_codes
    assert "core_importer_dry_run_not_passed" in blocker_codes


def test_build_release_bundle_rejects_manual_pass_statuses_in_formal_mode(tmp_path: Path):
    package_dir = tmp_path / "pkg-manual-formal"
    package_dir.mkdir()
    table_path = package_dir / "fa_test.csv"
    table_path.write_text("id\n1\n", encoding="utf-8")
    (package_dir / "quality_report.json").write_text('{"errors":[],"warnings":[]}\n', encoding="utf-8")
    (package_dir / "manifest.json").write_text(json.dumps({
        "package_id": "pkg-manual-formal",
        "built_at": "now",
        "tables": [{"name": "fa_test", "file": "fa_test.csv"}],
        "files": ["fa_test.csv"],
        "hashes": {"fa_test.csv": hashlib.sha256(table_path.read_bytes()).hexdigest()},
        "quality_report": "quality_report.json",
    }), encoding="utf-8")

    result = build_release_bundle(
        package_dirs=[package_dir],
        output=tmp_path / "release_bundle.json",
        load_modes={"pkg-manual-formal": "upsert_or_replace_package"},
        readiness_statuses={"pkg-manual-formal": "passed"},
        review_statuses={"pkg-manual-formal": "passed"},
        dry_run_statuses={"pkg-manual-formal": "passed"},
    )

    bundle = json.loads((tmp_path / "release_bundle.json").read_text(encoding="utf-8"))
    blocker_codes = [blocker["code"] for blocker in result["blockers"]]
    package = bundle["packages"][0]
    assert result["ready_for_core_import"] is False
    assert bundle["release_mode"] == "formal"
    assert bundle["manual_status_policy"]["manual_pass_status_allowed"] is False
    assert blocker_codes.count("manual_pass_status_not_allowed") == 3
    assert package["readiness"]["formal_core_import_evidence"] is False
    assert package["review_reconciliation"]["formal_core_import_evidence"] is False
    assert package["core_importer_dry_run"]["formal_core_import_evidence"] is False


def test_build_release_bundle_allows_manual_pass_statuses_only_for_smoke_bundle(tmp_path: Path):
    package_dir = tmp_path / "pkg-manual-smoke"
    package_dir.mkdir()
    table_path = package_dir / "fa_test.csv"
    table_path.write_text("id\n1\n", encoding="utf-8")
    (package_dir / "quality_report.json").write_text('{"errors":[],"warnings":[]}\n', encoding="utf-8")
    (package_dir / "manifest.json").write_text(json.dumps({
        "package_id": "pkg-manual-smoke",
        "built_at": "now",
        "tables": [{"name": "fa_test", "file": "fa_test.csv"}],
        "files": ["fa_test.csv"],
        "hashes": {"fa_test.csv": hashlib.sha256(table_path.read_bytes()).hexdigest()},
        "quality_report": "quality_report.json",
    }), encoding="utf-8")

    result = build_release_bundle(
        package_dirs=[package_dir],
        output=tmp_path / "release_bundle.json",
        load_modes={"pkg-manual-smoke": "upsert_or_replace_package"},
        readiness_statuses={"pkg-manual-smoke": "passed"},
        review_statuses={"pkg-manual-smoke": "passed"},
        dry_run_statuses={"pkg-manual-smoke": "passed"},
        release_mode="smoke",
    )

    bundle = json.loads((tmp_path / "release_bundle.json").read_text(encoding="utf-8"))
    blocker_codes = [blocker["code"] for blocker in result["blockers"]]
    package = bundle["packages"][0]
    assert result["ready_for_core_import"] is False
    assert result["formal_core_import_allowed"] is False
    assert bundle["release_mode"] == "smoke"
    assert bundle["manual_status_policy"]["manual_pass_status_allowed"] is True
    assert "manual_pass_status_not_allowed" not in blocker_codes
    assert "non_formal_release_mode" in blocker_codes
    assert package["readiness"]["status"] == "passed"
    assert package["review_reconciliation"]["status"] == "passed"
    assert package["core_importer_dry_run"]["status"] == "passed"


def test_config_json_files_do_not_have_duplicate_keys(tmp_path: Path):
    duplicate_paths = []
    for path in sorted(Path("config").glob("*.json")):
        duplicates = []

        def reject_duplicate_keys(pairs):
            seen = set()
            result = {}
            for key, value in pairs:
                if key in seen:
                    duplicates.append(key)
                seen.add(key)
                result[key] = value
            return result

        json.loads(path.read_text(encoding="utf-8"), object_pairs_hook=reject_duplicate_keys)
        if duplicates:
            duplicate_paths.append(f"{path}: {sorted(set(duplicates))}")

    assert duplicate_paths == []

    duplicate_json = tmp_path / "duplicate.json"
    duplicate_json.write_text('{"version":"1","version":"2"}\n', encoding="utf-8")
    with pytest.raises(ValueError, match="duplicate JSON key"):
        load_json_config(duplicate_json)


def test_build_local_package_from_cleaned_csv(tmp_path: Path):
    source = tmp_path / "cleaned.csv"
    with source.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["院校代码", "院校名称", "专业代码", "专业名称", "批次", "科类", "年份", "计划数"])
        writer.writeheader()
        writer.writerow({
            "院校代码": "0142",
            "院校名称": "沈阳工业大学",
            "专业代码": "F2",
            "专业名称": "土木工程",
            "批次": "本科批",
            "科类": "物理类",
            "年份": "2026",
            "计划数": "12",
        })

    result = build_local_package(
        source_key="ln_admission_plan",
        table_name="fa_dim_ln_admission_plan",
        input_path=source,
        output_root=tmp_path / "exports",
        package_id="pkg-local-test",
        source_version="fixture",
    )
    package_dir = Path(result["package_dir"])
    manifest_report = validate_manifest(package_dir / "manifest.json")
    assert manifest_report["errors"] == []
    assert (package_dir / "fa_dim_ln_admission_plan.csv").exists()

    quality = json.loads((package_dir / "quality_report.json").read_text(encoding="utf-8"))
    assert quality["errors"] == []
    assert quality["year_coverage"] == [2026]
    assert result["rows"] == 1


def test_build_local_package_includes_intake_lineage(tmp_path: Path):
    source = tmp_path / "cleaned.csv"
    with source.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["院校代码", "院校名称", "专业代码", "专业名称", "批次", "科类", "年份"])
        writer.writeheader()
        writer.writerow({
            "院校代码": "0140",
            "院校名称": "辽宁大学",
            "专业代码": "01",
            "专业名称": "法学",
            "批次": "本科批",
            "科类": "历史类",
            "年份": "2026",
        })
    intake_manifest = tmp_path / "_intake_manifest.json"
    intake_manifest.write_text(json.dumps({
        "source_key": "ln_admission_plan",
        "source_name": "辽宁招生计划",
        "source_kind": "controlled_manual_export",
        "source_date": "2026-06-20",
        "intake_at": "2026-06-21T00:00:00",
        "acquired_by": "fixture",
        "official_distribution": "网报志愿系统",
        "evidence_urls": ["https://example.edu/evidence"],
        "target_tables": ["fa_dim_ln_admission_plan"],
        "files": [
            {
                "file_name": "raw_plan.xlsx",
                "path": "/tmp/raw_plan.xlsx",
                "size_bytes": 128,
                "sha256": "abc123",
            }
        ],
    }, ensure_ascii=False), encoding="utf-8")

    result = build_local_package(
        source_key="ln_admission_plan",
        table_name="fa_dim_ln_admission_plan",
        input_path=source,
        output_root=tmp_path / "exports",
        package_id="pkg-local-lineage-test",
        source_version="fixture-lineage",
        intake_manifest=intake_manifest,
    )
    package_dir = Path(result["package_dir"])
    manifest = json.loads((package_dir / "manifest.json").read_text(encoding="utf-8"))
    assert manifest["source_lineage"]["acquired_by"] == "fixture"
    assert manifest["source_lineage"]["files"][0]["sha256"] == "abc123"
    assert manifest["source_lineage"]["evidence_urls"] == ["https://example.edu/evidence"]
    assert result["source_lineage"]["source_date"] == "2026-06-20"
    assert result["source_lineage"]["target_source_key"] == "ln_admission_plan"

    duplicate_key_manifest = tmp_path / "_duplicate_intake_manifest.json"
    duplicate_key_manifest.write_text(
        (
            '{"source_key":"ln_admission_plan",'
            '"source_key":"shadow_source",'
            '"target_tables":["fa_dim_ln_admission_plan"],'
            '"files":[{"file_name":"raw_plan.xlsx","sha256":"abc123"}]}\n'
        ),
        encoding="utf-8",
    )
    try:
        build_local_package(
            source_key="ln_admission_plan",
            table_name="fa_dim_ln_admission_plan",
            input_path=source,
            output_root=tmp_path / "exports",
            package_id="pkg-duplicate-intake-lineage-test",
            source_version="fixture-lineage",
            intake_manifest=duplicate_key_manifest,
        )
        duplicate_manifest_rejected = False
    except ValueError as exc:
        duplicate_manifest_rejected = "duplicate JSON key" in str(exc)
    assert duplicate_manifest_rejected

    malformed_files_manifest = tmp_path / "_malformed_files_intake_manifest.json"
    malformed_files_manifest.write_text(json.dumps({
        "source_key": "ln_admission_plan",
        "source_name": "辽宁招生计划",
        "source_kind": "controlled_manual_export",
        "source_date": "2026-06-20",
        "intake_at": "2026-06-21T00:00:00",
        "acquired_by": "fixture",
        "official_distribution": "网报志愿系统",
        "evidence_urls": ["https://example.edu/evidence"],
        "target_tables": ["fa_dim_ln_admission_plan"],
        "files": {"file_name": "raw_plan.xlsx", "sha256": "abc123"},
    }, ensure_ascii=False), encoding="utf-8")
    try:
        build_local_package(
            source_key="ln_admission_plan",
            table_name="fa_dim_ln_admission_plan",
            input_path=source,
            output_root=tmp_path / "exports",
            package_id="pkg-malformed-intake-lineage-test",
            source_version="fixture-lineage",
            intake_manifest=malformed_files_manifest,
        )
        malformed_files_rejected = False
    except ValueError as exc:
        malformed_files_rejected = "intake manifest files must be a list" in str(exc)
    assert malformed_files_rejected


def test_build_local_package_accepts_configured_shared_intake_source(tmp_path: Path):
    source = tmp_path / "score_history.csv"
    with source.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["school_code", "major_code", "batch", "subject_cat", "score_year", "min_score", "min_rank"],
        )
        writer.writeheader()
        writer.writerow({
            "school_code": "0140",
            "major_code": "01",
            "batch": "本科批",
            "subject_cat": "历史类",
            "score_year": "2025",
            "min_score": "620",
            "min_rank": "1200",
        })
    intake_manifest = tmp_path / "_intake_manifest.json"
    intake_manifest.write_text(json.dumps({
        "source_key": "ln_application_workbook",
        "source_name": "辽宁本地报考工作簿",
        "source_kind": "controlled_manual_cleaned_workbook",
        "source_date": "2025-08-27",
        "intake_at": "2026-05-13T00:00:00",
        "acquired_by": "fixture",
        "official_distribution": "本地报考工作簿",
        "evidence_urls": ["https://example.edu/evidence"],
        "target_tables": ["fa_dim_ln_admission_plan", "fa_fact_ln_score_history"],
        "files": [
            {
                "file_name": "application_workbook.xlsx",
                "path": "/tmp/application_workbook.xlsx",
                "size_bytes": 128,
                "sha256": "shared-source-sha256",
            }
        ],
    }, ensure_ascii=False), encoding="utf-8")

    result = build_local_package(
        source_key="ln_score_history",
        table_name="fa_fact_ln_score_history",
        input_path=source,
        output_root=tmp_path / "exports",
        package_id="pkg-shared-intake-test",
        source_version="fixture-shared-intake",
        intake_manifest=intake_manifest,
    )

    package_dir = Path(result["package_dir"])
    manifest = json.loads((package_dir / "manifest.json").read_text(encoding="utf-8"))
    assert manifest["source_lineage"]["source_key"] == "ln_application_workbook"
    assert manifest["source_lineage"]["target_source_key"] == "ln_score_history"
    assert manifest["source_lineage"]["files"][0]["sha256"] == "shared-source-sha256"
    assert validate_manifest(package_dir / "manifest.json")["errors"] == []


def test_build_local_score_distribution_package_from_transcript_with_image_lineage(tmp_path: Path):
    source = tmp_path / "score_distribution.csv"
    with source.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["科类", "年份", "分数", "人数", "累计", "来源日期"],
        )
        writer.writeheader()
        writer.writerow({"科类": "物理类", "年份": "2024", "分数": "676", "人数": "12", "累计": "12", "来源日期": "2024-06-25"})
        writer.writerow({"科类": "物理类", "年份": "2024", "分数": "675", "人数": "2", "累计": "14", "来源日期": "2024-06-25"})
        writer.writerow({"科类": "物理类", "年份": "2024", "分数": "672", "人数": "5", "累计": "19", "来源日期": "2024-06-25"})
    intake_manifest = tmp_path / "_page_images_manifest.json"
    intake_manifest.write_text(json.dumps({
        "source_key": "ln_score_distribution",
        "source_name": "辽宁一分一段表图片",
        "source_kind": "official_page_images",
        "source_date": "2024-06-25",
        "intake_at": "2026-05-13T00:00:00",
        "acquired_by": "datahub.download_page_images",
        "official_distribution": "辽宁省教育厅官网图片",
        "evidence_urls": ["https://jyt.ln.gov.cn/example/index.shtml"],
        "target_tables": ["fa_fact_ln_score_distribution"],
        "files": [
            {
                "file_name": "ln_score_distribution_2024_001.png",
                "path": "/tmp/ln_score_distribution_2024_001.png",
                "size_bytes": 1024,
                "sha256": "image-sha256-fixture",
            }
        ],
    }, ensure_ascii=False), encoding="utf-8")

    result = build_local_package(
        source_key="ln_score_distribution",
        table_name="fa_fact_ln_score_distribution",
        input_path=source,
        output_root=tmp_path / "exports",
        package_id="pkg-score-distribution-transcript-test",
        source_version="fixture-score-distribution-transcript",
        intake_manifest=intake_manifest,
    )
    package_dir = Path(result["package_dir"])
    assert validate_manifest(package_dir / "manifest.json")["errors"] == []
    assert result["quality_report"]["errors"] == []
    assert any("has few score rows" in warning for warning in result["quality_report"]["warnings"])

    manifest = json.loads((package_dir / "manifest.json").read_text(encoding="utf-8"))
    assert manifest["source_lineage"]["source_kind"] == "official_page_images"
    assert manifest["source_lineage"]["files"][0]["sha256"] == "image-sha256-fixture"

    with (package_dir / "fa_fact_ln_score_distribution.csv").open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    assert rows[0]["subject_cat"] == "物理类"
    assert rows[0]["cumulative_rank"] == "12"


def test_build_local_score_distribution_package_rejects_cumulative_mismatch(tmp_path: Path):
    source = tmp_path / "score_distribution_bad.csv"
    with source.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["subject_cat", "score_year", "score", "score_count", "cumulative_rank", "source_date"],
        )
        writer.writeheader()
        writer.writerow({
            "subject_cat": "历史类",
            "score_year": "2023",
            "score": "665",
            "score_count": "3",
            "cumulative_rank": "3",
            "source_date": "2023-06-24",
        })
        writer.writerow({
            "subject_cat": "历史类",
            "score_year": "2023",
            "score": "664",
            "score_count": "2",
            "cumulative_rank": "6",
            "source_date": "2023-06-24",
        })

    try:
        build_local_package(
            source_key="ln_score_distribution",
            table_name="fa_fact_ln_score_distribution",
            input_path=source,
            output_root=tmp_path / "exports",
            package_id="pkg-score-distribution-bad-test",
            source_version="fixture-score-distribution-bad",
        )
        rejected = False
    except ValueError as exc:
        rejected = "cumulative mismatch" in str(exc)
    assert rejected


def test_discover_assets_from_source_config(tmp_path: Path):
    raw_dir = tmp_path / "raw" / "ln_admission_plan" / "2026"
    raw_dir.mkdir(parents=True)
    source = raw_dir / "2026-05-12_cleaned.xlsx"
    source.write_bytes(b"placeholder")

    assets = discover_assets("ln_admission_plan", project_root=tmp_path)
    assert len(assets) == 1
    assert assets[0].source_key == "ln_admission_plan"
    assert assets[0].source_date == "2026-05-12"
    assert assets[0].path == source


def test_intake_manual_assets_records_provenance(tmp_path: Path):
    source = tmp_path / "2026_plan.xlsx"
    source.write_bytes(b"manual excel placeholder")
    digest = hashlib.sha256(source.read_bytes()).hexdigest()

    result = intake_manual_assets(
        "ln_admission_plan",
        [source],
        tmp_path / "raw",
        source_date="2026-06-20",
        acquired_by="fixture",
        official_distribution="网报志愿系统",
        evidence_urls=["https://example.edu/evidence"],
        notes="fixture intake",
    )

    target = tmp_path / "raw" / "ln_admission_plan" / "2026-06-20" / "2026_plan.xlsx"
    manifest = target.parent / "_intake_manifest.json"
    assert target.exists()
    assert result["file_count"] == 1
    assert result["files"][0]["sha256"] == digest

    manifest_data = json.loads(manifest.read_text(encoding="utf-8"))
    assert manifest_data["source_key"] == "ln_admission_plan"
    assert manifest_data["acquired_by"] == "fixture"
    assert manifest_data["official_distribution"] == "网报志愿系统"
    assert manifest_data["files"][0]["sha256"] == digest

    assets = discover_assets("ln_admission_plan", project_root=tmp_path)
    assert any(asset.path == target for asset in assets)


def test_intake_manual_assets_rejects_remote_configured_source(tmp_path: Path):
    source = tmp_path / "catalog.pdf"
    source.write_bytes(b"placeholder")
    try:
        intake_manual_assets(
            "moe_major_catalog",
            [source],
            tmp_path / "raw",
            source_date="2025-04-22",
            acquired_by="fixture",
        )
        rejected = False
    except ValueError as exc:
        rejected = "not configured for manual intake" in str(exc)
    assert rejected


def test_probe_source_candidates_reports_accessible_file(tmp_path: Path, monkeypatch):
    source = tmp_path / "candidate.html"
    source.write_text("fixture candidate", encoding="utf-8")

    def fake_sources():
        return {
            "sources": {
                "fixture_source": {
                    "name": "Fixture Source",
                    "research_candidates": [
                        {
                            "label": "local fixture",
                            "kind": "fixture_page",
                            "url": source.resolve().as_uri(),
                            "source_date": "2026-05-13",
                            "expected_table": "fa_fact_fixture",
                        }
                    ],
                }
            }
        }

    monkeypatch.setattr("datahub.connectors.source_candidates.load_sources", fake_sources)

    output = tmp_path / "probe.json"
    report = probe_source_candidates("fixture_source", output=output)
    assert report["candidate_count"] == 1
    assert report["accessible_count"] == 1
    assert report["candidates"][0]["probe_status"] == "accessible"
    assert report["candidates"][0]["size_bytes"] == len("fixture candidate")
    assert report["candidates"][0]["sha256"] == hashlib.sha256(b"fixture candidate").hexdigest()
    assert output.exists()


def test_probe_source_candidates_detects_configured_antibot_marker(tmp_path: Path, monkeypatch):
    source = tmp_path / "challenge.html"
    body = "<html><script>$_ts={};</script></html>"
    source.write_text(body, encoding="utf-8")

    def fake_sources():
        return {
            "sources": {
                "fixture_source": {
                    "name": "Fixture Source",
                    "probe": {"blocked_content_markers": ["$_ts"]},
                    "research_candidates": [
                        {
                            "label": "local challenge",
                            "kind": "fixture_page",
                            "url": source.resolve().as_uri(),
                            "source_date": "2026-05-13",
                            "expected_table": "fa_fact_fixture",
                        }
                    ],
                }
            }
        }

    monkeypatch.setattr("datahub.connectors.source_candidates.load_sources", fake_sources)

    report = probe_source_candidates("fixture_source")
    assert report["candidate_count"] == 1
    assert report["accessible_count"] == 0
    assert report["blocked_by_antibot_count"] == 1
    assert report["candidates"][0]["probe_status"] == "blocked_by_antibot"
    assert report["candidates"][0]["blocked_marker"] == "$_ts"
    assert report["candidates"][0]["sha256"] == hashlib.sha256(body.encode("utf-8")).hexdigest()


def test_probe_source_candidates_detects_configured_antibot_http_status(monkeypatch):
    def fake_sources():
        return {
            "sources": {
                "fixture_source": {
                    "name": "Fixture Source",
                    "probe": {"blocked_http_statuses": [412]},
                    "research_candidates": [
                        {
                            "label": "blocked candidate",
                            "kind": "fixture_page",
                            "url": "https://example.test/challenge",
                            "source_date": "2026-05-13",
                            "expected_table": "fa_fact_fixture",
                        }
                    ],
                }
            }
        }

    def fake_urlopen(request, timeout=60):
        headers = Message()
        headers["Content-Type"] = "text/html"
        raise HTTPError(request.full_url, 412, "Precondition Failed", headers, None)

    monkeypatch.setattr("datahub.connectors.source_candidates.load_sources", fake_sources)
    monkeypatch.setattr("datahub.connectors.source_candidates.urlopen", fake_urlopen)

    report = probe_source_candidates("fixture_source")
    assert report["candidate_count"] == 1
    assert report["accessible_count"] == 0
    assert report["inaccessible_count"] == 0
    assert report["blocked_by_antibot_count"] == 1
    assert report["candidates"][0]["probe_status"] == "blocked_by_antibot"
    assert report["candidates"][0]["http_status"] == 412
    assert report["candidates"][0]["blocked_http_status"] == 412


def test_audit_sources_marks_admission_plan_manual():
    report = audit_sources()
    by_key = {item["source_key"]: item for item in report["sources"]}
    assert by_key["ln_admission_plan"]["status"] == "manual_required"
    assert by_key["ln_application_workbook"]["status"] == "manual_required"
    assert by_key["ln_application_workbook"]["target_tables"] == [
        "fa_dim_ln_admission_plan",
        "fa_fact_ln_score_history",
    ]
    assert by_key["ln_admission_plan"]["official_distribution"]
    assert by_key["ln_projection_score"]["status"] == "remote_configured"
    assert by_key["ln_projection_score"]["research_candidate_count"] >= 4
    assert by_key["ln_score_history"]["status"] == "partial_official_derivation_configured"
    assert by_key["ln_score_distribution"]["status"] == "remote_configured"
    assert by_key["ln_score_distribution"]["page_image_source_count"] == 5
    assert by_key["ln_score_distribution"]["research_candidate_count"] >= 3
    assert by_key["ln_score_distribution"]["ocr_engine"] == "macos_vision"
    assert by_key["major_mapping_review"]["status"] == "local_db_configured"
    assert by_key["school_profile"]["status"] == "remote_configured"
    assert by_key["school_profile_supplemental"]["status"] == "manual_review_configured"
    assert by_key["school_profile_supplemental"]["target_tables"] == ["fa_dim_school_profile"]
    assert by_key["school_location_geocode"]["status"] == "web_api_configured_requires_connector"
    assert by_key["school_location_geocode"]["target_tables"] == ["fa_dim_school_location"]
    assert by_key["region_profile_geocode"]["status"] == "web_api_configured_requires_connector"
    assert by_key["region_profile_geocode"]["target_tables"] == ["fa_dim_region_profile"]
    assert by_key["entity_normalization_registry"]["status"] == "derived_from_datahub_signals"
    assert "fa_dim_entity_registry" in by_key["entity_normalization_registry"]["target_tables"]
    assert by_key["data_update_governance"]["status"] == "derived_from_datahub_signals"
    assert "fa_meta_source_snapshot" in by_key["data_update_governance"]["target_tables"]
    assert by_key["campus_surrounding_poi"]["status"] == "web_api_configured_requires_connector"
    assert by_key["campus_surrounding_poi"]["target_tables"] == ["fa_fact_campus_surrounding_poi"]
    assert by_key["campus_housing_market"]["status"] == "source_collection_required"
    assert by_key["campus_housing_market"]["target_tables"] == ["fa_fact_campus_housing_market"]
    assert by_key["region_living_cost"]["status"] == "source_collection_required"
    assert by_key["region_living_cost"]["target_tables"] == ["fa_fact_region_living_cost"]
    assert by_key["campus_living_score"]["status"] == "derived_from_datahub_signals"
    assert by_key["campus_living_score"]["target_tables"] == ["fa_mart_campus_living_score"]
    assert by_key["school_identity_bridge"]["status"] == "local_db_configured"
    assert by_key["school_outcome"]["target_tables"] == ["fa_fact_school_outcome"]
    assert by_key["major_outcome"]["status"] == "source_collection_required"
    assert by_key["school_recruitment_event"]["status"] == "source_collection_required"
    assert by_key["school_recruitment_event"]["target_tables"] == ["fa_fact_school_recruitment_event"]
    assert by_key["school_research_industry_link"]["status"] == "source_collection_required"
    assert by_key["school_research_industry_link"]["target_tables"] == ["fa_fact_school_research_industry_link"]
    assert by_key["school_local_employment"]["status"] == "source_collection_required"
    assert by_key["school_local_employment"]["target_tables"] == ["fa_fact_school_local_employment"]
    assert by_key["city_industry_zone"]["status"] == "source_collection_required"
    assert by_key["city_industry_zone"]["target_tables"] == ["fa_dim_city_industry_zone"]
    assert by_key["school_city_industry_fit"]["status"] == "derived_from_datahub_signals"
    assert by_key["school_city_industry_fit"]["target_tables"] == ["fa_mart_school_city_industry_fit"]
    assert by_key["city_economic_indicator"]["status"] == "source_collection_required"
    assert by_key["city_economic_indicator"]["target_tables"] == ["fa_fact_city_economic_indicator"]
    assert by_key["city_public_resource"]["status"] == "source_collection_required"
    assert by_key["city_public_resource"]["target_tables"] == ["fa_fact_city_public_resource"]
    assert by_key["city_listed_company_signal"]["status"] == "source_collection_required"
    assert by_key["city_listed_company_signal"]["target_tables"] == ["fa_fact_city_listed_company_signal"]
    assert by_key["city_ranking_signal"]["status"] == "source_collection_required"
    assert by_key["city_ranking_signal"]["target_tables"] == ["fa_fact_city_ranking_signal"]
    assert by_key["city_development_score"]["status"] == "derived_from_datahub_signals"
    assert by_key["city_development_score"]["target_tables"] == ["fa_mart_city_development_score"]
    assert by_key["major_employment_role_map"]["status"] == "curation_required"
    assert by_key["major_employment_role_map"]["target_tables"] == ["fa_bridge_major_employment_role"]
    assert by_key["company_role_demand_signal"]["status"] == "source_collection_required"
    assert by_key["company_role_demand_signal"]["target_tables"] == ["fa_fact_company_role_demand_signal"]
    assert by_key["major_city_employment_fit"]["status"] == "derived_from_datahub_signals"
    assert by_key["major_city_employment_fit"]["target_tables"] == ["fa_mart_major_city_employment_fit"]
    assert by_key["career_occupation_catalog"]["target_tables"] == ["fa_dim_career_occupation"]
    assert by_key["career_civil_service_posts"]["status"] == "configured_web_api_raw_intake"
    assert by_key["career_civil_service_posts"]["target_tables"] == [
        "fa_fact_civil_service_position",
        "fa_fact_career_signal",
    ]
    assert by_key["career_signal"]["status"] == "source_collection_required"
    assert by_key["career_score"]["status"] == "derived_from_datahub_signals"
    assert by_key["policy_industry_map"]["status"] == "curated_seed_configured"
    assert by_key["policy_plan_history"]["status"] == "curated_seed_configured"


def test_evidence_domain_schemas_are_package_ready():
    schemas = load_source_schemas()["tables"]
    for table_name in [
        "fa_dim_school_profile",
        "fa_dim_school_location",
        "fa_dim_region_profile",
        "fa_dim_entity_registry",
        "fa_bridge_entity_alias",
        "fa_dim_metric_registry",
        "fa_bridge_metric_alias",
        "fa_meta_source_snapshot",
        "fa_meta_source_health",
        "fa_meta_update_run",
        "fa_meta_update_run_step",
        "fa_meta_nonstandard_review_queue",
        "fa_fact_civil_service_position",
        "fa_fact_campus_surrounding_poi",
        "fa_fact_campus_housing_market",
        "fa_fact_region_living_cost",
        "fa_mart_campus_living_score",
        "fa_bridge_school_identity",
        "fa_fact_school_outcome",
        "fa_fact_major_outcome",
        "fa_fact_school_recruitment_event",
        "fa_fact_school_research_industry_link",
        "fa_fact_school_local_employment",
        "fa_dim_city_industry_zone",
        "fa_mart_school_city_industry_fit",
        "fa_fact_city_economic_indicator",
        "fa_fact_city_public_resource",
        "fa_fact_city_listed_company_signal",
        "fa_fact_city_ranking_signal",
        "fa_mart_city_development_score",
        "fa_bridge_major_employment_role",
        "fa_fact_company_role_demand_signal",
        "fa_mart_major_city_employment_fit",
        "fa_dim_career_occupation",
        "fa_fact_career_signal",
        "fa_mart_career_score",
        "fa_dim_policy_industry_map",
        "fa_dim_policy_plan_history",
    ]:
        schema = schemas[table_name]
        assert table_name.startswith("fa_")
        assert "source_date" in schema["columns"]
        assert "availability_date" in schema["columns"]
        assert "built_at" in schema["columns"]
        assert set(schema["required"]).issubset(set(schema["columns"]))
        assert schema["primary_key"]

    score_distribution = schemas["fa_fact_ln_score_distribution"]
    assert score_distribution["source_key"] == "ln_score_distribution"
    assert score_distribution["primary_key"] == ["subject_cat", "score_year", "score"]
    assert "cumulative_rank" in score_distribution["columns"]
    quality_config = score_distribution["quality_checks"]["score_distribution"]
    assert quality_config["score_max"] == 750
    assert quality_config["require_cumulative_sum"] is True


def test_parse_ln_score_distribution_lines():
    rows = parse_ln_score_distribution_lines(
        [
            "2025年辽宁省普通高校招生考试成绩统计表(物理学科类)",
            "分数 人数 累计",
            "707       11      11及以上",
            "706        2      13",
            "665       57   1,048",
            "分数 人数 累计 664 62 1,110 663 73 1,183",
        ],
        score_year=2025,
        subject_cat="物理类",
        source_date="2025-06-24",
    )

    by_score = {row["score"]: row for row in rows}
    assert by_score[707]["score_count"] == 11
    assert by_score[707]["cumulative_rank"] == 11
    assert by_score[665]["cumulative_rank"] == 1048
    assert by_score[663]["subject_cat"] == "物理类"


def test_parse_score_distribution_grid_images_preserves_score_gaps(tmp_path: Path, monkeypatch):
    image = tmp_path / "page.png"
    image.write_bytes(b"placeholder")
    row_images = [
        grid_parser.GridRowImage("page.png", 1, 1, 0.88, tmp_path / "r1.png"),
        grid_parser.GridRowImage("page.png", 1, 2, 0.86, tmp_path / "r2.png"),
        grid_parser.GridRowImage("page.png", 1, 3, 0.84, tmp_path / "r3.png"),
    ]
    ocr_results = [
        {
            "image_path": str(row_images[0].path),
            "observations": [
                {"text": "665及以上 12", "x": 0.17},
                {"text": "12", "x": 0.82},
            ],
        },
        {
            "image_path": str(row_images[1].path),
            "observations": [
                {"text": "663", "x": 0.17},
                {"text": "2", "x": 0.55},
                {"text": "1414", "x": 0.82},
            ],
        },
        {
            "image_path": str(row_images[2].path),
            "observations": [
                {"text": "662", "x": 0.17},
                {"text": "17", "x": 0.82},
            ],
        },
    ]
    monkeypatch.setattr(grid_parser, "_build_row_images", lambda *args, **kwargs: row_images)
    monkeypatch.setattr(grid_parser, "_run_row_ocr", lambda *args, **kwargs: ocr_results)

    rows, report = grid_parser.parse_score_distribution_grid_images(
        [image],
        subject_cat="历史类",
        score_year=2022,
        source_date="2022-06-23",
        work_dir=tmp_path / "rows",
    )

    by_score = {row["score"]: row for row in rows}
    assert sorted(by_score, reverse=True) == [665, 663, 662]
    assert by_score[663]["score_count"] == 2
    assert by_score[663]["cumulative_rank"] == 14
    assert by_score[662]["score_count"] == 3
    assert by_score[662]["cumulative_rank"] == 17
    assert report["repair_counts"]["cumulative_repaired_by_score_count_jump"] == 1
    assert report["quality_errors"] == []


def test_parse_score_distribution_grid_images_uses_smaller_count_from_noisy_cell():
    payload = grid_parser._parse_one_row(
        [
            {"text": "632", "x": 0.17},
            {"text": "30 ~1563", "x": 0.55},
        ],
        {
            "score_x_max": 0.36,
            "score_count_x_min": 0.32,
            "score_count_x_max": 0.68,
            "cumulative_x_min": 0.68,
            "max_score_count": 6000,
            "score_count_multi_number_strategy": "smallest_positive",
        },
    )
    assert payload["score_ocr"] == 632
    assert payload["score_count"] == 30
    assert payload["cumulative_rank"] is None


def test_parse_score_distribution_grid_images_score_count_strategy_is_configurable():
    observations = [
        {"text": "632", "x": 0.17},
        {"text": "30 ~1563", "x": 0.55},
    ]
    base_config = {
        "score_x_max": 0.36,
        "score_count_x_min": 0.32,
        "score_count_x_max": 0.68,
        "cumulative_x_min": 0.68,
        "max_score_count": 6000,
    }

    payload = grid_parser._parse_one_row(
        observations,
        {**base_config, "score_count_multi_number_strategy": "last_positive"},
    )

    assert payload["score_count"] == 1563


def test_audit_score_distribution_csvs_blocks_drift_and_missing_rows(tmp_path: Path):
    baseline = tmp_path / "baseline.csv"
    candidate = tmp_path / "candidate.csv"
    columns = ["subject_cat", "score_year", "score", "score_count", "cumulative_rank", "source_date"]
    with baseline.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=columns)
        writer.writeheader()
        writer.writerows([
            {
                "subject_cat": "历史类",
                "score_year": "2024",
                "score": "676",
                "score_count": "12",
                "cumulative_rank": "12",
                "source_date": "2024-06-24",
            },
            {
                "subject_cat": "历史类",
                "score_year": "2024",
                "score": "675",
                "score_count": "3",
                "cumulative_rank": "15",
                "source_date": "2024-06-24",
            },
        ])
    with candidate.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=columns)
        writer.writeheader()
        writer.writerow({
            "subject_cat": "历史类",
            "score_year": "2024",
            "score": "676",
            "score_count": "11",
            "cumulative_rank": "11",
            "source_date": "2024-06-25",
        })

    report = audit_score_distribution_csvs(
        candidate_csvs=[candidate],
        baseline_csvs=[baseline],
        report_path=tmp_path / "audit.json",
    )

    assert report["counts"]["matched_rows"] == 1
    assert report["counts"]["baseline_only_rows"] == 1
    assert report["counts"]["different_rows"] == 1
    assert report["decision"]["safe_to_promote_without_review"] is False
    assert report["decision"]["reconciliation_required"] is True
    assert report["samples"]["different_rows"][0]["diffs"]["score_count"]["candidate"] == 11
    assert report["sequence_summary"]["baseline"][0]["missing_score_count"] == 0


def test_parse_score_distribution_image_groups_uses_configured_manifest_indexes(tmp_path: Path, monkeypatch):
    manifest = tmp_path / "_page_images_index.json"
    files = [
        {
            "file_name": f"ln_score_distribution_2024_{index:03d}.jpg",
            "path": str(tmp_path / f"page_{index:03d}.jpg"),
        }
        for index in range(1, 9)
    ]
    manifest.write_text(
        json.dumps(
            {
                "source_key": "ln_score_distribution",
                "source_date": "2024-06-25",
                "page_url": "https://jyt.ln.gov.cn/jyt/jyzx/jyyw/2024062510394164694/index.shtml",
                "files": files,
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    calls = []

    def fake_parse(image_paths, *, subject_cat, score_year, source_date, work_dir, swiftc):
        calls.append({
            "image_paths": [path.name for path in image_paths],
            "subject_cat": subject_cat,
            "score_year": score_year,
            "source_date": source_date,
            "work_dir": work_dir.name,
            "swiftc": swiftc,
        })
        return (
            [
                {
                    "subject_cat": subject_cat,
                    "score_year": score_year,
                    "score": 676,
                    "score_count": 12,
                    "cumulative_rank": 12,
                    "source_date": source_date,
                }
            ],
            {"quality_errors": [], "output_rows": 1},
        )

    monkeypatch.setattr(image_group_parser, "parse_score_distribution_grid_images", fake_parse)

    result = image_group_parser.parse_score_distribution_image_groups(
        manifest_path=manifest,
        output_dir=tmp_path / "out",
        work_dir=tmp_path / "rows",
        group_keys=["ordinary_physics"],
        swiftc="swiftc-test",
        summary_report_path=tmp_path / "summary.json",
    )

    assert result["group_count"] == 1
    assert calls == [
        {
            "image_paths": [f"page_{index:03d}.jpg" for index in range(5, 9)],
            "subject_cat": "物理类",
            "score_year": 2024,
            "source_date": "2024-06-25",
            "work_dir": "ordinary_physics",
            "swiftc": "swiftc-test",
        }
    ]
    output_csv = tmp_path / "out" / "ln_score_distribution_2024_ordinary_physics_official_grid_candidate.csv"
    assert output_csv.exists()
    assert (tmp_path / "summary.json").exists()

    duplicate_manifest = tmp_path / "_page_images_duplicate.json"
    duplicate_manifest.write_text(
        (
            '{"source_key":"ln_score_distribution",'
            '"source_key":"shadow_source",'
            '"source_date":"2024-06-25",'
            '"page_url":"https://jyt.ln.gov.cn/jyt/jyzx/jyyw/2024062510394164694/index.shtml",'
            '"files":[]}\n'
        ),
        encoding="utf-8",
    )
    try:
        image_group_parser.parse_score_distribution_image_groups(
            manifest_path=duplicate_manifest,
            output_dir=tmp_path / "dup_out",
            work_dir=tmp_path / "dup_rows",
            group_keys=["ordinary_physics"],
            swiftc="swiftc-test",
        )
        duplicate_manifest_rejected = False
    except ValueError as exc:
        duplicate_manifest_rejected = "duplicate JSON key" in str(exc)
    assert duplicate_manifest_rejected


def test_parse_ln_score_distribution_ocr_jsonl_candidates(tmp_path: Path):
    ocr_jsonl = tmp_path / "ocr.jsonl"
    image_result = {
        "image_path": "/tmp/ln_score_distribution_2024_001.jpg",
        "observations": [
            {"text": "2024年辽宁省普通高校招生考试成绩统计表（历史学科类）", "confidence": 1, "x": 0.1, "y": 0.95, "width": 0.7, "height": 0.02},
            {"text": "676及以上12", "confidence": 0.8, "x": 0.08, "y": 0.88, "width": 0.1, "height": 0.01},
            {"text": "12", "confidence": 1, "x": 0.22, "y": 0.88, "width": 0.03, "height": 0.01},
            {"text": "675", "confidence": 1, "x": 0.08, "y": 0.86, "width": 0.03, "height": 0.01},
            {"text": "3", "confidence": 1, "x": 0.16, "y": 0.86, "width": 0.01, "height": 0.01},
            {"text": "15", "confidence": 1, "x": 0.22, "y": 0.86, "width": 0.02, "height": 0.01},
            {"text": "674及以上17", "confidence": 0.8, "x": 0.08, "y": 0.84, "width": 0.1, "height": 0.01},
            {"text": "2 19", "confidence": 0.9, "x": 0.08, "y": 0.82, "width": 0.1, "height": 0.01},
            {"text": "404 463 100,014", "confidence": 0.9, "x": 0.30, "y": 0.80, "width": 0.1, "height": 0.01},
            {"text": "150 22,006", "confidence": 0.7, "x": 0.60, "y": 0.50, "width": 0.1, "height": 0.01},
        ],
    }
    ocr_jsonl.write_text(json.dumps(image_result, ensure_ascii=False) + "\n", encoding="utf-8")

    rows, report = parse_ln_score_distribution_ocr_jsonl(ocr_jsonl, source_date="2024-06-25")
    by_score = {row["score"]: row for row in rows if row["parse_status"] == "parsed"}
    assert by_score[676]["score_count"] == 12
    assert by_score[675]["cumulative_rank"] == 15
    assert by_score[674]["cumulative_rank"] == 17
    assert by_score[404]["cumulative_rank"] == 100014
    inferred = next(row for row in rows if row["parse_status"] == "inferred_score")
    assert inferred["score"] == 673
    assert inferred["math_status"] == "ok"
    assert by_score[676]["math_status"] == "ok"
    assert by_score[675]["math_status"] == "ok"
    assert report["inferred_score_rows"] == 1
    assert report["subjects"] == ["历史类"]
    assert report["candidate_rows"] >= 2

    output = tmp_path / "candidates.csv"
    write_candidate_csv(output, rows)
    with output.open(encoding="utf-8", newline="") as f:
        written = list(csv.DictReader(f))
    assert written[0]["parse_status"] == "parsed"
    assert "raw_text" in written[0]


def test_parse_ln_score_distribution_ocr_infers_single_number_rows(tmp_path: Path):
    ocr_jsonl = tmp_path / "ocr_single_number.jsonl"
    image_result = {
        "image_path": "/tmp/ln_score_distribution_2024_001.jpg",
        "observations": [
            {"text": "2024年辽宁省普通高校招生考试成绩统计表（历史学科类）", "confidence": 1, "x": 0.1, "y": 0.95, "width": 0.7, "height": 0.02},
            {"text": "676", "confidence": 1, "x": 0.08, "y": 0.88, "width": 0.03, "height": 0.01},
            {"text": "12", "confidence": 1, "x": 0.16, "y": 0.88, "width": 0.02, "height": 0.01},
            {"text": "12", "confidence": 1, "x": 0.22, "y": 0.88, "width": 0.02, "height": 0.01},
            {"text": "3", "confidence": 1, "x": 0.16, "y": 0.86, "width": 0.01, "height": 0.01},
            {"text": "17", "confidence": 1, "x": 0.22, "y": 0.84, "width": 0.02, "height": 0.01},
            {"text": "673", "confidence": 1, "x": 0.08, "y": 0.82, "width": 0.03, "height": 0.01},
            {"text": "4", "confidence": 1, "x": 0.16, "y": 0.82, "width": 0.01, "height": 0.01},
            {"text": "21", "confidence": 1, "x": 0.22, "y": 0.82, "width": 0.02, "height": 0.01},
        ],
    }
    ocr_jsonl.write_text(json.dumps(image_result, ensure_ascii=False) + "\n", encoding="utf-8")

    rows, report = parse_ln_score_distribution_ocr_jsonl(ocr_jsonl, source_date="2024-06-25")
    by_score = {row["score"]: row for row in rows}
    assert by_score[675]["parse_status"] == "inferred_row"
    assert by_score[675]["score_count"] == 3
    assert by_score[675]["cumulative_rank"] == 15
    assert by_score[674]["parse_status"] == "inferred_row"
    assert by_score[674]["score_count"] == 2
    assert by_score[674]["cumulative_rank"] == 17
    assert by_score[674]["math_status"] == "ok"
    assert report["inferred_row_rows"] == 2


def test_build_score_distribution_review_tasks_from_candidates(tmp_path: Path):
    candidates = tmp_path / "candidates.csv"
    rows = [
        {
            "subject_cat": "历史类",
            "score_year": "2024",
            "score": "676",
            "score_count": "12",
            "cumulative_rank": "12",
            "source_date": "2024-06-25",
            "image_file": "page1.jpg",
            "block_index": "1",
            "row_y": "0.88",
            "ocr_confidence": "0.8",
            "parse_status": "parsed",
            "math_status": "ok",
            "raw_text": "676及以上12",
        },
        {
            "subject_cat": "历史类",
            "score_year": "2024",
            "score": "675",
            "score_count": "3",
            "cumulative_rank": "16",
            "source_date": "2024-06-25",
            "image_file": "page1.jpg",
            "block_index": "1",
            "row_y": "0.86",
            "ocr_confidence": "1.0",
            "parse_status": "parsed",
            "math_status": "cumulative_mismatch",
            "raw_text": "675 3 16",
        },
        {
            "subject_cat": "历史类",
            "score_year": "2024",
            "score": "93",
            "score_count": "4874",
            "cumulative_rank": "",
            "source_date": "2024-06-25",
            "image_file": "page1.jpg",
            "block_index": "4",
            "row_y": "0.87",
            "ocr_confidence": "1.0",
            "parse_status": "incomplete",
            "math_status": "not_checked",
            "raw_text": "93 4,874",
        },
    ]
    write_candidate_csv(candidates, rows)

    tasks, report = build_score_distribution_review_tasks(candidates)
    assert report["candidate_rows"] == 3
    assert report["review_task_rows"] == 2
    assert report["issue_counts"]["cumulative_mismatch"] == 1
    assert tasks[0]["issue_type"] == "cumulative_mismatch"
    assert tasks[0]["priority"] == 2
    assert tasks[0]["review_status"] == "todo"
    assert tasks[0]["suggested_score"] == ""
    assert "累计校验" in tasks[0]["suggested_action"]

    output = tmp_path / "review.csv"
    write_review_task_csv(output, tasks)
    with output.open(encoding="utf-8", newline="") as f:
        written = list(csv.DictReader(f))
    assert written[0]["corrected_score"] == ""
    assert "suggested_score" in written[0]
    assert written[0]["issue_type"] == "cumulative_mismatch"


def test_build_score_distribution_review_tasks_prefills_sequence_suggestions(tmp_path: Path):
    candidates = tmp_path / "candidates.csv"
    rows = [
        {
            "subject_cat": "物理类",
            "score_year": "2022",
            "score": "487",
            "score_count": "300",
            "cumulative_rank": "51198",
            "source_date": "2022-06-23",
            "image_file": "page-1.png",
            "block_index": "4",
            "row_y": "0.08",
            "ocr_confidence": "0.95",
            "parse_status": "parsed",
            "math_status": "ok",
            "raw_text": "487 300 51,198",
        },
        {
            "subject_cat": "物理类",
            "score_year": "2022",
            "score": "51562",
            "score_count": "",
            "cumulative_rank": "",
            "source_date": "2022-06-23",
            "image_file": "page-2.png",
            "block_index": "1",
            "row_y": "0.88",
            "ocr_confidence": "0.70",
            "parse_status": "invalid_score",
            "math_status": "not_checked",
            "raw_text": "51,562",
        },
        {
            "subject_cat": "物理类",
            "score_year": "2022",
            "score": "5201",
            "score_count": "1",
            "cumulative_rank": "",
            "source_date": "2022-06-23",
            "image_file": "page-2.png",
            "block_index": "1",
            "row_y": "0.86",
            "ocr_confidence": "0.60",
            "parse_status": "invalid_score",
            "math_status": "not_checked",
            "raw_text": "52,01 1",
        },
        {
            "subject_cat": "物理类",
            "score_year": "2022",
            "score": "484",
            "score_count": "452",
            "cumulative_rank": "52463",
            "source_date": "2022-06-23",
            "image_file": "page-2.png",
            "block_index": "1",
            "row_y": "0.84",
            "ocr_confidence": "0.96",
            "parse_status": "parsed",
            "math_status": "cumulative_mismatch",
            "raw_text": "484 452 52,463",
        },
    ]
    write_candidate_csv(candidates, rows)

    tasks, report = build_score_distribution_review_tasks(candidates)

    assert report["review_task_rows"] == 3
    assert report["suggested_review_rows"] == 3
    by_text = {task["raw_text"]: task for task in tasks}
    assert by_text["51,562"]["suggested_score"] == 486
    assert by_text["51,562"]["suggested_score_count"] == 364
    assert by_text["51,562"]["suggested_cumulative_rank"] == 51562
    assert by_text["52,01 1"]["suggested_score"] == 485
    assert by_text["52,01 1"]["suggested_score_count"] == 449
    assert by_text["52,01 1"]["suggested_cumulative_rank"] == 52011
    assert by_text["484 452 52,463"]["suggested_score"] == 484
    assert by_text["484 452 52,463"]["suggested_score_count"] == 452
    assert by_text["484 452 52,463"]["suggested_cumulative_rank"] == 52463


def test_build_score_distribution_review_tasks_prefills_single_boundary_suggestions(tmp_path: Path):
    candidates = tmp_path / "candidates.csv"
    rows = []
    for index in range(9):
        rows.append({
            "subject_cat": "物理类",
            "score_year": "2024",
            "score": str(index + 2),
            "score_count": str((index + 2) * 2),
            "cumulative_rank": "",
            "source_date": "2024-06-25",
            "image_file": "page-low-score.jpg",
            "block_index": "3",
            "row_y": f"{0.90 - index * 0.01:.2f}",
            "ocr_confidence": "0.8",
            "parse_status": "incomplete",
            "math_status": "not_checked",
            "raw_text": f"{index + 2} {(index + 2) * 2}",
        })
    rows.append({
        "subject_cat": "物理类",
        "score_year": "2024",
        "score": "491",
        "score_count": "12",
        "cumulative_rank": "108",
        "source_date": "2024-06-25",
        "image_file": "page-low-score.jpg",
        "block_index": "3",
        "row_y": "0.81",
        "ocr_confidence": "0.95",
        "parse_status": "parsed",
        "math_status": "ok",
        "raw_text": "491 12 108",
    })
    write_candidate_csv(candidates, rows)

    tasks, report = build_score_distribution_review_tasks(candidates)

    assert report["review_task_rows"] == 9
    assert report["suggested_review_rows"] == 9
    first = next(task for task in tasks if task["raw_text"] == "2 4")
    assert first["review_status"] == "todo"
    assert first["corrected_score"] == ""
    assert first["suggested_score"] == 500
    assert first["suggested_score_count"] == 2
    assert first["suggested_cumulative_rank"] == 4


def test_prefill_score_distribution_review_suggestions_requires_human_approval(tmp_path: Path):
    candidates = tmp_path / "candidates.csv"
    rows = []
    for index in range(9):
        rows.append({
            "subject_cat": "物理类",
            "score_year": "2024",
            "score": str(index + 2),
            "score_count": str((index + 2) * 2),
            "cumulative_rank": "",
            "source_date": "2024-06-25",
            "image_file": "page-low-score.jpg",
            "block_index": "3",
            "row_y": f"{0.90 - index * 0.01:.2f}",
            "ocr_confidence": "0.8",
            "parse_status": "incomplete",
            "math_status": "not_checked",
            "raw_text": f"{index + 2} {(index + 2) * 2}",
        })
    rows.append({
        "subject_cat": "物理类",
        "score_year": "2024",
        "score": "491",
        "score_count": "12",
        "cumulative_rank": "108",
        "source_date": "2024-06-25",
        "image_file": "page-low-score.jpg",
        "block_index": "3",
        "row_y": "0.81",
        "ocr_confidence": "0.95",
        "parse_status": "parsed",
        "math_status": "ok",
        "raw_text": "491 12 108",
    })
    write_candidate_csv(candidates, rows)
    tasks, _ = build_score_distribution_review_tasks(candidates)
    review = tmp_path / "review.csv"
    write_review_task_csv(review, tasks)

    prefilled_rows, report = prefill_score_distribution_review_suggestions(review)

    assert report["prefilled_rows"] == 9
    first = next(row for row in prefilled_rows if row["raw_text"] == "2 4")
    assert first["corrected_score"] == first["suggested_score"]
    assert first["corrected_score_count"] == first["suggested_score_count"]
    assert first["corrected_cumulative_rank"] == first["suggested_cumulative_rank"]
    assert first["review_status"] == "todo"
    assert "requires_image_check" in first["reviewer_notes"]

    prefilled_review = tmp_path / "prefilled_review.csv"
    write_review_task_csv(prefilled_review, prefilled_rows)
    try:
        apply_score_distribution_review(candidates, prefilled_review)
        accepted_without_review = True
    except ValueError as exc:
        accepted_without_review = False
        assert "unresolved review rows" in str(exc)
    assert accepted_without_review is False


def test_build_score_distribution_review_tasks_skips_conflicting_boundary_suggestions(tmp_path: Path):
    candidates = tmp_path / "candidates.csv"
    rows = [
        {
            "subject_cat": "物理类",
            "score_year": "2024",
            "score": "530",
            "score_count": "1",
            "cumulative_rank": "1",
            "source_date": "2024-06-25",
            "image_file": "page-conflict.jpg",
            "block_index": "3",
            "row_y": "0.90",
            "ocr_confidence": "0.95",
            "parse_status": "parsed",
            "math_status": "ok",
            "raw_text": "530 1 1",
        }
    ]
    for index in range(8):
        rows.append({
            "subject_cat": "物理类",
            "score_year": "2024",
            "score": str(index + 2),
            "score_count": str((index + 2) * 2),
            "cumulative_rank": "",
            "source_date": "2024-06-25",
            "image_file": "page-conflict.jpg",
            "block_index": "3",
            "row_y": f"{0.89 - index * 0.01:.2f}",
            "ocr_confidence": "0.8",
            "parse_status": "incomplete",
            "math_status": "not_checked",
            "raw_text": f"{index + 2} {(index + 2) * 2}",
        })
    rows.append({
        "subject_cat": "物理类",
        "score_year": "2024",
        "score": "491",
        "score_count": "12",
        "cumulative_rank": "108",
        "source_date": "2024-06-25",
        "image_file": "page-conflict.jpg",
        "block_index": "3",
        "row_y": "0.81",
        "ocr_confidence": "0.95",
        "parse_status": "parsed",
        "math_status": "ok",
        "raw_text": "491 12 108",
    })
    write_candidate_csv(candidates, rows)

    tasks, report = build_score_distribution_review_tasks(candidates)

    assert report["review_task_rows"] == 8
    assert report["suggested_review_rows"] == 0
    assert all(task["suggested_score"] == "" for task in tasks)


def test_build_and_merge_score_distribution_review_workspace(tmp_path: Path):
    review = tmp_path / "review.csv"
    image_path = tmp_path / "page1.jpg"
    image_path.write_bytes(b"not-a-real-image")
    manifest = tmp_path / "_page_images_index.json"
    manifest.write_text(json.dumps({
        "files": [
            {"file_name": "page1.jpg", "path": str(image_path)},
            {"file_name": "page2.jpg", "path": str(tmp_path / "page2.jpg")},
        ]
    }), encoding="utf-8")
    rows = [
        {
            "review_id": "r1",
            "priority": "2",
            "issue_type": "cumulative_mismatch",
            "suggested_action": "核对累计",
            "subject_cat": "历史类",
            "score_year": "2024",
            "score": "675",
            "score_count": "3",
            "cumulative_rank": "16",
            "source_date": "2024-06-25",
            "image_file": "page1.jpg",
            "block_index": "1",
            "row_y": "0.86",
            "ocr_confidence": "1.0",
            "parse_status": "parsed",
            "math_status": "cumulative_mismatch",
            "raw_text": "675 3 16",
            "review_status": "todo",
            "reviewer_notes": "",
            "corrected_score": "",
            "corrected_score_count": "",
            "corrected_cumulative_rank": "",
        },
        {
            "review_id": "r2",
            "priority": "5",
            "issue_type": "incomplete",
            "suggested_action": "补齐",
            "subject_cat": "历史类",
            "score_year": "2024",
            "score": "93",
            "score_count": "4874",
            "cumulative_rank": "",
            "source_date": "2024-06-25",
            "image_file": "page2.jpg",
            "block_index": "4",
            "row_y": "0.87",
            "ocr_confidence": "1.0",
            "parse_status": "incomplete",
            "math_status": "not_checked",
            "raw_text": "93 4,874",
            "review_status": "todo",
            "reviewer_notes": "",
            "corrected_score": "",
            "corrected_score_count": "",
            "corrected_cumulative_rank": "",
        },
    ]
    write_review_task_csv(review, rows)

    workspace = tmp_path / "workspace"
    report = build_score_distribution_review_workspace(
        review_csv=review,
        output_dir=workspace,
        image_manifest=manifest,
    )
    assert report["task_rows"] == 2
    assert report["unresolved_rows"] == 2
    assert len(report["batches"]) == 2
    assert report["batches"][0]["locator_rows"] == 1
    assert (workspace / "index.html").exists()
    assert (workspace / "review_workspace_manifest.json").exists()
    html_text = (workspace / "index.html").read_text(encoding="utf-8")
    assert "row-marker cumulative_mismatch" in html_text
    assert 'data-review-id="r1"' in html_text
    assert "image-stage" in html_text

    page1_batch = Path(report["batches"][0]["csv"])
    with page1_batch.open(encoding="utf-8", newline="") as f:
        batch_rows = list(csv.DictReader(f))
    batch_rows[0]["review_status"] = "approved"
    batch_rows[0]["corrected_cumulative_rank"] = "15"
    batch_rows[0]["issue_type"] = "changed-but-not-merged"
    write_review_task_csv(page1_batch, batch_rows)

    merged = tmp_path / "merged_review.csv"
    merge_report = merge_score_distribution_review_workspace(
        review_csv=review,
        workspace_dir=workspace,
        output=merged,
    )
    assert merge_report["updated_rows"] == 1
    assert merge_report["unresolved_rows"] == 1

    with merged.open(encoding="utf-8", newline="") as f:
        merged_rows = {row["review_id"]: row for row in csv.DictReader(f)}
    assert merged_rows["r1"]["review_status"] == "approved"
    assert merged_rows["r1"]["corrected_cumulative_rank"] == "15"
    assert merged_rows["r1"]["issue_type"] == "cumulative_mismatch"
    assert merged_rows["r2"]["review_status"] == "todo"

    duplicate_manifest = tmp_path / "_page_images_duplicate.json"
    duplicate_manifest.write_text(
        (
            '{"files":[],'
            f'"files":[{{"file_name":"page1.jpg","path":"{image_path}"}}]}}\n'
        ),
        encoding="utf-8",
    )
    try:
        build_score_distribution_review_workspace(
            review_csv=review,
            output_dir=tmp_path / "duplicate_workspace",
            image_manifest=duplicate_manifest,
        )
        duplicate_manifest_rejected = False
    except ValueError as exc:
        duplicate_manifest_rejected = "duplicate JSON key" in str(exc)
    assert duplicate_manifest_rejected


def test_apply_score_distribution_review_writes_cleaned_rows(tmp_path: Path):
    candidates = tmp_path / "candidates.csv"
    candidate_rows = [
        {
            "subject_cat": "历史类",
            "score_year": "2024",
            "score": "676",
            "score_count": "12",
            "cumulative_rank": "12",
            "source_date": "2024-06-25",
            "image_file": "page1.jpg",
            "block_index": "1",
            "row_y": "0.88",
            "ocr_confidence": "0.8",
            "parse_status": "parsed",
            "math_status": "ok",
            "raw_text": "676及以上12",
        },
        {
            "subject_cat": "历史类",
            "score_year": "2024",
            "score": "675",
            "score_count": "3",
            "cumulative_rank": "16",
            "source_date": "2024-06-25",
            "image_file": "page1.jpg",
            "block_index": "1",
            "row_y": "0.86",
            "ocr_confidence": "1.0",
            "parse_status": "parsed",
            "math_status": "cumulative_mismatch",
            "raw_text": "675 3 16",
        },
        {
            "subject_cat": "历史类",
            "score_year": "2024",
            "score": "93",
            "score_count": "4874",
            "cumulative_rank": "",
            "source_date": "2024-06-25",
            "image_file": "page1.jpg",
            "block_index": "4",
            "row_y": "0.87",
            "ocr_confidence": "1.0",
            "parse_status": "incomplete",
            "math_status": "not_checked",
            "raw_text": "93 4,874",
        },
    ]
    write_candidate_csv(candidates, candidate_rows)
    tasks, _ = build_score_distribution_review_tasks(candidates)
    for task in tasks:
        task["review_status"] = "approved"
        if task["raw_text"] == "675 3 16":
            task["corrected_cumulative_rank"] = "15"
        if task["raw_text"] == "93 4,874":
            task["corrected_score"] = "674"
            task["corrected_score_count"] = "2"
            task["corrected_cumulative_rank"] = "17"
    review = tmp_path / "review.csv"
    write_review_task_csv(review, tasks)

    cleaned_rows, report = apply_score_distribution_review(candidates, review)
    assert report["unresolved_rows"] == 0
    assert report["quality_errors"] == []
    assert report["applied_review_rows"] == 2
    assert [row["score"] for row in cleaned_rows] == [676, 675, 674]

    cleaned = tmp_path / "cleaned.csv"
    write_cleaned_score_distribution_csv(cleaned, cleaned_rows)
    with cleaned.open(encoding="utf-8", newline="") as f:
        written = list(csv.DictReader(f))
    assert written[2]["cumulative_rank"] == "17"


def test_audit_score_distribution_readiness_reports_review_blockers(tmp_path: Path):
    candidates = tmp_path / "candidates.csv"
    candidate_rows = [
        {
            "subject_cat": "历史类",
            "score_year": "2024",
            "score": "676",
            "score_count": "12",
            "cumulative_rank": "12",
            "source_date": "2024-06-25",
            "image_file": "page1.jpg",
            "block_index": "1",
            "row_y": "0.88",
            "ocr_confidence": "0.8",
            "parse_status": "parsed",
            "math_status": "ok",
            "raw_text": "676及以上12",
        },
        {
            "subject_cat": "历史类",
            "score_year": "2024",
            "score": "675",
            "score_count": "3",
            "cumulative_rank": "16",
            "source_date": "2024-06-25",
            "image_file": "page1.jpg",
            "block_index": "1",
            "row_y": "0.86",
            "ocr_confidence": "1.0",
            "parse_status": "parsed",
            "math_status": "cumulative_mismatch",
            "raw_text": "675 3 16",
        },
    ]
    write_candidate_csv(candidates, candidate_rows)

    no_review = audit_score_distribution_readiness(candidate_csv=candidates)
    assert no_review["required_review"]["review_task_rows"] == 1
    assert "review_csv_required" in no_review["ready"]["blocking_reasons"]

    tasks, _ = build_score_distribution_review_tasks(candidates)
    review = tmp_path / "review.csv"
    write_review_task_csv(review, tasks)
    pending = audit_score_distribution_readiness(candidate_csv=candidates, review_csv=review)
    assert pending["review_progress"]["unresolved_rows"] == 1
    assert pending["strict_apply"]["ok"] is False

    tasks[0]["review_status"] = "corrected"
    tasks[0]["corrected_cumulative_rank"] = "15"
    write_review_task_csv(review, tasks)
    cleaned = tmp_path / "cleaned.csv"
    cleaned_rows, _ = apply_score_distribution_review(candidates, review)
    write_cleaned_score_distribution_csv(cleaned, cleaned_rows)
    ready = audit_score_distribution_readiness(
        candidate_csv=candidates,
        review_csv=review,
        cleaned_csv=cleaned,
    )
    assert ready["strict_apply"]["ok"] is True
    assert ready["ready"]["review_complete"] is True
    assert ready["ready"]["cleaned_package_ready"] is True
    assert ready["ready"]["blocking_reasons"] == []


def test_apply_score_distribution_review_rejects_unresolved_rows(tmp_path: Path):
    candidates = tmp_path / "candidates.csv"
    write_candidate_csv(candidates, [
        {
            "subject_cat": "历史类",
            "score_year": "2024",
            "score": "676",
            "score_count": "12",
            "cumulative_rank": "12",
            "source_date": "2024-06-25",
            "image_file": "page1.jpg",
            "block_index": "1",
            "row_y": "0.88",
            "ocr_confidence": "0.8",
            "parse_status": "parsed",
            "math_status": "ok",
            "raw_text": "676及以上12",
        },
        {
            "subject_cat": "历史类",
            "score_year": "2024",
            "score": "675",
            "score_count": "3",
            "cumulative_rank": "16",
            "source_date": "2024-06-25",
            "image_file": "page1.jpg",
            "block_index": "1",
            "row_y": "0.86",
            "ocr_confidence": "1.0",
            "parse_status": "parsed",
            "math_status": "cumulative_mismatch",
            "raw_text": "675 3 16",
        },
    ])
    tasks, _ = build_score_distribution_review_tasks(candidates)
    review = tmp_path / "review.csv"
    write_review_task_csv(review, tasks)

    try:
        apply_score_distribution_review(candidates, review)
        rejected = False
    except ValueError as exc:
        rejected = "unresolved review rows" in str(exc)
    assert rejected


def test_apply_score_distribution_review_rejects_duplicate_primary_keys(tmp_path: Path):
    candidates = tmp_path / "candidates.csv"
    write_candidate_csv(candidates, [
        {
            "subject_cat": "历史类",
            "score_year": "2024",
            "score": "676",
            "score_count": "12",
            "cumulative_rank": "12",
            "source_date": "2024-06-25",
            "image_file": "page1.jpg",
            "block_index": "1",
            "row_y": "0.88",
            "ocr_confidence": "0.8",
            "parse_status": "parsed",
            "math_status": "ok",
            "raw_text": "676及以上12",
        },
        {
            "subject_cat": "历史类",
            "score_year": "2024",
            "score": "676",
            "score_count": "12",
            "cumulative_rank": "12",
            "source_date": "2024-06-25",
            "image_file": "page1.jpg",
            "block_index": "1",
            "row_y": "0.87",
            "ocr_confidence": "0.8",
            "parse_status": "parsed",
            "math_status": "ok",
            "raw_text": "676 12 12",
        },
    ])
    review = tmp_path / "review.csv"
    write_review_task_csv(review, [])

    try:
        apply_score_distribution_review(candidates, review)
        rejected = False
    except ValueError as exc:
        rejected = "duplicate primary keys" in str(exc)
    assert rejected


def test_build_score_history_from_projection_package(tmp_path: Path):
    projection = tmp_path / "projection.csv"
    with projection.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "school_code",
                "school_name",
                "major_code",
                "major_full",
                "batch",
                "subject_cat",
                "score_year",
                "min_score",
                "tie_breaker_json",
                "source_date",
            ],
        )
        writer.writeheader()
        writer.writerow({
            "school_code": "0378",
            "school_name": "安徽财经大学",
            "major_code": "13",
            "major_full": "计算机类",
            "batch": "本科批",
            "subject_cat": "物理类",
            "score_year": "2025",
            "min_score": "574",
            "tie_breaker_json": "{}",
            "source_date": "2025-07-20",
        })
    distribution = tmp_path / "distribution.csv"
    with distribution.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["subject_cat", "score_year", "score", "score_count", "cumulative_rank", "source_date"],
        )
        writer.writeheader()
        writer.writerow({
            "subject_cat": "物理类",
            "score_year": "2025",
            "score": "574",
            "score_count": "364",
            "cumulative_rank": "22820",
            "source_date": "2025-06-24",
        })

    result = build_score_history_from_projection_package(
        projection_csv=projection,
        score_distribution_csv=distribution,
        output_root=tmp_path / "exports",
        package_id="pkg-score-history-derived-test",
    )
    package_dir = Path(result["package_dir"])
    assert validate_manifest(package_dir / "manifest.json")["errors"] == []
    assert result["rows"] == 1
    assert result["quality_report"]["warnings"][0]["code"] == "rank_is_score_cumulative_rank"

    with (package_dir / "fa_fact_ln_score_history.csv").open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    assert rows[0]["min_rank"] == "22820"
    assert rows[0]["school_code"] == "0378"


def test_audit_score_history_package_against_core_reports_overlap_drift(tmp_path: Path):
    db = tmp_path / "core.duckdb"
    con = duckdb.connect(str(db))
    try:
        con.execute("""
            CREATE TABLE fa_fact_ln_score_history (
                school_code VARCHAR,
                major_code VARCHAR,
                batch VARCHAR,
                subject_cat VARCHAR,
                score_year INTEGER,
                min_score DOUBLE,
                min_rank INTEGER
            )
        """)
        con.execute("""
            INSERT INTO fa_fact_ln_score_history VALUES
                ('1001', '01', '本科批', '物理类', 2024, 600, 1000),
                ('1002', '02', '本科批', '物理类', 2024, 580, 2000),
                ('1003', '03', '本科批', '物理类', 2024, 570, 3000),
                ('2001', '01', '本科批', '物理类', 2023, 610, 900)
        """)
    finally:
        con.close()

    package_dir = tmp_path / "exports" / "pkg-score-history-audit"
    package_dir.mkdir(parents=True)
    table_path = package_dir / "fa_fact_ln_score_history.csv"
    with table_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "school_code",
                "major_code",
                "batch",
                "subject_cat",
                "score_year",
                "min_score",
                "min_rank",
                "plan_count",
            ],
        )
        writer.writeheader()
        writer.writerows([
            {
                "school_code": "1001",
                "major_code": "01",
                "batch": "本科批",
                "subject_cat": "物理类",
                "score_year": "2024",
                "min_score": "600",
                "min_rank": "1000",
                "plan_count": "10",
            },
            {
                "school_code": "1002",
                "major_code": "02",
                "batch": "本科批",
                "subject_cat": "物理类",
                "score_year": "2024",
                "min_score": "580",
                "min_rank": "1990",
                "plan_count": "8",
            },
            {
                "school_code": "1003",
                "major_code": "04",
                "batch": "本科批",
                "subject_cat": "物理类",
                "score_year": "2024",
                "min_score": "570",
                "min_rank": "3000",
                "plan_count": "6",
            },
        ])
    (package_dir / "quality_report.json").write_text('{"errors":[]}\n', encoding="utf-8")
    (package_dir / "manifest.json").write_text(json.dumps({
        "package_id": "pkg-score-history-audit",
        "built_at": "2026-05-13T00:00:00",
        "source_version": "fixture",
        "tables": [{"name": "fa_fact_ln_score_history", "file": "fa_fact_ln_score_history.csv"}],
        "files": ["fa_fact_ln_score_history.csv"],
        "hashes": {},
        "quality_report": "quality_report.json",
    }, ensure_ascii=False), encoding="utf-8")

    report = audit_score_history_package_against_core(
        core_db=db,
        package_dirs=[package_dir],
        sample_limit=5,
    )

    assert report["errors"] == []
    assert report["counts"]["package_rows"] == 3
    assert report["counts"]["core_scoped_rows"] == 3
    assert report["counts"]["matched_rows"] == 2
    assert report["counts"]["package_only_rows"] == 1
    assert report["counts"]["core_only_rows"] == 1
    assert report["counts"]["different_rows"] == 1
    assert report["decision"]["reconciliation_required"] is True
    assert report["decision"]["safe_to_import_without_reconciliation"] is False
    assert report["reconciliation_hints"]["same_values_different_key_candidates"]["candidate_pairs"] == 1
    assert (
        report["reconciliation_hints"]["same_values_different_key_candidates"]["samples"][0]["variant_differences"]
        == [{"column": "major_code", "package_value": "04", "core_value": "03"}]
    )
    assert report["samples"]["different_rows"][0]["differences"] == [
        {"column": "min_rank", "package_value": 1990, "core_value": 2000}
    ]

    (package_dir / "quality_report.json").write_text('{"errors":["manual review is not complete"]}\n', encoding="utf-8")
    bad_quality_report = audit_score_history_package_against_core(
        core_db=db,
        package_dirs=[package_dir],
        sample_limit=5,
    )
    assert any("quality_report error" in error for error in bad_quality_report["errors"])
    assert bad_quality_report["decision"]["safe_to_import_without_reconciliation"] is False
    with pytest.raises(ValueError, match="quality_report error"):
        build_score_history_reconciliation_plan(
            core_db=db,
            package_dirs=[package_dir],
            output_dir=tmp_path / "blocked_reconciliation",
        )


def test_build_score_history_reconciliation_plan_from_audit_inputs(tmp_path: Path):
    db = tmp_path / "core.duckdb"
    con = duckdb.connect(str(db))
    try:
        con.execute("""
            CREATE TABLE fa_fact_ln_score_history (
                school_code VARCHAR,
                major_code VARCHAR,
                batch VARCHAR,
                subject_cat VARCHAR,
                score_year INTEGER,
                min_score DOUBLE,
                min_rank INTEGER
            )
        """)
        con.execute("""
            INSERT INTO fa_fact_ln_score_history VALUES
                ('1001', '01', '本科批', '物理类', 2024, 600, 1000),
                ('1002', '02', '本科批', '物理类', 2024, 580, 2000),
                ('1003', '03', '本科批', '物理类', 2024, 570, 3000),
                ('1007', '07', '本科批', '物理类', 2024, 0, 0),
                ('2001', '01', '本科批', '物理类', 2023, 610, 900)
        """)
    finally:
        con.close()

    package_dir = tmp_path / "exports" / "pkg-score-history-reconciliation"
    package_dir.mkdir(parents=True)
    with (package_dir / "fa_fact_ln_score_history.csv").open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "school_code",
                "major_code",
                "batch",
                "subject_cat",
                "score_year",
                "min_score",
                "min_rank",
                "plan_count",
            ],
        )
        writer.writeheader()
        writer.writerows([
            {
                "school_code": "1001",
                "major_code": "01",
                "batch": "本科批",
                "subject_cat": "物理类",
                "score_year": "2024",
                "min_score": "600",
                "min_rank": "1000",
                "plan_count": "10",
            },
            {
                "school_code": "1002",
                "major_code": "02",
                "batch": "本科批",
                "subject_cat": "物理类",
                "score_year": "2024",
                "min_score": "580",
                "min_rank": "1990",
                "plan_count": "8",
            },
            {
                "school_code": "1003",
                "major_code": "04",
                "batch": "本科批",
                "subject_cat": "物理类",
                "score_year": "2024",
                "min_score": "570",
                "min_rank": "3000",
                "plan_count": "6",
            },
            {
                "school_code": "1006",
                "major_code": "06",
                "batch": "本科批",
                "subject_cat": "物理类",
                "score_year": "2024",
                "min_score": "550",
                "min_rank": "6000",
                "plan_count": "4",
            },
        ])
    (package_dir / "quality_report.json").write_text('{"errors":[]}\n', encoding="utf-8")
    (package_dir / "manifest.json").write_text(json.dumps({
        "package_id": "pkg-score-history-reconciliation",
        "built_at": "2026-05-13T00:00:00",
        "source_version": "fixture",
        "tables": [{"name": "fa_fact_ln_score_history", "file": "fa_fact_ln_score_history.csv"}],
        "files": ["fa_fact_ln_score_history.csv"],
        "hashes": {},
        "quality_report": "quality_report.json",
    }, ensure_ascii=False), encoding="utf-8")

    result = build_score_history_reconciliation_plan(
        core_db=db,
        package_dirs=[package_dir],
        output_dir=tmp_path / "reconciliation",
    )

    assert result["rows"] == 4
    assert result["issue_counts"] == {
        "core_only_zero_placeholder": 1,
        "major_code_drift_candidate": 1,
        "package_only_unmatched": 1,
        "value_drift": 1,
    }
    with Path(result["csv"]).open(encoding="utf-8", newline="") as f:
        tasks = list(csv.DictReader(f))
    by_type = {task["issue_type"]: task for task in tasks}
    assert by_type["major_code_drift_candidate"]["status"] == "todo"
    assert by_type["major_code_drift_candidate"]["suggested_action"] == "review_major_code_alignment"
    assert by_type["major_code_drift_candidate"]["package_major_code"] == "04"
    assert by_type["major_code_drift_candidate"]["core_major_code"] == "03"
    assert by_type["value_drift"]["differences_json"] == json.dumps([
        {"column": "min_rank", "package_value": 1990, "core_value": 2000}
    ], ensure_ascii=False, sort_keys=True)
    assert by_type["core_only_zero_placeholder"]["suggested_action"] == "review_core_zero_placeholder_for_delete_plan"
    assert by_type["core_only_zero_placeholder"]["core_min_score"] == "0"
    assert by_type["core_only_zero_placeholder"]["core_min_rank"] == "0"
    manifest = json.loads(Path(result["manifest"]).read_text(encoding="utf-8"))
    assert manifest["notes"].startswith("Review plan only")


def test_audit_score_history_reconciliation_plan_reports_progress(tmp_path: Path):
    plan = tmp_path / "score_history_reconciliation_plan.csv"
    fieldnames = [
        "task_id",
        "issue_type",
        "priority",
        "status",
        "suggested_action",
        "match_confidence",
        "score_year",
        "batch",
        "subject_cat",
        "school_code",
        "package_major_code",
        "core_major_code",
        "package_min_score",
        "core_min_score",
        "package_min_rank",
        "core_min_rank",
        "package_key_json",
        "core_key_json",
        "core_candidates_json",
        "matching_values_json",
        "differences_json",
        "review_decision",
        "reviewer",
        "reviewed_at",
        "notes",
    ]
    with plan.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerow({
            "task_id": "t1",
            "issue_type": "major_code_drift_candidate",
            "priority": "1",
            "status": "reviewed",
            "suggested_action": "review_major_code_alignment",
            "match_confidence": "high",
            "score_year": "2024",
            "batch": "本科批",
            "subject_cat": "物理类",
            "school_code": "1003",
            "package_major_code": "04",
            "core_major_code": "03",
            "package_min_score": "570",
            "core_min_score": "570",
            "package_min_rank": "3000",
            "core_min_rank": "3000",
            "package_key_json": "{}",
            "core_key_json": "{}",
            "core_candidates_json": "[]",
            "matching_values_json": "{}",
            "differences_json": "[]",
            "review_decision": "map_package_to_core_major_code",
            "reviewer": "tester",
            "reviewed_at": "2026-05-13",
            "notes": "fixture",
        })
        writer.writerow({
            "task_id": "t2",
            "issue_type": "value_drift",
            "priority": "2",
            "status": "todo",
            "suggested_action": "review_source_value_conflict",
            "match_confidence": "primary_key_match",
            "score_year": "2024",
            "batch": "本科批",
            "subject_cat": "物理类",
            "school_code": "1002",
            "package_major_code": "02",
            "core_major_code": "02",
            "package_min_score": "580",
            "core_min_score": "580",
            "package_min_rank": "1990",
            "core_min_rank": "2000",
            "package_key_json": "{}",
            "core_key_json": "{}",
            "core_candidates_json": "[]",
            "matching_values_json": "{}",
            "differences_json": "[]",
            "review_decision": "",
            "reviewer": "",
            "reviewed_at": "",
            "notes": "",
        })
        writer.writerow({
            "task_id": "t3",
            "issue_type": "value_drift",
            "priority": "2",
            "status": "todo",
            "suggested_action": "review_source_value_conflict",
            "match_confidence": "primary_key_match",
            "score_year": "2023",
            "batch": "本科批",
            "subject_cat": "历史类",
            "school_code": "1001",
            "package_major_code": "03",
            "core_major_code": "03",
            "package_min_score": "520",
            "core_min_score": "",
            "package_min_rank": "12000",
            "core_min_rank": "0",
            "package_key_json": "{}",
            "core_key_json": "{}",
            "core_candidates_json": "[]",
            "matching_values_json": "{}",
            "differences_json": "[]",
            "review_decision": "",
            "reviewer": "",
            "reviewed_at": "",
            "notes": "",
        })

    report = audit_score_history_reconciliation_plan(plan)

    assert report["errors"] == []
    assert report["rows"] == 3
    assert report["status_counts"] == {"reviewed": 1, "todo": 2}
    assert report["decision_counts"] == {"map_package_to_core_major_code": 1}
    assert report["progress"]["ready_rows"] == 1
    assert report["progress"]["pending_rows"] == 2
    assert report["pending_diagnostics"]["subject_counts"] == [
        {"issue_type": "value_drift", "subject_cat": "物理类", "rows": 1},
        {"issue_type": "value_drift", "subject_cat": "历史类", "rows": 1},
    ]
    assert report["pending_diagnostics"]["candidate_count_counts"] == [
        {"issue_type": "value_drift", "core_candidate_count": 0, "rows": 2}
    ]
    assert report["pending_diagnostics"]["top_school_counts"] == [
        {"issue_type": "value_drift", "school_code": "1002", "rows": 1},
        {"issue_type": "value_drift", "school_code": "1001", "rows": 1},
    ]
    value_drift = report["pending_diagnostics"]["value_drift"]
    assert value_drift["rows"] == 2
    assert value_drift["year_counts"] == [
        {"score_year": "2023", "rows": 1},
        {"score_year": "2024", "rows": 1},
    ]
    assert value_drift["score_delta_buckets"] == {"0": 1, "core_missing": 1}
    assert value_drift["rank_delta_buckets"] == {"<= 100": 1, "> 1000": 1}
    assert value_drift["delta_bucket_counts"] == [
        {
            "score_year": "2024",
            "subject_cat": "物理类",
            "score_delta_bucket": "0",
            "rank_delta_bucket": "<= 100",
            "rows": 1,
        },
        {
            "score_year": "2023",
            "subject_cat": "历史类",
            "score_delta_bucket": "core_missing",
            "rank_delta_bucket": "> 1000",
            "rows": 1,
        },
    ]
    assert value_drift["core_blank_or_zero_counts"] == {
        "core_min_rank_blank_or_zero": 1,
        "core_min_score_blank_or_zero": 1,
    }
    assert report["ready"]["review_complete"] is False
    assert report["ready"]["package_ready"] is False


def test_audit_score_history_reconciliation_plan_blocks_source_research_decision(tmp_path: Path):
    plan = tmp_path / "score_history_reconciliation_plan.csv"
    with plan.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=PLAN_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerow({
            "task_id": "research-1",
            "issue_type": "value_drift",
            "priority": "2",
            "status": "reviewed",
            "suggested_action": "review_source_value_conflict",
            "match_confidence": "primary_key_match",
            "score_year": "2024",
            "batch": "本科批",
            "subject_cat": "物理类",
            "school_code": "1002",
            "package_major_code": "02",
            "core_major_code": "02",
            "package_min_score": "580",
            "core_min_score": "580",
            "package_min_rank": "1990",
            "core_min_rank": "2000",
            "package_key_json": "{}",
            "core_key_json": "{}",
            "core_candidates_json": "[]",
            "matching_values_json": "{}",
            "differences_json": "[]",
            "review_decision": "needs_source_research",
            "reviewer": "tester",
            "reviewed_at": "2026-05-13",
            "notes": "needs official source check",
        })

    report = audit_score_history_reconciliation_plan(plan)

    assert report["errors"] == []
    assert report["ready"]["review_complete"] is True
    assert report["progress"]["blocking_decision_rows"] == 1
    assert report["ready"]["package_ready"] is False


def test_audit_score_history_reconciliation_plan_rejects_decision_side_mismatch(tmp_path: Path):
    plan = tmp_path / "score_history_reconciliation_plan.csv"
    rows = [
        {
            "task_id": "core-only-as-package",
            "issue_type": "core_only_unmatched",
            "priority": "4",
            "status": "reviewed",
            "suggested_action": "review_core_only_row",
            "match_confidence": "none",
            "score_year": "2024",
            "batch": "本科批",
            "subject_cat": "物理类",
            "school_code": "1001",
            "core_major_code": "01",
            "core_min_score": "500",
            "core_min_rank": "12000",
            "package_key_json": "{}",
            "core_key_json": '{"school_code":"1001","major_code":"01"}',
            "core_candidates_json": "[]",
            "matching_values_json": "{}",
            "differences_json": "[]",
            "review_decision": "use_package_row",
            "reviewer": "tester",
            "reviewed_at": "2026-05-14",
        },
        {
            "task_id": "package-only-as-core",
            "issue_type": "package_only_unmatched",
            "priority": "3",
            "status": "reviewed",
            "suggested_action": "review_package_only_row",
            "match_confidence": "none",
            "score_year": "2024",
            "batch": "本科批",
            "subject_cat": "历史类",
            "school_code": "1002",
            "package_major_code": "02",
            "package_min_score": "510",
            "package_min_rank": "9000",
            "package_key_json": '{"school_code":"1002","major_code":"02"}',
            "core_key_json": "{}",
            "core_candidates_json": "[]",
            "matching_values_json": "{}",
            "differences_json": "[]",
            "review_decision": "keep_core_row",
            "reviewer": "tester",
            "reviewed_at": "2026-05-14",
        },
        {
            "task_id": "package-only-as-map",
            "issue_type": "major_code_drift_candidate",
            "priority": "1",
            "status": "reviewed",
            "suggested_action": "review_major_code_alignment",
            "match_confidence": "high",
            "score_year": "2024",
            "batch": "本科批",
            "subject_cat": "物理类",
            "school_code": "1003",
            "package_major_code": "03",
            "package_min_score": "520",
            "package_min_rank": "8000",
            "package_key_json": '{"school_code":"1003","major_code":"03"}',
            "core_key_json": "{}",
            "core_candidates_json": "[]",
            "matching_values_json": "{}",
            "differences_json": "[]",
            "review_decision": "map_package_to_core_major_code",
            "reviewer": "tester",
            "reviewed_at": "2026-05-14",
        },
    ]
    with plan.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=PLAN_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)

    report = audit_score_history_reconciliation_plan(plan)

    assert "row 2 use_package_row without package side" in report["errors"]
    assert "row 3 keep_core_row without core side" in report["errors"]
    assert "row 4 map_package_to_core_major_code without core side" in report["errors"]


def test_apply_score_history_reconciliation_auto_decisions_marks_zero_placeholders(tmp_path: Path):
    plan = tmp_path / "score_history_reconciliation_plan.csv"
    rows = [
        {
            "task_id": "zero-1",
            "issue_type": "core_only_zero_placeholder",
            "priority": "5",
            "status": "todo",
            "suggested_action": "review_core_zero_placeholder_for_delete_plan",
            "match_confidence": "zero_placeholder",
            "score_year": "2024",
            "batch": "本科批",
            "subject_cat": "物理类",
            "school_code": "1001",
            "package_major_code": "",
            "core_major_code": "01",
            "package_min_score": "",
            "core_min_score": "0",
            "package_min_rank": "",
            "core_min_rank": "0",
            "package_key_json": "{}",
            "core_key_json": "{\"school_code\":\"1001\"}",
            "core_candidates_json": "[]",
            "matching_values_json": "{}",
            "differences_json": "[]",
            "review_decision": "",
            "reviewer": "",
            "reviewed_at": "",
            "notes": "",
        },
        {
            "task_id": "nonzero-1",
            "issue_type": "core_only_unmatched",
            "priority": "4",
            "status": "todo",
            "suggested_action": "review_core_only_row",
            "match_confidence": "none",
            "score_year": "2024",
            "batch": "本科批",
            "subject_cat": "物理类",
            "school_code": "1002",
            "package_major_code": "",
            "core_major_code": "02",
            "package_min_score": "",
            "core_min_score": "510",
            "package_min_rank": "",
            "core_min_rank": "40000",
            "package_key_json": "{}",
            "core_key_json": "{\"school_code\":\"1002\"}",
            "core_candidates_json": "[]",
            "matching_values_json": "{}",
            "differences_json": "[]",
            "review_decision": "",
            "reviewer": "",
            "reviewed_at": "",
            "notes": "",
        },
    ]
    with plan.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=PLAN_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)

    output = tmp_path / "score_history_reconciliation_plan_auto.csv"
    report = apply_score_history_reconciliation_auto_decisions(plan_csv=plan, output=output)

    assert report["updated_rows"] == 1
    assert report["rule_counts"] == {"core_zero_placeholder_to_delete_plan": 1}
    with output.open(encoding="utf-8", newline="") as f:
        by_id = {row["task_id"]: row for row in csv.DictReader(f)}
    zero_row = by_id["zero-1"]
    assert zero_row["status"] == "reviewed"
    assert zero_row["review_decision"] == "exclude_row"
    assert zero_row["reviewer"] == "datahub_auto_rule"
    assert "auto_rule=core_zero_placeholder_to_delete_plan" in zero_row["notes"]
    assert by_id["nonzero-1"]["status"] == "todo"

    audit = audit_score_history_reconciliation_plan(output)
    assert audit["errors"] == []
    assert audit["progress"]["ready_rows"] == 1
    assert audit["progress"]["pending_rows"] == 1
    assert audit["ready"]["package_ready"] is False


def test_apply_score_history_major_name_reference_decisions_resolves_exact_candidate(tmp_path: Path):
    plan = tmp_path / "score_history_reconciliation_plan.csv"
    candidate_11 = {
        "key": {
            "score_year": 2022,
            "batch": "本科批",
            "subject_cat": "历史类",
            "school_code": "0651",
            "major_code": "11",
        },
        "variant_differences": [{"column": "major_code", "package_value": "0B", "core_value": "11"}],
    }
    candidate_13 = {
        "key": {
            "score_year": 2022,
            "batch": "本科批",
            "subject_cat": "历史类",
            "school_code": "0651",
            "major_code": "13",
        },
        "variant_differences": [{"column": "major_code", "package_value": "0B", "core_value": "13"}],
    }
    rows = [
        {
            "task_id": "drift-1",
            "issue_type": "major_code_drift_candidate",
            "priority": "1",
            "status": "todo",
            "suggested_action": "review_major_code_alignment",
            "match_confidence": "high",
            "score_year": "2022",
            "batch": "本科批",
            "subject_cat": "历史类",
            "school_code": "0651",
            "package_major_code": "0B",
            "core_major_code": "11|13",
            "package_min_score": "616",
            "core_min_score": "616",
            "package_min_rank": "578",
            "core_min_rank": "578",
            "package_key_json": "{}",
            "core_key_json": json.dumps(candidate_11["key"], ensure_ascii=False),
            "core_candidates_json": json.dumps([candidate_11, candidate_13], ensure_ascii=False),
            "matching_values_json": "{}",
            "differences_json": "[]",
            "review_decision": "",
            "reviewer": "",
            "reviewed_at": "",
            "notes": "",
        },
        {
            "task_id": "drift-no-match",
            "issue_type": "major_code_drift_candidate",
            "priority": "1",
            "status": "todo",
            "suggested_action": "review_major_code_alignment",
            "match_confidence": "high",
            "score_year": "2022",
            "batch": "本科批",
            "subject_cat": "历史类",
            "school_code": "0652",
            "package_major_code": "01",
            "core_major_code": "02|03",
            "package_min_score": "600",
            "core_min_score": "600",
            "package_min_rank": "1200",
            "core_min_rank": "1200",
            "package_key_json": "{}",
            "core_key_json": "{}",
            "core_candidates_json": json.dumps([
                {"key": {"score_year": 2022, "batch": "本科批", "subject_cat": "历史类", "school_code": "0652", "major_code": "02"}},
                {"key": {"score_year": 2022, "batch": "本科批", "subject_cat": "历史类", "school_code": "0652", "major_code": "03"}},
            ], ensure_ascii=False),
            "matching_values_json": "{}",
            "differences_json": "[]",
            "review_decision": "",
            "reviewer": "",
            "reviewed_at": "",
            "notes": "",
        },
    ]
    with plan.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=PLAN_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)

    projection = tmp_path / "ln_projection_score_2022_official.csv"
    with projection.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["school_code", "school_name", "major_code", "major_full", "batch", "subject_cat", "score_year", "min_score"],
        )
        writer.writeheader()
        writer.writerow({
            "school_code": "0651",
            "school_name": "西南财经大学",
            "major_code": "0B",
            "major_full": "工商管理类(会计学院)",
            "batch": "本科批",
            "subject_cat": "历史类",
            "score_year": "2022",
            "min_score": "616",
        })
        writer.writerow({
            "school_code": "0652",
            "school_name": "测试大学",
            "major_code": "01",
            "major_full": "没有匹配的专业",
            "batch": "本科批",
            "subject_cat": "历史类",
            "score_year": "2022",
            "min_score": "600",
        })

    core_db = tmp_path / "university.db"
    con = duckdb.connect(str(core_db))
    con.execute("""
        CREATE TABLE fa_dim_ln_admission_plan (
            school_code VARCHAR,
            major_code VARCHAR,
            subject_cat VARCHAR,
            batch VARCHAR,
            year INTEGER,
            major_full VARCHAR,
            major_short VARCHAR
        )
    """)
    con.executemany(
        "INSERT INTO fa_dim_ln_admission_plan VALUES (?, ?, ?, ?, ?, ?, ?)",
        [
            ("0651", "11", "历史类", "本科批", 2026, "工商管理类(会计学院)", ""),
            ("0651", "13", "历史类", "本科批", 2026, "会计学(双语实验班)", ""),
            ("0652", "02", "历史类", "本科批", 2026, "候选一", ""),
            ("0652", "03", "历史类", "本科批", 2026, "候选二", ""),
        ],
    )
    con.close()

    output = tmp_path / "score_history_reconciliation_plan_name_reference.csv"
    report = apply_score_history_major_name_reference_decisions(
        plan_csv=plan,
        projection_csv=projection,
        core_db=core_db,
        output=output,
        reviewed_at="2026-05-14",
    )

    assert report["updated_rows"] == 1
    assert report["match_counts"] == {"no_match": 1, "single_exact": 1}
    with output.open(encoding="utf-8", newline="") as f:
        by_id = {row["task_id"]: row for row in csv.DictReader(f)}
    resolved = by_id["drift-1"]
    assert resolved["status"] == "reviewed"
    assert resolved["review_decision"] == "map_package_to_core_major_code"
    assert resolved["reviewer"] == "datahub_major_name_reference"
    assert resolved["core_major_code"] == "11"
    assert json.loads(resolved["core_key_json"])["major_code"] == "11"
    assert len(json.loads(resolved["core_candidates_json"])) == 1
    assert "major_name_reference=exact" in resolved["notes"]
    assert by_id["drift-no-match"]["status"] == "todo"

    audit = audit_score_history_reconciliation_plan(output)
    assert audit["errors"] == []
    assert audit["progress"]["ready_rows"] == 1
    assert audit["progress"]["pending_rows"] == 1


def test_apply_score_history_pair_name_reference_decisions_maps_package_to_core_code(tmp_path: Path):
    plan = tmp_path / "score_history_reconciliation_plan.csv"
    package_key = {
        "score_year": 2022,
        "batch": "本科批",
        "subject_cat": "物理类",
        "school_code": "0162",
        "major_code": "H1",
    }
    core_key = {
        "score_year": 2022,
        "batch": "本科批",
        "subject_cat": "物理类",
        "school_code": "0162",
        "major_code": "0L",
    }
    rows = [
        {
            "task_id": "package-name-pair",
            "issue_type": "package_only_unmatched",
            "priority": "3",
            "status": "reviewed",
            "suggested_action": "review_package_only_row",
            "match_confidence": "none",
            "score_year": "2022",
            "batch": "本科批",
            "subject_cat": "物理类",
            "school_code": "0162",
            "package_major_code": "H1",
            "core_major_code": "",
            "package_min_score": "498",
            "core_min_score": "",
            "package_min_rank": "46215",
            "core_min_rank": "",
            "package_key_json": json.dumps(package_key, ensure_ascii=False),
            "core_key_json": "{}",
            "core_candidates_json": "[]",
            "matching_values_json": "{}",
            "differences_json": "[]",
            "review_decision": "use_package_row",
            "reviewer": "datahub_reference_rule",
            "reviewed_at": "2026-05-14",
            "notes": "official package row",
        },
        {
            "task_id": "core-name-pair",
            "issue_type": "core_only_unmatched",
            "priority": "4",
            "status": "todo",
            "suggested_action": "review_core_only_row",
            "match_confidence": "none",
            "score_year": "2022",
            "batch": "本科批",
            "subject_cat": "物理类",
            "school_code": "0162",
            "package_major_code": "",
            "core_major_code": "0L",
            "package_min_score": "",
            "core_min_score": "490",
            "package_min_rank": "",
            "core_min_rank": "50000",
            "package_key_json": "{}",
            "core_key_json": json.dumps(core_key, ensure_ascii=False),
            "core_candidates_json": "[]",
            "matching_values_json": "{}",
            "differences_json": "[]",
            "review_decision": "",
            "reviewer": "",
            "reviewed_at": "",
            "notes": "",
        },
    ]
    with plan.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=PLAN_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)

    projection = tmp_path / "ln_projection_score_2022_official.csv"
    with projection.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["school_code", "major_code", "major_full", "batch", "subject_cat", "score_year"],
        )
        writer.writeheader()
        writer.writerow({
            "school_code": "0162",
            "major_code": "H1",
            "major_full": "医学信息工程",
            "batch": "本科批",
            "subject_cat": "物理类",
            "score_year": "2022",
        })

    core_db = tmp_path / "university.db"
    con = duckdb.connect(str(core_db))
    con.execute("""
        CREATE TABLE fa_dim_ln_admission_plan (
            school_code VARCHAR,
            major_code VARCHAR,
            subject_cat VARCHAR,
            batch VARCHAR,
            year INTEGER,
            major_full VARCHAR,
            major_short VARCHAR
        )
    """)
    con.execute(
        "INSERT INTO fa_dim_ln_admission_plan VALUES (?, ?, ?, ?, ?, ?, ?)",
        ("0162", "0L", "物理类", "本科批", 2026, "医学信息工程", ""),
    )
    con.close()

    output = tmp_path / "score_history_reconciliation_plan_pair_name_reference.csv"
    report = apply_score_history_pair_name_reference_decisions(
        plan_csv=plan,
        projection_csv=projection,
        core_db=core_db,
        output=output,
        reviewed_at="2026-05-14",
    )

    assert report["updated_pairs"] == 1
    assert report["match_counts"] == {"single_exact_pair": 1}
    with output.open(encoding="utf-8", newline="") as f:
        by_id = {row["task_id"]: row for row in csv.DictReader(f)}
    package_row = by_id["package-name-pair"]
    core_row = by_id["core-name-pair"]
    assert package_row["review_decision"] == "map_package_to_core_major_code"
    assert package_row["core_major_code"] == "0L"
    assert json.loads(package_row["core_key_json"])["major_code"] == "0L"
    assert len(json.loads(package_row["core_candidates_json"])) == 1
    assert core_row["review_decision"] == "covered_by_mapped_package_row"

    audit = audit_score_history_reconciliation_plan(output)
    assert audit["errors"] == []
    assert audit["ready"]["package_ready"] is True

    package = build_score_history_package_from_reconciliation_plan(
        plan_csv=output,
        output_root=tmp_path / "exports",
        package_id="pkg-paired-score-history",
    )
    assert package["rows"] == 1
    assert package["skipped_rows"] == 1
    with (Path(package["package_dir"]) / "fa_fact_ln_score_history.csv").open(encoding="utf-8", newline="") as f:
        output_rows = list(csv.DictReader(f))
    assert output_rows[0]["major_code"] == "0L"
    assert output_rows[0]["min_score"] == "498"
    assert output_rows[0]["min_rank"] == "46215"

    delete_plan = build_score_history_delete_plan_from_reconciliation_plan(
        plan_csv=output,
        output_dir=tmp_path / "delete_plan",
    )
    assert delete_plan["rows"] == 0


def test_apply_score_history_pair_name_reference_remaps_value_drift_and_deletes_original_core(tmp_path: Path):
    plan = tmp_path / "score_history_reconciliation_plan.csv"
    old_core_key = {
        "score_year": 2022,
        "batch": "本科批",
        "subject_cat": "物理类",
        "school_code": "1607",
        "major_code": "0X",
    }
    target_core_key = {
        "score_year": 2022,
        "batch": "本科批",
        "subject_cat": "物理类",
        "school_code": "1607",
        "major_code": "0G",
    }
    rows = [
        {
            "task_id": "value-existing-package-key",
            "issue_type": "value_drift",
            "priority": "2",
            "status": "reviewed",
            "suggested_action": "review_source_value_conflict",
            "match_confidence": "primary_key_match",
            "score_year": "2022",
            "batch": "本科批",
            "subject_cat": "物理类",
            "school_code": "1607",
            "package_major_code": "0X",
            "core_major_code": "0X",
            "package_min_score": "500",
            "core_min_score": "499",
            "package_min_rank": "30000",
            "core_min_rank": "30100",
            "package_key_json": json.dumps(old_core_key, ensure_ascii=False),
            "core_key_json": json.dumps(old_core_key, ensure_ascii=False),
            "core_candidates_json": "[]",
            "matching_values_json": "{}",
            "differences_json": "[]",
            "review_decision": "use_package_row",
            "reviewer": "tester",
            "reviewed_at": "2026-05-14",
            "notes": "",
        },
        {
            "task_id": "core-only-same-name",
            "issue_type": "core_only_unmatched",
            "priority": "4",
            "status": "todo",
            "suggested_action": "review_core_only_row",
            "match_confidence": "none",
            "score_year": "2022",
            "batch": "本科批",
            "subject_cat": "物理类",
            "school_code": "1607",
            "package_major_code": "",
            "core_major_code": "0G",
            "package_min_score": "",
            "core_min_score": "500",
            "package_min_rank": "",
            "core_min_rank": "30000",
            "package_key_json": "{}",
            "core_key_json": json.dumps(target_core_key, ensure_ascii=False),
            "core_candidates_json": "[]",
            "matching_values_json": "{}",
            "differences_json": "[]",
            "review_decision": "",
            "reviewer": "",
            "reviewed_at": "",
            "notes": "",
        },
    ]
    with plan.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=PLAN_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)

    projection = tmp_path / "ln_projection_score_2022_official.csv"
    with projection.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["school_code", "major_code", "major_full", "batch", "subject_cat", "score_year"],
        )
        writer.writeheader()
        writer.writerow({
            "school_code": "1607",
            "major_code": "0X",
            "major_full": "自动化",
            "batch": "本科批",
            "subject_cat": "物理类",
            "score_year": "2022",
        })

    core_db = tmp_path / "university.db"
    con = duckdb.connect(str(core_db))
    con.execute("""
        CREATE TABLE fa_dim_ln_admission_plan (
            school_code VARCHAR,
            major_code VARCHAR,
            subject_cat VARCHAR,
            batch VARCHAR,
            year INTEGER,
            major_full VARCHAR,
            major_short VARCHAR
        )
    """)
    con.execute(
        "INSERT INTO fa_dim_ln_admission_plan VALUES (?, ?, ?, ?, ?, ?, ?)",
        ("1607", "0G", "物理类", "本科批", 2026, "自动化", ""),
    )
    con.close()

    output = tmp_path / "pair_value_drift.csv"
    report = apply_score_history_pair_name_reference_decisions(
        plan_csv=plan,
        projection_csv=projection,
        core_db=core_db,
        output=output,
        reviewed_at="2026-05-14",
    )

    assert report["updated_pairs"] == 1
    assert report["match_counts"] == {"single_exact_value_drift_pair": 1}
    with output.open(encoding="utf-8", newline="") as f:
        by_id = {row["task_id"]: row for row in csv.DictReader(f)}
    value_row = by_id["value-existing-package-key"]
    core_row = by_id["core-only-same-name"]
    assert value_row["review_decision"] == "map_package_to_core_major_code_delete_original_core"
    assert json.loads(value_row["core_key_json"])["major_code"] == "0X"
    assert json.loads(value_row["core_candidates_json"])[0]["key"]["major_code"] == "0G"
    assert value_row["core_major_code"] == "0G"
    assert core_row["review_decision"] == "covered_by_mapped_package_row"

    audit = audit_score_history_reconciliation_plan(output)
    assert audit["errors"] == []
    assert audit["ready"]["package_ready"] is True

    package = build_score_history_package_from_reconciliation_plan(
        plan_csv=output,
        output_root=tmp_path / "exports",
        package_id="pkg-value-drift-paired-score-history",
    )
    assert package["rows"] == 1
    assert package["skipped_rows"] == 1
    with (Path(package["package_dir"]) / "fa_fact_ln_score_history.csv").open(encoding="utf-8", newline="") as f:
        output_rows = list(csv.DictReader(f))
    assert output_rows[0]["major_code"] == "0G"
    assert output_rows[0]["min_score"] == "500"
    assert output_rows[0]["min_rank"] == "30000"

    delete_plan = build_score_history_delete_plan_from_reconciliation_plan(
        plan_csv=output,
        output_dir=tmp_path / "delete_plan",
    )
    assert delete_plan["rows"] == 1
    with Path(delete_plan["csv"]).open(encoding="utf-8", newline="") as f:
        delete_rows = list(csv.DictReader(f))
    assert delete_rows[0]["major_code"] == "0X"


def test_apply_score_history_reconciliation_auto_decisions_uses_reference_package(tmp_path: Path):
    plan = tmp_path / "score_history_reconciliation_plan.csv"
    rows = [
        {
            "task_id": "core-official",
            "issue_type": "core_only_unmatched",
            "priority": "4",
            "status": "todo",
            "suggested_action": "review_core_only_row",
            "match_confidence": "none",
            "score_year": "2025",
            "batch": "本科批",
            "subject_cat": "物理类",
            "school_code": "1001",
            "package_major_code": "",
            "core_major_code": "91",
            "package_min_score": "",
            "core_min_score": "521",
            "package_min_rank": "",
            "core_min_rank": "11362",
            "package_key_json": "{}",
            "core_key_json": "{}",
            "core_candidates_json": "[]",
            "matching_values_json": "{}",
            "differences_json": "[]",
            "review_decision": "",
            "reviewer": "",
            "reviewed_at": "",
            "notes": "",
        },
        {
            "task_id": "package-official",
            "issue_type": "package_only_unmatched",
            "priority": "3",
            "status": "todo",
            "suggested_action": "review_package_only_row",
            "match_confidence": "none",
            "score_year": "2025",
            "batch": "本科批",
            "subject_cat": "历史类",
            "school_code": "1002",
            "package_major_code": "NA",
            "core_major_code": "",
            "package_min_score": "453",
            "core_min_score": "",
            "package_min_rank": "80263",
            "core_min_rank": "",
            "package_key_json": "{}",
            "core_key_json": "{}",
            "core_candidates_json": "[]",
            "matching_values_json": "{}",
            "differences_json": "[]",
            "review_decision": "",
            "reviewer": "",
            "reviewed_at": "",
            "notes": "",
        },
        {
            "task_id": "core-not-official",
            "issue_type": "core_only_unmatched",
            "priority": "4",
            "status": "todo",
            "suggested_action": "review_core_only_row",
            "match_confidence": "none",
            "score_year": "2025",
            "batch": "本科批",
            "subject_cat": "物理类",
            "school_code": "1003",
            "package_major_code": "",
            "core_major_code": "92",
            "package_min_score": "",
            "core_min_score": "500",
            "package_min_rank": "",
            "core_min_rank": "20000",
            "package_key_json": "{}",
            "core_key_json": "{}",
            "core_candidates_json": "[]",
            "matching_values_json": "{}",
            "differences_json": "[]",
            "review_decision": "",
            "reviewer": "",
            "reviewed_at": "",
            "notes": "",
        },
        {
            "task_id": "major-single-official",
            "issue_type": "major_code_drift_candidate",
            "priority": "1",
            "status": "todo",
            "suggested_action": "review_major_code_alignment",
            "match_confidence": "high",
            "score_year": "2025",
            "batch": "本科批",
            "subject_cat": "物理类",
            "school_code": "1004",
            "package_major_code": "35",
            "core_major_code": "26",
            "package_min_score": "473",
            "core_min_score": "473",
            "package_min_rank": "57476",
            "core_min_rank": "57476",
            "package_key_json": "{}",
            "core_key_json": "{}",
            "core_candidates_json": json.dumps([{"key": {"school_code": "1004", "major_code": "26"}}], ensure_ascii=False),
            "matching_values_json": "{}",
            "differences_json": "[]",
            "review_decision": "",
            "reviewer": "",
            "reviewed_at": "",
            "notes": "",
        },
        {
            "task_id": "major-multi-official",
            "issue_type": "major_code_drift_candidate",
            "priority": "1",
            "status": "todo",
            "suggested_action": "review_major_code_alignment",
            "match_confidence": "high",
            "score_year": "2025",
            "batch": "本科批",
            "subject_cat": "物理类",
            "school_code": "1005",
            "package_major_code": "35",
            "core_major_code": "26|27",
            "package_min_score": "473",
            "core_min_score": "473",
            "package_min_rank": "57476",
            "core_min_rank": "57476",
            "package_key_json": "{}",
            "core_key_json": "{}",
            "core_candidates_json": json.dumps([
                {"key": {"school_code": "1005", "major_code": "26"}},
                {"key": {"school_code": "1005", "major_code": "27"}},
            ], ensure_ascii=False),
            "matching_values_json": "{}",
            "differences_json": "[]",
            "review_decision": "",
            "reviewer": "",
            "reviewed_at": "",
            "notes": "",
        },
    ]
    with plan.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=PLAN_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    reference_dir = tmp_path / "official_reference"
    reference_dir.mkdir()
    with (reference_dir / "fa_fact_ln_score_history.csv").open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["school_code", "major_code", "batch", "subject_cat", "score_year", "min_score", "min_rank"],
        )
        writer.writeheader()
        writer.writerows([
            {
                "school_code": "1001",
                "major_code": "91",
                "batch": "本科批",
                "subject_cat": "物理类",
                "score_year": "2025",
                "min_score": "521",
                "min_rank": "11362",
            },
            {
                "school_code": "1002",
                "major_code": "NA",
                "batch": "本科批",
                "subject_cat": "历史类",
                "score_year": "2025",
                "min_score": "453",
                "min_rank": "80263",
            },
            {
                "school_code": "1004",
                "major_code": "35",
                "batch": "本科批",
                "subject_cat": "物理类",
                "score_year": "2025",
                "min_score": "473",
                "min_rank": "57476",
            },
            {
                "school_code": "1005",
                "major_code": "35",
                "batch": "本科批",
                "subject_cat": "物理类",
                "score_year": "2025",
                "min_score": "473",
                "min_rank": "57476",
            },
        ])
    (reference_dir / "quality_report.json").write_text('{"errors":[],"warnings":[]}\n', encoding="utf-8")
    (reference_dir / "manifest.json").write_text(json.dumps({
        "package_id": "official-reference",
        "built_at": "2026-05-14T00:00:00",
        "source_version": "official-reference-fixture",
        "tables": [{"name": "fa_fact_ln_score_history", "file": "fa_fact_ln_score_history.csv"}],
        "files": ["fa_fact_ln_score_history.csv"],
        "hashes": {},
        "quality_report": "quality_report.json",
    }), encoding="utf-8")

    output = tmp_path / "score_history_reconciliation_plan_reference.csv"
    report = apply_score_history_reconciliation_auto_decisions(
        plan_csv=plan,
        output=output,
        reference_package_dirs=[reference_dir],
    )

    assert report["updated_rows"] == 3
    assert report["reference_rows"] == 4
    assert report["rule_counts"] == {
        "official_reference_maps_single_major_code_drift": 1,
        "official_reference_keeps_core_only_row": 1,
        "official_reference_uses_package_only_row": 1,
    }
    with output.open(encoding="utf-8", newline="") as f:
        by_id = {row["task_id"]: row for row in csv.DictReader(f)}
    assert by_id["core-official"]["review_decision"] == "keep_core_row"
    assert by_id["core-official"]["reviewer"] == "datahub_reference_rule"
    assert by_id["package-official"]["review_decision"] == "use_package_row"
    assert by_id["major-single-official"]["review_decision"] == "map_package_to_core_major_code"
    assert by_id["major-multi-official"]["status"] == "todo"
    assert by_id["core-not-official"]["status"] == "todo"

    bad_reference_dir = tmp_path / "bad_official_reference"
    bad_reference_dir.mkdir()
    (bad_reference_dir / "fa_fact_ln_score_history.csv").write_text(
        (reference_dir / "fa_fact_ln_score_history.csv").read_text(encoding="utf-8"),
        encoding="utf-8",
    )
    (bad_reference_dir / "quality_report.json").write_text(
        '{"errors":["reference source review is not complete"],"warnings":[]}\n',
        encoding="utf-8",
    )
    (bad_reference_dir / "manifest.json").write_text(json.dumps({
        "package_id": "bad-official-reference",
        "built_at": "2026-05-14T00:00:00",
        "source_version": "bad-official-reference-fixture",
        "tables": [{"name": "fa_fact_ln_score_history", "file": "fa_fact_ln_score_history.csv"}],
        "files": ["fa_fact_ln_score_history.csv"],
        "hashes": {},
        "quality_report": "quality_report.json",
    }), encoding="utf-8")
    try:
        apply_score_history_reconciliation_auto_decisions(
            plan_csv=plan,
            output=tmp_path / "bad_reference_output.csv",
            reference_package_dirs=[bad_reference_dir],
        )
        rejected_bad_reference = False
    except ValueError as exc:
        rejected_bad_reference = (
            "reference package manifest errors" in str(exc)
            and "quality_report has errors" in str(exc)
        )
    assert rejected_bad_reference


def test_build_score_history_reconciliation_review_batch_limits_pending_rows(tmp_path: Path):
    plan = tmp_path / "score_history_reconciliation_plan.csv"
    rows = []
    for index in range(2):
        rows.append({
            "task_id": f"major-{index}",
            "issue_type": "major_code_drift_candidate",
            "priority": "1",
            "status": "todo",
            "suggested_action": "review_major_code_alignment",
            "match_confidence": "high",
            "score_year": "2024",
            "batch": "本科批",
            "subject_cat": "物理类",
            "school_code": f"100{index}",
            "package_major_code": "04",
            "core_major_code": "03",
            "package_min_score": "570",
            "core_min_score": "570",
            "package_min_rank": "3000",
            "core_min_rank": "3000",
            "package_key_json": "{}",
            "core_key_json": "{}",
            "core_candidates_json": "[]",
            "matching_values_json": "{}",
            "differences_json": "[]",
            "review_decision": "",
            "reviewer": "",
            "reviewed_at": "",
            "notes": "",
        })
    rows.append({
        "task_id": "value-1",
        "issue_type": "value_drift",
        "priority": "2",
        "status": "todo",
        "suggested_action": "review_source_value_conflict",
        "match_confidence": "primary_key_match",
        "score_year": "2024",
        "batch": "本科批",
        "subject_cat": "物理类",
        "school_code": "2001",
        "package_major_code": "02",
        "core_major_code": "02",
        "package_min_score": "580",
        "core_min_score": "580",
        "package_min_rank": "1990",
        "core_min_rank": "2000",
        "package_key_json": "{}",
        "core_key_json": "{}",
        "core_candidates_json": "[]",
        "matching_values_json": "{}",
        "differences_json": "[]",
        "review_decision": "",
        "reviewer": "",
        "reviewed_at": "",
        "notes": "",
    })
    rows.append({
        **rows[-1],
        "task_id": "value-reviewed",
        "status": "reviewed",
        "review_decision": "use_package_row",
        "reviewer": "tester",
        "reviewed_at": "2026-05-13",
    })
    with plan.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=PLAN_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)

    result = build_score_history_reconciliation_review_batch(
        plan_csv=plan,
        output_dir=tmp_path / "batch",
        limit_per_issue=1,
    )

    assert result["rows"] == 2
    assert result["issue_counts"] == {"major_code_drift_candidate": 1, "value_drift": 1}
    with Path(result["csv"]).open(encoding="utf-8", newline="") as f:
        batch_rows = list(csv.DictReader(f))
    assert {row["status"] for row in batch_rows} == {"todo"}
    assert [row["issue_type"] for row in batch_rows] == ["major_code_drift_candidate", "value_drift"]
    value_row = [row for row in batch_rows if row["issue_type"] == "value_drift"][0]
    assert value_row["score_delta"] == "0"
    assert value_row["rank_delta"] == "-10"
    assert value_row["score_delta_bucket"] == "0"
    assert value_row["rank_delta_bucket"] == "<= 100"
    assert value_row["core_score_state"] == "present"
    assert value_row["core_rank_state"] == "present"
    manifest = json.loads(Path(result["manifest"]).read_text(encoding="utf-8"))
    assert manifest["value_drift_context"]["rows"] == 1
    assert manifest["value_drift_context"]["score_delta_buckets"] == {"0": 1}


def test_build_score_history_reconciliation_review_batch_filters_score_year(tmp_path: Path):
    plan = tmp_path / "score_history_reconciliation_plan.csv"
    rows = []
    for year in ("2023", "2024"):
        rows.append({
            "task_id": f"major-{year}",
            "issue_type": "major_code_drift_candidate",
            "priority": "1",
            "status": "todo",
            "suggested_action": "review_major_code_alignment",
            "match_confidence": "high",
            "score_year": year,
            "batch": "本科批",
            "subject_cat": "物理类",
            "school_code": "1001",
            "package_major_code": "04",
            "core_major_code": "03",
            "package_min_score": "570",
            "core_min_score": "570",
            "package_min_rank": "3000",
            "core_min_rank": "3000",
            "package_key_json": "{}",
            "core_key_json": "{}",
            "core_candidates_json": "[]",
            "matching_values_json": "{}",
            "differences_json": "[]",
            "review_decision": "",
            "reviewer": "",
            "reviewed_at": "",
            "notes": "",
        })
    with plan.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=PLAN_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)

    result = build_score_history_reconciliation_review_batch(
        plan_csv=plan,
        output_dir=tmp_path / "batch_2024",
        limit_per_issue=10,
        score_year=2024,
    )

    assert result["rows"] == 1
    assert result["score_year"] == 2024
    with Path(result["csv"]).open(encoding="utf-8", newline="") as f:
        batch_rows = list(csv.DictReader(f))
    assert [row["score_year"] for row in batch_rows] == ["2024"]
    manifest = json.loads(Path(result["manifest"]).read_text(encoding="utf-8"))
    assert manifest["score_year"] == 2024


def test_build_score_history_reconciliation_review_batch_filters_subject_cat(tmp_path: Path):
    plan = tmp_path / "score_history_reconciliation_plan.csv"
    rows = []
    for task_id, subject_cat in [
        ("physics", "物理类"),
        ("history", "历史类"),
    ]:
        rows.append({
            "task_id": task_id,
            "issue_type": "value_drift",
            "priority": "2",
            "status": "todo",
            "suggested_action": "review_source_value_conflict",
            "match_confidence": "primary_key_match",
            "score_year": "2024",
            "batch": "本科批",
            "subject_cat": subject_cat,
            "school_code": "1001",
            "package_major_code": "04",
            "core_major_code": "04",
            "package_min_score": "570",
            "core_min_score": "520",
            "package_min_rank": "3000",
            "core_min_rank": "12000",
            "package_key_json": "{}",
            "core_key_json": "{}",
            "core_candidates_json": "[]",
            "matching_values_json": "{}",
            "differences_json": "[]",
            "review_decision": "",
            "reviewer": "",
            "reviewed_at": "",
            "notes": "",
        })
    with plan.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=PLAN_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)

    result = build_score_history_reconciliation_review_batch(
        plan_csv=plan,
        output_dir=tmp_path / "batch_physics",
        limit_per_issue=10,
        subject_cat="物理类",
    )

    assert result["rows"] == 1
    assert result["subject_cat"] == "物理类"
    with Path(result["csv"]).open(encoding="utf-8", newline="") as f:
        batch_rows = list(csv.DictReader(f))
    assert [row["task_id"] for row in batch_rows] == ["physics"]
    assert batch_rows[0]["subject_cat"] == "物理类"
    manifest = json.loads(Path(result["manifest"]).read_text(encoding="utf-8"))
    assert manifest["subject_cat"] == "物理类"


def test_build_score_history_reconciliation_review_batch_filters_school_code(tmp_path: Path):
    plan = tmp_path / "score_history_reconciliation_plan.csv"
    rows = []
    for task_id, school_code in [
        ("target", "0183"),
        ("other", "0140"),
    ]:
        rows.append({
            "task_id": task_id,
            "issue_type": "value_drift",
            "priority": "2",
            "status": "todo",
            "suggested_action": "review_source_value_conflict",
            "match_confidence": "primary_key_match",
            "score_year": "2024",
            "batch": "本科批",
            "subject_cat": "物理类",
            "school_code": school_code,
            "package_major_code": "04",
            "core_major_code": "04",
            "package_min_score": "570",
            "core_min_score": "520",
            "package_min_rank": "3000",
            "core_min_rank": "12000",
            "package_key_json": "{}",
            "core_key_json": "{}",
            "core_candidates_json": "[]",
            "matching_values_json": "{}",
            "differences_json": "[]",
            "review_decision": "",
            "reviewer": "",
            "reviewed_at": "",
            "notes": "",
        })
    with plan.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=PLAN_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)

    result = build_score_history_reconciliation_review_batch(
        plan_csv=plan,
        output_dir=tmp_path / "batch_school",
        limit_per_issue=10,
        school_code="0183",
    )

    assert result["rows"] == 1
    assert result["school_code"] == "0183"
    with Path(result["csv"]).open(encoding="utf-8", newline="") as f:
        batch_rows = list(csv.DictReader(f))
    assert [row["task_id"] for row in batch_rows] == ["target"]
    assert batch_rows[0]["school_code"] == "0183"
    manifest = json.loads(Path(result["manifest"]).read_text(encoding="utf-8"))
    assert manifest["school_code"] == "0183"


def test_build_score_history_reconciliation_review_batch_filters_value_drift_core_state(tmp_path: Path):
    plan = tmp_path / "score_history_reconciliation_plan.csv"
    rows = []
    for task_id, core_score, core_rank in [
        ("missing", "", "0"),
        ("present", "520", "12000"),
    ]:
        rows.append({
            "task_id": task_id,
            "issue_type": "value_drift",
            "priority": "2",
            "status": "todo",
            "suggested_action": "review_source_value_conflict",
            "match_confidence": "primary_key_match",
            "score_year": "2024",
            "batch": "本科批",
            "subject_cat": "物理类",
            "school_code": "1001",
            "package_major_code": "04",
            "core_major_code": "04",
            "package_min_score": "570",
            "core_min_score": core_score,
            "package_min_rank": "3000",
            "core_min_rank": core_rank,
            "package_key_json": "{}",
            "core_key_json": "{}",
            "core_candidates_json": "[]",
            "matching_values_json": "{}",
            "differences_json": "[]",
            "review_decision": "",
            "reviewer": "",
            "reviewed_at": "",
            "notes": "",
        })
    with plan.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=PLAN_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)

    result = build_score_history_reconciliation_review_batch(
        plan_csv=plan,
        output_dir=tmp_path / "batch_core_missing",
        limit_per_issue=10,
        value_drift_core_state="missing_or_zero",
    )

    assert result["rows"] == 1
    assert result["value_drift_core_state"] == "missing_or_zero"
    with Path(result["csv"]).open(encoding="utf-8", newline="") as f:
        batch_rows = list(csv.DictReader(f))
    assert [row["task_id"] for row in batch_rows] == ["missing"]
    assert batch_rows[0]["core_score_state"] == "missing"
    assert batch_rows[0]["core_rank_state"] == "zero"


def test_build_score_history_reconciliation_review_batch_filters_value_drift_delta_buckets(tmp_path: Path):
    plan = tmp_path / "score_history_reconciliation_plan.csv"
    rows = []
    for task_id, package_score, core_score, package_rank, core_rank in [
        ("small", "580", "579", "1990", "2000"),
        ("large", "570", "520", "3000", "12000"),
    ]:
        rows.append({
            "task_id": task_id,
            "issue_type": "value_drift",
            "priority": "2",
            "status": "todo",
            "suggested_action": "review_source_value_conflict",
            "match_confidence": "primary_key_match",
            "score_year": "2024",
            "batch": "本科批",
            "subject_cat": "物理类",
            "school_code": "1001",
            "package_major_code": "04",
            "core_major_code": "04",
            "package_min_score": package_score,
            "core_min_score": core_score,
            "package_min_rank": package_rank,
            "core_min_rank": core_rank,
            "package_key_json": "{}",
            "core_key_json": "{}",
            "core_candidates_json": "[]",
            "matching_values_json": "{}",
            "differences_json": "[]",
            "review_decision": "",
            "reviewer": "",
            "reviewed_at": "",
            "notes": "",
        })
    with plan.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=PLAN_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)

    result = build_score_history_reconciliation_review_batch(
        plan_csv=plan,
        output_dir=tmp_path / "batch_small_delta",
        limit_per_issue=10,
        value_drift_score_delta_bucket="<=1",
        value_drift_rank_delta_bucket="<=100",
    )

    assert result["rows"] == 1
    assert result["value_drift_score_delta_bucket"] == "<= 1"
    assert result["value_drift_rank_delta_bucket"] == "<= 100"
    with Path(result["csv"]).open(encoding="utf-8", newline="") as f:
        batch_rows = list(csv.DictReader(f))
    assert [row["task_id"] for row in batch_rows] == ["small"]
    assert batch_rows[0]["score_delta_bucket"] == "<= 1"
    assert batch_rows[0]["rank_delta_bucket"] == "<= 100"


def test_build_score_history_reconciliation_review_batch_adds_reference_context(tmp_path: Path):
    plan = tmp_path / "score_history_reconciliation_plan.csv"
    candidates = [
        {
            "key": {
                "score_year": 2022,
                "batch": "本科批",
                "subject_cat": "物理类",
                "school_code": "0162",
                "major_code": "0L",
            }
        },
        {
            "key": {
                "score_year": 2022,
                "batch": "本科批",
                "subject_cat": "物理类",
                "school_code": "0162",
                "major_code": "0P",
            }
        },
    ]
    row = {
        "task_id": "major-context",
        "issue_type": "major_code_drift_candidate",
        "priority": "1",
        "status": "todo",
        "suggested_action": "review_major_code_alignment",
        "match_confidence": "high",
        "score_year": "2022",
        "batch": "本科批",
        "subject_cat": "物理类",
        "school_code": "0162",
        "package_major_code": "H1",
        "core_major_code": "0L|0P",
        "package_min_score": "498",
        "core_min_score": "498",
        "package_min_rank": "46215",
        "core_min_rank": "46215",
        "package_key_json": "{}",
        "core_key_json": "{}",
        "core_candidates_json": json.dumps(candidates, ensure_ascii=False),
        "matching_values_json": "{}",
        "differences_json": "[]",
        "review_decision": "",
        "reviewer": "",
        "reviewed_at": "",
        "notes": "",
    }
    with plan.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=PLAN_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerow(row)

    projection = tmp_path / "ln_projection_score_2022_official.csv"
    with projection.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["school_code", "major_code", "major_full", "batch", "subject_cat", "score_year"],
        )
        writer.writeheader()
        writer.writerow({
            "school_code": "0162",
            "major_code": "H1",
            "major_full": "医学信息工程",
            "batch": "本科批",
            "subject_cat": "物理类",
            "score_year": "2022",
        })

    core_db = tmp_path / "university.db"
    con = duckdb.connect(str(core_db))
    con.execute("""
        CREATE TABLE fa_dim_ln_admission_plan (
            school_code VARCHAR,
            major_code VARCHAR,
            subject_cat VARCHAR,
            batch VARCHAR,
            year INTEGER,
            major_full VARCHAR,
            major_short VARCHAR
        )
    """)
    con.executemany(
        "INSERT INTO fa_dim_ln_admission_plan VALUES (?, ?, ?, ?, ?, ?, ?)",
        [
            ("0162", "0L", "物理类", "本科批", 2026, "医学信息工程(非英语语种考生慎报)", ""),
            ("0162", "0P", "物理类", "本科批", 2026, "运动康复", ""),
        ],
    )
    con.close()

    result = build_score_history_reconciliation_review_batch(
        plan_csv=plan,
        output_dir=tmp_path / "batch_with_context",
        limit_per_issue=1,
        projection_csv=projection,
        core_db=core_db,
        core_plan_year=2026,
    )

    assert result["reference_context"]["hint_counts"] == {"single_contains": 1}
    assert result["reference_context"]["issue_hint_counts"] == [
        {"issue_type": "major_code_drift_candidate", "major_name_match_hint": "single_contains", "rows": 1}
    ]
    assert result["reference_context"]["issue_package_hint_counts"] == [
        {"issue_type": "major_code_drift_candidate", "package_name_match_hint": "single_contains", "rows": 1}
    ]
    assert result["reference_context"]["hint_combo_counts"] == [
        {
            "issue_type": "major_code_drift_candidate",
            "major_name_match_hint": "single_contains",
            "package_name_match_hint": "single_contains",
            "rows": 1,
        }
    ]
    with Path(result["csv"]).open(encoding="utf-8", newline="") as f:
        batch_rows = list(csv.DictReader(f))
    batch_row = batch_rows[0]
    assert batch_row["package_major_full"] == "医学信息工程"
    assert batch_row["major_name_match_hint"] == "single_contains"
    assert batch_row["suggested_core_major_code"] == "0L"
    assert batch_row["package_name_match_hint"] == "single_contains"
    assert batch_row["suggested_package_major_code"] == "H1"
    candidate_names = json.loads(batch_row["core_candidate_names_json"])
    assert candidate_names[0]["major_full"] == "医学信息工程(非英语语种考生慎报)"
    assert candidate_names[0]["match_kind"] == "contains"
    package_candidates = json.loads(batch_row["package_candidate_names_json"])
    assert package_candidates == [
        {
            "major_code": "H1",
            "major_full": "医学信息工程",
            "match_kind": "contains",
            "matched_core_major_code": "0L",
        }
    ]


def test_build_score_history_reconciliation_review_batch_adds_token_overlap_context(tmp_path: Path):
    plan = tmp_path / "score_history_reconciliation_plan.csv"
    candidates = [
        {
            "key": {
                "score_year": 2022,
                "batch": "本科批",
                "subject_cat": "历史类",
                "school_code": "0002",
                "major_code": "H1",
            }
        },
        {
            "key": {
                "score_year": 2022,
                "batch": "本科批",
                "subject_cat": "历史类",
                "school_code": "0002",
                "major_code": "J1",
            }
        },
    ]
    row = {
        "task_id": "major-token-overlap",
        "issue_type": "major_code_drift_candidate",
        "priority": "1",
        "status": "todo",
        "suggested_action": "review_major_code_alignment",
        "match_confidence": "high",
        "score_year": "2022",
        "batch": "本科批",
        "subject_cat": "历史类",
        "school_code": "0002",
        "package_major_code": "03",
        "core_major_code": "H1|J1",
        "package_min_score": "651",
        "core_min_score": "651",
        "package_min_rank": "47",
        "core_min_rank": "47",
        "package_key_json": "{}",
        "core_key_json": "{}",
        "core_candidates_json": json.dumps(candidates, ensure_ascii=False),
        "matching_values_json": "{}",
        "differences_json": "[]",
        "review_decision": "",
        "reviewer": "",
        "reviewed_at": "",
        "notes": "",
    }
    with plan.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=PLAN_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerow(row)

    projection = tmp_path / "ln_projection_score_2022_official.csv"
    with projection.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["school_code", "major_code", "major_full", "batch", "subject_cat", "score_year"],
        )
        writer.writeheader()
        writer.writerow({
            "school_code": "0002",
            "major_code": "03",
            "major_full": "经济学类(含双学士学位项目，拔尖人才培养基地)(经济学、国民经济管理、能源经济、国际经济与贸易、数字经济)",
            "batch": "本科批",
            "subject_cat": "历史类",
            "score_year": "2022",
        })

    core_db = tmp_path / "university.db"
    con = duckdb.connect(str(core_db))
    con.execute("""
        CREATE TABLE fa_dim_ln_admission_plan (
            school_code VARCHAR,
            major_code VARCHAR,
            subject_cat VARCHAR,
            batch VARCHAR,
            year INTEGER,
            major_full VARCHAR,
            major_short VARCHAR
        )
    """)
    con.executemany(
        "INSERT INTO fa_dim_ln_admission_plan VALUES (?, ?, ?, ?, ?, ?, ?)",
        [
            ("0002", "H1", "历史类", "本科批", 2026, "金融学类(金融学、保险学)", ""),
            ("0002", "J1", "历史类", "本科批", 2026, "经济学类(经济学、国民经济管理、能源经济、国际经济与贸易)", ""),
        ],
    )
    con.close()

    result = build_score_history_reconciliation_review_batch(
        plan_csv=plan,
        output_dir=tmp_path / "batch_with_token_context",
        limit_per_issue=1,
        projection_csv=projection,
        core_db=core_db,
        core_plan_year=2026,
    )

    assert result["reference_context"]["hint_counts"] == {"single_token_overlap": 1}
    assert result["reference_context"]["package_hint_counts"] == {"single_token_overlap": 1}
    assert result["reference_context"]["token_overlap"] == {
        "enabled": True,
        "min_score": 0.6,
        "min_shared_tokens": 2,
        "stop_token_count": 13,
    }
    with Path(result["csv"]).open(encoding="utf-8", newline="") as f:
        batch_rows = list(csv.DictReader(f))
    batch_row = batch_rows[0]
    assert batch_row["major_name_match_hint"] == "single_token_overlap"
    assert batch_row["suggested_core_major_code"] == "J1"
    assert batch_row["package_name_match_hint"] == "single_token_overlap"
    assert batch_row["suggested_package_major_code"] == "03"
    candidate_names = json.loads(batch_row["core_candidate_names_json"])
    matched = [item for item in candidate_names if item["major_code"] == "J1"][0]
    assert matched["match_kind"] == "token_overlap"
    assert "经济学类" in matched["shared_tokens"]
    package_candidates = json.loads(batch_row["package_candidate_names_json"])
    assert package_candidates[0]["match_kind"] == "token_overlap"
    assert package_candidates[0]["matched_core_major_code"] == "J1"


def test_merge_score_history_reconciliation_review_batch_updates_only_editable_columns(tmp_path: Path):
    plan = tmp_path / "score_history_reconciliation_plan.csv"
    rows = [
        {
            "task_id": "major-1",
            "issue_type": "major_code_drift_candidate",
            "priority": "1",
            "status": "todo",
            "suggested_action": "review_major_code_alignment",
            "match_confidence": "high",
            "score_year": "2024",
            "batch": "本科批",
            "subject_cat": "物理类",
            "school_code": "1001",
            "package_major_code": "04",
            "core_major_code": "03",
            "package_min_score": "570",
            "core_min_score": "570",
            "package_min_rank": "3000",
            "core_min_rank": "3000",
            "package_key_json": "{}",
            "core_key_json": "{}",
            "core_candidates_json": "[]",
            "matching_values_json": "{}",
            "differences_json": "[]",
            "review_decision": "",
            "reviewer": "",
            "reviewed_at": "",
            "notes": "",
        },
        {
            "task_id": "value-1",
            "issue_type": "value_drift",
            "priority": "2",
            "status": "todo",
            "suggested_action": "review_source_value_conflict",
            "match_confidence": "primary_key_match",
            "score_year": "2024",
            "batch": "本科批",
            "subject_cat": "物理类",
            "school_code": "1002",
            "package_major_code": "02",
            "core_major_code": "02",
            "package_min_score": "580",
            "core_min_score": "580",
            "package_min_rank": "1990",
            "core_min_rank": "2000",
            "package_key_json": "{}",
            "core_key_json": "{}",
            "core_candidates_json": "[]",
            "matching_values_json": "{}",
            "differences_json": "[]",
            "review_decision": "",
            "reviewer": "",
            "reviewed_at": "",
            "notes": "",
        },
    ]
    with plan.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=PLAN_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)

    batch = tmp_path / "score_history_reconciliation_review_batch.csv"
    edited_rows = [dict(rows[0])]
    edited_rows[0].update({
        "status": "reviewed",
        "review_decision": "map_package_to_core_major_code",
        "reviewer": "tester",
        "reviewed_at": "2026-05-13",
        "notes": "matched by official source",
        "school_code": "9999",
    })
    with batch.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=PLAN_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(edited_rows)

    output = tmp_path / "merged.csv"
    report = merge_score_history_reconciliation_review_batch(
        plan_csv=plan,
        batch_csv=batch,
        output=output,
    )

    assert report["updated_rows"] == 1
    with output.open(encoding="utf-8", newline="") as f:
        merged = {row["task_id"]: row for row in csv.DictReader(f)}
    assert merged["major-1"]["status"] == "reviewed"
    assert merged["major-1"]["review_decision"] == "map_package_to_core_major_code"
    assert merged["major-1"]["school_code"] == "1001"
    assert merged["value-1"]["status"] == "todo"


def test_build_score_history_from_reconciliation_plan_rejects_unready(tmp_path: Path):
    plan = tmp_path / "score_history_reconciliation_plan.csv"
    with plan.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=PLAN_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerow({
            "task_id": "todo-1",
            "issue_type": "value_drift",
            "priority": "2",
            "status": "todo",
            "suggested_action": "review_source_value_conflict",
            "match_confidence": "primary_key_match",
            "score_year": "2024",
            "batch": "本科批",
            "subject_cat": "物理类",
            "school_code": "1002",
            "package_major_code": "02",
            "core_major_code": "02",
            "package_min_score": "580",
            "core_min_score": "580",
            "package_min_rank": "1990",
            "core_min_rank": "2000",
            "package_key_json": "{}",
            "core_key_json": "{}",
            "core_candidates_json": "[]",
            "matching_values_json": "{}",
            "differences_json": "[]",
            "review_decision": "",
            "reviewer": "",
            "reviewed_at": "",
            "notes": "",
        })

    try:
        build_score_history_package_from_reconciliation_plan(
            plan_csv=plan,
            output_root=tmp_path / "exports",
            package_id="pkg-reconciled-reject",
        )
        rejected = False
    except ValueError as exc:
        rejected = "not package-ready" in str(exc)
    assert rejected


def test_build_score_history_from_reconciliation_plan_exports_reviewed_rows(tmp_path: Path):
    plan = tmp_path / "score_history_reconciliation_plan.csv"
    rows = [
        {
            "task_id": "value-1",
            "issue_type": "value_drift",
            "priority": "2",
            "status": "reviewed",
            "suggested_action": "review_source_value_conflict",
            "match_confidence": "primary_key_match",
            "score_year": "2024",
            "batch": "本科批",
            "subject_cat": "物理类",
            "school_code": "1002",
            "package_major_code": "02",
            "core_major_code": "02",
            "package_min_score": "580",
            "core_min_score": "580",
            "package_min_rank": "1990",
            "core_min_rank": "2000",
            "package_key_json": "{}",
            "core_key_json": json.dumps({"school_code": "1002", "major_code": "02"}, ensure_ascii=False),
            "core_candidates_json": "[]",
            "matching_values_json": "{}",
            "differences_json": "[]",
            "review_decision": "keep_core_row",
            "reviewer": "tester",
            "reviewed_at": "2026-05-13",
            "notes": "core value verified",
        },
        {
            "task_id": "package-1",
            "issue_type": "package_only_unmatched",
            "priority": "3",
            "status": "reviewed",
            "suggested_action": "review_package_only_row",
            "match_confidence": "none",
            "score_year": "2024",
            "batch": "本科批",
            "subject_cat": "物理类",
            "school_code": "1006",
            "package_major_code": "06",
            "core_major_code": "",
            "package_min_score": "550",
            "core_min_score": "",
            "package_min_rank": "6000",
            "core_min_rank": "",
            "package_key_json": json.dumps({"school_code": "1006", "major_code": "06"}, ensure_ascii=False),
            "core_key_json": "{}",
            "core_candidates_json": "[]",
            "matching_values_json": "{}",
            "differences_json": "[]",
            "review_decision": "use_package_row",
            "reviewer": "tester",
            "reviewed_at": "2026-05-13",
            "notes": "package row verified",
        },
        {
            "task_id": "major-1",
            "issue_type": "major_code_drift_candidate",
            "priority": "1",
            "status": "reviewed",
            "suggested_action": "review_major_code_alignment",
            "match_confidence": "high",
            "score_year": "2024",
            "batch": "本科批",
            "subject_cat": "物理类",
            "school_code": "1003",
            "package_major_code": "04",
            "core_major_code": "03",
            "package_min_score": "570",
            "core_min_score": "570",
            "package_min_rank": "3000",
            "core_min_rank": "3000",
            "package_key_json": json.dumps({"school_code": "1003", "major_code": "04"}, ensure_ascii=False),
            "core_key_json": json.dumps({"school_code": "1003", "major_code": "03"}, ensure_ascii=False),
            "core_candidates_json": json.dumps([{"key": {"school_code": "1003", "major_code": "03"}}], ensure_ascii=False),
            "matching_values_json": "{}",
            "differences_json": "[]",
            "review_decision": "map_package_to_core_major_code",
            "reviewer": "tester",
            "reviewed_at": "2026-05-13",
            "notes": "major code aligned",
        },
        {
            "task_id": "package-exclude-1",
            "issue_type": "package_only_unmatched",
            "priority": "3",
            "status": "reviewed",
            "suggested_action": "review_package_only_row",
            "match_confidence": "none",
            "score_year": "2024",
            "batch": "本科批",
            "subject_cat": "物理类",
            "school_code": "1008",
            "package_major_code": "08",
            "core_major_code": "",
            "package_min_score": "530",
            "core_min_score": "",
            "package_min_rank": "8000",
            "core_min_rank": "",
            "package_key_json": json.dumps({"school_code": "1008", "major_code": "08"}, ensure_ascii=False),
            "core_key_json": "{}",
            "core_candidates_json": "[]",
            "matching_values_json": "{}",
            "differences_json": "[]",
            "review_decision": "exclude_row",
            "reviewer": "tester",
            "reviewed_at": "2026-05-13",
            "notes": "excluded from patch package",
        },
    ]
    with plan.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=PLAN_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)

    result = build_score_history_package_from_reconciliation_plan(
        plan_csv=plan,
        output_root=tmp_path / "exports",
        package_id="pkg-reconciled-score-history",
    )

    package_dir = Path(result["package_dir"])
    assert validate_manifest(package_dir / "manifest.json")["errors"] == []
    assert result["rows"] == 3
    assert result["skipped_rows"] == 1
    assert result["quality_report"]["decision_counts"]["exclude_row"] == 1
    with (package_dir / "fa_fact_ln_score_history.csv").open(encoding="utf-8", newline="") as f:
        output_rows = list(csv.DictReader(f))
    by_key = {(row["school_code"], row["major_code"]): row for row in output_rows}
    expected_id = hashlib.md5("1006||06||本科批||物理类||2024".encode()).hexdigest()[:16]
    assert by_key[("1002", "02")]["min_rank"] == "2000"
    assert by_key[("1006", "06")]["min_rank"] == "6000"
    assert by_key[("1006", "06")]["id"] == expected_id
    assert by_key[("1006", "06")]["score_type"] == "最低分"
    assert by_key[("1003", "03")]["min_rank"] == "3000"


def test_build_score_history_from_reconciliation_plan_rejects_core_delete_semantics(tmp_path: Path):
    plan = tmp_path / "score_history_reconciliation_plan.csv"
    with plan.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=PLAN_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerow({
            "task_id": "core-delete-1",
            "issue_type": "core_only_unmatched",
            "priority": "4",
            "status": "reviewed",
            "suggested_action": "review_core_only_row",
            "match_confidence": "none",
            "score_year": "2024",
            "batch": "本科批",
            "subject_cat": "物理类",
            "school_code": "1007",
            "package_major_code": "",
            "core_major_code": "07",
            "package_min_score": "",
            "core_min_score": "540",
            "package_min_rank": "",
            "core_min_rank": "7000",
            "package_key_json": "{}",
            "core_key_json": json.dumps({"school_code": "1007", "major_code": "07"}, ensure_ascii=False),
            "core_candidates_json": "[]",
            "matching_values_json": "{}",
            "differences_json": "[]",
            "review_decision": "exclude_row",
            "reviewer": "tester",
            "reviewed_at": "2026-05-13",
            "notes": "delete required",
        })

    try:
        build_score_history_package_from_reconciliation_plan(
            plan_csv=plan,
            output_root=tmp_path / "exports",
            package_id="pkg-reconciled-core-delete",
        )
        rejected = False
    except ValueError as exc:
        rejected = "cannot delete existing core rows" in str(exc)
    assert rejected


def test_build_score_history_from_reconciliation_plan_can_skip_core_delete_rows_when_explicit(tmp_path: Path):
    plan = tmp_path / "score_history_reconciliation_plan.csv"
    rows = [
        {
            "task_id": "core-delete-1",
            "issue_type": "core_only_unmatched",
            "priority": "4",
            "status": "reviewed",
            "suggested_action": "review_core_only_row",
            "match_confidence": "none",
            "score_year": "2024",
            "batch": "本科批",
            "subject_cat": "物理类",
            "school_code": "1007",
            "package_major_code": "",
            "core_major_code": "07",
            "package_min_score": "",
            "core_min_score": "540",
            "package_min_rank": "",
            "core_min_rank": "7000",
            "package_key_json": "{}",
            "core_key_json": json.dumps({"school_code": "1007", "major_code": "07"}, ensure_ascii=False),
            "core_candidates_json": "[]",
            "matching_values_json": "{}",
            "differences_json": "[]",
            "review_decision": "exclude_row",
            "reviewer": "tester",
            "reviewed_at": "2026-05-13",
            "notes": "delete required",
        },
        {
            "task_id": "package-1",
            "issue_type": "package_only_unmatched",
            "priority": "3",
            "status": "reviewed",
            "suggested_action": "review_package_only_row",
            "match_confidence": "none",
            "score_year": "2024",
            "batch": "本科批",
            "subject_cat": "物理类",
            "school_code": "1008",
            "package_major_code": "08",
            "core_major_code": "",
            "package_min_score": "530",
            "core_min_score": "",
            "package_min_rank": "8000",
            "core_min_rank": "",
            "package_key_json": json.dumps({"school_code": "1008", "major_code": "08"}, ensure_ascii=False),
            "core_key_json": "{}",
            "core_candidates_json": "[]",
            "matching_values_json": "{}",
            "differences_json": "[]",
            "review_decision": "use_package_row",
            "reviewer": "tester",
            "reviewed_at": "2026-05-13",
            "notes": "insert required",
        },
    ]
    with plan.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=PLAN_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)

    result = build_score_history_package_from_reconciliation_plan(
        plan_csv=plan,
        output_root=tmp_path / "exports",
        package_id="pkg-reconciled-core-delete-skipped",
        allow_core_exclude_rows=True,
    )

    assert result["rows"] == 1
    assert result["skipped_rows"] == 1
    assert result["allow_core_exclude_rows"] is True
    with (Path(result["package_dir"]) / "fa_fact_ln_score_history.csv").open(encoding="utf-8", newline="") as f:
        output_rows = list(csv.DictReader(f))
    assert [(row["school_code"], row["major_code"]) for row in output_rows] == [("1008", "08")]
    assert output_rows[0]["id"]
    assert output_rows[0]["score_type"] == "最低分"


def test_build_score_history_delete_plan_rejects_unready(tmp_path: Path):
    plan = tmp_path / "score_history_reconciliation_plan.csv"
    with plan.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=PLAN_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerow({
            "task_id": "todo-delete-1",
            "issue_type": "core_only_unmatched",
            "priority": "4",
            "status": "todo",
            "suggested_action": "review_core_only_row",
            "match_confidence": "none",
            "score_year": "2024",
            "batch": "本科批",
            "subject_cat": "物理类",
            "school_code": "1007",
            "package_major_code": "",
            "core_major_code": "07",
            "package_min_score": "",
            "core_min_score": "540",
            "package_min_rank": "",
            "core_min_rank": "7000",
            "package_key_json": "{}",
            "core_key_json": json.dumps({"school_code": "1007", "major_code": "07"}, ensure_ascii=False),
            "core_candidates_json": "[]",
            "matching_values_json": "{}",
            "differences_json": "[]",
            "review_decision": "",
            "reviewer": "",
            "reviewed_at": "",
            "notes": "",
        })

    try:
        build_score_history_delete_plan_from_reconciliation_plan(
            plan_csv=plan,
            output_dir=tmp_path / "delete_plan",
        )
        rejected = False
    except ValueError as exc:
        rejected = "not ready for delete planning" in str(exc)
    assert rejected


def test_build_score_history_delete_plan_from_reviewed_core_excludes(tmp_path: Path):
    plan = tmp_path / "score_history_reconciliation_plan.csv"
    rows = [
        {
            "task_id": "core-delete-1",
            "issue_type": "core_only_unmatched",
            "priority": "4",
            "status": "reviewed",
            "suggested_action": "review_core_only_row",
            "match_confidence": "none",
            "score_year": "2024",
            "batch": "本科批",
            "subject_cat": "物理类",
            "school_code": "1007",
            "package_major_code": "",
            "core_major_code": "07",
            "package_min_score": "",
            "core_min_score": "540",
            "package_min_rank": "",
            "core_min_rank": "7000",
            "package_key_json": "{}",
            "core_key_json": json.dumps({
                "school_code": "1007",
                "major_code": "07",
                "batch": "本科批",
                "subject_cat": "物理类",
                "score_year": 2024,
            }, ensure_ascii=False),
            "core_candidates_json": "[]",
            "matching_values_json": "{}",
            "differences_json": "[]",
            "review_decision": "exclude_row",
            "reviewer": "tester",
            "reviewed_at": "2026-05-13",
            "notes": "delete after source verification",
        },
        {
            "task_id": "package-exclude-1",
            "issue_type": "package_only_unmatched",
            "priority": "3",
            "status": "reviewed",
            "suggested_action": "review_package_only_row",
            "match_confidence": "none",
            "score_year": "2024",
            "batch": "本科批",
            "subject_cat": "物理类",
            "school_code": "1008",
            "package_major_code": "08",
            "core_major_code": "",
            "package_min_score": "530",
            "core_min_score": "",
            "package_min_rank": "8000",
            "core_min_rank": "",
            "package_key_json": json.dumps({"school_code": "1008", "major_code": "08"}, ensure_ascii=False),
            "core_key_json": "{}",
            "core_candidates_json": "[]",
            "matching_values_json": "{}",
            "differences_json": "[]",
            "review_decision": "exclude_row",
            "reviewer": "tester",
            "reviewed_at": "2026-05-13",
            "notes": "do not add package row",
        },
    ]
    with plan.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=PLAN_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)

    result = build_score_history_delete_plan_from_reconciliation_plan(
        plan_csv=plan,
        output_dir=tmp_path / "delete_plan",
    )

    assert result["rows"] == 1
    with Path(result["csv"]).open(encoding="utf-8", newline="") as f:
        delete_rows = list(csv.DictReader(f))
    assert delete_rows[0]["school_code"] == "1007"
    assert delete_rows[0]["major_code"] == "07"
    assert delete_rows[0]["score_year"] == "2024"
    manifest = json.loads(Path(result["manifest"]).read_text(encoding="utf-8"))
    assert manifest["csv"] == "score_history_delete_plan.csv"
    assert manifest["notes"].startswith("Delete migration plan only")


def test_build_policy_industry_map_package_from_config(tmp_path: Path):
    config = tmp_path / "policy_industry_map.json"
    config.write_text(json.dumps({
        "version": "fixture-policy-map",
        "source_date": "2026-03-16",
        "availability_date": "2026-05-13",
        "policy_period": "十五五(2026-2030)",
        "source_lineage": {
            "source_key": "policy_industry_map",
            "source_kind": "curated_policy_config",
            "evidence_urls": ["https://example.gov/policy"],
        },
        "validation": {
            "allowed_policy_labels": ["重点扶持", "中性", "收缩"],
            "policy_intensity_min": 1,
            "policy_intensity_max": 3,
        },
        "rows": [
            {
                "tdx_l2": "T1205",
                "tdx_l2_name": "软件服务",
                "tdx_l1_name": "信息产业",
                "policy_label": "重点扶持",
                "policy_intensity": 3,
                "key_themes": ["人工智能", "国产软件替代"],
                "rationale": "fixture rationale",
            }
        ],
    }, ensure_ascii=False), encoding="utf-8")

    result = build_policy_industry_map_package(
        output_root=tmp_path / "exports",
        config_path=config,
        package_id="pkg-policy-map-test",
    )
    package_dir = Path(result["package_dir"])
    assert validate_manifest(package_dir / "manifest.json")["errors"] == []
    assert result["rows"] == 1
    assert result["source_lineage"]["config_file"] == "config/policy_industry_map.json"

    with (package_dir / "fa_dim_policy_industry_map.csv").open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    assert rows[0]["policy_period"] == "十五五(2026-2030)"
    assert json.loads(rows[0]["key_themes_json"]) == ["人工智能", "国产软件替代"]


def test_build_policy_package_rejects_duplicate_json_keys(tmp_path: Path):
    config = tmp_path / "policy_industry_map.json"
    config.write_text('{"version":"1","version":"2","rows":[]}\n', encoding="utf-8")

    with pytest.raises(ValueError, match="duplicate JSON key"):
        build_policy_industry_map_package(
            output_root=tmp_path / "exports",
            config_path=config,
            package_id="pkg-policy-map-duplicate-json-key",
        )


def test_build_policy_plan_history_package_rejects_duplicate_keys(tmp_path: Path):
    config = tmp_path / "policy_plan_history.json"
    config.write_text(json.dumps({
        "version": "fixture-policy-history",
        "source_date": "2026-05-13",
        "availability_date": "2026-05-13",
        "source_lineage": {
            "source_key": "policy_plan_history",
            "source_kind": "curated_policy_backtest_config",
            "evidence_urls": ["https://example.gov/history"],
        },
        "validation": {
            "allowed_policy_labels": ["重点扶持", "中性", "收缩"],
            "allowed_actual_outcomes": ["超预期兑现", "基本兑现", "部分兑现", "未兑现"],
            "outcome_score_min": 1,
            "outcome_score_max": 5,
        },
        "rows": [
            {
                "plan_period": "十四五(2021-2025)",
                "tdx_l2": "T0706",
                "tdx_l2_name": "电气设备",
                "tdx_l1_name": "装备制造",
                "policy_label": "重点扶持",
                "actual_outcome": "超预期兑现",
                "outcome_score": 5,
                "evidence": "fixture evidence",
            },
            {
                "plan_period": "十四五(2021-2025)",
                "tdx_l2": "T0706",
                "tdx_l2_name": "电气设备",
                "tdx_l1_name": "装备制造",
                "policy_label": "重点扶持",
                "actual_outcome": "基本兑现",
                "outcome_score": 4,
                "evidence": "duplicate fixture evidence",
            },
        ],
    }, ensure_ascii=False), encoding="utf-8")

    try:
        build_policy_plan_history_package(
            output_root=tmp_path / "exports",
            config_path=config,
            package_id="pkg-policy-history-bad-test",
        )
        rejected = False
    except ValueError as exc:
        rejected = "duplicate primary keys" in str(exc)
    assert rejected


def test_build_policy_plan_history_package_from_config(tmp_path: Path):
    config = tmp_path / "policy_plan_history.json"
    config.write_text(json.dumps({
        "version": "fixture-policy-history",
        "source_date": "2026-05-13",
        "availability_date": "2026-05-13",
        "source_lineage": {
            "source_key": "policy_plan_history",
            "source_kind": "curated_policy_backtest_config",
            "evidence_urls": ["https://example.gov/history"],
        },
        "validation": {
            "allowed_policy_labels": ["重点扶持", "中性", "收缩"],
            "allowed_actual_outcomes": ["超预期兑现", "基本兑现", "部分兑现", "未兑现"],
            "outcome_score_min": 1,
            "outcome_score_max": 5,
        },
        "rows": [
            {
                "plan_period": "十四五(2021-2025)",
                "tdx_l2": "T0706",
                "tdx_l2_name": "电气设备",
                "tdx_l1_name": "装备制造",
                "policy_label": "重点扶持",
                "actual_outcome": "超预期兑现",
                "outcome_score": 5,
                "evidence": "fixture evidence",
            }
        ],
    }, ensure_ascii=False), encoding="utf-8")

    result = build_policy_plan_history_package(
        output_root=tmp_path / "exports",
        config_path=config,
        package_id="pkg-policy-history-test",
    )
    package_dir = Path(result["package_dir"])
    assert validate_manifest(package_dir / "manifest.json")["errors"] == []
    assert result["rows"] == 1

    with (package_dir / "fa_dim_policy_plan_history.csv").open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    assert rows[0]["plan_period"] == "十四五(2021-2025)"
    assert rows[0]["outcome_score"] == "5"


def test_parse_moe_school_profile_rows():
    rows = parse_moe_school_profile_rows(
        [
            ["附件1：", "", "", "", "", "", ""],
            ["全国普通高等学校名单\n（截至2025年6月20日）", "", "", "", "", "", ""],
            ["序号", "学校名称", "学校标识码", "主管部门", "所在地", "办学层次", "备注"],
            ["北京市（92所）", "", "", "", "", "", ""],
            [1, "北京大学", 4111010001, "教育部", "北京市", "本科", ""],
            [2, "北京城市学院", 4111011418, "北京市教委", "北京市", "本科", "民办"],
            ["辽宁省（116所）", "", "", "", "", "", ""],
            [180, "东北大学", 4121010145, "教育部", "沈阳市", "本科", ""],
        ],
        source_date="2025-06-20",
        availability_date="2025-06-27",
        built_at="2026-05-13T00:00:00",
    )

    assert rows[0]["national_school_code"] == "4111010001"
    assert rows[0]["province"] == "北京市"
    assert rows[1]["ownership"] == "民办"
    assert rows[2]["school_name"] == "东北大学"
    assert rows[2]["province"] == "辽宁省"
    assert rows[2]["competent_authority"] == "教育部"


def test_build_school_outcome_package_from_cleaned_csv(tmp_path: Path):
    source = tmp_path / "school_outcome.csv"
    with source.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "院校代码",
                "院校名称",
                "指标键",
                "指标名称",
                "指标值",
                "单位",
                "指标年份",
                "统计口径",
                "来源标题",
                "来源链接",
                "证据摘录",
                "来源日期",
                "可用日期",
                "构建时间",
            ],
        )
        writer.writeheader()
        writer.writerow({
            "院校代码": "10145",
            "院校名称": "东北大学",
            "指标键": "postgrad_rate",
            "指标名称": "深造率",
            "指标值": "46.2%",
            "单位": "ratio",
            "指标年份": "2025",
            "统计口径": "本科毕业生",
            "来源标题": "2025届毕业生就业质量报告",
            "来源链接": "https://example.edu/report.pdf",
            "证据摘录": "本科毕业生深造率为46.2%。",
            "来源日期": "2025-12-31",
            "可用日期": "2026-01-05",
            "构建时间": "2026-05-13T00:00:00",
        })

    result = build_local_package(
        source_key="school_outcome",
        table_name="fa_fact_school_outcome",
        input_path=source,
        output_root=tmp_path / "exports",
        package_id="pkg-school-outcome-test",
        source_version="fixture-school-outcome",
    )
    package_dir = Path(result["package_dir"])
    assert validate_manifest(package_dir / "manifest.json")["errors"] == []
    assert result["rows"] == 1

    schema = get_table_schema("fa_fact_school_outcome")
    with (package_dir / "fa_fact_school_outcome.csv").open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    assert rows[0]["metric_key"] == "postgrad_rate"
    assert float(rows[0]["metric_value"]) == 0.462
    assert set(rows[0]).issuperset(schema["required"])


def test_outcome_metric_registry_rejects_unknown_keys(tmp_path: Path):
    metrics = load_outcome_metrics()
    assert "postgrad_rate" in metrics["domains"]["school"]
    source = tmp_path / "school_outcome_bad.csv"
    with source.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "院校代码",
                "指标键",
                "指标名称",
                "指标值",
                "单位",
                "指标年份",
                "来源链接",
                "来源日期",
                "可用日期",
                "构建时间",
            ],
        )
        writer.writeheader()
        writer.writerow({
            "院校代码": "10145",
            "指标键": "made_up_metric",
            "指标名称": "任意指标",
            "指标值": "1",
            "单位": "ratio",
            "指标年份": "2025",
            "来源链接": "https://example.edu/report.pdf",
            "来源日期": "2025-12-31",
            "可用日期": "2026-01-05",
            "构建时间": "2026-05-13T00:00:00",
        })
    try:
        build_local_package(
            source_key="school_outcome",
            table_name="fa_fact_school_outcome",
            input_path=source,
            output_root=tmp_path / "exports",
            package_id="pkg-school-outcome-bad-test",
            source_version="fixture-school-outcome-bad",
        )
        rejected = False
    except ValueError as exc:
        rejected = "unregistered metric_key" in str(exc)
    assert rejected


def test_outcome_metric_registry_requires_metric_scope(tmp_path: Path):
    source = tmp_path / "school_outcome_missing_scope.csv"
    with source.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "院校代码",
                "院校名称",
                "指标键",
                "指标名称",
                "指标值",
                "单位",
                "指标年份",
                "统计口径",
                "来源标题",
                "来源链接",
                "证据摘录",
                "来源日期",
                "可用日期",
                "构建时间",
            ],
        )
        writer.writeheader()
        writer.writerow({
            "院校代码": "10145",
            "院校名称": "东北大学",
            "指标键": "civil_service_rate",
            "指标名称": "体制内去向比例",
            "指标值": "34.6%",
            "单位": "ratio",
            "指标年份": "2025",
            "统计口径": "",
            "来源标题": "2025届毕业生就业质量报告",
            "来源链接": "https://example.edu/report.pdf",
            "证据摘录": "国有企业签约比例为34.6%。",
            "来源日期": "2025-12-31",
            "可用日期": "2026-01-05",
            "构建时间": "2026-05-13T00:00:00",
        })

    with pytest.raises(ValueError, match="missing metric_scope"):
        build_local_package(
            source_key="school_outcome",
            table_name="fa_fact_school_outcome",
            input_path=source,
            output_root=tmp_path / "exports",
            package_id="pkg-school-outcome-missing-scope",
            source_version="fixture-school-outcome",
        )


def test_build_career_source_plan_from_config(tmp_path: Path):
    config = load_career_data_sources()
    assert "salary_median" in config["metrics"]
    assert "shortage_rank" in config["metrics"]
    occupations = tmp_path / "occupations.csv"
    with occupations.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["occupation_code", "occupation_name", "tdx_l2", "tdx_l2_name"],
        )
        writer.writeheader()
        writer.writerow({
            "occupation_code": "15-102",
            "occupation_name": "软件工程师",
            "tdx_l2": "T1205",
            "tdx_l2_name": "软件服务",
        })

    result = build_career_source_plan(
        output_dir=tmp_path / "career_plan",
        source_keys=["career_recruitment_snapshot", "career_civil_service_posts"],
        metric_year=2026,
        city="沈阳",
        occupation_input=occupations,
    )

    assert result["rows"] == 7
    with Path(result["csv"]).open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    assert set(rows[0]).issuperset(CAREER_PLAN_COLUMNS)
    assert {row["source_key"] for row in rows} == {"career_recruitment_snapshot", "career_civil_service_posts"}
    assert any(row["metric_key"] == "salary_median" for row in rows)
    assert any(row["target_table"] == "fa_fact_career_signal" for row in rows)
    assert all(row["city"] == "沈阳" for row in rows)
    assert all(row["occupation_name"] == "软件工程师" for row in rows)
    assert "软件工程师 招聘 薪资 沈阳 2026" in rows[0]["search_queries"]
    manifest = json.loads(Path(result["manifest"]).read_text(encoding="utf-8"))
    assert manifest["notes"].startswith("Collection plan only")


def test_build_career_source_plan_from_core_catalog(tmp_path: Path):
    core_db = tmp_path / "core.duckdb"
    con = duckdb.connect(str(core_db))
    try:
        con.execute("""
            CREATE TABLE fa_dim_career_occupation (
              occupation_code VARCHAR,
              occupation_name VARCHAR,
              tdx_l2 VARCHAR,
              tdx_l2_name VARCHAR
            )
        """)
        con.execute(
            "INSERT INTO fa_dim_career_occupation VALUES (?, ?, ?, ?), (?, ?, ?, ?)",
            [
                "4-04-05-05",
                "互联网营销师",
                "T0502",
                "商贸代理",
                "2-02-10-03",
                "软件工程技术人员",
                "T1205",
                "软件服务",
            ],
        )
    finally:
        con.close()

    result = build_career_source_plan(
        output_dir=tmp_path / "career_plan_core",
        source_keys=["career_recruitment_snapshot"],
        metric_year=2026,
        city="沈阳",
        core_db=core_db,
        occupation_limit=1,
    )

    assert result["rows"] == 5
    with Path(result["csv"]).open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    assert {row["occupation_name"] for row in rows} == {"互联网营销师"}
    assert {row["tdx_l2"] for row in rows} == {"T0502"}
    assert {row["metric_key"] for row in rows} == {
        "job_posting_count",
        "shortage_rank",
        "salary_median",
        "salary_p75",
        "work_intensity_index",
    }
    manifest = json.loads(Path(result["manifest"]).read_text(encoding="utf-8"))
    assert manifest["core_db"] == str(core_db)
    assert manifest["occupation_limit"] == 1


def test_parse_digital_occupation_catalog_html_builds_catalog_package(tmp_path: Path):
    html = """
    <html><body><table>
      <tr><td>序号</td><td>职业编码</td><td>职业名称</td></tr>
      <tr><td>1</td><td>2-02-01-02</td><td>地球物理地球化学与遥感勘查工程技术人员</td></tr>
      <tr><td>2</td><td>4-04-05-05</td><td>互联网营销师</td></tr>
    </table></body></html>
    """
    rows = parse_digital_occupation_catalog_html(
        html,
        source_title="国家职业分类大典首次标识数字职业",
        source_url="https://chinajob.mohrss.gov.cn/c/2022-10-28/363399.shtml",
        source_date="2022-10-28",
        availability_date="2022-10-28",
    )
    assert len(rows) == 2
    assert rows[0]["occupation_family"] == "专业技术人员"
    assert rows[1]["occupation_family"] == "社会生产服务和生活服务类"
    assert rows[0]["occupation_level"] == 4
    assert rows[0]["tdx_l2"] == "T1301"
    assert rows[0]["tdx_l2_name"] == "综合类"
    assert rows[1]["tdx_l2"] == "T0502"
    assert rows[1]["tdx_l2_name"] == "商贸代理"
    assert rows[0]["major_keywords_json"] == "[]"

    cleaned = tmp_path / "digital_occupation_catalog.csv"
    write_digital_occupation_catalog_csv(cleaned, rows)
    result = build_local_package(
        source_key="career_occupation_catalog",
        table_name="fa_dim_career_occupation",
        input_path=cleaned,
        output_root=tmp_path / "exports",
        package_id="digital-occupation-catalog-test",
        source_version="fixture-digital-occupation",
    )
    package_dir = Path(result["package_dir"])
    assert result["rows"] == 2
    assert result["quality_report"]["errors"] == []
    assert validate_manifest(package_dir / "manifest.json")["errors"] == []


def test_build_broad_occupation_catalog_seed_rows(tmp_path: Path):
    seeds = tmp_path / "broad_occupation_seeds.json"
    seeds.write_text(
        json.dumps({
            "source_title": "国家职业分类大典（2022年版）社会工作专业人员相关公示材料",
            "source_url": "https://example.com/occupation",
            "source_date": "2022-07-25",
            "availability_date": "2026-05-18",
            "seeds": [
                {
                    "occupation_code": "2-07-10-03",
                    "occupation_name": "心理咨询师",
                    "occupation_family": "专业技术人员",
                    "tdx_l2": "T1301",
                    "tdx_l2_name": "综合类",
                    "major_keywords_json": "[\"心理学\"]",
                    "skill_keywords_json": "[\"心理咨询\"]",
                    "evidence_quote": "本小类包括下列职业：2-07-10-03心理咨询师",
                }
            ],
        }, ensure_ascii=False),
        encoding="utf-8",
    )

    rows = build_broad_occupation_catalog_seed_rows(seeds)

    assert rows[0]["occupation_code"] == "2-07-10-03"
    assert rows[0]["occupation_name"] == "心理咨询师"
    assert rows[0]["occupation_family"] == "专业技术人员"
    assert rows[0]["occupation_level"] == 4
    assert rows[0]["major_keywords_json"] == "[\"心理学\"]"

    cleaned = tmp_path / "broad_occupation_catalog.csv"
    write_digital_occupation_catalog_csv(cleaned, rows)
    result = build_local_package(
        source_key="career_occupation_catalog",
        table_name="fa_dim_career_occupation",
        input_path=cleaned,
        output_root=tmp_path / "exports",
        package_id="broad-occupation-catalog-test",
        source_version="fixture-broad-occupation",
    )
    assert result["rows"] == 1
    assert result["quality_report"]["errors"] == []


def test_merge_occupation_catalog_csvs_rejects_duplicate_codes(tmp_path: Path):
    first = tmp_path / "first.csv"
    second = tmp_path / "second.csv"
    write_digital_occupation_catalog_csv(first, [{
        "occupation_code": "2-02-30-09",
        "occupation_name": "数据分析处理工程技术人员",
        "occupation_family": "专业技术人员",
        "occupation_level": 4,
        "tdx_l2": "T1205",
        "tdx_l2_name": "软件服务",
        "major_keywords_json": "[]",
        "skill_keywords_json": "[]",
        "source_title": "source A",
        "source_url": "https://example.com/a",
        "evidence_quote": "职业编码 2-02-30-09，职业名称 数据分析处理工程技术人员",
        "source_date": "2022-10-28",
        "availability_date": "2022-10-28",
        "built_at": "2026-05-18T00:00:00",
    }])
    write_digital_occupation_catalog_csv(second, [{
        "occupation_code": "2-07-10-03",
        "occupation_name": "心理咨询师",
        "occupation_family": "专业技术人员",
        "occupation_level": 4,
        "tdx_l2": "T1301",
        "tdx_l2_name": "综合类",
        "major_keywords_json": "[]",
        "skill_keywords_json": "[]",
        "source_title": "source B",
        "source_url": "https://example.com/b",
        "evidence_quote": "职业编码 2-07-10-03，职业名称 心理咨询师",
        "source_date": "2022-07-25",
        "availability_date": "2026-05-18",
        "built_at": "2026-05-18T00:00:00",
    }])

    merged = merge_occupation_catalog_csvs([first, second])
    assert [row["occupation_code"] for row in merged] == ["2-02-30-09", "2-07-10-03"]

    try:
        merge_occupation_catalog_csvs([first, first])
    except ValueError as exc:
        assert "duplicate occupation_code" in str(exc)
    else:
        raise AssertionError("duplicate occupation_code should be rejected")


def test_audit_career_source_plan_reports_progress_and_errors(tmp_path: Path):
    plan = tmp_path / "career_source_plan.csv"
    with plan.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=CAREER_PLAN_COLUMNS)
        writer.writeheader()
        base = {column: "" for column in CAREER_PLAN_COLUMNS}
        writer.writerow({
            **base,
            "source_key": "career_recruitment_snapshot",
            "source_name": "招聘需求与薪资快照",
            "source_kind": "controlled_market_snapshot",
            "target_table": "fa_fact_career_signal",
            "occupation_code": "15-102",
            "occupation_name": "软件工程师",
            "tdx_l2": "T1205",
            "tdx_l2_name": "软件服务",
            "metric_key": "salary_median",
            "metric_label": "月薪中位数",
            "metric_unit": "cny_month",
            "metric_value": "12000",
            "metric_scope": "公开招聘样本",
            "metric_year": "2026",
            "city": "沈阳",
            "collection_methods": '["manual_platform_export"]',
            "evidence_urls": "[]",
            "search_queries": '["软件工程师 招聘 薪资 沈阳 2026"]',
            "source_title": "招聘快照",
            "source_url": "https://example.com/jobs/software",
            "evidence_quote": "软件工程师薪资中位数约12000元/月。",
            "source_date": "2026-05-13",
            "availability_date": "2026-05-13",
            "status": "verified",
            "reviewer": "fixture",
            "reviewed_at": "2026-05-13T00:00:00",
        })
        writer.writerow({
            **base,
            "source_key": "career_recruitment_snapshot",
            "source_name": "招聘需求与薪资快照",
            "source_kind": "controlled_market_snapshot",
            "target_table": "fa_fact_career_signal",
            "occupation_code": "15-102",
            "occupation_name": "软件工程师",
            "metric_key": "made_up_metric",
            "metric_label": "未知指标",
            "metric_unit": "score",
            "metric_year": "2026",
            "city": "沈阳",
            "collection_methods": '["manual_platform_export"]',
            "evidence_urls": "[]",
            "search_queries": '["fixture"]',
            "status": "verified",
        })

    report = audit_career_source_plan(plan)
    assert report["rows"] == 2
    assert report["progress"]["complete_rows"] == 2
    assert report["evidence_counts"]["rows_with_source_url"] == 1
    assert any("unregistered career metric_key" in error for error in report["errors"])
    assert any("complete status missing evidence" in error for error in report["errors"])


def test_audit_career_source_coverage_maps_metrics_to_sources(tmp_path: Path):
    report_path = tmp_path / "career_source_coverage.json"
    report = audit_career_source_coverage(report_path=report_path)

    assert report_path.exists()
    assert report["uncovered_metrics"] == []
    assert report["warnings"] == []
    assert "career_recruitment_snapshot" in report["metric_to_sources"]["salary_median"]
    assert "career_salary_survey" in report["metric_to_sources"]["salary_median"]
    assert "career_civil_service_posts" in report["metric_to_sources"]["civil_service_post_count"]
    source_by_key = {row["source_key"]: row for row in report["source_rows"]}
    assert source_by_key["career_occupation_catalog"]["coverage_status"] == "official_seed_ready"
    assert source_by_key["career_recruitment_snapshot"]["coverage_status"] == "manual_snapshot_required"
    assert report["summary"]["covered_metric_count"] == report["metric_count"]


def test_career_platform_source_policy_blocks_outcome_publication():
    config = load_career_data_sources()
    policy = config["platform_source_policy"]
    tiers = policy["tiers"]

    assert "fa_fact_school_outcome" in policy["disallowed_targets"]
    assert "fa_fact_major_outcome" in policy["disallowed_targets"]
    assert "fa_fact_career_signal" in tiers["public_research_report"]["allowed_evidence_targets"]
    assert "fa_mart_major_city_employment_fit" in tiers["government_market_report"]["allowed_evidence_targets"]
    assert tiers["community_scraper"]["allowed_evidence_targets"] == []
    assert tiers["community_scraper"]["allowed_metrics"] == []
    assert "授权" in tiers["licensed_api_or_terminal"]["publish_rule"]
    assert any("平台样本偏差" in control for control in policy["bias_controls"])

    recruitment = config["source_plan"]["sources"]["career_recruitment_snapshot"]
    assert "public_research_report_metric" in recruitment["collection_methods"]
    assert "licensed_data_terminal_export" in recruitment["collection_methods"]
    assert "fa_fact_career_signal" in recruitment["target_tables"]


def test_apply_career_shortage_page_updates_plan_candidates(tmp_path: Path):
    html = """
    <html><body><p>
    排行前3个紧缺职业分别为市场营销专业人员、计算机网络工程技术人员、计算机软件工程技术人员。
    </p></body></html>
    """
    ranking = parse_shortage_ranking(html)
    assert ranking[1] == {"rank": 2, "occupation_name": "计算机网络工程技术人员"}

    html_file = tmp_path / "guangzhou_shortage.html"
    html_file.write_text(html, encoding="utf-8")
    plan = tmp_path / "career_source_plan.csv"
    network = _career_plan_row(
        "career_recruitment_snapshot",
        "fa_fact_career_signal",
        "shortage_rank",
        status="todo",
    )
    network.update({
        "occupation_code": "2-02-10-04",
        "occupation_name": "计算机网络工程技术人员",
        "metric_year": "2025",
        "city": "广州",
    })
    software = _career_plan_row(
        "career_recruitment_snapshot",
        "fa_fact_career_signal",
        "shortage_rank",
        status="todo",
    )
    software.update({
        "occupation_code": "2-02-10-03",
        "occupation_name": "计算机软件工程技术人员",
        "metric_year": "2025",
        "city": "广州",
    })
    salary = _career_plan_row(
        "career_recruitment_snapshot",
        "fa_fact_career_signal",
        "salary_median",
        status="todo",
    )
    _write_career_plan(plan, [network, software, salary])

    output = tmp_path / "career_source_plan_shortage.csv"
    report = apply_career_shortage_page_to_plan(
        plan_csv=plan,
        html_file=html_file,
        output=output,
        source_title="广州市2025年第四季度人力资源市场供求状况分析",
        source_url="https://www.gz.gov.cn/example.html",
        source_date="2026-04-08",
        availability_date="2026-04-08",
    )

    assert report["ranked_item_count"] == 3
    assert report["matched_rows"] == 2
    assert report["updated_rows"] == 2
    assert report["unmatched_ranked_items"] == [{"rank": 1, "occupation_name": "市场营销专业人员"}]
    with output.open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    by_metric_name = {(row["metric_key"], row["occupation_name"]): row for row in rows}
    network_row = by_metric_name[("shortage_rank", "计算机网络工程技术人员")]
    assert network_row["status"] == "in_progress"
    assert network_row["metric_value"] == "2"
    assert network_row["source_title"] == "广州市2025年第四季度人力资源市场供求状况分析"
    assert network_row["evidence_quote"] == "计算机网络工程技术人员排名2。"
    assert by_metric_name[("salary_median", "软件工程师")]["metric_value"] == ""


def test_download_scs_resources_writes_raw_manifest(tmp_path: Path, monkeypatch):
    api_body = json.dumps({
        "resList": [
            {
                "fileType": ".zip",
                "resResourceId": "resource-main",
                "resourceName": "中央机关及其直属机构2026年度考试录用公务员招考简章.zip",
                "resourceComment": "2026年度考试录用公务员招考简章",
            },
            {
                "fileType": ".docx",
                "resResourceId": "resource-form",
                "resourceName": "2026年度考试录用公务员报名登记表.docx",
                "resourceComment": "报名登记表",
            },
        ]
    }, ensure_ascii=False).encode("utf-8")
    zip_body = b"PK\x03\x04 fixture"

    monkeypatch.setattr(
        "datahub.connectors.scs_resources.load_career_data_sources",
        lambda: {
            "source_plan": {
                "sources": {
                    "career_civil_service_posts": {
                        "name": "公考与事业编岗位目录",
                        "kind": "official_attachment",
                        "target_tables": ["fa_fact_career_signal"],
                        "official_distribution": "国考职位表",
                        "evidence_urls": ["https://bm.scs.gov.cn/kl2026"],
                        "resource_api": {
                            "api_url": "https://example.gov/api/resources",
                            "download_base_url": "https://example.gov/download/",
                            "source_date": "2025-10-14",
                            "availability_date": "2025-10-14",
                            "include_resource_keywords": ["考试录用公务员招考简章"],
                            "exclude_resource_keywords": ["报名登记表"],
                            "allowed_file_types": [".zip"],
                            "headers": {"User-Agent": "fixture"},
                        },
                    }
                }
            }
        },
    )

    class FakeResponse:
        def __init__(self, body: bytes):
            self._body = body

        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, tb):
            return False

        def read(self):
            return self._body

    def fake_urlopen(request, timeout=60):
        if request.full_url == "https://example.gov/api/resources":
            assert request.headers["User-agent"] == "fixture"
            return FakeResponse(api_body)
        if request.full_url == "https://example.gov/download/resource-main":
            return FakeResponse(zip_body)
        raise AssertionError(request.full_url)

    monkeypatch.setattr("datahub.connectors.scs_resources.urlopen", fake_urlopen)

    result = download_scs_resources(
        source_key="career_civil_service_posts",
        output_root=tmp_path / "raw",
    )

    assert result["resource_count"] == 2
    assert result["selected_resource_count"] == 1
    assert result["downloaded_files"] == 1
    output_dir = tmp_path / "raw" / "career_civil_service_posts" / "2025-10-14"
    assert (output_dir / "_scs_resource_api_response.json").read_bytes() == api_body
    resource_file = output_dir / "中央机关及其直属机构2026年度考试录用公务员招考简章.zip"
    assert resource_file.read_bytes() == zip_body
    manifest = json.loads((output_dir / "_scs_resource_manifest.json").read_text(encoding="utf-8"))
    assert manifest["source_key"] == "career_civil_service_posts"
    assert manifest["target_tables"] == ["fa_fact_career_signal"]
    assert manifest["api_response_sha256"] == hashlib.sha256(api_body).hexdigest()
    assert manifest["files"][0]["resource_id"] == "resource-main"
    assert manifest["files"][0]["sha256"] == hashlib.sha256(zip_body).hexdigest()


def test_parse_scs_position_workbook_from_zip(tmp_path: Path, monkeypatch):
    header = [
        "部门代码",
        "部门名称",
        "用人司局",
        "机构性质",
        "招考职位",
        "职位属性",
        "职位分布",
        "职位简介",
        "职位代码",
        "机构层级",
        "考试类别",
        "招考人数",
        "专业",
        "学历",
        "学位",
        "政治面貌",
        "基层工作最低年限",
        "服务基层项目工作经历",
        "是否在面试阶段组织专业能力测试",
        "面试人员比例",
        "工作地点",
        "落户地点",
        "备注",
        "部门网站",
        "咨询电话1",
        "咨询电话2",
        "咨询电话3",
    ]
    data_row = [
        "002000",
        "中央办公厅",
        "警卫局",
        "中央党群机关",
        "财务管理岗位",
        "普通职位",
        "其他职位",
        "从事财务管理相关工作",
        "100110001002",
        "中央",
        "综合管理类",
        1,
        "本科：120203K会计学、120204财务管理；研究生：1253会计",
        "本科或硕士研究生",
        "与最高学历相对应的学位",
        "中共党员",
        "二年",
        "无限制",
        "否",
        "5:1",
        "北京市",
        "北京市",
        "",
        "https://example.gov",
        "010-1",
        "",
        "",
    ]

    class FakeSheet:
        name = "中央党群机关"
        nrows = 3
        ncols = len(header)

        def cell_value(self, row, col):
            if row == 0:
                return "说明" if col == 0 else ""
            if row == 1:
                return header[col]
            if row == 2:
                return data_row[col]
            return ""

    class FakeBook:
        def sheets(self):
            return [FakeSheet()]

    def fake_open_workbook(*, file_contents):
        assert file_contents == b"fixture-xls"
        return FakeBook()

    monkeypatch.setattr("datahub.parsers.scs_position_workbook.xlrd.open_workbook", fake_open_workbook)

    zip_path = tmp_path / "positions.zip"
    with ZipFile(zip_path, "w") as zf:
        zf.writestr("positions.xls", b"fixture-xls")

    rows = parse_scs_position_workbook(
        input_path=zip_path,
        source_title="中央机关及其直属机构2026年度考试录用公务员招考简章",
        source_url="http://dl.scs.gov.cn/download/resource-main",
        source_date="2025-10-14",
        availability_date="2025-10-14",
    )

    assert len(rows) == 1
    row = rows[0]
    assert row["source_key"] == "career_civil_service_posts"
    assert row["sheet_name"] == "中央党群机关"
    assert row["row_number"] == "3"
    assert row["department_code"] == "002000"
    assert row["position_code"] == "100110001002"
    assert row["recruit_count"] == "1"
    assert "会计学" in row["major_requirement"]
    assert row["work_location"] == "北京市"
    assert row["built_at"]

    output = tmp_path / "scs_positions.csv"
    write_scs_position_csv(output, rows)
    with output.open(encoding="utf-8", newline="") as f:
        written = list(csv.DictReader(f))
    assert written[0]["position_name"] == "财务管理岗位"

    result = build_local_package(
        source_key="career_civil_service_posts",
        table_name="fa_fact_civil_service_position",
        input_path=output,
        output_root=tmp_path / "exports",
        package_id="scs-position-test",
        source_version="fixture-scs-position",
    )
    package_dir = Path(result["package_dir"])
    assert result["rows"] == 1
    assert result["quality_report"]["errors"] == []
    assert validate_manifest(package_dir / "manifest.json")["errors"] == []


def test_build_civil_service_signal_plan_from_positions(tmp_path: Path):
    positions = tmp_path / "scs_positions.csv"
    position_rows = [
        {
            "source_key": "career_civil_service_posts",
            "source_title": "中央机关及其直属机构2026年度考试录用公务员招考简章",
            "source_url": "http://dl.scs.gov.cn/download/resource-main",
            "source_date": "2025-10-14",
            "availability_date": "2025-10-14",
            "sheet_name": "中央国家行政机关",
            "row_number": "3",
            "position_code": "100110001002",
            "position_name": "信息化建设岗位",
            "position_description": "从事软件系统建设和数据治理工作",
            "recruit_count": "2",
            "major_requirement": "本科：080901计算机科学与技术、080902软件工程",
            "work_location": "北京市",
            "remarks": "",
        },
        {
            "source_key": "career_civil_service_posts",
            "source_title": "中央机关及其直属机构2026年度考试录用公务员招考简章",
            "source_url": "http://dl.scs.gov.cn/download/resource-main",
            "source_date": "2025-10-14",
            "availability_date": "2025-10-14",
            "sheet_name": "中央国家行政机关",
            "row_number": "5",
            "position_code": "100110001004",
            "position_name": "信息公开岗位",
            "position_description": "从事政务公开和综合文字工作",
            "recruit_count": "1",
            "major_requirement": "本科：030101K法学、050101汉语言文学",
            "work_location": "北京市",
            "remarks": "",
        },
        {
            "source_key": "career_civil_service_posts",
            "source_title": "中央机关及其直属机构2026年度考试录用公务员招考简章",
            "source_url": "http://dl.scs.gov.cn/download/resource-main",
            "source_date": "2025-10-14",
            "availability_date": "2025-10-14",
            "sheet_name": "中央党群机关",
            "row_number": "4",
            "position_code": "100110001003",
            "position_name": "财务管理岗位",
            "position_description": "从事财务管理相关工作",
            "recruit_count": "1",
            "major_requirement": "本科：120203K会计学、120204财务管理",
            "work_location": "北京市",
            "remarks": "",
        },
    ]
    with positions.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=position_rows[0].keys())
        writer.writeheader()
        writer.writerows(position_rows)

    occupations = tmp_path / "occupations.csv"
    occupation_rows = [
        {
            "occupation_code": "2-02-10-03",
            "occupation_name": "计算机软件工程技术人员",
            "tdx_l2": "T1205",
            "tdx_l2_name": "软件服务",
            "major_keywords_json": "[]",
            "skill_keywords_json": "[]",
        },
        {
            "occupation_code": "2-06-04-00",
            "occupation_name": "会计专业人员",
            "tdx_l2": "T1001",
            "tdx_l2_name": "银行",
            "major_keywords_json": "[\"会计\", \"财务\"]",
            "skill_keywords_json": "[]",
        },
        {
            "occupation_code": "4-04-02-01",
            "occupation_name": "信息通信网络机务员",
            "tdx_l2": "T1202",
            "tdx_l2_name": "通信设备",
            "major_keywords_json": "[]",
            "skill_keywords_json": "[]",
        },
    ]
    with occupations.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=occupation_rows[0].keys())
        writer.writeheader()
        writer.writerows(occupation_rows)

    result = build_civil_service_signal_plan(
        positions_csv=positions,
        occupation_input=occupations,
        output_dir=tmp_path / "civil_service_signal",
        metric_year=2026,
    )

    assert result["rows"] == 2
    plan_csv = Path(result["csv"])
    with plan_csv.open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    by_code = {row["occupation_code"]: row for row in rows}
    assert by_code["2-02-10-03"]["metric_key"] == "civil_service_post_count"
    assert by_code["2-02-10-03"]["metric_value"] == "1"
    assert "招考人数2人" in by_code["2-02-10-03"]["metric_scope"]
    assert "信息化建设岗位" in by_code["2-02-10-03"]["evidence_quote"]
    assert "命中：" in by_code["2-02-10-03"]["evidence_quote"]
    assert by_code["2-06-04-00"]["status"] == "in_progress"
    assert "4-04-02-01" not in by_code

    audit = audit_career_source_plan(plan_csv)
    assert audit["errors"] == []
    assert audit["status_counts"] == {"in_progress": 2}


def test_civil_service_signal_plan_prioritizes_specific_evidence(tmp_path: Path):
    positions = tmp_path / "scs_positions.csv"
    position_rows = [
        {
            "source_key": "career_civil_service_posts",
            "source_title": "中央机关及其直属机构2026年度考试录用公务员招考简章",
            "source_url": "http://dl.scs.gov.cn/download/resource-main",
            "source_date": "2025-10-14",
            "availability_date": "2025-10-14",
            "sheet_name": "中央国家行政机关",
            "row_number": "3",
            "position_code": "100110001001",
            "position_name": "宽泛信息岗位",
            "position_description": "从事信息化建设",
            "recruit_count": "1",
            "major_requirement": "本科：0809计算机类",
            "work_location": "北京市",
            "remarks": "",
        },
        {
            "source_key": "career_civil_service_posts",
            "source_title": "中央机关及其直属机构2026年度考试录用公务员招考简章",
            "source_url": "http://dl.scs.gov.cn/download/resource-main",
            "source_date": "2025-10-14",
            "availability_date": "2025-10-14",
            "sheet_name": "中央国家行政机关",
            "row_number": "4",
            "position_code": "100110001002",
            "position_name": "具体软件岗位",
            "position_description": "从事软件系统建设",
            "recruit_count": "1",
            "major_requirement": "本科：080901计算机科学与技术、080902软件工程、计算机软件工程",
            "work_location": "北京市",
            "remarks": "",
        },
    ]
    with positions.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=position_rows[0].keys())
        writer.writeheader()
        writer.writerows(position_rows)

    occupations = tmp_path / "occupations.csv"
    occupation_rows = [{
        "occupation_code": "2-02-10-03",
        "occupation_name": "计算机软件工程技术人员",
        "tdx_l2": "T1205",
        "tdx_l2_name": "软件服务",
        "major_keywords_json": "[]",
        "skill_keywords_json": "[]",
    }]
    with occupations.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=occupation_rows[0].keys())
        writer.writeheader()
        writer.writerows(occupation_rows)

    result = build_civil_service_signal_plan(
        positions_csv=positions,
        occupation_input=occupations,
        output_dir=tmp_path / "civil_service_signal",
        metric_year=2026,
    )

    with Path(result["csv"]).open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    assert rows[0]["metric_value"] == "2"
    assert rows[0]["evidence_quote"].startswith("具体软件岗位")
    assert "命中：计算机/软件/计算机软件" in rows[0]["evidence_quote"]


def test_civil_service_signal_plan_exposes_hidden_match_context(tmp_path: Path):
    positions = tmp_path / "scs_positions.csv"
    position_rows = [{
        "source_key": "career_civil_service_posts",
        "source_title": "中央机关及其直属机构2026年度考试录用公务员招考简章",
        "source_url": "http://dl.scs.gov.cn/download/resource-main",
        "source_date": "2025-10-14",
        "availability_date": "2025-10-14",
        "sheet_name": "中央国家行政机关",
        "row_number": "3",
        "position_code": "100110001001",
        "position_name": "平台监管岗位",
        "position_description": "从事跨境电子商务监管与数据核验",
        "recruit_count": "1",
        "major_requirement": "本科：1202工商管理类",
        "work_location": "北京市",
        "remarks": "",
    }]
    with positions.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=position_rows[0].keys())
        writer.writeheader()
        writer.writerows(position_rows)

    occupations = tmp_path / "occupations.csv"
    occupation_rows = [{
        "occupation_code": "4-01-06-01",
        "occupation_name": "电子商务师",
        "tdx_l2": "T0502",
        "tdx_l2_name": "商贸代理",
        "major_keywords_json": "[]",
        "skill_keywords_json": "[]",
    }]
    with occupations.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=occupation_rows[0].keys())
        writer.writeheader()
        writer.writerows(occupation_rows)

    result = build_civil_service_signal_plan(
        positions_csv=positions,
        occupation_input=occupations,
        output_dir=tmp_path / "civil_service_signal",
        metric_year=2026,
    )

    with Path(result["csv"]).open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    assert rows[0]["metric_value"] == "1"
    assert "专业：本科：1202工商管理类" in rows[0]["evidence_quote"]
    assert "简介：从事跨境电子商务监管与数据核验" in rows[0]["evidence_quote"]
    assert "命中：电子商务" in rows[0]["evidence_quote"]


def test_build_career_source_review_batch_limits_pending_rows(tmp_path: Path):
    plan = tmp_path / "career_source_plan.csv"
    rows = [
        _career_plan_row("career_recruitment_snapshot", "fa_fact_career_signal", "salary_median", status="todo"),
        _career_plan_row("career_recruitment_snapshot", "fa_fact_career_signal", "work_intensity_index", status="in_progress"),
        _career_plan_row("career_recruitment_snapshot", "fa_fact_career_signal", "salary_p75", status="verified"),
        _career_plan_row("career_civil_service_posts", "fa_fact_career_signal", "civil_service_post_count", status="todo"),
    ]
    rows[0]["metric_value"] = "12000"
    rows[1]["metric_value"] = "65"
    rows[3]["metric_value"] = "80"
    _write_career_plan(plan, rows)

    result = build_career_source_review_batch(
        plan_csv=plan,
        output_dir=tmp_path / "career_batch",
        limit_per_source=1,
    )

    assert result["rows"] == 2
    assert result["source_counts"] == {
        "career_civil_service_posts": 1,
        "career_recruitment_snapshot": 1,
    }
    with Path(result["csv"]).open(encoding="utf-8", newline="") as f:
        batch_rows = list(csv.DictReader(f))
    assert {row["status"] for row in batch_rows} <= {"todo", "in_progress"}
    assert any(row["metric_key"] == "salary_median" for row in batch_rows)
    assert not any(row["metric_key"] == "work_intensity_index" for row in batch_rows)
    assert all(row["metric_key"] != "salary_p75" for row in batch_rows)
    manifest = json.loads(Path(result["manifest"]).read_text(encoding="utf-8"))
    assert manifest["task_key_columns"] == [
        "source_key",
        "target_table",
        "occupation_code",
        "occupation_name",
        "metric_key",
        "metric_year",
        "city",
    ]
    assert "source_url" in manifest["editable_columns"]
    assert manifest["sort"][1] == {"field": "metric_value", "type": "number", "direction": "desc"}


def test_apply_career_source_review_seeds_updates_matching_rows(tmp_path: Path):
    audit = audit_career_source_review_seeds()
    assert audit["errors"] == []
    assert audit["seed_count"] >= 20
    assert audit["status_counts"]["verified"] >= 20

    plan = tmp_path / "career_source_plan.csv"
    seeded = _career_plan_row(
        "career_civil_service_posts",
        "fa_fact_career_signal",
        "civil_service_post_count",
        status="in_progress",
    )
    seeded.update({
        "occupation_code": "2-02-10-03",
        "occupation_name": "计算机软件工程技术人员",
        "city": "全国",
        "source_title": "中央机关及其直属机构2026年度考试录用公务员招考简章",
        "source_url": "http://dl.scs.gov.cn/download/resource-main",
        "evidence_quote": "专业：软件工程，命中：软件",
        "source_date": "2025-10-14",
        "availability_date": "2025-10-14",
    })
    pending = _career_plan_row(
        "career_civil_service_posts",
        "fa_fact_career_signal",
        "civil_service_post_count",
        status="in_progress",
    )
    pending.update({
        "occupation_code": "4-04-05-02",
        "occupation_name": "计算机软件测试员",
        "city": "全国",
    })
    salary_seeded = _career_plan_row(
        "career_salary_survey",
        "fa_fact_career_signal",
        "salary_median",
        status="todo",
    )
    salary_seeded.update({
        "occupation_code": "2-02-10-03",
        "occupation_name": "计算机软件工程技术人员",
        "city": "宁波",
        "metric_year": "2024",
    })
    shortage_seeded = _career_plan_row(
        "career_recruitment_snapshot",
        "fa_fact_career_signal",
        "shortage_rank",
        status="todo",
    )
    shortage_seeded.update({
        "occupation_code": "2-02-10-04",
        "occupation_name": "计算机网络工程技术人员",
        "city": "广州",
        "metric_year": "2025",
    })
    _write_career_plan(plan, [seeded, salary_seeded, shortage_seeded, pending])

    output = tmp_path / "career_source_plan_seeded.csv"
    report = apply_career_source_review_seeds(plan_csv=plan, output=output)
    assert report["matched_rows"] == 3
    assert report["updated_rows"] == 3
    assert report["unmatched_seeds"] == audit["seed_count"] - 3

    with output.open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    by_key = {(row["source_key"], row["occupation_code"], row["city"], row["metric_key"]): row for row in rows}
    software_scs = by_key[("career_civil_service_posts", "2-02-10-03", "全国", "civil_service_post_count")]
    assert software_scs["status"] == "verified"
    assert software_scs["reviewer"] == "codex"
    assert "seed_review=" in software_scs["notes"]
    software_salary = by_key[("career_salary_survey", "2-02-10-03", "宁波", "salary_median")]
    assert software_salary["status"] == "verified"
    assert software_salary["metric_value"] == "13584"
    assert software_salary["source_title"] == "宁波市2024年度薪酬调查信息"
    assert "年度50%位163009元折算月薪" in software_salary["metric_scope"]
    network_shortage = by_key[("career_recruitment_snapshot", "2-02-10-04", "广州", "shortage_rank")]
    assert network_shortage["status"] == "verified"
    assert network_shortage["metric_value"] == "12"
    assert network_shortage["metric_unit"] == "rank"
    assert network_shortage["source_title"] == "广州市2025年第四季度人力资源市场供求状况分析"
    assert by_key[("career_civil_service_posts", "4-04-05-02", "全国", "civil_service_post_count")]["status"] == "in_progress"


def test_audit_career_source_review_seeds_requires_direct_metric_seed_evidence(monkeypatch: pytest.MonkeyPatch):
    seed = {
        "seed_id": "missing_salary_evidence",
        "source_key": "career_salary_survey",
        "target_table": "fa_fact_career_signal",
        "occupation_code": "2-02-10-03",
        "occupation_name": "计算机软件工程技术人员",
        "metric_key": "salary_median",
        "metric_year": 2024,
        "city": "宁波",
        "status": "verified",
        "reviewer": "codex",
        "reviewed_at": "2026-05-14",
        "review_note": "测试缺失直接指标证据。",
    }
    monkeypatch.setattr(
        "datahub.builders.career_source_seed_merge.load_career_source_review_seeds",
        lambda: {"seeds": [seed]},
    )

    audit = audit_career_source_review_seeds()

    assert audit["errors"] == [
        "seed 1 missing source-required fields for career_salary_survey: "
        "metric_value, metric_scope, source_title, source_url, evidence_quote, source_date, availability_date"
    ]


def test_audit_career_source_review_seeds_allows_plan_replayed_seed_without_copy_fields(monkeypatch: pytest.MonkeyPatch):
    seed = {
        "seed_id": "scs_seed_without_copy_fields",
        "source_key": "career_civil_service_posts",
        "target_table": "fa_fact_career_signal",
        "occupation_code": "2-02-10-03",
        "occupation_name": "计算机软件工程技术人员",
        "metric_key": "civil_service_post_count",
        "metric_year": 2026,
        "city": "全国",
        "status": "verified",
        "reviewer": "codex",
        "reviewed_at": "2026-05-14",
        "review_note": "官方职位表证据由计划重新生成。",
    }
    monkeypatch.setattr(
        "datahub.builders.career_source_seed_merge.load_career_source_review_seeds",
        lambda: {"seeds": [seed]},
    )

    audit = audit_career_source_review_seeds()

    assert audit["errors"] == []


def test_audit_career_source_review_seeds_validates_seed_metadata(monkeypatch: pytest.MonkeyPatch):
    seed = {
        "seed_id": "bad_seed_metadata",
        "source_key": "career_civil_service_posts",
        "target_table": "fa_fact_career_signal",
        "occupation_code": "2-02-10-03",
        "occupation_name": "计算机软件工程技术人员",
        "metric_key": "civil_service_post_count",
        "metric_year": "2026.5",
        "city": "全国",
        "status": "verified",
        "reviewer": "codex",
        "reviewed_at": "2026/05/14",
        "review_note": "测试基础元数据校验。",
    }
    monkeypatch.setattr(
        "datahub.builders.career_source_seed_merge.load_career_source_review_seeds",
        lambda: {"seeds": [seed]},
    )

    audit = audit_career_source_review_seeds()

    assert "seed 1 metric_year is not an integer" in audit["errors"]
    assert "seed 1 reviewed_at must use YYYY-MM-DD" in audit["errors"]


def test_audit_career_source_review_seeds_validates_direct_metric_evidence(monkeypatch: pytest.MonkeyPatch):
    seed = {
        "seed_id": "bad_direct_metric_evidence",
        "source_key": "career_recruitment_snapshot",
        "target_table": "fa_fact_career_signal",
        "occupation_code": "2-02-10-04",
        "occupation_name": "计算机网络工程技术人员",
        "metric_key": "shortage_rank",
        "metric_year": 2025,
        "city": "广州",
        "status": "verified",
        "reviewer": "codex",
        "reviewed_at": "2024-12-31",
        "review_note": "测试直接指标证据校验。",
        "metric_value": 0,
        "metric_scope": "公开供求报告排行",
        "source_title": "广州市2025年第四季度人力资源市场供求状况分析",
        "source_url": "ftp://example.com/report",
        "evidence_quote": "网络工程技术人员排行第0。",
        "source_date": "2025-02-01",
        "availability_date": "2025-01-01",
    }
    monkeypatch.setattr(
        "datahub.builders.career_source_seed_merge.load_career_source_review_seeds",
        lambda: {"seeds": [seed]},
    )

    audit = audit_career_source_review_seeds()

    assert "seed 1 source_date must not be after availability_date" in audit["errors"]
    assert "seed 1 reviewed_at must not be before availability_date" in audit["errors"]
    assert "seed 1 source_url must be an http(s) URL" in audit["errors"]
    assert "seed 1 metric_value is below min_value 1: 0.0" in audit["errors"]


def test_merge_career_source_review_batch_updates_only_editable_columns(tmp_path: Path):
    plan = tmp_path / "career_source_plan.csv"
    rows = [
        _career_plan_row("career_recruitment_snapshot", "fa_fact_career_signal", "salary_median", status="todo"),
        _career_plan_row("career_civil_service_posts", "fa_fact_career_signal", "civil_service_post_count", status="todo"),
    ]
    _write_career_plan(plan, rows)
    batch = tmp_path / "career_source_review_batch.csv"
    edited_row = {**rows[0]}
    edited_row.update({
        "source_name": "被篡改的来源名",
        "status": "verified",
        "metric_value": "12000",
        "metric_scope": "公开招聘样本，税前月薪",
        "source_title": "招聘薪资快照",
        "source_url": "https://example.com/jobs/software",
        "evidence_quote": "软件工程师薪资中位数约12000元/月。",
        "source_date": "2026-05-13",
        "availability_date": "2026-05-13",
        "reviewer": "fixture",
        "reviewed_at": "2026-05-13T00:00:00",
        "notes": "核对完成",
    })
    _write_career_plan(batch, [edited_row])

    output = tmp_path / "career_source_plan_merged.csv"
    report = merge_career_source_review_batch(
        plan_csv=plan,
        batch_csv=batch,
        output=output,
    )

    assert report["updated_rows"] == 1
    assert report["status_counts"] == {"todo": 1, "verified": 1}
    with output.open(encoding="utf-8", newline="") as f:
        merged_rows = list(csv.DictReader(f))
    assert merged_rows[0]["source_name"] == "招聘需求与薪资快照"
    assert merged_rows[0]["status"] == "verified"
    assert merged_rows[0]["source_url"] == "https://example.com/jobs/software"
    assert merged_rows[0]["notes"] == "核对完成"
    assert merged_rows[1]["status"] == "todo"


def test_build_career_signal_package_from_verified_source_plan(tmp_path: Path):
    plan = tmp_path / "career_source_plan.csv"
    verified_row = _career_plan_row(
        "career_recruitment_snapshot",
        "fa_fact_career_signal",
        "salary_median",
        status="verified",
    )
    verified_row.update({
        "metric_value": "12000",
        "metric_scope": "公开招聘样本，税前月薪",
        "source_title": "招聘薪资快照",
        "source_url": "https://example.com/jobs/software",
        "evidence_quote": "软件工程师薪资中位数约12000元/月。",
        "source_date": "2026-05-13",
        "availability_date": "2026-05-13",
    })
    rows = [
        verified_row,
        _career_plan_row("career_recruitment_snapshot", "fa_fact_career_signal", "salary_p75", status="todo"),
    ]
    _write_career_plan(plan, rows)

    result = build_career_signal_package_from_source_plan(
        plan_csv=plan,
        output_root=tmp_path / "exports",
        package_id="career-signal-from-plan-test",
        source_version="fixture-career-plan",
    )

    package = result["package"]
    package_dir = Path(package["package_dir"])
    assert result["rows"] == 1
    assert package["rows"] == 1
    assert package["quality_report"]["errors"] == []
    assert validate_manifest(package_dir / "manifest.json")["errors"] == []
    with (package_dir / "fa_fact_career_signal.csv").open(encoding="utf-8", newline="") as f:
        signal_rows = list(csv.DictReader(f))
    assert signal_rows[0]["metric_name"] == "月薪中位数"
    assert signal_rows[0]["source_url"] == "https://example.com/jobs/software"
    manifest = json.loads((package_dir / "manifest.json").read_text(encoding="utf-8"))
    assert manifest["source_lineage"]["source_kind"] == "verified_career_source_plan"
    assert manifest["source_lineage"]["metric_keys"] == ["salary_median"]


def test_build_career_signal_package_from_cleaned_csv(tmp_path: Path):
    source = tmp_path / "career_signal.csv"
    with source.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "职业代码",
                "职业名称",
                "通达信二级行业代码",
                "通达信二级行业",
                "城市",
                "指标键",
                "指标名称",
                "指标值",
                "单位",
                "指标年份",
                "统计口径",
                "来源标题",
                "来源链接",
                "证据摘录",
                "来源日期",
                "可用日期",
                "构建时间",
            ],
        )
        writer.writeheader()
        writer.writerow({
            "职业代码": "15-102",
            "职业名称": "软件工程师",
            "通达信二级行业代码": "T1205",
            "通达信二级行业": "软件服务",
            "城市": "沈阳",
            "指标键": "salary_median",
            "指标名称": "月薪中位数",
            "指标值": "12000",
            "单位": "cny_month",
            "指标年份": "2026",
            "统计口径": "公开招聘样本",
            "来源标题": "招聘快照",
            "来源链接": "https://example.com/jobs/software",
            "证据摘录": "软件工程师薪资中位数约12000元/月。",
            "来源日期": "2026-05-13",
            "可用日期": "2026-05-13",
            "构建时间": "2026-05-13T00:00:00",
        })

    result = build_local_package(
        source_key="career_signal",
        table_name="fa_fact_career_signal",
        input_path=source,
        output_root=tmp_path / "exports",
        package_id="pkg-career-signal-test",
        source_version="fixture-career-signal",
    )
    package_dir = Path(result["package_dir"])
    assert validate_manifest(package_dir / "manifest.json")["errors"] == []
    assert result["rows"] == 1
    with (package_dir / "fa_fact_career_signal.csv").open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    assert rows[0]["metric_key"] == "salary_median"
    assert rows[0]["metric_unit"] == "cny_month"
    assert rows[0]["occupation_name"] == "软件工程师"


def test_career_metric_registry_rejects_unknown_keys(tmp_path: Path):
    source = tmp_path / "career_signal_bad.csv"
    with source.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "occupation_code",
                "occupation_name",
                "city",
                "metric_key",
                "metric_name",
                "metric_value",
                "metric_unit",
                "metric_year",
                "source_title",
                "source_url",
                "evidence_quote",
                "source_date",
                "availability_date",
                "built_at",
            ],
        )
        writer.writeheader()
        writer.writerow({
            "occupation_code": "15-102",
            "occupation_name": "软件工程师",
            "city": "沈阳",
            "metric_key": "made_up_career_metric",
            "metric_name": "任意指标",
            "metric_value": "1",
            "metric_unit": "score",
            "metric_year": "2026",
            "source_title": "fixture",
            "source_url": "https://example.com",
            "evidence_quote": "fixture quote",
            "source_date": "2026-05-13",
            "availability_date": "2026-05-13",
            "built_at": "2026-05-13T00:00:00",
        })
    try:
        build_local_package(
            source_key="career_signal",
            table_name="fa_fact_career_signal",
            input_path=source,
            output_root=tmp_path / "exports",
            package_id="pkg-career-signal-bad-test",
            source_version="fixture-career-signal-bad",
        )
        rejected = False
    except ValueError as exc:
        rejected = "unregistered career metric_key" in str(exc)
    assert rejected


def test_build_career_score_package_from_signals(tmp_path: Path):
    source = tmp_path / "career_signal.csv"
    with source.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "occupation_code",
                "occupation_name",
                "tdx_l2",
                "tdx_l2_name",
                "city",
                "metric_key",
                "metric_name",
                "metric_value",
                "metric_unit",
                "metric_year",
                "metric_scope",
                "source_title",
                "source_url",
                "evidence_quote",
                "source_date",
                "availability_date",
                "built_at",
            ],
        )
        writer.writeheader()
        base = {
            "occupation_code": "15-102",
            "occupation_name": "软件工程师",
            "tdx_l2": "T1205",
            "tdx_l2_name": "软件服务",
            "city": "沈阳",
            "metric_year": "2026",
            "metric_scope": "公开样本",
            "source_title": "fixture",
            "source_date": "2026-05-13",
            "availability_date": "2026-05-13",
            "built_at": "2026-05-13T00:00:00",
        }
        rows = [
            ("salary_median", "月薪中位数", "12000", "cny_month", "https://example.com/salary", "薪资中位数约12000元/月。"),
            ("job_posting_count", "招聘岗位数", "1200", "count", "https://example.com/jobs", "近30天公开岗位1200个。"),
            ("civil_service_post_count", "公考岗位数", "80", "count", "https://example.com/civil", "可匹配职位80个。"),
            ("work_intensity_index", "工作强度指数", "65", "score", "https://example.com/intensity", "强度指数65分。"),
        ]
        for metric_key, metric_name, value, unit, url, quote in rows:
            writer.writerow({
                **base,
                "metric_key": metric_key,
                "metric_name": metric_name,
                "metric_value": value,
                "metric_unit": unit,
                "source_url": url,
                "evidence_quote": quote,
            })

    result = build_career_score_package(
        signal_input=source,
        output_root=tmp_path / "exports",
        package_id="pkg-career-score-test",
        source_version="fixture-career-score",
    )

    package_dir = Path(result["package_dir"])
    assert validate_manifest(package_dir / "manifest.json")["errors"] == []
    assert result["rows"] == 1
    with (package_dir / "fa_mart_career_score.csv").open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    assert rows[0]["occupation_code"] == "15-102"
    assert rows[0]["score_profile"] == "career_default_v1"
    assert float(rows[0]["friendly_35_score"]) > 0
    assert int(float(rows[0]["signal_count"])) == 4
    lineage = json.loads(rows[0]["pit_lineage_json"])
    assert "fa_fact_career_signal" in lineage["tables"]


def test_build_campus_living_score_package(tmp_path: Path):
    location_input = tmp_path / "school_location.csv"
    with location_input.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "national_school_code",
                "local_school_code",
                "school_name",
                "campus_key",
                "campus_name",
                "campus_type",
                "address",
                "province",
                "city",
                "district",
                "adcode",
                "longitude",
                "latitude",
                "coordinate_system",
                "geocode_level",
                "geocode_confidence",
                "source_address_url",
                "source_date",
                "availability_date",
                "built_at",
            ],
        )
        writer.writeheader()
        writer.writerow({
            "national_school_code": "4121010145",
            "local_school_code": "0145",
            "school_name": "东北大学",
            "campus_key": "0145_main",
            "campus_name": "南湖校区",
            "campus_type": "main",
            "address": "辽宁省沈阳市和平区文化路三巷",
            "province": "辽宁",
            "city": "沈阳",
            "district": "和平区",
            "adcode": "210102",
            "longitude": "123.425",
            "latitude": "41.774",
            "coordinate_system": "GCJ-02",
            "geocode_level": "门牌号",
            "geocode_confidence": "0.95",
            "source_address_url": "https://example.com/school-location",
            "source_date": "2026-05-13",
            "availability_date": "2026-05-13",
            "built_at": "2026-05-13T00:00:00",
        })

    poi_input = tmp_path / "campus_poi.csv"
    with poi_input.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "national_school_code",
                "local_school_code",
                "school_name",
                "campus_key",
                "campus_name",
                "poi_id",
                "poi_name",
                "poi_category",
                "category_group",
                "distance_m",
                "walking_minutes",
                "longitude",
                "latitude",
                "coordinate_system",
                "province",
                "city",
                "district",
                "adcode",
                "address",
                "source_provider",
                "source_url",
                "raw_response_hash",
                "snapshot_date",
                "source_date",
                "availability_date",
                "built_at",
            ],
        )
        writer.writeheader()
        base = {
            "national_school_code": "4121010145",
            "local_school_code": "0145",
            "school_name": "东北大学",
            "campus_key": "0145_main",
            "campus_name": "南湖校区",
            "province": "辽宁",
            "city": "沈阳",
            "district": "和平区",
            "adcode": "210102",
            "coordinate_system": "GCJ-02",
            "source_provider": "fixture_amap",
            "source_url": "https://example.com/poi",
            "raw_response_hash": "hash-poi",
            "snapshot_date": "2026-05-14",
            "source_date": "2026-05-13",
            "availability_date": "2026-05-13",
            "built_at": "2026-05-13T00:00:00",
        }
        for poi_id, poi_name, poi_category, category_group, distance_m in [
            ("metro-1", "地铁A", "地铁站", "subway_station", 820),
            ("metro-2", "地铁B", "地铁站", "subway_station", 1600),
            ("bus-1", "公交A", "公交站", "bus_stop", 300),
            ("bus-2", "公交B", "公交站", "bus_stop", 650),
            ("market-1", "超市A", "超市", "supermarket", 500),
            ("food-1", "快餐A", "餐饮", "restaurant", 420),
            ("hospital-1", "医院A", "医院", "hospital", 2200),
            ("park-1", "公园A", "公园", "park", 1300),
        ]:
            writer.writerow({
                **base,
                "poi_id": poi_id,
                "poi_name": poi_name,
                "poi_category": poi_category,
                "category_group": category_group,
                "distance_m": distance_m,
                "walking_minutes": round(distance_m / 80, 1),
                "longitude": "123.42",
                "latitude": "41.77",
                "address": "fixture address",
            })

    housing_input = tmp_path / "campus_housing.csv"
    with housing_input.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "national_school_code",
                "local_school_code",
                "school_name",
                "campus_key",
                "campus_name",
                "radius_m",
                "listing_type",
                "housing_metric_key",
                "housing_metric_name",
                "metric_value",
                "metric_unit",
                "sample_count",
                "source_platform",
                "source_url",
                "source_scope",
                "raw_response_hash",
                "snapshot_date",
                "source_date",
                "availability_date",
                "built_at",
            ],
        )
        writer.writeheader()
        base = {
            "national_school_code": "4121010145",
            "local_school_code": "0145",
            "school_name": "东北大学",
            "campus_key": "0145_main",
            "campus_name": "南湖校区",
            "radius_m": "3000",
            "listing_type": "rent",
            "source_platform": "fixture_market",
            "source_url": "https://example.com/housing",
            "source_scope": "3km rental snapshot",
            "raw_response_hash": "hash-housing",
            "snapshot_date": "2026-05-14",
            "source_date": "2026-05-13",
            "availability_date": "2026-05-13",
            "built_at": "2026-05-13T00:00:00",
        }
        for metric_key, metric_name, value, unit, sample_count in [
            ("median_monthly_rent", "月租金中位数", "1800", "元/月", "64"),
            ("median_rent_per_sqm", "每平米租金中位数", "42", "元/月/平", "64"),
        ]:
            writer.writerow({
                **base,
                "housing_metric_key": metric_key,
                "housing_metric_name": metric_name,
                "metric_value": value,
                "metric_unit": unit,
                "sample_count": sample_count,
            })

    region_cost_input = tmp_path / "region_living_cost.csv"
    with region_cost_input.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "adcode",
                "region_name",
                "region_level",
                "province",
                "city",
                "district",
                "metric_key",
                "metric_name",
                "metric_value",
                "metric_unit",
                "metric_year",
                "sample_count",
                "source_provider",
                "source_title",
                "source_url",
                "source_scope",
                "evidence_quote",
                "raw_response_hash",
                "snapshot_date",
                "source_date",
                "availability_date",
                "built_at",
            ],
        )
        writer.writeheader()
        base = {
            "adcode": "210102",
            "region_name": "和平区",
            "region_level": "district",
            "province": "辽宁",
            "city": "沈阳",
            "district": "和平区",
            "metric_year": "2026",
            "sample_count": "20",
            "source_provider": "fixture_cost",
            "source_title": "fixture living cost",
            "source_url": "https://example.com/living-cost",
            "source_scope": "district fixture",
            "raw_response_hash": "hash-cost",
            "snapshot_date": "2026-05-14",
            "source_date": "2026-05-13",
            "availability_date": "2026-05-13",
            "built_at": "2026-05-13T00:00:00",
        }
        for metric_key, metric_name, value, unit in [
            ("student_meal_baseline", "学生餐费基准", "18", "元/餐"),
            ("living_service_density", "生活服务密度", "72", "score"),
            ("commute_time_to_city_center", "到核心区通勤时间", "32", "分钟"),
        ]:
            writer.writerow({
                **base,
                "metric_key": metric_key,
                "metric_name": metric_name,
                "metric_value": value,
                "metric_unit": unit,
                "evidence_quote": f"{metric_name}{value}",
            })

    result = build_campus_living_score_package(
        location_input=location_input,
        poi_input=poi_input,
        housing_input=housing_input,
        region_cost_input=region_cost_input,
        output_root=tmp_path / "exports",
        package_id="pkg-campus-living-score-test",
        source_version="fixture-campus-living",
    )

    package_dir = Path(result["package_dir"])
    assert validate_manifest(package_dir / "manifest.json")["errors"] == []
    assert result["rows"] == 1
    with (package_dir / "fa_mart_campus_living_score.csv").open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    row = rows[0]
    assert row["national_school_code"] == "4121010145"
    assert row["campus_name"] == "南湖校区"
    assert row["score_profile"] == "student_living_default"
    assert float(row["overall_score"]) > 0
    contributions = json.loads(row["signal_contribution_json"])
    assert "housing_cost_score" in contributions["components"]
    lineage = json.loads(row["pit_lineage_json"])
    assert "fa_fact_campus_housing_market" in lineage["tables"]
    assert result["quality_report"]["input_quality"]["errors"] == []


def test_audit_campus_living_score_inputs_reports_missing_files(tmp_path: Path):
    report_path = tmp_path / "campus_living_readiness.json"

    report = audit_campus_living_score_inputs(output=report_path)

    assert report_path.exists()
    assert report["ready_for_build"] is False
    assert report["location_rows"] == 0
    assert report["poi_rows"] == 0
    assert report["housing_rows"] == 0
    assert report["region_living_cost_rows"] == 0
    assert report["errors"] == [
        "location_input_missing",
        "poi_input_missing",
        "housing_input_missing",
        "region_cost_input_missing",
    ]


def test_build_campus_living_score_rejects_bad_input_metadata(tmp_path: Path):
    location_input = tmp_path / "school_location.csv"
    with location_input.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "national_school_code",
                "school_name",
                "campus_key",
                "address",
                "province",
                "city",
                "district",
                "adcode",
                "longitude",
                "latitude",
                "coordinate_system",
                "geocode_level",
                "geocode_confidence",
                "source_address_url",
                "source_date",
                "availability_date",
                "built_at",
            ],
        )
        writer.writeheader()
        writer.writerow({
            "national_school_code": "4121010145",
            "school_name": "东北大学",
            "campus_key": "0145_main",
            "address": "辽宁省沈阳市和平区文化路三巷",
            "province": "辽宁",
            "city": "沈阳",
            "district": "和平区",
            "adcode": "210102",
            "longitude": "123.425",
            "latitude": "41.774",
            "coordinate_system": "GCJ-02",
            "geocode_level": "门牌号",
            "geocode_confidence": "1.5",
            "source_address_url": "ftp://example.com/location",
            "source_date": "2026-05-14",
            "availability_date": "2026-05-13",
            "built_at": "2026-05-13T00:00:00",
        })

    poi_input = tmp_path / "campus_poi.csv"
    with poi_input.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "national_school_code",
                "school_name",
                "campus_key",
                "poi_id",
                "poi_name",
                "category_group",
                "distance_m",
                "source_provider",
                "source_url",
                "snapshot_date",
                "source_date",
                "availability_date",
                "built_at",
            ],
        )
        writer.writeheader()
        writer.writerow({
            "national_school_code": "4121010145",
            "school_name": "东北大学",
            "campus_key": "0145_main",
            "poi_id": "bad-poi",
            "poi_name": "坏设施",
            "category_group": "unknown_poi",
            "distance_m": "-1",
            "source_provider": "fixture",
            "source_url": "https://example.com/poi",
            "snapshot_date": "2026-05-14",
            "source_date": "2026-05-13",
            "availability_date": "2026-05-13",
            "built_at": "2026-05-13T00:00:00",
        })

    housing_input = tmp_path / "campus_housing.csv"
    with housing_input.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "national_school_code",
                "school_name",
                "campus_key",
                "radius_m",
                "listing_type",
                "housing_metric_key",
                "metric_value",
                "metric_unit",
                "sample_count",
                "source_platform",
                "source_url",
                "snapshot_date",
                "source_date",
                "availability_date",
                "built_at",
            ],
        )
        writer.writeheader()
        writer.writerow({
            "national_school_code": "4121010145",
            "school_name": "东北大学",
            "campus_key": "0145_main",
            "radius_m": "3000",
            "listing_type": "lease",
            "housing_metric_key": "unknown_metric",
            "metric_value": "1800",
            "metric_unit": "元/月",
            "sample_count": "10.5",
            "source_platform": "fixture",
            "source_url": "https://example.com/housing",
            "snapshot_date": "2026-05-14",
            "source_date": "2026-05-13",
            "availability_date": "2026-05-13",
            "built_at": "2026-05-13T00:00:00",
        })

    region_cost_input = tmp_path / "region_living_cost.csv"
    with region_cost_input.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "adcode",
                "region_name",
                "region_level",
                "city",
                "metric_key",
                "metric_name",
                "metric_value",
                "metric_unit",
                "metric_year",
                "source_provider",
                "source_url",
                "snapshot_date",
                "source_date",
                "availability_date",
                "built_at",
            ],
        )
        writer.writeheader()
        writer.writerow({
            "adcode": "210102",
            "region_name": "和平区",
            "region_level": "district",
            "city": "沈阳",
            "metric_key": "unknown_cost",
            "metric_name": "未知成本",
            "metric_value": "18",
            "metric_unit": "元",
            "metric_year": "2026.5",
            "source_provider": "fixture",
            "source_url": "ftp://example.com/cost",
            "snapshot_date": "2026-05-14",
            "source_date": "2026-05-14",
            "availability_date": "2026-05-13",
            "built_at": "2026-05-13T00:00:00",
        })

    with pytest.raises(ValueError) as exc:
        build_campus_living_score_package(
            location_input=location_input,
            poi_input=poi_input,
            housing_input=housing_input,
            region_cost_input=region_cost_input,
            output_root=tmp_path / "exports",
            package_id="pkg-campus-living-score-bad-input",
            source_version="fixture-campus-living",
        )

    message = str(exc.value)
    assert "location row 1 source_address_url must be an http(s) URL" in message
    assert "location row 1 geocode_confidence outside 0-1: 1.5" in message
    assert "poi row 1 unregistered category_group: unknown_poi" in message
    assert "poi row 1 distance_m below 0: -1" in message
    assert "housing row 1 unregistered listing_type: lease" in message
    assert "housing row 1 unregistered housing_metric_key: unknown_metric" in message
    assert "housing row 1 sample_count is not an integer" in message
    assert "region_living_cost row 1 source_url must be an http(s) URL" in message
    assert "region_living_cost row 1 metric_year is not an integer" in message
    assert "region_living_cost row 1 source_date must not be after availability_date" in message


def test_build_school_city_industry_fit_package(tmp_path: Path):
    recruitment_input = tmp_path / "school_recruitment.csv"
    with recruitment_input.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "national_school_code",
                "local_school_code",
                "school_name",
                "campus_key",
                "city",
                "event_id",
                "event_title",
                "event_type",
                "event_date",
                "event_year",
                "employer_name",
                "employer_canonical_name",
                "employer_city",
                "employer_industry_tdx_l2",
                "employer_industry_tdx_l2_name",
                "target_majors_json",
                "job_roles_json",
                "source_title",
                "source_url",
                "raw_response_hash",
                "source_date",
                "availability_date",
                "built_at",
            ],
        )
        writer.writeheader()
        base = {
            "national_school_code": "4121010145",
            "local_school_code": "0145",
            "school_name": "东北大学",
            "campus_key": "0145_main",
            "city": "沈阳",
            "event_year": "2026",
            "employer_city": "沈阳",
            "employer_industry_tdx_l2": "T1205",
            "employer_industry_tdx_l2_name": "软件服务",
            "target_majors_json": json.dumps(["计算机科学与技术", "软件工程"], ensure_ascii=False),
            "source_title": "fixture recruitment",
            "source_url": "https://example.com/recruitment",
            "raw_response_hash": "hash-recruitment",
            "source_date": "2026-05-13",
            "availability_date": "2026-05-13",
            "built_at": "2026-05-13T00:00:00",
        }
        for event_id, event_type, employer, roles in [
            ("r1", "campus_talk", "沈阳软件A", ["后端工程师", "测试工程师"]),
            ("r2", "internship", "沈阳软件B", ["实习工程师"]),
            ("r3", "career_fair", "沈阳软件C", ["产品经理"]),
        ]:
            writer.writerow({
                **base,
                "event_id": event_id,
                "event_title": f"{employer} 宣讲会",
                "event_type": event_type,
                "event_date": "2026-04-10",
                "employer_name": employer,
                "employer_canonical_name": employer,
                "job_roles_json": json.dumps(roles, ensure_ascii=False),
            })

    research_input = tmp_path / "school_research.csv"
    with research_input.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "national_school_code",
                "local_school_code",
                "school_name",
                "platform_name",
                "platform_level",
                "discipline",
                "research_keywords_json",
                "tdx_l2",
                "tdx_l2_name",
                "city",
                "linked_company_name",
                "industry_zone_name",
                "evidence_quote",
                "source_title",
                "source_url",
                "source_date",
                "availability_date",
                "built_at",
            ],
        )
        writer.writeheader()
        writer.writerow({
            "national_school_code": "4121010145",
            "local_school_code": "0145",
            "school_name": "东北大学",
            "platform_name": "软件工程国家重点实验室",
            "platform_level": "national",
            "discipline": "软件工程",
            "research_keywords_json": json.dumps(["软件", "工业互联网"], ensure_ascii=False),
            "tdx_l2": "T1205",
            "tdx_l2_name": "软件服务",
            "city": "沈阳",
            "linked_company_name": "沈阳软件A",
            "industry_zone_name": "沈阳软件园",
            "evidence_quote": "实验室与本地软件企业合作。",
            "source_title": "fixture research",
            "source_url": "https://example.com/research",
            "source_date": "2026-05-13",
            "availability_date": "2026-05-13",
            "built_at": "2026-05-13T00:00:00",
        })

    employment_input = tmp_path / "school_employment.csv"
    with employment_input.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "national_school_code",
                "local_school_code",
                "school_name",
                "province",
                "city",
                "metric_key",
                "metric_name",
                "metric_value",
                "metric_unit",
                "metric_year",
                "metric_scope",
                "employer_name",
                "industry_tdx_l2",
                "industry_tdx_l2_name",
                "source_title",
                "source_url",
                "evidence_quote",
                "source_date",
                "availability_date",
                "built_at",
            ],
        )
        writer.writeheader()
        base = {
            "national_school_code": "4121010145",
            "local_school_code": "0145",
            "school_name": "东北大学",
            "province": "辽宁",
            "city": "沈阳",
            "metric_year": "2024",
            "metric_scope": "本科毕业生",
            "employer_name": "沈阳软件A",
            "industry_tdx_l2": "T1205",
            "industry_tdx_l2_name": "软件服务",
            "source_title": "fixture employment",
            "source_url": "https://example.com/employment",
            "source_date": "2026-05-13",
            "availability_date": "2026-05-13",
            "built_at": "2026-05-13T00:00:00",
        }
        for metric_key, metric_name, value, unit in [
            ("local_city_retention_rate", "留沈比例", "0.46", "ratio"),
            ("top_employer_hire_count", "重点雇主录用人数", "80", "count"),
            ("top_industry_share", "重点行业占比", "0.31", "ratio"),
        ]:
            writer.writerow({
                **base,
                "metric_key": metric_key,
                "metric_name": metric_name,
                "metric_value": value,
                "metric_unit": unit,
                "evidence_quote": f"{metric_name}{value}",
            })

    zone_input = tmp_path / "city_zone.csv"
    with zone_input.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "zone_id",
                "zone_name",
                "zone_level",
                "province",
                "city",
                "district",
                "adcode",
                "longitude",
                "latitude",
                "coordinate_system",
                "tdx_l2",
                "tdx_l2_name",
                "anchor_companies_json",
                "source_title",
                "source_url",
                "evidence_quote",
                "source_date",
                "availability_date",
                "built_at",
            ],
        )
        writer.writeheader()
        writer.writerow({
            "zone_id": "sy-software-park",
            "zone_name": "沈阳软件园",
            "zone_level": "municipal",
            "province": "辽宁",
            "city": "沈阳",
            "district": "浑南区",
            "adcode": "210112",
            "longitude": "123.46",
            "latitude": "41.73",
            "coordinate_system": "GCJ-02",
            "tdx_l2": "T1205",
            "tdx_l2_name": "软件服务",
            "anchor_companies_json": json.dumps(["沈阳软件A"], ensure_ascii=False),
            "source_title": "fixture zone",
            "source_url": "https://example.com/zone",
            "evidence_quote": "沈阳软件园聚集软件服务企业。",
            "source_date": "2026-05-13",
            "availability_date": "2026-05-13",
            "built_at": "2026-05-13T00:00:00",
        })

    location_input = tmp_path / "school_location.csv"
    with location_input.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "national_school_code",
                "local_school_code",
                "school_name",
                "campus_key",
                "campus_name",
                "address",
                "province",
                "city",
                "district",
                "adcode",
                "longitude",
                "latitude",
                "coordinate_system",
                "geocode_level",
                "source_address_url",
                "source_date",
                "availability_date",
                "built_at",
            ],
        )
        writer.writeheader()
        writer.writerow({
            "national_school_code": "4121010145",
            "local_school_code": "0145",
            "school_name": "东北大学",
            "campus_key": "0145_main",
            "campus_name": "南湖校区",
            "address": "辽宁省沈阳市和平区文化路三巷",
            "province": "辽宁",
            "city": "沈阳",
            "district": "和平区",
            "adcode": "210102",
            "longitude": "123.425",
            "latitude": "41.774",
            "coordinate_system": "GCJ-02",
            "geocode_level": "门牌号",
            "source_address_url": "https://example.com/location",
            "source_date": "2026-05-13",
            "availability_date": "2026-05-13",
            "built_at": "2026-05-13T00:00:00",
        })

    result = build_school_city_industry_fit_package(
        recruitment_input=recruitment_input,
        research_input=research_input,
        employment_input=employment_input,
        zone_input=zone_input,
        location_input=location_input,
        output_root=tmp_path / "exports",
        package_id="pkg-school-city-industry-fit-test",
        source_version="fixture-school-city-industry",
    )

    package_dir = Path(result["package_dir"])
    assert validate_manifest(package_dir / "manifest.json")["errors"] == []
    assert result["rows"] == 1
    with (package_dir / "fa_mart_school_city_industry_fit.csv").open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    row = rows[0]
    assert row["national_school_code"] == "4121010145"
    assert row["city"] == "沈阳"
    assert row["tdx_l2"] == "T1205"
    assert row["score_profile"] == "city_industry_default"
    assert float(row["overall_score"]) > 0
    assert float(row["zone_proximity_score"]) > 0
    contributions = json.loads(row["signal_contribution_json"])
    assert "recruitment_score" in contributions["components"]
    lineage = json.loads(row["pit_lineage_json"])
    assert "fa_dim_city_industry_zone" in lineage["tables"]
    assert result["quality_report"]["input_quality"]["errors"] == []


def test_audit_school_city_industry_fit_inputs_reports_missing_files(tmp_path: Path):
    report_path = tmp_path / "school_city_industry_readiness.json"

    report = audit_school_city_industry_fit_inputs(output=report_path)

    assert report_path.exists()
    assert report["ready_for_build"] is False
    assert report["recruitment_rows"] == 0
    assert report["research_rows"] == 0
    assert report["employment_rows"] == 0
    assert report["zone_rows"] == 0
    assert report["location_rows"] == 0
    assert report["errors"] == [
        "recruitment_input_missing",
        "research_input_missing",
        "employment_input_missing",
        "zone_input_missing",
        "location_input_missing",
    ]


def test_build_school_city_industry_fit_rejects_bad_input_metadata(tmp_path: Path):
    recruitment_input = tmp_path / "school_recruitment.csv"
    with recruitment_input.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "national_school_code",
                "school_name",
                "city",
                "event_id",
                "event_type",
                "event_date",
                "event_year",
                "employer_name",
                "employer_industry_tdx_l2",
                "source_title",
                "source_url",
                "source_date",
                "availability_date",
                "built_at",
            ],
        )
        writer.writeheader()
        writer.writerow({
            "national_school_code": "4121010145",
            "school_name": "东北大学",
            "city": "沈阳",
            "event_id": "bad-event",
            "event_type": "unsupported",
            "event_date": "2026/04/10",
            "event_year": "2026.5",
            "employer_name": "沈阳软件A",
            "employer_industry_tdx_l2": "T1205",
            "source_title": "fixture recruitment",
            "source_url": "ftp://example.com/recruitment",
            "source_date": "2026-05-14",
            "availability_date": "2026-05-13",
            "built_at": "2026-05-13T00:00:00",
        })

    research_input = tmp_path / "school_research.csv"
    with research_input.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "national_school_code",
                "school_name",
                "platform_name",
                "platform_level",
                "tdx_l2",
                "tdx_l2_name",
                "city",
                "source_url",
                "source_date",
                "availability_date",
                "built_at",
            ],
        )
        writer.writeheader()
        writer.writerow({
            "national_school_code": "4121010145",
            "school_name": "东北大学",
            "platform_name": "平台",
            "platform_level": "unknown_new",
            "tdx_l2": "",
            "tdx_l2_name": "",
            "city": "沈阳",
            "source_url": "https://example.com/research",
            "source_date": "2026-05-13",
            "availability_date": "2026-05-13",
            "built_at": "2026-05-13T00:00:00",
        })

    employment_input = tmp_path / "school_employment.csv"
    with employment_input.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "national_school_code",
                "school_name",
                "city",
                "metric_key",
                "metric_name",
                "metric_value",
                "metric_unit",
                "metric_year",
                "industry_tdx_l2",
                "source_title",
                "source_url",
                "source_date",
                "availability_date",
                "built_at",
            ],
        )
        writer.writeheader()
        writer.writerow({
            "national_school_code": "4121010145",
            "school_name": "东北大学",
            "city": "沈阳",
            "metric_key": "unknown_metric",
            "metric_name": "未知指标",
            "metric_value": "abc",
            "metric_unit": "ratio",
            "metric_year": "2026.5",
            "industry_tdx_l2": "",
            "source_title": "fixture employment",
            "source_url": "https://example.com/employment",
            "source_date": "2026-05-13",
            "availability_date": "2026-05-13",
            "built_at": "2026-05-13T00:00:00",
        })

    zone_input = tmp_path / "city_zone.csv"
    with zone_input.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "zone_id",
                "zone_name",
                "city",
                "longitude",
                "latitude",
                "tdx_l2",
                "tdx_l2_name",
                "source_url",
                "source_date",
                "availability_date",
                "built_at",
            ],
        )
        writer.writeheader()
        writer.writerow({
            "zone_id": "bad-zone",
            "zone_name": "坏园区",
            "city": "沈阳",
            "longitude": "bad",
            "latitude": "100",
            "tdx_l2": "T1205",
            "tdx_l2_name": "软件服务",
            "source_url": "https://example.com/zone",
            "source_date": "2026-05-13",
            "availability_date": "2026-05-13",
            "built_at": "2026-05-13T00:00:00",
        })

    location_input = tmp_path / "school_location.csv"
    with location_input.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "national_school_code",
                "school_name",
                "campus_key",
                "address",
                "province",
                "city",
                "district",
                "adcode",
                "longitude",
                "latitude",
                "coordinate_system",
                "geocode_level",
                "source_address_url",
                "source_date",
                "availability_date",
                "built_at",
            ],
        )
        writer.writeheader()
        writer.writerow({
            "national_school_code": "4121010145",
            "school_name": "东北大学",
            "campus_key": "0145_main",
            "address": "辽宁省沈阳市和平区文化路三巷",
            "province": "辽宁",
            "city": "沈阳",
            "district": "和平区",
            "adcode": "210102",
            "longitude": "123.425",
            "latitude": "41.774",
            "coordinate_system": "GCJ-02",
            "geocode_level": "门牌号",
            "source_address_url": "ftp://example.com/location",
            "source_date": "2026-05-13",
            "availability_date": "2026-05-13",
            "built_at": "2026-05-13T00:00:00",
        })

    with pytest.raises(ValueError) as exc:
        build_school_city_industry_fit_package(
            recruitment_input=recruitment_input,
            research_input=research_input,
            employment_input=employment_input,
            zone_input=zone_input,
            location_input=location_input,
            output_root=tmp_path / "exports",
            package_id="pkg-school-city-industry-fit-bad-input",
            source_version="fixture-school-city-industry",
        )

    message = str(exc.value)
    assert "recruitment row 1 source_url must be an http(s) URL" in message
    assert "recruitment row 1 source_date must not be after availability_date" in message
    assert "recruitment row 1 unregistered event_type: unsupported" in message
    assert "recruitment row 1 event_date must use YYYY-MM-DD" in message
    assert "recruitment row 1 event_year is not an integer" in message
    assert "research row 1 missing tdx_l2" in message
    assert "employment row 1 unregistered metric_key: unknown_metric" in message
    assert "employment row 1 metric_year is not an integer" in message
    assert "employment row 1 metric_value is not numeric: None" in message
    assert "employment row 1 missing industry_tdx_l2" in message
    assert "zone row 1 latitude outside range: 100" in message
    assert "location row 1 source_address_url must be an http(s) URL" in message


def test_build_city_development_score_package(tmp_path: Path):
    economic_input = tmp_path / "city_economic.csv"
    with economic_input.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "adcode",
                "province",
                "city",
                "region_level",
                "metric_key",
                "metric_name",
                "metric_value",
                "metric_unit",
                "metric_year",
                "metric_scope",
                "source_title",
                "source_url",
                "evidence_quote",
                "source_date",
                "availability_date",
                "built_at",
            ],
        )
        writer.writeheader()
        base = {
            "adcode": "210100",
            "province": "辽宁",
            "city": "沈阳",
            "region_level": "city",
            "metric_year": "2025",
            "metric_scope": "年度统计",
            "source_title": "fixture economic bulletin",
            "source_date": "2026-05-13",
            "availability_date": "2026-05-13",
            "built_at": "2026-05-13T00:00:00",
        }
        for metric_key, metric_name, value, unit in [
            ("gdp", "地区生产总值", 9000, "亿元"),
            ("gdp_per_capita", "人均地区生产总值", 120000, "元"),
            ("urban_avg_wage", "城镇平均工资", 105000, "元"),
            ("tertiary_industry_share", "第三产业占比", 0.58, "ratio"),
        ]:
            writer.writerow({
                **base,
                "metric_key": metric_key,
                "metric_name": metric_name,
                "metric_value": value,
                "metric_unit": unit,
                "source_url": f"https://example.com/economic/{metric_key}",
                "evidence_quote": f"{metric_name}{value}",
            })

    public_input = tmp_path / "city_public.csv"
    with public_input.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "adcode",
                "province",
                "city",
                "region_level",
                "resource_domain",
                "metric_key",
                "metric_name",
                "metric_value",
                "metric_unit",
                "metric_year",
                "metric_scope",
                "source_title",
                "source_url",
                "evidence_quote",
                "source_date",
                "availability_date",
                "built_at",
            ],
        )
        writer.writeheader()
        base = {
            "adcode": "210100",
            "province": "辽宁",
            "city": "沈阳",
            "region_level": "city",
            "metric_year": "2025",
            "metric_scope": "年度统计",
            "source_title": "fixture public resource",
            "source_date": "2026-05-13",
            "availability_date": "2026-05-13",
            "built_at": "2026-05-13T00:00:00",
        }
        for domain, metric_key, metric_name, value, unit in [
            ("medical", "hospital_beds_per_1000", "每千人口床位数", 7.2, "张/千人"),
            ("medical", "licensed_doctors_per_1000", "每千人口医师数", 4.1, "人/千人"),
            ("education", "higher_education_institution_count", "普通高校数量", 45, "所"),
            ("transport", "rail_transit_mileage", "轨道交通运营里程", 220, "公里"),
        ]:
            writer.writerow({
                **base,
                "resource_domain": domain,
                "metric_key": metric_key,
                "metric_name": metric_name,
                "metric_value": value,
                "metric_unit": unit,
                "source_url": f"https://example.com/public/{metric_key}",
                "evidence_quote": f"{metric_name}{value}",
            })

    listed_input = tmp_path / "city_listed.csv"
    with listed_input.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "province",
                "city",
                "tdx_l2",
                "tdx_l2_name",
                "metric_key",
                "metric_name",
                "metric_value",
                "metric_unit",
                "metric_year",
                "source_system",
                "source_scope",
                "source_date",
                "availability_date",
                "built_at",
            ],
        )
        writer.writeheader()
        base = {
            "province": "辽宁",
            "city": "沈阳",
            "tdx_l2": "all",
            "tdx_l2_name": "全部行业",
            "metric_year": "2025",
            "source_system": "fixture_snapshot",
            "source_scope": "城市上市公司聚合",
            "source_date": "2026-05-13",
            "availability_date": "2026-05-13",
            "built_at": "2026-05-13T00:00:00",
        }
        for metric_key, metric_name, value, unit in [
            ("listed_company_count", "上市公司数量", 92, "count"),
            ("listed_company_tdx_l2_count", "上市公司覆盖行业数", 26, "count"),
            ("listed_company_revenue_proxy", "上市公司营收代理", 320000000000, "cny"),
        ]:
            writer.writerow({
                **base,
                "metric_key": metric_key,
                "metric_name": metric_name,
                "metric_value": value,
                "metric_unit": unit,
            })

    result = build_city_development_score_package(
        economic_input=economic_input,
        public_resource_input=public_input,
        listed_company_input=listed_input,
        output_root=tmp_path / "exports",
        package_id="pkg-city-development-score-test",
        source_version="fixture-city-development",
    )

    package_dir = Path(result["package_dir"])
    assert validate_manifest(package_dir / "manifest.json")["errors"] == []
    assert result["rows"] == 1
    with (package_dir / "fa_mart_city_development_score.csv").open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    row = rows[0]
    assert row["adcode"] == "210100"
    assert row["city"] == "沈阳"
    assert row["score_profile"] == "city_development_default"
    assert float(row["overall_score"]) > 0
    contributions = json.loads(row["signal_contribution_json"])
    assert "industry_depth_score" in contributions["components"]
    lineage = json.loads(row["pit_lineage_json"])
    assert "fa_fact_city_public_resource" in lineage["tables"]
    assert result["quality_report"]["input_quality"]["errors"] == []


def test_build_city_development_score_rejects_bad_input_metadata(tmp_path: Path):
    economic_input = tmp_path / "city_economic.csv"
    with economic_input.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "adcode",
                "province",
                "city",
                "region_level",
                "metric_key",
                "metric_name",
                "metric_value",
                "metric_unit",
                "metric_year",
                "metric_scope",
                "source_title",
                "source_url",
                "evidence_quote",
                "source_date",
                "availability_date",
                "built_at",
            ],
        )
        writer.writeheader()
        writer.writerow({
            "adcode": "210100",
            "province": "辽宁",
            "city": "沈阳",
            "region_level": "city",
            "metric_key": "unknown_metric",
            "metric_name": "未知指标",
            "metric_value": "9000",
            "metric_unit": "亿元",
            "metric_year": "2025.5",
            "metric_scope": "年度统计",
            "source_title": "fixture economic bulletin",
            "source_url": "ftp://example.com/bad-economic",
            "evidence_quote": "未知指标9000",
            "source_date": "2026-05-14",
            "availability_date": "2026-05-13",
            "built_at": "2026-05-13T00:00:00",
        })

    public_input = tmp_path / "city_public.csv"
    with public_input.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "adcode",
                "province",
                "city",
                "region_level",
                "resource_domain",
                "metric_key",
                "metric_name",
                "metric_value",
                "metric_unit",
                "metric_year",
                "metric_scope",
                "source_title",
                "source_url",
                "evidence_quote",
                "source_date",
                "availability_date",
                "built_at",
            ],
        )
        writer.writeheader()

    listed_input = tmp_path / "city_listed.csv"
    with listed_input.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "province",
                "city",
                "tdx_l2",
                "tdx_l2_name",
                "metric_key",
                "metric_name",
                "metric_value",
                "metric_unit",
                "metric_year",
                "source_system",
                "source_scope",
                "source_date",
                "availability_date",
                "built_at",
            ],
        )
        writer.writeheader()

    with pytest.raises(ValueError) as exc:
        build_city_development_score_package(
            economic_input=economic_input,
            public_resource_input=public_input,
            listed_company_input=listed_input,
            output_root=tmp_path / "exports",
            package_id="pkg-city-development-score-bad-input",
            source_version="fixture-city-development",
        )

    message = str(exc.value)
    assert "economic row 1 unregistered metric_key: unknown_metric" in message
    assert "economic row 1 metric_year is not an integer" in message
    assert "economic row 1 source_url must be an http(s) URL" in message
    assert "economic row 1 source_date must not be after availability_date" in message


def test_entity_normalization_registry_config_and_schemas():
    config = load_entity_normalization()
    assert config["source_key"] == "entity_normalization_registry"
    assert "city" in config["entity_types"]
    assert "school" in config["entity_types"]
    assert config["model_input_policy"]["require_canonical_entity_id"] is True
    stages = [stage["stage"] for stage in config["pipeline_stages"]]
    assert stages[:4] == ["raw_intake", "parse", "normalize", "canonicalize"]

    schemas = load_source_schemas()["tables"]
    assert schemas["fa_dim_entity_registry"]["primary_key"] == ["entity_id"]
    assert schemas["fa_bridge_entity_alias"]["primary_key"] == [
        "entity_type",
        "alias_name",
        "source_system",
        "alias_scope",
    ]
    assert schemas["fa_dim_metric_registry"]["primary_key"] == ["metric_domain", "metric_key"]
    assert schemas["fa_bridge_metric_alias"]["primary_key"] == ["metric_domain", "alias_name", "source_system"]


def test_data_update_policy_config_and_schemas():
    config = load_data_update_policy()
    assert config["source_key"] == "data_update_governance"
    assert "manual_review_promote" in config["update_modes"]
    assert "append_snapshot" in config["update_modes"]
    assert config["nonstandard_policy"]["old_data_policy"]["never_delete_without_delete_plan"] is True
    assert config["update_mode_runbook"]["manual_review_promote"]["old_data_handling"]
    assert config["validity_check_catalog"]["required_evidence_present"]["block_on_fail"] is True
    assert config["scheduler"]["failure_policy"]["block_dependents_on_source_failure"] is True
    assert config["state_management"]["delete_policy"]["require_delete_plan"] is True
    assert "schema_changed" in config["source_health_policy"]["statuses"]
    assert config["scheduler"]["batching"]["target_table_locking"] is True
    assert "amap_api_limited" in config["scheduler"]["serial_groups"]
    assert "city_collection_parallel" in config["scheduler"]["parallel_groups"]
    assert config["source_policies"]["ln_admission_plan"]["depends_on"] == ["ln_application_workbook"]
    assert config["source_policies"]["ln_score_history"]["depends_on"] == [
        "ln_projection_score",
        "ln_score_distribution",
    ]
    assert config["source_policies"]["school_profile_supplemental"]["validity_profile"] == "manual_file"
    assert "school_profile_supplemental" in config["source_policies"]["school_identity_bridge"]["depends_on"]
    assert config["source_policies"]["city_economic_indicator"]["parallelizable"] is True
    assert config["source_policies"]["region_profile_geocode"]["parallelizable"] is False
    assert "city_listed_company_signal" in config["source_policies"]["city_development_score"]["depends_on"]
    assert "school_city_industry_fit" in config["source_policies"]
    assert "major_city_employment_fit" in config["source_policies"]
    assert config["source_policies"]["career_civil_service_posts"]["validity_profile"] == "web_api"
    assert "career_civil_service_posts" in config["source_policies"]["career_signal"]["depends_on"]

    schemas = load_source_schemas()["tables"]
    assert schemas["fa_meta_source_snapshot"]["primary_key"] == ["source_key", "snapshot_id"]
    assert schemas["fa_meta_source_health"]["primary_key"] == ["source_key", "check_at", "check_type"]
    assert schemas["fa_meta_update_run"]["primary_key"] == ["update_run_id"]
    assert schemas["fa_meta_update_run_step"]["primary_key"] == ["update_run_id", "source_key", "step_key"]
    assert schemas["fa_meta_nonstandard_review_queue"]["primary_key"] == ["review_id"]
    assert schemas["fa_fact_civil_service_position"]["source_key"] == "career_civil_service_posts"
    assert schemas["fa_fact_civil_service_position"]["primary_key"] == ["source_date", "sheet_name", "row_number"]


def test_pipeline_error_policy_governs_modular_tools_and_llm_escalation():
    config = load_pipeline_error_policy()
    assert config["source_key"] == "pipeline_error_policy"

    adapter_contract = config["architecture"]["tool_adapter_contract"]
    assert "core_db_write" in adapter_contract["forbidden_outputs"]
    assert "api_key_plaintext" in adapter_contract["forbidden_outputs"]
    assert "raw_snapshot" in adapter_contract["output_artifacts"]
    assert "candidate_csv" in adapter_contract["output_artifacts"]

    scheduler_contract = config["architecture"]["unified_scheduler_contract"]
    assert "config/data_update_policy.json" in scheduler_contract["config_inputs"]
    assert "fa_meta_update_run_step" in scheduler_contract["runtime_outputs"]
    assert "core_dry_run_ok" in scheduler_contract["gates"]

    layers = config["adapter_layers"]
    assert layers["raw_intake"]["can_publish_package"] is False
    assert layers["parse_extract"]["can_publish_package"] is False
    assert layers["review_promote"]["can_publish_package"] is False
    assert layers["package_build"]["can_publish_package"] is True

    errors = config["error_classes"]
    assert errors["captcha_or_login_required"]["severity"] == "blocking"
    assert errors["captcha_or_login_required"]["automatic_action"] == "mark_manual_intake_required"
    assert errors["core_dry_run_failed"]["severity"] == "critical"
    assert errors["core_dry_run_failed"]["automatic_action"] == "do_not_execute_import"
    assert errors["metric_value_invalid"]["llm_escalation"] == "forbidden_for_value_decision"
    assert errors["secret_leak_risk"]["llm_escalation"] == "forbidden"

    llm_policy = config["llm_escalation_policy"]
    assert llm_policy["default"] == "deterministic_first"
    assert "parser_patch_draft" in llm_policy["allowed_tasks"]
    assert "invent_metric_value" in llm_policy["forbidden_tasks"]
    assert "bypass_captcha_or_login" in llm_policy["forbidden_tasks"]
    assert "include_current_audit_report" in llm_policy["handoff_requirements"]

    command_center = config["llm_command_center"]
    assert command_center["concept_name"] == "metadata_driven_data_control_plane"
    assert command_center["normal_mode"] == "deterministic_pipeline_runs_without_llm"
    assert command_center["incident_mode"] == "llm_reads_artifacts_and_proposes_repair_plan"
    assert "audit_report" in command_center["required_context"]
    assert "readiness_report" in command_center["required_context"]
    assert "draft_parser_patch" in command_center["allowed_commands"]
    assert "generate_review_batch_plan" in command_center["allowed_commands"]
    assert "execute_core_import" in command_center["blocked_commands"]
    assert "publish_unreviewed_package" in command_center["blocked_commands"]
    assert "source_card" in command_center["knowledge_cells"]
    assert "pitfall_card" in command_center["knowledge_cells"]
    assert "core_write_requires_dry_run_ok" in command_center["decision_gates"]

    external_patterns = config["external_tool_patterns"]
    assert "great_expectations" in external_patterns
    assert "pandera" in external_patterns
    assert "dagster" in external_patterns
    assert "prefect" in external_patterns


def test_build_data_update_plan_with_dependencies(tmp_path: Path):
    result = build_data_update_plan(
        output_dir=tmp_path / "update_plan",
        source_keys=["city_development_score"],
        update_run_id="fixture_city_update",
    )

    assert result["blocked_steps"] == 0
    with Path(result["csv"]).open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    by_key = {row["source_key"]: row for row in rows}
    assert set(by_key) == {
        "region_profile_geocode",
        "city_economic_indicator",
        "city_public_resource",
        "city_listed_company_signal",
        "city_ranking_signal",
        "city_development_score",
    }

    assert int(by_key["region_profile_geocode"]["phase"]) == 1
    assert int(by_key["city_economic_indicator"]["phase"]) > int(by_key["region_profile_geocode"]["phase"])
    assert int(by_key["city_public_resource"]["phase"]) == int(by_key["city_economic_indicator"]["phase"])
    assert int(by_key["city_listed_company_signal"]["phase"]) == int(by_key["city_economic_indicator"]["phase"])
    assert int(by_key["city_development_score"]["phase"]) > int(by_key["city_economic_indicator"]["phase"])
    assert by_key["region_profile_geocode"]["execution_group"] == "serial:amap_api_limited"
    assert by_key["city_economic_indicator"]["execution_group"] == "parallel:city_collection_parallel"
    assert by_key["city_listed_company_signal"]["execution_group"] == "parallel:city_collection_parallel"
    assert "metric_registered" in json.loads(by_key["city_ranking_signal"]["validity_checks"])
    assert json.loads(by_key["city_development_score"]["depends_on"]) == [
        "city_economic_indicator",
        "city_public_resource",
        "city_listed_company_signal",
        "city_ranking_signal",
    ]

    manifest = json.loads(Path(result["manifest"]).read_text(encoding="utf-8"))
    assert manifest["update_run_id"] == "fixture_city_update"
    assert manifest["include_dependencies"] is True
    assert "2" in manifest["phases"]
    assert manifest["phases"]["2"]["execution_groups"]["parallel:city_collection_parallel"]


def test_build_data_update_plan_without_dependencies_marks_blocked(tmp_path: Path):
    result = build_data_update_plan(
        output_dir=tmp_path / "partial_update_plan",
        source_keys=["city_development_score"],
        include_dependencies=False,
        update_run_id="fixture_partial_update",
    )

    with Path(result["csv"]).open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    assert len(rows) == 1
    assert rows[0]["source_key"] == "city_development_score"
    assert rows[0]["step_status"] == "blocked"
    assert "dependency_not_in_plan" in rows[0]["block_reason"]
    assert set(json.loads(rows[0]["missing_dependencies"])) == {
        "city_economic_indicator",
        "city_public_resource",
        "city_listed_company_signal",
        "city_ranking_signal",
    }


def test_build_data_update_readiness_plan_expands_validity_checks(tmp_path: Path):
    result = build_data_update_readiness_plan(
        output_dir=tmp_path / "readiness_plan",
        source_keys=["city_development_score"],
        update_run_id="fixture_city_readiness",
    )

    assert result["blocking_check_rows"] == 0
    with Path(result["csv"]).open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    assert result["rows"] == len(rows)
    by_source_check = {(row["source_key"], row["check_key"]): row for row in rows}
    city_metric = by_source_check[("city_economic_indicator", "required_evidence_present")]
    assert city_metric["current_status"] == "awaiting_collection_review"
    assert "source_url" in city_metric["expected_evidence"]
    assert "只替换本次分区" in city_metric["incremental_strategy"]

    derived = by_source_check[("city_development_score", "input_package_lineage_present")]
    assert derived["current_status"] == "planned_after_dependencies"
    assert "从已发布标准表重建 mart" in derived["incremental_strategy"]


def test_build_data_update_readiness_plan_blocks_missing_dependencies(tmp_path: Path):
    result = build_data_update_readiness_plan(
        output_dir=tmp_path / "partial_readiness_plan",
        source_keys=["city_development_score"],
        include_dependencies=False,
        update_run_id="fixture_partial_readiness",
    )

    assert result["blocking_check_rows"] > 0
    with Path(result["csv"]).open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    assert {row["current_status"] for row in rows} == {"blocked_by_dependency"}
    assert "dependency_not_in_plan" in rows[0]["notes"]


def test_build_data_update_batch_plan_groups_parallel_sources(tmp_path: Path):
    result = build_data_update_batch_plan(
        output_dir=tmp_path / "batch_plan",
        source_keys=["city_development_score"],
        update_run_id="fixture_city_batches",
    )

    assert result["blocked_batches"] == 0
    with Path(result["csv"]).open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    by_group = {row["execution_group"]: row for row in rows}
    city_batch = by_group["parallel:city_collection_parallel"]
    assert city_batch["concurrency_mode"] == "parallel"
    assert city_batch["max_parallel"] == "3"
    assert set(json.loads(city_batch["source_keys"])) == {
        "city_economic_indicator",
        "city_public_resource",
        "city_listed_company_signal",
        "city_ranking_signal",
    }

    derived_batch = by_group["parallel:derived_readonly_mart"]
    assert json.loads(derived_batch["source_keys"]) == ["city_development_score"]
    assert "phase_complete_without_blocking_failure" in derived_batch["dependency_gate"]


def test_audit_score_source_coverage_tracks_derivation_gaps(tmp_path: Path):
    report_path = tmp_path / "score_source_coverage.json"
    report = audit_score_source_coverage(report_path=report_path)

    assert report_path.exists()
    by_year = {row["score_year"]: row for row in report["coverage_by_year"]}
    assert by_year[2025]["derivation_status"] == "official_remote_derivable"
    assert by_year[2025]["projection_status"] == "official_remote_ready"
    assert by_year[2025]["distribution_status"] == "official_remote_ready"
    assert by_year[2024]["derivation_status"] == "derivable_with_mirror_inputs"
    assert by_year[2024]["score_distribution"]["official_page_image_count"] >= 1
    assert by_year[2024]["score_distribution"]["official_grid_image_group_count"] == 2
    assert by_year[2024]["score_distribution"]["research_candidate_count"] >= 3
    assert "https://www.centv.cn/p/514016.html" in by_year[2024]["score_distribution"]["candidate_urls"]
    assert (
        "https://www.centv.cn/a/10001/202406/fa32ce4a7f0c38dd31301a067c430714.pdf"
        in by_year[2024]["score_distribution"]["candidate_urls"]
    )
    assert by_year[2023]["projection_status"] == "official_remote_ready"
    assert by_year[2023]["score_distribution"]["official_grid_image_group_count"] == 2
    assert by_year[2023]["score_distribution"]["research_candidate_count"] >= 3
    assert "https://www.centv.cn/p/467520.html" in by_year[2023]["score_distribution"]["candidate_urls"]
    assert (
        "https://gaokao.chsi.com.cn/gkxx/zc/ss/202306/20230625/2293096350.html"
        in by_year[2023]["score_distribution"]["candidate_urls"]
    )
    assert (
        "https://gaokao.chsi.com.cn/news/file.do?method=downFile&id=2293096351&attach=true&hist=false"
        in by_year[2023]["score_distribution"]["candidate_urls"]
    )
    assert (
        "https://epaper.lnd.com.cn/lswbepaper/pc/layout/202306/25/node_A02.html"
        in by_year[2023]["score_distribution"]["candidate_urls"]
    )
    assert (
        "https://epaper.lnd.com.cn/lswbepaper/pad/con/202306/25/content_195929.html"
        in by_year[2023]["score_distribution"]["candidate_urls"]
    )
    assert (
        "https://epaper.lnd.com.cn/lswbepaper/pad/pic/202306/25/783c9b12-c9ca-406f-908b-7c427eecb636.jpg.2"
        in by_year[2023]["score_distribution"]["candidate_urls"]
    )
    assert (
        "https://epaper.lnd.com.cn/lswbepaper/pc/layout/202406/25/node_A02.html"
        in by_year[2024]["score_distribution"]["candidate_urls"]
    )
    assert (
        "https://gaokao.chsi.com.cn/gkxx/zc/ss/202406/20240625/2293298978.html"
        in by_year[2024]["score_distribution"]["candidate_urls"]
    )
    assert by_year[2022]["projection_status"] == "official_remote_ready"
    assert by_year[2022]["score_distribution"]["official_grid_image_count"] == 2
    assert by_year[2022]["derivation_status"] == "official_image_derivable"
    assert report["summary"]["derivable_years"] == [2022, 2023, 2024, 2025]
    assert report["summary"]["publication_ready_years"] == [2022, 2025]
    assert report["summary"]["blocked_or_review_years"] == [2023, 2024]


def test_audit_data_update_policy_has_no_errors():
    report = audit_data_update_policy()
    assert report["errors"] == []
    assert report["status"] == "ok"
    assert report["policy_count"] >= 18


def test_build_entity_normalization_registry_package(tmp_path: Path):
    region_profile = tmp_path / "region_profile.csv"
    with region_profile.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "adcode",
                "region_name",
                "region_level",
                "parent_adcode",
                "province",
                "city",
                "district",
                "citycode",
                "center_longitude",
                "center_latitude",
                "coordinate_system",
                "source_provider",
                "source_date",
                "availability_date",
                "built_at",
            ],
        )
        writer.writeheader()
        writer.writerow({
            "adcode": "210100",
            "region_name": "沈阳市",
            "region_level": "city",
            "parent_adcode": "210000",
            "province": "辽宁省",
            "city": "沈阳市",
            "district": "",
            "citycode": "024",
            "center_longitude": "123.4",
            "center_latitude": "41.8",
            "coordinate_system": "GCJ-02",
            "source_provider": "fixture",
            "source_date": "2026-05-13",
            "availability_date": "2026-05-13",
            "built_at": "now",
        })
    school_profile = tmp_path / "school_profile.csv"
    with school_profile.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "national_school_code",
                "school_name",
                "province",
                "city",
                "school_tier",
                "source_date",
                "availability_date",
                "built_at",
            ],
        )
        writer.writeheader()
        writer.writerow({
            "national_school_code": "4121010145",
            "school_name": "东北大学",
            "province": "辽宁",
            "city": "沈阳",
            "school_tier": "985",
            "source_date": "2025-06-24",
            "availability_date": "2026-05-13",
            "built_at": "now",
        })
    major_catalog = tmp_path / "major_catalog.csv"
    with major_catalog.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["major_code", "major_name", "major_category", "major_class", "degree_type", "study_years"],
        )
        writer.writeheader()
        writer.writerow({
            "major_code": "080901",
            "major_name": "计算机科学与技术",
            "major_category": "工学",
            "major_class": "计算机类",
            "degree_type": "工学",
            "study_years": "4",
        })
    career_occupation = tmp_path / "career_occupation.csv"
    with career_occupation.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "occupation_code",
                "occupation_name",
                "occupation_family",
                "occupation_level",
                "tdx_l2",
                "tdx_l2_name",
                "source_title",
                "source_url",
                "source_date",
                "availability_date",
                "built_at",
            ],
        )
        writer.writeheader()
        writer.writerow({
            "occupation_code": "2-02-10-03",
            "occupation_name": "软件工程师",
            "occupation_family": "专业技术人员",
            "occupation_level": "4",
            "tdx_l2": "T1205",
            "tdx_l2_name": "软件服务",
            "source_title": "国家职业分类大典数字职业",
            "source_url": "https://example.com/occupation",
            "source_date": "2022-10-28",
            "availability_date": "2026-05-13",
            "built_at": "now",
        })
    policy_industry = tmp_path / "policy_industry.csv"
    with policy_industry.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "tdx_l2",
                "tdx_l2_name",
                "tdx_l1_name",
                "policy_label",
                "policy_intensity",
                "key_themes_json",
                "rationale",
                "policy_period",
                "source_date",
                "availability_date",
                "built_at",
            ],
        )
        writer.writeheader()
        writer.writerow({
            "tdx_l2": "T1205",
            "tdx_l2_name": "软件服务",
            "tdx_l1_name": "信息产业",
            "policy_label": "重点扶持",
            "policy_intensity": "3",
            "key_themes_json": "[]",
            "rationale": "fixture",
            "policy_period": "十五五",
            "source_date": "2026-05-13",
            "availability_date": "2026-05-13",
            "built_at": "now",
        })

    result = build_entity_normalization_registry_package(
        output_root=tmp_path / "exports",
        region_profile_input=region_profile,
        school_profile_input=school_profile,
        major_catalog_input=major_catalog,
        career_occupation_input=career_occupation,
        policy_industry_input=policy_industry,
        package_id="pkg-entity-normalization-test",
        source_version="fixture-entity-normalization",
    )

    package_dir = Path(result["package_dir"])
    assert validate_manifest(package_dir / "manifest.json")["errors"] == []
    assert result["quality_report"]["errors"] == []
    with (package_dir / "fa_dim_entity_registry.csv").open(encoding="utf-8", newline="") as f:
        entities = list(csv.DictReader(f))
    by_id = {row["entity_id"]: row for row in entities}
    assert by_id["geo_city_210100"]["display_name"] == "沈阳"
    assert by_id["school_4121010145"]["display_name"] == "东北大学"
    assert by_id["major_080901"]["display_name"] == "计算机科学与技术"
    assert by_id["occupation_2_02_10_03"]["display_name"] == "软件工程师"
    assert by_id["tdx_l2_T1205"]["display_name"] == "软件服务"

    with (package_dir / "fa_bridge_entity_alias.csv").open(encoding="utf-8", newline="") as f:
        aliases = list(csv.DictReader(f))
    assert any(row["alias_name"] == "沈阳市" and row["canonical_name"] == "沈阳" for row in aliases)

    with (package_dir / "fa_dim_metric_registry.csv").open(encoding="utf-8", newline="") as f:
        metrics = list(csv.DictReader(f))
    metric_keys = {(row["metric_domain"], row["metric_key"]) for row in metrics}
    assert ("city_context.economic", "gdp") in metric_keys
    assert any(row["metric_key"] == "salary_median" for row in metrics)


def test_build_city_listed_company_signal_package(tmp_path: Path):
    company_input = tmp_path / "company_city.csv"
    with company_input.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "stock_code",
                "company_name",
                "hq_province",
                "hq_city",
                "tdx_l2",
                "tdx_l2_name",
                "revenue_proxy",
            ],
        )
        writer.writeheader()
        for row in [
            ("600001", "沈阳软件A", "辽宁", "沈阳市", "T1205", "软件服务", 1200000000),
            ("600002", "沈阳装备B", "辽宁", "沈阳", "T1102", "专用设备", 3000000000),
            ("600003", "沈阳装备C", "辽宁", "沈阳", "T1102", "专用设备", 1800000000),
            ("600004", "大连软件D", "辽宁", "大连", "T1205", "软件服务", 900000000),
            ("", "缺少代码", "辽宁", "沈阳", "T1205", "软件服务", 500000000),
        ]:
            writer.writerow({
                "stock_code": row[0],
                "company_name": row[1],
                "hq_province": row[2],
                "hq_city": row[3],
                "tdx_l2": row[4],
                "tdx_l2_name": row[5],
                "revenue_proxy": row[6],
            })

    result = build_city_listed_company_signal_package(
        company_input=company_input,
        output_root=tmp_path / "exports",
        package_id="pkg-city-listed-company-signal-test",
        source_version="fixture-city-listed",
        metric_year=2025,
        source_date="2026-05-13",
        availability_date="2026-05-13",
        source_system="fixture_company_city",
    )

    package_dir = Path(result["package_dir"])
    assert validate_manifest(package_dir / "manifest.json")["errors"] == []
    with (package_dir / "fa_fact_city_listed_company_signal.csv").open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    by_city_metric = {(row["city"], row["metric_key"]): row for row in rows}
    assert int(float(by_city_metric[("沈阳", "listed_company_count")]["metric_value"])) == 3
    assert int(float(by_city_metric[("沈阳", "listed_company_tdx_l2_count")]["metric_value"])) == 2
    assert int(float(by_city_metric[("沈阳", "listed_company_revenue_proxy")]["metric_value"])) == 6000000000
    assert by_city_metric[("沈阳", "listed_company_count")]["source_scope"] == "city_total"
    assert result["quality_report"]["errors"] == []


def test_build_and_audit_city_context_collection_plan(tmp_path: Path):
    city_input = tmp_path / "cities.csv"
    with city_input.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["adcode", "province", "city", "region_level", "priority_rank"])
        writer.writeheader()
        writer.writerow({"adcode": "210100", "province": "辽宁", "city": "沈阳", "region_level": "city", "priority_rank": "1"})
        writer.writerow({"adcode": "210200", "province": "辽宁", "city": "大连", "region_level": "city", "priority_rank": "2"})

    result = build_city_context_collection_plan(
        city_input=city_input,
        output_dir=tmp_path / "city_context",
        domains=["economic", "public_resource"],
        metric_year=2025,
        limit=1,
    )

    assert result["rows"] == 13
    with Path(result["csv"]).open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    assert {row["domain"] for row in rows} == {"economic", "public_resource"}
    assert all(json.loads(row["search_queries"]) for row in rows)
    assert next(row for row in rows if row["metric_key"] == "hospital_beds_per_1000")["resource_domain"] == "medical"

    audit = audit_city_context_collection_plan(Path(result["csv"]))
    assert audit["errors"] == []
    assert audit["progress"]["pending_rows"] == 13

    batch = build_city_context_review_batch(
        plan_csv=Path(result["csv"]),
        output_dir=tmp_path / "city_context_batch",
        domains=["economic"],
        limit_per_domain=2,
    )
    assert batch["rows"] == 2
    with Path(batch["csv"]).open(encoding="utf-8", newline="") as f:
        batch_rows = list(csv.DictReader(f))
    batch_rows[0]["status"] = "verified"
    batch_rows[0]["metric_value"] = "9000"
    batch_rows[0]["source_title"] = "fixture bulletin"
    batch_rows[0]["source_url"] = "https://example.com/shenyang-gdp"
    batch_rows[0]["evidence_quote"] = "地区生产总值9000亿元"
    batch_rows[0]["source_date"] = "2026-05-13"
    batch_rows[0]["availability_date"] = "2026-05-13"
    edited_batch = tmp_path / "city_context_batch_edited.csv"
    with edited_batch.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=batch_rows[0].keys())
        writer.writeheader()
        writer.writerows(batch_rows)
    merge = merge_city_context_review_batch(
        plan_csv=Path(result["csv"]),
        batch_csv=edited_batch,
        output=tmp_path / "city_context_merged.csv",
    )
    assert merge["updated_rows"] == 1
    merged_audit = audit_city_context_collection_plan(Path(merge["output"]))
    assert merged_audit["progress"]["complete_rows"] == 1
    assert merged_audit["errors"] == []
    package_result = build_city_context_packages_from_collection_plan(
        plan_csv=Path(merge["output"]),
        output_root=tmp_path / "exports",
        domains=["economic"],
        package_id="pkg-city-context-{domain}",
        source_version="fixture-city-context",
    )
    assert len(package_result["packages"]) == 1
    package_dir = Path(package_result["packages"][0]["package_dir"])
    assert validate_manifest(package_dir / "manifest.json")["errors"] == []
    with (package_dir / "fa_fact_city_economic_indicator.csv").open(encoding="utf-8", newline="") as f:
        package_rows = list(csv.DictReader(f))
    assert package_rows[0]["metric_key"] == batch_rows[0]["metric_key"]
    assert package_rows[0]["city"] == "沈阳"

    reviewed = tmp_path / "city_context_reviewed.csv"
    rows[0]["status"] = "verified"
    rows[0]["metric_value"] = "9000"
    rows[0]["source_title"] = "fixture bulletin"
    rows[0]["source_url"] = "https://example.com/shenyang-gdp"
    rows[0]["evidence_quote"] = "地区生产总值9000亿元"
    rows[0]["source_date"] = "2026-05-13"
    rows[0]["availability_date"] = "2026-05-13"
    with reviewed.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)
    reviewed_audit = audit_city_context_collection_plan(reviewed)
    assert reviewed_audit["progress"]["complete_rows"] == 1
    assert reviewed_audit["errors"] == []

    rows[1]["status"] = "verified"
    rows[1]["metric_value"] = "120"
    bad = tmp_path / "city_context_bad.csv"
    with bad.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)
    bad_audit = audit_city_context_collection_plan(bad)
    assert any("complete status missing evidence" in error for error in bad_audit["errors"])


def test_audit_city_context_collection_plan_validates_source_metadata(tmp_path: Path):
    city_input = tmp_path / "cities.csv"
    with city_input.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["adcode", "province", "city", "region_level", "priority_rank"])
        writer.writeheader()
        writer.writerow({"adcode": "210100", "province": "辽宁", "city": "沈阳", "region_level": "city", "priority_rank": "1"})

    result = build_city_context_collection_plan(
        city_input=city_input,
        output_dir=tmp_path / "city_context",
        domains=["economic"],
        metric_year=2025,
    )
    with Path(result["csv"]).open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    rows[0]["status"] = "verified"
    rows[0]["metric_year"] = "2025.5"
    rows[0]["metric_value"] = "9000"
    rows[0]["source_title"] = "fixture bulletin"
    rows[0]["source_url"] = "ftp://example.com/shenyang-gdp"
    rows[0]["evidence_quote"] = "地区生产总值9000亿元"
    rows[0]["source_date"] = "2026-05-14"
    rows[0]["availability_date"] = "2026-05-13"
    rows[0]["reviewed_at"] = "2026/05/14"
    reviewed = tmp_path / "city_context_bad_metadata.csv"
    with reviewed.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)

    audit = audit_city_context_collection_plan(reviewed)

    assert "row 1 metric_year is not an integer" in audit["errors"]
    assert "row 1 source_url must be an http(s) URL" in audit["errors"]
    assert "row 1 reviewed_at must use YYYY-MM-DD" in audit["errors"]
    assert "row 1 source_date must not be after availability_date" in audit["errors"]


def test_build_city_ranking_collection_plan_and_package(tmp_path: Path):
    city_input = tmp_path / "cities.csv"
    with city_input.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["adcode", "province", "city", "region_level", "priority_rank"])
        writer.writeheader()
        writer.writerow({"adcode": "210100", "province": "辽宁", "city": "沈阳", "region_level": "city", "priority_rank": "1"})

    result = build_city_context_collection_plan(
        city_input=city_input,
        output_dir=tmp_path / "city_ranking",
        domains=["city_ranking"],
        metric_year=2025,
    )

    with Path(result["csv"]).open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    assert result["rows"] == 5
    assert {row["domain"] for row in rows} == {"city_ranking"}
    assert {row["dimension_key"] for row in rows} >= {"commercial_lifestyle_vitality", "research_output"}

    rows[0]["status"] = "verified"
    rows[0]["rank_value"] = "12"
    rows[0]["tier_label"] = "二线城市"
    rows[0]["source_title"] = "fixture city ranking"
    rows[0]["source_url"] = "https://example.com/city-ranking"
    rows[0]["evidence_quote"] = "沈阳位列二线城市"
    rows[0]["metric_scope"] = "中国内地城市"
    rows[0]["source_date"] = "2026-05-13"
    rows[0]["availability_date"] = "2026-05-13"
    reviewed = tmp_path / "city_ranking_reviewed.csv"
    with reviewed.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)

    audit = audit_city_context_collection_plan(reviewed)
    assert audit["errors"] == []
    assert audit["progress"]["complete_rows"] == 1

    package_result = build_city_context_packages_from_collection_plan(
        plan_csv=reviewed,
        output_root=tmp_path / "exports",
        domains=["city_ranking"],
        package_id="pkg-city-ranking",
        source_version="fixture-city-ranking",
    )
    package_dir = Path(package_result["packages"][0]["package_dir"])
    assert validate_manifest(package_dir / "manifest.json")["errors"] == []
    with (package_dir / "fa_fact_city_ranking_signal.csv").open(encoding="utf-8", newline="") as f:
        package_rows = list(csv.DictReader(f))
    assert package_rows[0]["city"] == "沈阳"
    assert package_rows[0]["ranking_source_key"] == rows[0]["ranking_source_key"]
    assert package_rows[0]["dimension_key"] == rows[0]["dimension_key"]
    assert package_rows[0]["rank_value"] == "12"
    assert package_rows[0]["tier_label"] == "二线城市"

    rows[0]["rank_value"] = ""
    rows[0]["score_value"] = ""
    rows[0]["tier_label"] = ""
    missing_value = tmp_path / "city_ranking_missing_value.csv"
    with missing_value.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)
    missing_value_audit = audit_city_context_collection_plan(missing_value)
    assert any("complete status missing value" in error for error in missing_value_audit["errors"])


def test_build_city_context_target_cities_from_core(tmp_path: Path):
    core_db = tmp_path / "core.duckdb"
    con = duckdb.connect(str(core_db))
    con.execute("""
        CREATE TABLE fa_dim_ln_admission_plan (
            school_code VARCHAR,
            school_name VARCHAR,
            major_code VARCHAR,
            major_full VARCHAR,
            batch VARCHAR,
            subject_cat VARCHAR,
            region VARCHAR,
            plan_count INTEGER
        )
    """)
    con.executemany(
        "INSERT INTO fa_dim_ln_admission_plan VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
        [
            ("10145", "东北大学", "01", "计算机类", "本科批", "物理类", "辽宁沈阳", 10),
            ("10145", "东北大学", "02", "自动化类", "本科批", "物理类", "辽宁沈阳", 8),
            ("10141", "大连理工大学", "01", "软件工程", "本科批", "物理类", "辽宁大连", 12),
            ("10001", "北京大学", "01", "法学", "本科批", "历史类", "北京", 2),
            ("10200", "东北师范大学", "01", "汉语言文学", "专科批", "历史类", "吉林长春", 3),
        ],
    )
    con.execute("""
        CREATE TABLE fa_dim_region_profile (
            adcode VARCHAR,
            region_name VARCHAR,
            region_level VARCHAR,
            parent_adcode VARCHAR,
            province VARCHAR,
            city VARCHAR,
            district VARCHAR,
            citycode VARCHAR,
            center_longitude DOUBLE,
            center_latitude DOUBLE,
            coordinate_system VARCHAR,
            source_provider VARCHAR,
            source_date VARCHAR,
            availability_date VARCHAR,
            built_at VARCHAR
        )
    """)
    con.executemany(
        """
        INSERT INTO fa_dim_region_profile VALUES
        (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        [
            ("210100", "沈阳市", "city", "210000", "辽宁省", "沈阳市", "", "024", 123.4, 41.8, "GCJ-02", "fixture", "2026-05-13", "2026-05-13", "now"),
            ("210200", "大连市", "city", "210000", "辽宁省", "大连市", "", "0411", 121.6, 38.9, "GCJ-02", "fixture", "2026-05-13", "2026-05-13", "now"),
        ],
    )
    con.close()

    result = build_city_context_target_cities(
        core_db=core_db,
        output_dir=tmp_path / "city_context",
    )

    assert result["rows"] == 3
    assert result["ready_rows"] == 2
    with Path(result["ready_csv"]).open(encoding="utf-8", newline="") as f:
        ready_rows = list(csv.DictReader(f))
    assert [row["city"] for row in ready_rows] == ["沈阳", "大连"]
    assert {row["adcode"] for row in ready_rows} == {"210100", "210200"}
    assert ready_rows[0]["priority_rank"] == "1"

    with Path(result["review_csv"]).open(encoding="utf-8", newline="") as f:
        review_rows = list(csv.DictReader(f))
    blocked = [row for row in review_rows if row["match_status"] == "blocked"]
    assert len(blocked) == 1
    assert blocked[0]["city"] == "北京"
    assert blocked[0]["blocking_reason"] == "missing_region_profile_adcode"

    plan_result = build_city_context_collection_plan(
        city_input=Path(result["ready_csv"]),
        output_dir=tmp_path / "city_context_plan",
        domains=["city_ranking"],
        metric_year=2025,
    )
    assert plan_result["rows"] == 10


def test_build_major_city_employment_fit_package(tmp_path: Path):
    role_input = tmp_path / "major_roles.csv"
    with role_input.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "major_code",
                "major_name",
                "major_class",
                "role_key",
                "role_name",
                "role_family",
                "role_type",
                "occupation_code",
                "occupation_name",
                "tdx_l2",
                "tdx_l2_name",
                "public_sector_fit",
                "private_sector_fit",
                "listed_company_fit",
                "confidence",
                "rationale",
                "source_title",
                "source_url",
                "source_date",
                "availability_date",
                "built_at",
            ],
        )
        writer.writeheader()
        base = {
            "major_code": "120203K",
            "major_name": "会计学",
            "major_class": "工商管理类",
            "tdx_l2": "T1001",
            "tdx_l2_name": "银行",
            "source_title": "fixture role map",
            "source_url": "https://example.com/role-map",
            "source_date": "2026-05-13",
            "availability_date": "2026-05-13",
            "built_at": "2026-05-13T00:00:00",
        }
        for row in [
            ("accountant", "会计", "财务", "direct", "2-06-03-00", "会计专业人员", 70, 85, 78, "high"),
            ("hr_generalist", "人力资源专员", "组织职能", "generalist", "", "", 45, 72, 62, "medium"),
            ("public_finance", "财政财务岗位", "公共部门", "public_sector", "", "", 82, 45, 35, "medium"),
        ]:
            writer.writerow({
                **base,
                "role_key": row[0],
                "role_name": row[1],
                "role_family": row[2],
                "role_type": row[3],
                "occupation_code": row[4],
                "occupation_name": row[5],
                "public_sector_fit": row[6],
                "private_sector_fit": row[7],
                "listed_company_fit": row[8],
                "confidence": row[9],
                "rationale": "fixture",
            })

    demand_input = tmp_path / "company_role_demand.csv"
    with demand_input.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "company_id",
                "stock_code",
                "company_name",
                "listed_company_flag",
                "province",
                "city",
                "tdx_l2",
                "tdx_l2_name",
                "role_key",
                "role_name",
                "role_family",
                "metric_key",
                "metric_name",
                "metric_value",
                "metric_unit",
                "metric_year",
                "metric_scope",
                "source_title",
                "source_url",
                "evidence_quote",
                "source_date",
                "availability_date",
                "built_at",
            ],
        )
        writer.writeheader()
        base = {
            "province": "辽宁",
            "city": "沈阳",
            "tdx_l2": "T1001",
            "tdx_l2_name": "银行",
            "metric_year": "2026",
            "metric_scope": "公开样本",
            "source_title": "fixture demand",
            "source_date": "2026-05-13",
            "availability_date": "2026-05-13",
            "built_at": "2026-05-13T00:00:00",
        }
        rows = [
            ("bank-a", "600001", "沈阳银行A", "true", "accountant", "会计", "财务", "job_posting_count", "岗位数量", 180, "count"),
            ("bank-a", "600001", "沈阳银行A", "true", "accountant", "会计", "财务", "internship_post_count", "实习数量", 32, "count"),
            ("industry-b", "600002", "沈阳制造B", "true", "hr_generalist", "人力资源专员", "组织职能", "job_posting_count", "岗位数量", 42, "count"),
            ("public-c", "", "公共部门C", "false", "public_finance", "财政财务岗位", "公共部门", "public_sector_post_count", "公共部门岗位", 24, "count"),
        ]
        for company_id, stock_code, company_name, listed, role_key, role_name, family, metric_key, metric_name, value, unit in rows:
            writer.writerow({
                **base,
                "company_id": company_id,
                "stock_code": stock_code,
                "company_name": company_name,
                "listed_company_flag": listed,
                "role_key": role_key,
                "role_name": role_name,
                "role_family": family,
                "metric_key": metric_key,
                "metric_name": metric_name,
                "metric_value": value,
                "metric_unit": unit,
                "source_url": f"https://example.com/{company_id}/{metric_key}",
                "evidence_quote": f"{metric_name}{value}",
            })

    result = build_major_city_employment_fit_package(
        role_input=role_input,
        demand_input=demand_input,
        output_root=tmp_path / "exports",
        package_id="pkg-major-city-employment-fit-test",
        source_version="fixture-major-city-fit",
    )

    package_dir = Path(result["package_dir"])
    assert validate_manifest(package_dir / "manifest.json")["errors"] == []
    assert result["rows"] == 1
    with (package_dir / "fa_mart_major_city_employment_fit.csv").open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    row = rows[0]
    assert row["major_code"] == "120203K"
    assert row["city"] == "沈阳"
    assert row["score_profile"] == "major_city_employment_default"
    assert row["primary_role_key"] == "accountant"
    assert int(float(row["role_coverage_count"])) == 3
    assert int(float(row["listed_company_count"])) == 2
    assert float(row["overall_score"]) > 0
    role_mix = json.loads(row["role_mix_json"])
    assert "direct" in role_mix
    assert "generalist" in role_mix
    contributions = json.loads(row["signal_contribution_json"])
    assert contributions["listed_company_count"] == 2
    lineage = json.loads(row["pit_lineage_json"])
    assert "fa_fact_company_role_demand_signal" in lineage["tables"]
    assert result["quality_report"]["input_quality"]["errors"] == []


def test_build_major_city_employment_fit_rejects_bad_input_metadata(tmp_path: Path):
    role_input = tmp_path / "major_roles.csv"
    with role_input.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "major_code",
                "major_name",
                "role_key",
                "role_name",
                "role_type",
                "public_sector_fit",
                "private_sector_fit",
                "listed_company_fit",
                "confidence",
                "source_url",
                "source_date",
                "availability_date",
                "built_at",
            ],
        )
        writer.writeheader()
        writer.writerow({
            "major_code": "120203K",
            "major_name": "会计学",
            "role_key": "accountant",
            "role_name": "会计",
            "role_type": "direct",
            "public_sector_fit": "70",
            "private_sector_fit": "120",
            "listed_company_fit": "80",
            "confidence": "high",
            "source_url": "https://example.com/role-map",
            "source_date": "2026-05-13",
            "availability_date": "2026-05-13",
            "built_at": "2026-05-13T00:00:00",
        })

    demand_input = tmp_path / "company_role_demand.csv"
    with demand_input.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "company_id",
                "company_name",
                "listed_company_flag",
                "province",
                "city",
                "role_key",
                "role_name",
                "role_family",
                "metric_key",
                "metric_name",
                "metric_value",
                "metric_unit",
                "metric_year",
                "metric_scope",
                "source_title",
                "source_url",
                "evidence_quote",
                "source_date",
                "availability_date",
                "built_at",
            ],
        )
        writer.writeheader()
        writer.writerow({
            "company_id": "bank-a",
            "company_name": "沈阳银行A",
            "listed_company_flag": "true",
            "province": "辽宁",
            "city": "沈阳",
            "role_key": "accountant",
            "role_name": "会计",
            "role_family": "财务",
            "metric_key": "unsupported_metric",
            "metric_name": "未知指标",
            "metric_value": "12",
            "metric_unit": "count",
            "metric_year": "2026.5",
            "metric_scope": "公开样本",
            "source_title": "fixture demand",
            "source_url": "ftp://example.com/bad-demand",
            "evidence_quote": "未知指标12",
            "source_date": "2026-05-14",
            "availability_date": "2026-05-13",
            "built_at": "2026-05-13T00:00:00",
        })

    with pytest.raises(ValueError) as exc:
        build_major_city_employment_fit_package(
            role_input=role_input,
            demand_input=demand_input,
            output_root=tmp_path / "exports",
            package_id="pkg-major-city-employment-fit-bad-input",
            source_version="fixture-major-city-fit",
        )

    message = str(exc.value)
    assert "role row 1 private_sector_fit outside 0-100: 120" in message
    assert "demand row 1 source_url must be an http(s) URL" in message
    assert "demand row 1 source_date must not be after availability_date" in message
    assert "demand row 1 unregistered metric_key: unsupported_metric" in message
    assert "demand row 1 metric_year is not an integer" in message


def test_build_outcome_collection_plan_from_core_admission_plan(tmp_path: Path):
    db = tmp_path / "core.duckdb"
    con = duckdb.connect(str(db))
    try:
        con.execute("""
            CREATE TABLE fa_dim_ln_admission_plan (
                school_code VARCHAR,
                school_name VARCHAR,
                major_full VARCHAR,
                batch VARCHAR,
                subject_cat VARCHAR
            )
        """)
        con.execute("""
            INSERT INTO fa_dim_ln_admission_plan VALUES
                ('0140', '辽宁大学', '法学', '本科批', '历史类'),
                ('0140', '辽宁大学', '汉语言文学', '本科批', '历史类'),
                ('0183', '吉林大学', '计算机类', '本科批', '物理类'),
                ('0183', '吉林大学', '计算机类', '本科批', '物理类'),
                ('0300', '东北大学', '自动化', '本科批', '物理类'),
                ('0177', '沈阳音乐学院', '音乐表演', '艺术类本科批', '历史类')
        """)
    finally:
        con.close()

    result = build_outcome_collection_plan(
        core_db=db,
        output_dir=tmp_path / "collection",
        domains=["school", "major"],
        school_limit=2,
        major_limit=2,
        metric_year=2024,
    )

    assert result["rows"] == 16
    with Path(result["csv"]).open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    assert rows[0]["domain"] == "school"
    assert rows[0]["metric_key"] == "postgrad_rate"
    assert "就业质量报告" in rows[0]["search_queries"]
    assert "metric_value" in rows[0]
    assert "source_url" in rows[0]
    assert {row["metric_year"] for row in rows} == {"2024"}
    assert any(row["domain"] == "major" and row["entity_name"] == "计算机类" for row in rows)

    manifest = json.loads(Path(result["manifest"]).read_text(encoding="utf-8"))
    assert manifest["notes"].startswith("Collection plan only")
    assert manifest["rows"] == 16
    assert manifest["metric_year"] == 2024


def test_build_outcome_collection_plan_can_skip_covered_school_outcomes(tmp_path: Path):
    db = tmp_path / "core.duckdb"
    con = duckdb.connect(str(db))
    try:
        con.execute("""
            CREATE TABLE fa_dim_ln_admission_plan (
                school_code VARCHAR,
                school_name VARCHAR,
                major_full VARCHAR,
                batch VARCHAR,
                subject_cat VARCHAR
            )
        """)
        con.execute("""
            CREATE TABLE fa_fact_school_outcome (
                school_code VARCHAR,
                metric_key VARCHAR,
                metric_year INTEGER
            )
        """)
        con.execute("""
            INSERT INTO fa_dim_ln_admission_plan VALUES
                ('0140', '辽宁大学', '法学', '本科批', '历史类'),
                ('0140', '辽宁大学', '汉语言文学', '本科批', '历史类'),
                ('0183', '吉林大学', '计算机类', '本科批', '物理类'),
                ('0300', '东北大学', '自动化', '本科批', '物理类')
        """)
        con.execute("""
            INSERT INTO fa_fact_school_outcome VALUES
                ('0140', 'employment_rate', 2024),
                ('0300', 'employment_rate', 2023)
        """)
    finally:
        con.close()

    result = build_outcome_collection_plan(
        core_db=db,
        output_dir=tmp_path / "collection_missing",
        domains=["school"],
        school_limit=10,
        metric_year=2024,
        missing_school_outcome_only=True,
    )

    with Path(result["csv"]).open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    entity_codes = {row["entity_code"] for row in rows}
    assert "0140" not in entity_codes
    assert entity_codes == {"0183", "0300"}
    manifest = json.loads(Path(result["manifest"]).read_text(encoding="utf-8"))
    assert manifest["missing_school_outcome_only"] is True
    assert manifest["coverage_year"] == 2024


def test_build_outcome_collection_plan_includes_seeded_missing_schools_beyond_limit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
):
    db = tmp_path / "core.duckdb"
    con = duckdb.connect(str(db))
    try:
        con.execute("""
            CREATE TABLE fa_dim_ln_admission_plan (
                school_code VARCHAR,
                school_name VARCHAR,
                major_full VARCHAR,
                batch VARCHAR,
                subject_cat VARCHAR
            )
        """)
        con.execute("""
            CREATE TABLE fa_fact_school_outcome (
                school_code VARCHAR,
                metric_key VARCHAR,
                metric_year INTEGER
            )
        """)
        con.execute("""
            INSERT INTO fa_dim_ln_admission_plan VALUES
                ('0001', '高频学校', '法学', '本科批', '历史类'),
                ('0001', '高频学校', '汉语言文学', '本科批', '历史类'),
                ('0002', '中频学校', '计算机类', '本科批', '物理类'),
                ('0177', '沈阳音乐学院', '音乐表演', '本科批', '历史类')
        """)
    finally:
        con.close()

    monkeypatch.setattr(
        "datahub.builders.outcome_collection_plan.load_outcome_report_sources",
        lambda: {
            "seeds": [
                {
                    "domain": "school",
                    "entity_code": "0177",
                    "entity_name": "沈阳音乐学院",
                    "metric_year": 2024,
                    "report_scope": "undergraduate_teaching_quality_report",
                },
                {
                    "domain": "school",
                    "entity_code": "0002",
                    "entity_name": "中频学校",
                    "metric_year": 2024,
                    "report_scope": "undergraduate_teaching_quality_report",
                    "seed_status": "rejected",
                },
            ]
        },
    )

    result = build_outcome_collection_plan(
        core_db=db,
        output_dir=tmp_path / "collection_seeded_missing",
        domains=["school"],
        school_limit=1,
        metric_year=2024,
        missing_school_outcome_only=True,
    )

    with Path(result["csv"]).open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    entity_codes = {row["entity_code"] for row in rows}
    assert entity_codes == {"0001", "0177"}
    assert {row["priority_rank"] for row in rows if row["entity_code"] == "0177"} == {"3"}


def test_build_outcome_collection_plan_collapses_duplicate_school_codes(tmp_path: Path):
    db = tmp_path / "core.duckdb"
    con = duckdb.connect(str(db))
    try:
        con.execute("""
            CREATE TABLE fa_dim_ln_admission_plan (
                school_code VARCHAR,
                school_name VARCHAR,
                major_full VARCHAR,
                batch VARCHAR,
                subject_cat VARCHAR
            )
        """)
        con.execute("""
            INSERT INTO fa_dim_ln_admission_plan VALUES
                ('6407', '香港中文大学(深圳)', '计算机类', '本科批', '物理类'),
                ('6407', '香港中文大学(深圳)', '金融学', '本科批', '物理类'),
                ('6407', '香港中文大学', '经济学', '本科批', '历史类'),
                ('0140', '辽宁大学', '法学', '本科批', '历史类')
        """)
    finally:
        con.close()

    result = build_outcome_collection_plan(
        core_db=db,
        output_dir=tmp_path / "collection_duplicate_school_codes",
        domains=["school"],
        school_limit=10,
        metric_year=2024,
    )

    with Path(result["csv"]).open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    keys = [(row["domain"], row["entity_code"], row["metric_key"], row["metric_year"]) for row in rows]
    assert len(keys) == len(set(keys))
    assert {row["entity_code"] for row in rows} == {"6407", "0140"}
    assert {row["entity_name"] for row in rows if row["entity_code"] == "6407"} == {"香港中文大学(深圳)"}
    assert {row["plan_rows"] for row in rows if row["entity_code"] == "6407"} == {"3"}


def test_extract_outcome_report_candidates_from_lines(tmp_path: Path):
    rows = extract_outcome_metric_candidates_from_lines(
        [
            (3, "学校本科毕业生毕业去向落实率为 92.36%，其中继续深造比例为 24.18%。"),
            (8, "推荐免试研究生名额占本科毕业生人数比例约 6.40%。"),
            (9, "本科应届毕业生 4315 人，已就业人数为 3703 人，毕业去向落实率"),
            (9, "为 85.82%。其中，单位就业人数为 1352 人，占比 31.33%；升学人数为 1568 人，"),
            (9, "占比 36.34%；自由职业人数为 770 人，占比 17.84%。"),
            (10, "本科毕业生 4315 人中，攻读研究生 1549 人。其中，推荐免试攻读研究生"),
            (10, "582 人，占比 37.57%；考取本校研究生 150 人，占比 9.68%。"),
            (10, "2024 届本科毕业生考取研究生 876 人(其中推免生 143 人),占本科毕业生总数的 25.47%,"),
            (10, "截至 2024 年 8 月 31 日，学校应届本科毕业生总体就业率达 87.54%。毕业"),
            (10, "生最主要的毕业去向是企业，占 54.95%。升学 1160 人，占 21.70%，其中出国"),
            (10, "升学1443人（包含研究生591人，第二学士学位448人，境外留学404人）。2024届本科毕业生初次毕业去向落实率为88.64%。"),
            (10, "2024届本科毕业生（不含升学、出国、自由职业）在辽就业占比42.1%。"),
            (10, "其中出国境升学 371 人，占比 23.95%。"),
            (11, "2024 届毕业生初次就业率显示：42.1%毕业生在辽宁省就业。"),
            (12, "毕业授位 88 人，学位点就业率 100%。学员主要分布在国有企业单位。"),
            (12, "毕业生职业发展情况良好，毕业生近三年内 47.6%岗位得到晋升。"),
            (13, "师范生/非师范生毕业去向落实率分别为 77.47%、81.70%。"),
        ],
        domain="school",
        entity_code="0142",
        entity_name="沈阳工业大学",
        metric_year=2024,
        source_title="2023-2024年本科教学质量报告",
        source_url="https://www.sut.edu.cn/info/1584/67026.htm",
        source_date="2024-12-31",
        availability_date="2025-01-01",
    )

    assert {row["metric_key"] for row in rows} >= {"employment_rate", "postgrad_rate", "keep_research_rate"}
    assert any(row["candidate_value"] == "0.9236" for row in rows)
    assert any(row["candidate_value"] == "0.8582" for row in rows)
    assert any(row["match_alias"] == "升学人数" and row["candidate_value"] == "0.3634" for row in rows)
    assert any(row["match_alias"] == "考取研究生" and row["candidate_value"] == "0.2547" for row in rows)
    assert any(row["match_alias"] == "升学" and row["candidate_value"] == "0.217" for row in rows)
    assert any(row["match_alias"] == "推荐免试" for row in rows)
    assert not any(row["metric_key"] == "keep_research_rate" and row["candidate_value"] == "0.3757" for row in rows)
    assert not any(row["metric_key"] == "keep_research_rate" and row["candidate_value"] == "0.2547" for row in rows)
    assert not any(row["metric_key"] == "postgrad_rate" and row["candidate_value"] == "0.8864" for row in rows)
    assert not any(row["metric_key"] == "postgrad_rate" and row["candidate_value"] in {"0.421", "0.2395"} for row in rows)
    assert not any(row["metric_key"] == "employment_rate" and row["candidate_value"] in {"0.421", "1"} for row in rows)
    assert not any(row["metric_key"] == "employment_rate" and row["candidate_value"] == "0.7747" for row in rows)
    assert not any(row["metric_key"] == "civil_service_rate" and row["candidate_value"] == "0.476" for row in rows)
    assert all(row["review_status"] == "needs_review" for row in rows)

    output = tmp_path / "outcome_candidates.csv"
    write_outcome_metric_candidate_csv(output, rows)
    with output.open(encoding="utf-8", newline="") as f:
        written = list(csv.DictReader(f))
    assert set(written[0]).issuperset(CANDIDATE_COLUMNS)
    assert written[0]["source_url"] == "https://www.sut.edu.cn/info/1584/67026.htm"


def test_extract_outcome_report_candidates_from_ofd(tmp_path: Path):
    path = tmp_path / "dufe.ofd"
    with ZipFile(path, "w") as zf:
        zf.writestr(
            "Doc_0/Pages/Page_28/Content.xml",
            """<?xml version="1.0" encoding="UTF-8"?>
<ofd:Content xmlns:ofd="http://www.ofdspec.org/2016">
  <ofd:Layer>
    <ofd:TextObject ID="1" Boundary="10 30 60 5">
      <ofd:TextCode>2024届本科毕业生初次毕业去向落实率为</ofd:TextCode>
    </ofd:TextObject>
    <ofd:TextObject ID="2" Boundary="70 30.2 30 5">
      <ofd:TextCode>88.64%。</ofd:TextCode>
    </ofd:TextObject>
  </ofd:Layer>
</ofd:Content>
""",
        )

    rows = extract_outcome_metric_candidates_from_ofd(
        path,
        domain="school",
        entity_code="0173",
        entity_name="东北财经大学",
        metric_year=2024,
        source_title="东北财经大学2023-2024学年本科教学质量报告",
        source_url="https://xxgk.dufe.edu.cn/content_88732.html",
        source_date="2024-12-05",
        availability_date="2024-12-05",
    )

    assert len(rows) == 1
    assert rows[0]["metric_key"] == "employment_rate"
    assert rows[0]["candidate_value"] == "0.8864"
    assert rows[0]["page_number"] == 29
    assert "88.64%" in rows[0]["evidence_quote"]


def test_extract_outcome_report_candidates_from_docx(tmp_path: Path):
    path = tmp_path / "fuxin.docx"
    with ZipFile(path, "w") as zf:
        zf.writestr(
            "word/document.xml",
            """<?xml version="1.0" encoding="UTF-8"?>
<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
  <w:body>
    <w:p><w:r><w:t>2024届毕业生毕业去向落实率为97.80%。</w:t></w:r></w:p>
    <w:p><w:r><w:t>其中升学人数为77人，占比3.94%。</w:t></w:r></w:p>
  </w:body>
</w:document>
""",
        )

    rows = extract_outcome_metric_candidates_from_report(
        path,
        domain="school",
        entity_code="1250",
        entity_name="阜新高等专科学校",
        metric_year=2024,
        source_title="阜新高等专科学校2024年大专毕业生就业质量报告",
        source_url="https://www.fxgz.com.cn/showart/8804.html",
        source_date="2025-12-06",
        availability_date="2025-12-06",
    )

    assert any(row["metric_key"] == "employment_rate" and row["candidate_value"] == "0.978" for row in rows)
    assert any(row["metric_key"] == "postgrad_rate" and row["candidate_value"] == "0.0394" for row in rows)
    assert all(row["page_number"] == 1 for row in rows)


def test_extract_outcome_report_candidates_rejects_html_disguised_as_pdf(tmp_path: Path):
    path = tmp_path / "report.pdf"
    path.write_text("<!DOCTYPE html><html><body>captcha</body></html>", encoding="utf-8")

    with pytest.raises(ValueError, match="HTML, not PDF"):
        extract_outcome_metric_candidates_from_report(
            path,
            domain="school",
            entity_code="0169",
            entity_name="大连外国语大学",
            metric_year=2024,
            source_title="大连外国语大学2023-2024学年本科教学质量报告",
            source_url="https://xxgk.dlufl.edu.cn/info/1004/2001.htm",
            source_date="2024-12-31",
            availability_date="2024-12-31",
        )


def test_cli_extract_outcome_report_candidates_reports_bad_pdf(tmp_path: Path, monkeypatch, capsys):
    from datahub import cli

    bad_pdf = tmp_path / "report.pdf"
    bad_pdf.write_text("<!DOCTYPE html><html><body>captcha</body></html>", encoding="utf-8")
    output = tmp_path / "candidates.csv"

    monkeypatch.setattr("sys.argv", [
        "lifehack-datahub",
        "extract-outcome-report-candidates",
        "--input",
        str(bad_pdf),
        "--output",
        str(output),
        "--domain",
        "school",
        "--entity-code",
        "0169",
        "--entity-name",
        "大连外国语大学",
        "--metric-year",
        "2024",
        "--source-title",
        "大连外国语大学2023-2024学年本科教学质量报告",
        "--source-url",
        "https://xxgk.dlufl.edu.cn/info/1004/2001.htm",
        "--source-date",
        "2024-12-31",
        "--availability-date",
        "2024-12-31",
    ])

    assert cli.main() == 1
    captured = capsys.readouterr()
    report = json.loads(captured.out)
    assert report["rows"] == 0
    assert any("HTML, not PDF" in error for error in report["errors"])
    assert not output.exists()


def test_build_outcome_report_source_plan_groups_metric_tasks(tmp_path: Path):
    plan = tmp_path / "outcome_collection_plan.csv"
    rows = [
        _outcome_plan_row("school", "10140", "辽宁大学", "postgrad_rate", status="todo", priority_rank="1"),
        _outcome_plan_row("school", "10140", "辽宁大学", "employment_rate", status="todo", priority_rank="1"),
        _outcome_plan_row("major", "法学", "法学", "employment_rate", status="todo", priority_rank="2"),
    ]
    _write_outcome_plan(plan, rows)

    result = build_outcome_report_source_plan(
        plan_csv=plan,
        output_dir=tmp_path / "report_sources",
        domains=["school"],
    )

    assert result["rows"] == 3
    with Path(result["csv"]).open(encoding="utf-8", newline="") as f:
        source_rows = list(csv.DictReader(f))
    assert {row["report_scope"] for row in source_rows} == {
        "employment_quality_report",
        "higher_vocational_quality_report",
        "undergraduate_teaching_quality_report",
    }
    assert source_rows[0]["entity_name"] == "辽宁大学"
    assert json.loads(source_rows[0]["planned_metric_keys"]) == ["employment_rate", "postgrad_rate"]
    assert "辽宁大学 2025" in json.loads(source_rows[0]["search_queries"])[0]
    assert source_rows[0]["candidate_report_url"] == ""
    assert source_rows[0]["status"] == "todo"

    manifest = json.loads(Path(result["manifest"]).read_text(encoding="utf-8"))
    assert manifest["notes"].startswith("Report-source discovery plan only")
    assert manifest["rows"] == 3


def test_build_outcome_report_source_plan_keeps_seeded_scope_beyond_limit(tmp_path: Path):
    plan = tmp_path / "outcome_collection_plan.csv"
    first = _outcome_plan_row("school", "90001", "非种子学校", "employment_rate", status="todo", priority_rank="1")
    first["metric_year"] = "2024"
    seeded = _outcome_plan_row("school", "0152", "大连工业大学", "employment_rate", status="todo", priority_rank="2")
    seeded["metric_year"] = "2024"
    _write_outcome_plan(plan, [first, seeded])

    result = build_outcome_report_source_plan(
        plan_csv=plan,
        output_dir=tmp_path / "report_sources",
        domains=["school"],
        limit_per_domain=2,
    )

    assert result["rows"] == 3
    with Path(result["csv"]).open(encoding="utf-8", newline="") as f:
        source_rows = list(csv.DictReader(f))
    seeded_rows = [row for row in source_rows if row["entity_name"] == "大连工业大学"]
    assert [row["report_scope"] for row in seeded_rows] == ["undergraduate_teaching_quality_report"]
    assert {row["entity_name"] for row in source_rows} == {"非种子学校", "大连工业大学"}


def test_audit_outcome_report_source_plan_requires_confirmed_source(tmp_path: Path):
    plan = tmp_path / "outcome_collection_plan.csv"
    rows = [
        _outcome_plan_row("school", "10140", "辽宁大学", "postgrad_rate", status="todo", priority_rank="1"),
        _outcome_plan_row("school", "10140", "辽宁大学", "employment_rate", status="todo", priority_rank="1"),
    ]
    _write_outcome_plan(plan, rows)
    source_result = build_outcome_report_source_plan(
        plan_csv=plan,
        output_dir=tmp_path / "report_sources",
        domains=["school"],
    )
    report = audit_outcome_report_source_plan(Path(source_result["csv"]))
    assert report["errors"] == []
    assert report["pending_rows"] == 3
    assert report["ready_for_report_intake"] is False

    with Path(source_result["csv"]).open(encoding="utf-8", newline="") as f:
        source_rows = list(csv.DictReader(f))
    source_rows[0]["status"] = "verified"
    verified_plan = tmp_path / "outcome_report_source_plan_verified.csv"
    with verified_plan.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=source_rows[0].keys())
        writer.writeheader()
        writer.writerows(source_rows)

    missing_report = audit_outcome_report_source_plan(verified_plan)
    assert any("complete status missing candidate_report_title" in error for error in missing_report["errors"])

    source_rows[0].update({
        "candidate_report_title": "辽宁大学2025届毕业生就业质量报告",
        "candidate_report_url": "https://example.edu/lnu2025.pdf",
        "candidate_source_date": "2025-12-31",
        "availability_date": "2026-01-05",
    })
    with verified_plan.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=source_rows[0].keys())
        writer.writeheader()
        writer.writerows(source_rows)
    ready_report = audit_outcome_report_source_plan(verified_plan)
    assert ready_report["errors"] == []
    assert ready_report["complete_rows"] == 1
    assert ready_report["ready_for_report_intake"] is True


def test_outcome_report_source_review_batch_merge_is_surgical(tmp_path: Path):
    plan = tmp_path / "outcome_collection_plan.csv"
    rows = [
        _outcome_plan_row("school", "10140", "辽宁大学", "postgrad_rate", status="todo", priority_rank="1"),
        _outcome_plan_row("major", "法学", "法学", "employment_rate", status="todo", priority_rank="2"),
    ]
    _write_outcome_plan(plan, rows)
    source_result = build_outcome_report_source_plan(
        plan_csv=plan,
        output_dir=tmp_path / "report_sources",
    )
    batch_result = build_outcome_report_source_review_batch(
        plan_csv=Path(source_result["csv"]),
        output_dir=tmp_path / "source_batch",
        domains=["school"],
        limit_per_domain=1,
    )
    assert batch_result["rows"] == 1
    with Path(batch_result["csv"]).open(encoding="utf-8", newline="") as f:
        batch_rows = list(csv.DictReader(f))
    batch_rows[0].update({
        "entity_name": "被篡改的学校名",
        "status": "verified",
        "candidate_report_title": "辽宁大学2025届毕业生就业质量报告",
        "candidate_report_url": "https://example.edu/lnu2025.pdf",
        "candidate_source_date": "2025-12-31",
        "availability_date": "2026-01-05",
        "local_report_path": "/tmp/lnu2025.pdf",
        "reviewer": "fixture",
    })
    edited_batch = tmp_path / "source_batch" / "edited.csv"
    with edited_batch.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=batch_rows[0].keys())
        writer.writeheader()
        writer.writerows(batch_rows)

    output = tmp_path / "outcome_report_source_plan_merged.csv"
    report = merge_outcome_report_source_review_batch(
        plan_csv=Path(source_result["csv"]),
        batch_csv=edited_batch,
        output=output,
    )

    assert report["updated_rows"] == 1
    with output.open(encoding="utf-8", newline="") as f:
        merged_rows = list(csv.DictReader(f))
    school_row = next(row for row in merged_rows if row["domain"] == "school")
    assert school_row["entity_name"] == "辽宁大学"
    assert school_row["status"] == "verified"
    assert school_row["candidate_report_url"] == "https://example.edu/lnu2025.pdf"
    assert school_row["local_report_path"] == "/tmp/lnu2025.pdf"
    assert any(row["domain"] == "major" and row["status"] == "todo" for row in merged_rows)


def test_inherit_verified_outcome_collection_rows_reuses_matching_task_keys(tmp_path: Path):
    rebuilt_plan = tmp_path / "rebuilt_outcome_collection_plan.csv"
    rebuilt_rows = [
        _outcome_plan_row("school", "1001", "A大学", "employment_rate", status="todo", priority_rank="1"),
        _outcome_plan_row("school", "1002", "B大学", "employment_rate", status="todo", priority_rank="2"),
    ]
    for row in rebuilt_rows:
        row["metric_year"] = "2024"
    _write_outcome_plan(rebuilt_plan, rebuilt_rows)

    verified_plan = tmp_path / "verified_outcome_collection_plan.csv"
    verified_rows = [
        _outcome_plan_row("school", "1001", "A大学", "employment_rate", status="verified", priority_rank="9"),
        _outcome_plan_row("school", "9999", "旧学校", "employment_rate", status="verified", priority_rank="10"),
    ]
    for row in verified_rows:
        row["metric_year"] = "2024"
        row["metric_value"] = "0.91"
        row["source_title"] = "就业质量报告"
        row["source_url"] = "https://example.edu/report.pdf"
        row["evidence_quote"] = "毕业去向落实率为91%。"
        row["metric_scope"] = "2024届毕业生毕业去向落实率"
        row["source_date"] = "2025-01-01"
        row["availability_date"] = "2025-01-01"
        row["built_at"] = "2026-01-01T00:00:00"
        row["notes"] = "approved fixture"
    _write_outcome_plan(verified_plan, verified_rows)

    output = tmp_path / "rebuilt_outcome_collection_plan.inherited.csv"
    report = inherit_verified_outcome_collection_rows(
        plan_csv=rebuilt_plan,
        verified_plan_csv=verified_plan,
        output=output,
        report_path=tmp_path / "inherit_report.json",
    )

    assert report["reusable_verified_rows"] == 2
    assert report["inherited_rows"] == 1
    assert report["unmatched_verified_rows"] == 1
    assert report["status_counts"] == {"todo": 1, "verified": 1}
    with output.open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    inherited = next(row for row in rows if row["entity_code"] == "1001")
    untouched = next(row for row in rows if row["entity_code"] == "1002")
    assert inherited["status"] == "verified"
    assert inherited["metric_value"] == "0.91"
    assert inherited["evidence_quote"] == "毕业去向落实率为91%。"
    assert untouched["status"] == "todo"
    assert (tmp_path / "inherit_report.json").exists()


def test_apply_outcome_report_source_seeds_updates_matching_pending_rows(tmp_path: Path):
    seed_config = load_outcome_report_sources()
    assert seed_config["applied_status"] == "candidate_found"
    plan = tmp_path / "outcome_collection_plan.csv"
    rows = [
        _outcome_plan_row("school", "10140", "辽宁大学", "employment_rate", status="todo", priority_rank="1"),
        _outcome_plan_row("school", "10142", "沈阳工业大学", "employment_rate", status="todo", priority_rank="1"),
        _outcome_plan_row("school", "11258", "大连大学", "employment_rate", status="todo", priority_rank="1"),
        _outcome_plan_row("school", "0166", "沈阳师范大学", "employment_rate", status="todo", priority_rank="1"),
        _outcome_plan_row("major", "法学", "法学", "employment_rate", status="todo", priority_rank="2"),
    ]
    rows[0]["metric_year"] = "2022"
    rows[1]["metric_year"] = "2024"
    rows[2]["metric_year"] = "2024"
    rows[3]["metric_year"] = "2024"
    rows[4]["metric_year"] = "2022"
    _write_outcome_plan(plan, rows)
    source_result = build_outcome_report_source_plan(
        plan_csv=plan,
        output_dir=tmp_path / "report_sources",
    )

    output = tmp_path / "report_sources" / "outcome_report_source_plan_seeded.csv"
    report = apply_outcome_report_source_seeds(
        plan_csv=Path(source_result["csv"]),
        output=output,
        report_path=tmp_path / "report_sources" / "seed_merge.json",
    )

    assert report["updated_rows"] == 3
    assert "school|辽宁大学|2022|employment_quality_report" not in report["unmatched_seed_ids"]
    assert "school|沈阳工业大学|2024|undergraduate_teaching_quality_report" not in report["unmatched_seed_ids"]
    assert "school|大连大学|2024|undergraduate_teaching_quality_report" not in report["unmatched_seed_ids"]
    assert "school|0166|2024|undergraduate_teaching_quality_report" in report["inactive_seed_ids"]
    with output.open(encoding="utf-8", newline="") as f:
        merged_rows = list(csv.DictReader(f))
    lnu = next(row for row in merged_rows if row["entity_name"] == "辽宁大学" and row["report_scope"] == "employment_quality_report")
    assert lnu["status"] == "candidate_found"
    assert lnu["candidate_report_url"] == "https://www.lnu.edu.cn/info/15026/78891.htm"
    assert lnu["candidate_source_date"] == "2023-01-05"
    sut = next(row for row in merged_rows if row["entity_name"] == "沈阳工业大学" and row["report_scope"] == "undergraduate_teaching_quality_report")
    assert sut["candidate_report_url"] == "https://www.sut.edu.cn/info/1584/67026.htm"
    dlu = next(row for row in merged_rows if row["entity_name"] == "大连大学" and row["report_scope"] == "undergraduate_teaching_quality_report")
    assert dlu["candidate_report_url"] == "https://zgc.dlu.edu.cn/__local/5/C0/61/B92A4F9A28A99F089DF02DC8B08_93E673A2_121692.pdf"
    synu = next(row for row in merged_rows if row["entity_name"] == "沈阳师范大学" and row["report_scope"] == "undergraduate_teaching_quality_report")
    assert synu["status"] == "todo"
    assert synu["candidate_report_url"] == ""
    major = next(row for row in merged_rows if row["domain"] == "major")
    assert major["status"] == "todo"


def test_audit_outcome_report_source_seeds_validates_config(tmp_path: Path):
    report = audit_outcome_report_source_seeds(report_path=tmp_path / "seed_audit.json")

    assert report["errors"] == []
    assert report["seed_count"] >= 7
    assert report["inactive_seed_count"] >= 1
    assert report["applied_status"] == "candidate_found"
    assert not any("school seed missing entity_code" in warning for warning in report["warnings"])
    assert all(row["entity_code"] for row in report["seed_rows"] if row["domain"] == "school")
    assert any(row["entity_name"] == "辽宁大学" for row in report["seed_rows"])
    assert any(row["entity_name"] == "吉林大学" and row["metric_year"] == 2024 for row in report["seed_rows"])
    assert any(row["entity_name"] == "沈阳师范大学" and row["metric_year"] == 2024 and row["seed_status"] == "rejected" for row in report["seed_rows"])
    assert any(row["entity_name"] == "大连外国语大学" and row["metric_year"] == 2024 for row in report["seed_rows"])
    assert any(row["entity_name"] == "辽宁师范大学" and row["metric_year"] == 2024 for row in report["seed_rows"])
    assert any(row["entity_name"] == "大连大学" and row["metric_year"] == 2024 for row in report["seed_rows"])
    assert any(row["entity_name"] == "大连工业大学" and row["metric_year"] == 2024 for row in report["seed_rows"])
    assert any(row["entity_name"] == "大连民族大学" and row["metric_year"] == 2024 for row in report["seed_rows"])
    assert any(row["entity_name"] == "渤海大学" and row["metric_year"] == 2024 for row in report["seed_rows"])
    assert (tmp_path / "seed_audit.json").exists()


def test_audit_outcome_report_source_seeds_rejects_bad_metadata(monkeypatch: pytest.MonkeyPatch):
    seed = {
        "domain": "school",
        "entity_name": "辽宁大学",
        "metric_year": "2024.5",
        "report_scope": "bad_report_scope",
        "candidate_report_title": "辽宁大学本科教学质量报告",
        "candidate_report_url": "ftp://example.com/lnu.pdf",
        "candidate_source_date": "2025-01-01",
        "availability_date": "2024-12-31",
        "evidence_note": "测试报告来源元数据校验。",
    }
    monkeypatch.setattr(
        "datahub.builders.outcome_report_source_seed_merge.load_outcome_report_sources",
        lambda: {
            "version": "test",
            "applied_status": "candidate_found",
            "reviewer": "datahub_seed",
            "seeds": [seed],
        },
    )

    report = audit_outcome_report_source_seeds()

    assert "seed 1 report_scope is not configured for domain school: bad_report_scope" in report["errors"]
    assert "seed 1 metric_year is not an integer" in report["errors"]
    assert "seed 1 candidate_source_date must not be after availability_date" in report["errors"]
    assert "seed 1 candidate_report_url must be an http(s) URL: ftp://example.com/lnu.pdf" in report["errors"]


def test_build_outcome_report_intake_plan_requires_confirmed_url(tmp_path: Path):
    plan = tmp_path / "outcome_collection_plan.csv"
    rows = [
        _outcome_plan_row("school", "10140", "辽宁大学", "employment_rate", status="todo", priority_rank="1"),
        _outcome_plan_row("school", "10142", "沈阳工业大学", "employment_rate", status="todo", priority_rank="2"),
    ]
    rows[0]["metric_year"] = "2022"
    rows[1]["metric_year"] = "2024"
    _write_outcome_plan(plan, rows)
    source_result = build_outcome_report_source_plan(
        plan_csv=plan,
        output_dir=tmp_path / "report_sources",
        domains=["school"],
    )
    seeded_plan = tmp_path / "report_sources" / "outcome_report_source_plan_seeded.csv"
    apply_outcome_report_source_seeds(
        plan_csv=Path(source_result["csv"]),
        output=seeded_plan,
    )

    result = build_outcome_report_intake_plan(
        report_source_csv=seeded_plan,
        output_dir=tmp_path / "report_intake",
    )

    assert result["ready_rows"] == 2
    assert result["blocked_rows"] == 0
    with Path(result["csv"]).open(encoding="utf-8", newline="") as f:
        intake_rows = list(csv.DictReader(f))
    assert {row["intake_status"] for row in intake_rows} == {"ready_for_intake"}
    lnu = next(row for row in intake_rows if row["entity_name"] == "辽宁大学")
    assert lnu["suggested_local_report_path"].startswith("raw/outcome_report/2022/")
    assert lnu["candidate_report_url"] == "https://www.lnu.edu.cn/info/15026/78891.htm"


def test_download_outcome_report_intake_assets_follows_html_attachment(tmp_path: Path, monkeypatch):
    intake_csv = tmp_path / "outcome_report_intake_plan.csv"
    target_pdf = tmp_path / "raw" / "outcome_report" / "2024" / "report.pdf"
    row = {
        "domain": "school",
        "entity_code": "10140",
        "entity_name": "辽宁大学",
        "metric_year": "2024",
        "report_scope": "undergraduate_teaching_quality_report",
        "candidate_report_title": "辽宁大学2023-2024学年本科教学质量报告",
        "candidate_report_url": "https://example.edu/info/1.htm",
        "candidate_file_name": "辽宁大学2023-2024学年本科教学质量报告.pdf",
        "candidate_source_date": "2024-12-06",
        "availability_date": "2024-12-06",
        "suggested_local_report_path": str(target_pdf),
        "local_report_path": "",
        "intake_status": "ready_for_intake",
        "block_reason": "",
        "source_status": "candidate_found",
        "notes": "",
    }
    with intake_csv.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(row))
        writer.writeheader()
        writer.writerow(row)

    class FakeHeaders:
        def __init__(self, content_type: str):
            self._content_type = content_type

        def get(self, name: str, default=None):
            return self._content_type if name == "Content-Type" else default

    class FakeResponse:
        def __init__(self, url: str, body: bytes, content_type: str):
            self._url = url
            self._body = body
            self.headers = FakeHeaders(content_type)

        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, tb):
            return False

        def read(self):
            return self._body

        def geturl(self):
            return self._url

    def fake_urlopen(request, timeout=60):
        url = request.full_url
        if url.endswith("/info/1.htm"):
            html = '<html><body><a href="/system/download.jsp?id=1">辽宁大学2023-2024学年本科教学质量报告.pdf</a></body></html>'
            return FakeResponse(url, html.encode("utf-8"), "text/html; charset=utf-8")
        if url.endswith("/system/download.jsp?id=1"):
            return FakeResponse(url, b"%PDF-1.4 fixture", "application/x-download")
        raise AssertionError(url)

    monkeypatch.setattr("datahub.connectors.outcome_report_download.urlopen", fake_urlopen)

    output = tmp_path / "outcome_report_intake_plan.downloaded.csv"
    report = download_outcome_report_intake_assets(
        intake_csv=intake_csv,
        output=output,
    )

    assert report["downloaded_rows"] == 1
    assert report["failed_rows"] == 0
    assert report["failure_reason_counts"] == {}
    assert target_pdf.read_bytes() == b"%PDF-1.4 fixture"
    with output.open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    assert rows[0]["intake_status"] == "downloaded"
    assert rows[0]["local_report_path"] == str(target_pdf)
    assert rows[0]["download_url"] == "https://example.edu/system/download.jsp?id=1"
    assert rows[0]["download_sha256"] == hashlib.sha256(b"%PDF-1.4 fixture").hexdigest()


def test_download_outcome_report_intake_assets_extracts_cloud_zip_attachment(tmp_path: Path, monkeypatch):
    intake_csv = tmp_path / "outcome_report_intake_plan.csv"
    target_pdf = tmp_path / "raw" / "outcome_report" / "2024" / "bohai.pdf"
    row = {
        "domain": "school",
        "entity_code": "0167",
        "entity_name": "渤海大学",
        "metric_year": "2024",
        "report_scope": "undergraduate_teaching_quality_report",
        "candidate_report_title": "渤海大学2023-2024学年本科教学质量报告",
        "candidate_report_url": "https://example.edu/engine2/general/4172508/detail",
        "candidate_file_name": "渤海大学2023-2024学年本科教学质量报告.pdf",
        "candidate_source_date": "2024-12-19",
        "availability_date": "2024-12-19",
        "suggested_local_report_path": str(target_pdf),
        "local_report_path": "",
        "intake_status": "ready_for_intake",
        "block_reason": "",
        "source_status": "candidate_found",
        "notes": "",
    }
    with intake_csv.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(row))
        writer.writeheader()
        writer.writerow(row)

    pdf_body = b"%PDF-1.4 bohai fixture"
    archive_buffer = io.BytesIO()
    with ZipFile(archive_buffer, "w") as archive:
        archive.writestr("folder/bohai-report.pdf", pdf_body)
        archive.writestr("folder/support.xls", b"xls")
    payload = {
        "att_clouddisk": {
            "downPath": "/engine/upload/engine/2024-12/report.zip",
            "name": "渤海大学2023-2024学年本科教学质量报告.zip",
            "suffix": "zip",
        }
    }
    encoded = base64.b64encode(json.dumps(payload, ensure_ascii=False).encode("utf-8")).decode("ascii")

    class FakeHeaders:
        def __init__(self, content_type: str):
            self._content_type = content_type

        def get(self, name: str, default=None):
            return self._content_type if name == "Content-Type" else default

    class FakeResponse:
        def __init__(self, url: str, body: bytes, content_type: str):
            self._url = url
            self._body = body
            self.headers = FakeHeaders(content_type)

        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, tb):
            return False

        def read(self):
            return self._body

        def geturl(self):
            return self._url

    def fake_urlopen(request, timeout=60):
        url = request.full_url
        if url.endswith("/engine2/general/4172508/detail"):
            html = f'<html><iframe src="/engine2/assets/attachment/insertCloud.html" name="{encoded}"></iframe></html>'
            return FakeResponse(url, html.encode("utf-8"), "text/html; charset=utf-8")
        if url.endswith("/engine/upload/engine/2024-12/report.zip"):
            return FakeResponse(url, archive_buffer.getvalue(), "application/zip")
        raise AssertionError(url)

    monkeypatch.setattr("datahub.connectors.outcome_report_download.urlopen", fake_urlopen)

    output = tmp_path / "outcome_report_intake_plan.downloaded.csv"
    report = download_outcome_report_intake_assets(
        intake_csv=intake_csv,
        output=output,
    )

    assert report["downloaded_rows"] == 1
    assert target_pdf.read_bytes() == pdf_body
    with output.open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    assert rows[0]["download_url"].endswith("/engine/upload/engine/2024-12/report.zip#folder/bohai-report.pdf")
    assert rows[0]["download_sha256"] == hashlib.sha256(pdf_body).hexdigest()


def test_download_outcome_report_intake_assets_follows_vsb_pdf_iframe(tmp_path: Path, monkeypatch):
    intake_csv = tmp_path / "outcome_report_intake_plan.csv"
    target_pdf = tmp_path / "raw" / "outcome_report" / "2024" / "maple.pdf"
    row = {
        "domain": "school",
        "entity_code": "3961",
        "entity_name": "大连枫叶职业技术学院",
        "metric_year": "2024",
        "report_scope": "employment_quality_report",
        "candidate_report_title": "大连枫叶职业技术学院2024年质量年报",
        "candidate_report_url": "https://example.edu/info/1044/5895.htm",
        "candidate_file_name": "大连枫叶职业技术学院2024年质量年报.pdf",
        "candidate_source_date": "2025-02-16",
        "availability_date": "2025-02-16",
        "suggested_local_report_path": str(target_pdf),
        "local_report_path": "",
        "intake_status": "ready_for_intake",
        "block_reason": "",
        "source_status": "candidate_found",
        "notes": "",
    }
    with intake_csv.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(row))
        writer.writeheader()
        writer.writerow(row)

    class FakeHeaders:
        def __init__(self, content_type: str):
            self._content_type = content_type

        def get(self, name: str, default=None):
            return self._content_type if name == "Content-Type" else default

    class FakeResponse:
        def __init__(self, url: str, body: bytes, content_type: str):
            self._url = url
            self._body = body
            self.headers = FakeHeaders(content_type)

        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, tb):
            return False

        def read(self):
            return self._body

        def geturl(self):
            return self._url

    def fake_urlopen(request, timeout=60):
        url = request.full_url
        if url.endswith("/info/1044/5895.htm"):
            html = '<script>showVsbpdfIframe("/__local/A/D6/21/maple.pdf","100%","600","0","",vsb_pdf_image_data);</script>'
            return FakeResponse(url, html.encode("utf-8"), "text/html; charset=utf-8")
        if url.endswith("/__local/A/D6/21/maple.pdf"):
            return FakeResponse(url, b"%PDF-1.4 maple fixture", "application/pdf")
        raise AssertionError(url)

    monkeypatch.setattr("datahub.connectors.outcome_report_download.urlopen", fake_urlopen)

    output = tmp_path / "outcome_report_intake_plan.downloaded.csv"
    report = download_outcome_report_intake_assets(
        intake_csv=intake_csv,
        output=output,
    )

    assert report["downloaded_rows"] == 1
    assert report["failed_rows"] == 0
    assert target_pdf.read_bytes() == b"%PDF-1.4 maple fixture"
    with output.open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    assert rows[0]["download_url"] == "https://example.edu/__local/A/D6/21/maple.pdf"


def test_download_outcome_report_intake_assets_extracts_embedded_pdf_data(tmp_path: Path, monkeypatch):
    intake_csv = tmp_path / "outcome_report_intake_plan.csv"
    target_pdf = tmp_path / "raw" / "outcome_report" / "2024" / "dlufe.pdf"
    row = {
        "domain": "school",
        "entity_code": "3218",
        "entity_name": "大连财经学院",
        "metric_year": "2024",
        "report_scope": "undergraduate_teaching_quality_report",
        "candidate_report_title": "大连财经学院2023-2024学年本科教学质量报告",
        "candidate_report_url": "https://example.edu/pdfweb_567.shtml",
        "candidate_file_name": "大连财经学院2023-2024学年本科教学质量报告.pdf",
        "candidate_source_date": "2024-12-05",
        "availability_date": "2024-12-05",
        "suggested_local_report_path": str(target_pdf),
        "local_report_path": "",
        "intake_status": "ready_for_intake",
        "block_reason": "",
        "source_status": "candidate_found",
        "notes": "",
    }
    with intake_csv.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(row))
        writer.writeheader()
        writer.writerow(row)

    class FakeHeaders:
        def __init__(self, content_type: str):
            self._content_type = content_type

        def get(self, name: str, default=None):
            return self._content_type if name == "Content-Type" else default

    class FakeResponse:
        def __init__(self, url: str, body: bytes, content_type: str):
            self._url = url
            self._body = body
            self.headers = FakeHeaders(content_type)

        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, tb):
            return False

        def read(self):
            return self._body

        def geturl(self):
            return self._url

    pdf_body = b"%PDF-1.7 dlufe fixture"
    encoded = base64.b64encode(pdf_body).decode("ascii")

    def fake_urlopen(request, timeout=60):
        html = f"<html><script>var pdfData='{encoded}';</script></html>"
        return FakeResponse(request.full_url, html.encode("utf-8"), "text/html; charset=utf-8")

    monkeypatch.setattr("datahub.connectors.outcome_report_download.urlopen", fake_urlopen)

    output = tmp_path / "downloaded.csv"
    report = download_outcome_report_intake_assets(
        intake_csv=intake_csv,
        output=output,
    )

    assert report["downloaded_rows"] == 1
    assert report["failed_rows"] == 0
    assert target_pdf.read_bytes() == pdf_body
    with output.open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    assert rows[0]["download_url"] == "https://example.edu/pdfweb_567.shtml#embedded-pdf"


def test_download_outcome_report_intake_assets_flags_embedded_report_images(tmp_path: Path, monkeypatch):
    intake_csv = tmp_path / "outcome_report_intake_plan.csv"
    row = {
        "domain": "school",
        "entity_code": "3599",
        "entity_name": "大连艺术学院",
        "metric_year": "2024",
        "report_scope": "undergraduate_teaching_quality_report",
        "candidate_report_title": "大连艺术学院2023-2024学年本科教学质量报告",
        "candidate_report_url": "https://example.edu/detail/292_image_report.html",
        "candidate_file_name": "大连艺术学院2023-2024学年本科教学质量报告.pdf",
        "candidate_source_date": "2024-12-05",
        "availability_date": "2024-12-05",
        "suggested_local_report_path": str(tmp_path / "report.pdf"),
        "local_report_path": "",
        "intake_status": "ready_for_intake",
        "block_reason": "",
        "source_status": "candidate_found",
        "notes": "",
    }
    with intake_csv.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(row))
        writer.writeheader()
        writer.writerow(row)

    class FakeHeaders:
        def __init__(self, content_type: str):
            self._content_type = content_type

        def get(self, name: str, default=None):
            return self._content_type if name == "Content-Type" else default

    class FakeResponse:
        def __init__(self, url: str, body: bytes, content_type: str):
            self._url = url
            self._body = body
            self.headers = FakeHeaders(content_type)

        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, tb):
            return False

        def read(self):
            return self._body

        def geturl(self):
            return self._url

    def fake_urlopen(request, timeout=60):
        html = """
        <html><body>
          <a href="/detail/292_image_report.html">大连艺术学院2023-2024学年本科教学质量报告</a>
          <img alt="大连艺术学院2023-2024学年本科教学质量报告-第1页" src="/image/page1.jpg">
          <img alt="大连艺术学院2023-2024学年本科教学质量报告-第2页" src="/image/page2.jpg">
          <img alt="大连艺术学院2023-2024学年本科教学质量报告-第3页" src="/image/page3.jpg">
        </body></html>
        """
        return FakeResponse(request.full_url, html.encode("utf-8"), "text/html; charset=utf-8")

    monkeypatch.setattr("datahub.connectors.outcome_report_download.urlopen", fake_urlopen)

    output = tmp_path / "downloaded.csv"
    report = download_outcome_report_intake_assets(
        intake_csv=intake_csv,
        output=output,
    )

    assert report["downloaded_rows"] == 0
    assert report["failed_rows"] == 1
    assert report["failure_reason_counts"] == {
        "report rendered as embedded report images; OCR or manual intake required": 1
    }

    manual_queue = tmp_path / "manual_intake_queue.csv"
    manual_report = build_outcome_report_manual_intake_queue(
        intake_results_csv=output,
        output=manual_queue,
    )
    assert manual_report["reason_counts"] == {"image_pdf_ocr_required": 1}
    with manual_queue.open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    assert rows[0]["recommended_action"] == "ocr_or_manual_transcription"


def test_download_outcome_report_intake_assets_extracts_viewer_file_url(tmp_path: Path, monkeypatch):
    intake_csv = tmp_path / "outcome_report_intake_plan.csv"
    target_pdf = tmp_path / "raw" / "outcome_report" / "2024" / "neau.pdf"
    row = {
        "domain": "school",
        "entity_code": "0224",
        "entity_name": "东北农业大学",
        "metric_year": "2024",
        "report_scope": "employment_quality_report",
        "candidate_report_title": "东北农业大学2024届毕业生就业质量年度报告",
        "candidate_report_url": "https://neau.bysjy.com.cn/detail/news?id=1119250",
        "candidate_file_name": "东北农业大学2024届毕业生就业质量年度报告.pdf",
        "candidate_source_date": "2025-03-01",
        "availability_date": "2025-03-01",
        "suggested_local_report_path": str(target_pdf),
        "local_report_path": "",
        "intake_status": "ready_for_intake",
        "block_reason": "",
        "source_status": "candidate_found",
        "notes": "",
    }
    with intake_csv.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(row))
        writer.writeheader()
        writer.writerow(row)

    class FakeHeaders:
        def __init__(self, content_type: str):
            self._content_type = content_type

        def get(self, name: str, default=None):
            return self._content_type if name == "Content-Type" else default

    class FakeResponse:
        def __init__(self, url: str, body: bytes, content_type: str):
            self._url = url
            self._body = body
            self.headers = FakeHeaders(content_type)

        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, tb):
            return False

        def read(self):
            return self._body

        def geturl(self):
            return self._url

    def fake_urlopen(request, timeout=60):
        url = request.full_url
        if "detail/news" in url:
            html = (
                '<html><body><a href="http://js.bysjy.com.cn/default/quality_report/pdf.html?'
                'fileUrl=https%3A%2F%2Fo.bysjy.com.cn%2Fdocument%2F1735538975-4651.pdf">'
                "东北农业大学2024届毕业生就业质量年度报告</a></body></html>"
            )
            return FakeResponse(url, html.encode("utf-8"), "text/html; charset=utf-8")
        if url == "https://o.bysjy.com.cn/document/1735538975-4651.pdf":
            return FakeResponse(url, b"%PDF-1.4 neau fixture", "application/pdf")
        raise AssertionError(url)

    monkeypatch.setattr("datahub.connectors.outcome_report_download.urlopen", fake_urlopen)

    output = tmp_path / "downloaded.csv"
    report = download_outcome_report_intake_assets(
        intake_csv=intake_csv,
        output=output,
    )

    assert report["downloaded_rows"] == 1
    assert target_pdf.read_bytes() == b"%PDF-1.4 neau fixture"
    with output.open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    assert rows[0]["download_url"] == "https://o.bysjy.com.cn/document/1735538975-4651.pdf"


def test_download_outcome_report_intake_assets_summarizes_failure_reasons(tmp_path: Path, monkeypatch):
    intake_csv = tmp_path / "outcome_report_intake_plan.csv"
    row = {
        "domain": "school",
        "entity_code": "10142",
        "entity_name": "沈阳工业大学",
        "metric_year": "2024",
        "report_scope": "undergraduate_teaching_quality_report",
        "candidate_report_title": "2023-2024年沈阳工业大学本科教学质量报告",
        "candidate_report_url": "https://example.edu/info/2.htm",
        "candidate_file_name": "2023-2024年沈阳工业大学本科教学质量报告.pdf",
        "candidate_source_date": "2024-12-05",
        "availability_date": "2024-12-05",
        "suggested_local_report_path": str(tmp_path / "report.pdf"),
        "local_report_path": "",
        "intake_status": "ready_for_intake",
        "block_reason": "",
        "source_status": "candidate_found",
        "notes": "",
    }
    with intake_csv.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(row))
        writer.writeheader()
        writer.writerow(row)

    class FakeHeaders:
        def __init__(self, content_type: str):
            self._content_type = content_type

        def get(self, name: str, default=None):
            return self._content_type if name == "Content-Type" else default

    class FakeResponse:
        def __init__(self, url: str, body: bytes, content_type: str):
            self._url = url
            self._body = body
            self.headers = FakeHeaders(content_type)

        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, tb):
            return False

        def read(self):
            return self._body

        def geturl(self):
            return self._url

    def fake_urlopen(request, timeout=60):
        url = request.full_url
        if url.endswith("/info/2.htm"):
            html = '<html><body><a href="/system/download.jsp?id=2">2023-2024年沈阳工业大学本科教学质量报告.pdf</a></body></html>'
            return FakeResponse(url, html.encode("utf-8"), "text/html; charset=utf-8")
        if url.endswith("/system/download.jsp?id=2"):
            return FakeResponse(url, "<html>验证码 codeValue</html>".encode("utf-8"), "text/html; charset=utf-8")
        raise AssertionError(url)

    monkeypatch.setattr("datahub.connectors.outcome_report_download.urlopen", fake_urlopen)

    report = download_outcome_report_intake_assets(
        intake_csv=intake_csv,
        output=tmp_path / "downloaded.csv",
    )

    assert report["downloaded_rows"] == 0
    assert report["failed_rows"] == 1
    assert report["failure_reason_counts"] == {"attachment requires captcha or manual intake": 1}


def test_download_outcome_report_intake_assets_classifies_ssl_eof_failures(tmp_path: Path, monkeypatch):
    intake_csv = tmp_path / "outcome_report_intake_plan.csv"
    row = {
        "domain": "school",
        "entity_code": "0728",
        "entity_name": "西安音乐学院",
        "metric_year": "2024",
        "report_scope": "undergraduate_teaching_quality_report",
        "candidate_report_title": "西安音乐学院2023-2024学年本科教学质量报告",
        "candidate_report_url": "https://example.edu/report.pdf",
        "candidate_file_name": "西安音乐学院2023-2024学年本科教学质量报告.pdf",
        "candidate_source_date": "2024-12-31",
        "availability_date": "2024-12-31",
        "suggested_local_report_path": str(tmp_path / "report.pdf"),
        "local_report_path": "",
        "intake_status": "ready_for_intake",
        "block_reason": "",
        "source_status": "candidate_found",
        "notes": "",
    }
    with intake_csv.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(row))
        writer.writeheader()
        writer.writerow(row)

    def fake_urlopen(request, timeout=60):
        raise OSError("EOF occurred in violation of protocol (_ssl.c:1129)")

    monkeypatch.setattr("datahub.connectors.outcome_report_download.urlopen", fake_urlopen)

    report = download_outcome_report_intake_assets(
        intake_csv=intake_csv,
        output=tmp_path / "downloaded.csv",
    )

    assert report["downloaded_rows"] == 0
    assert report["failed_rows"] == 1
    assert report["failure_reason_counts"] == {"ssl handshake failed; manual intake required": 1}


def test_build_outcome_report_manual_intake_queue_classifies_failed_downloads(tmp_path: Path):
    intake_results_csv = tmp_path / "outcome_report_intake_results.csv"
    rows = [
        {
            "domain": "school",
            "entity_code": "0728",
            "entity_name": "西安音乐学院",
            "metric_year": "2024",
            "report_scope": "undergraduate_teaching_quality_report",
            "candidate_report_title": "西安音乐学院2023-2024学年本科教学质量报告",
            "candidate_report_url": "https://example.edu/xian.pdf",
            "candidate_file_name": "西安音乐学院2023-2024学年本科教学质量报告.pdf",
            "download_status": "failed",
            "download_error": "<urlopen error EOF occurred in violation of protocol (_ssl.c:1129)>",
        },
        {
            "domain": "school",
            "entity_code": "0157",
            "entity_name": "沈阳农业大学",
            "metric_year": "2024",
            "report_scope": "undergraduate_teaching_quality_report",
            "candidate_report_title": "沈阳农业大学2023-2024学年本科教学质量报告",
            "candidate_report_url": "https://example.edu/syau.htm",
            "candidate_file_name": "沈阳农业大学2023-2024学年本科教学质量报告.pdf",
            "download_status": "failed",
            "download_error": "attachment requires captcha or manual intake: https://example.edu/download",
        },
        {
            "domain": "school",
            "entity_code": "1258",
            "entity_name": "大连大学",
            "metric_year": "2024",
            "report_scope": "undergraduate_teaching_quality_report",
            "candidate_report_title": "大连大学2023-2024学年本科教学质量报告",
            "candidate_report_url": "https://example.edu/dlu.htm",
            "candidate_file_name": "大连大学2023-2024学年本科教学质量报告.pdf",
            "download_status": "failed",
            "download_error": "report rendered as PDF page images; OCR or manual intake required",
        },
        {
            "domain": "school",
            "entity_code": "4535",
            "entity_name": "浙江音乐学院",
            "metric_year": "2024",
            "report_scope": "undergraduate_teaching_quality_report",
            "candidate_report_title": "浙江音乐学院2023-2024学年本科教学质量报告",
            "candidate_report_url": "https://example.edu/zjcm.htm",
            "candidate_file_name": "浙江音乐学院2023-2024学年本科教学质量报告.pdf",
            "download_status": "downloaded",
            "download_error": "",
        },
    ]
    with intake_results_csv.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)

    output = tmp_path / "manual_intake_queue.csv"
    report = build_outcome_report_manual_intake_queue(
        intake_results_csv=intake_results_csv,
        output=output,
    )

    assert report["rows"] == 3
    assert report["reason_counts"] == {
        "captcha_required": 1,
        "image_pdf_ocr_required": 1,
        "ssl_handshake_failed": 1,
    }
    with output.open(encoding="utf-8", newline="") as f:
        output_rows = list(csv.DictReader(f))
    assert [row["entity_code"] for row in output_rows] == ["0728", "0157", "1258"]
    assert output_rows[0]["recommended_action"] == "manual_download_or_downloader_tls_fallback"
    assert output_rows[1]["recommended_action"] == "manual_browser_download"
    assert output_rows[2]["recommended_action"] == "ocr_or_manual_transcription"


def test_download_outcome_report_intake_assets_sends_referer_for_direct_attachment(tmp_path: Path, monkeypatch):
    intake_csv = tmp_path / "outcome_report_intake_plan.csv"
    row = {
        "domain": "school",
        "entity_code": "1688",
        "entity_name": "山东工商学院",
        "metric_year": "2024",
        "report_scope": "undergraduate_teaching_quality_report",
        "candidate_report_title": "山东工商学院2023-2024学年本科教学质量报告",
        "candidate_report_url": "https://example.edu/system/_content/download.jsp?id=1",
        "candidate_file_name": "山东工商学院2023-2024学年本科教学质量报告.pdf",
        "candidate_source_date": "2024-12-05",
        "availability_date": "2024-12-05",
        "suggested_local_report_path": str(tmp_path / "report.pdf"),
        "local_report_path": "",
        "intake_status": "ready_for_intake",
        "block_reason": "",
        "source_status": "candidate_found",
        "notes": "",
    }
    with intake_csv.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(row))
        writer.writeheader()
        writer.writerow(row)

    class FakeHeaders:
        def get(self, name: str, default=None):
            return "application/octet-stream" if name == "Content-Type" else default

    class FakeResponse:
        headers = FakeHeaders()

        def __init__(self, url: str):
            self._url = url

        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, tb):
            return False

        def read(self):
            return b"%PDF-1.7\nfixture"

        def geturl(self):
            return self._url

    seen_headers = {}

    def fake_urlopen(request, timeout=60):
        seen_headers.update(request.headers)
        assert request.headers["Referer"] == request.full_url
        return FakeResponse(request.full_url)

    monkeypatch.setattr("datahub.connectors.outcome_report_download.urlopen", fake_urlopen)

    report = download_outcome_report_intake_assets(
        intake_csv=intake_csv,
        output=tmp_path / "downloaded.csv",
    )

    assert report["downloaded_rows"] == 1
    assert seen_headers["Referer"] == row["candidate_report_url"]


def test_download_outcome_report_intake_assets_percent_encodes_chinese_direct_url(tmp_path: Path, monkeypatch):
    intake_csv = tmp_path / "outcome_report_intake_plan.csv"
    row = {
        "domain": "school",
        "entity_code": "0177",
        "entity_name": "沈阳音乐学院",
        "metric_year": "2024",
        "report_scope": "undergraduate_teaching_quality_report",
        "candidate_report_title": "沈阳音乐学院2023-2024学年本科教学质量报告",
        "candidate_report_url": "https://example.edu/video/沈阳音乐学院2023-2024学年本科教学质量报告.pdf",
        "candidate_file_name": "沈阳音乐学院2023-2024学年本科教学质量报告.pdf",
        "candidate_source_date": "2024-12-05",
        "availability_date": "2024-12-05",
        "suggested_local_report_path": str(tmp_path / "sycm.pdf"),
        "local_report_path": "",
        "intake_status": "ready_for_intake",
        "block_reason": "",
        "source_status": "candidate_found",
        "notes": "",
    }
    with intake_csv.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(row))
        writer.writeheader()
        writer.writerow(row)

    class FakeHeaders:
        def get(self, name: str, default=None):
            return "application/pdf" if name == "Content-Type" else default

    class FakeResponse:
        headers = FakeHeaders()

        def __init__(self, url: str):
            self._url = url

        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, tb):
            return False

        def read(self):
            return b"%PDF-1.7\nsycm fixture"

        def geturl(self):
            return self._url

    seen = {}

    def fake_urlopen(request, timeout=60):
        seen["url"] = request.full_url
        seen["referer"] = request.headers["Referer"]
        assert "%E6%B2%88%E9%98%B3%E9%9F%B3%E4%B9%90%E5%AD%A6%E9%99%A2" in request.full_url
        return FakeResponse(request.full_url)

    monkeypatch.setattr("datahub.connectors.outcome_report_download.urlopen", fake_urlopen)

    report = download_outcome_report_intake_assets(
        intake_csv=intake_csv,
        output=tmp_path / "downloaded.csv",
    )

    assert report["downloaded_rows"] == 1
    assert seen["url"] == seen["referer"]


def test_cli_download_outcome_report_intake_assets_can_allow_partial_failures(tmp_path: Path, monkeypatch):
    from datahub import cli
    from datahub.commands import outcome

    def fake_download(**kwargs):
        return {
            "intake_csv": str(kwargs["intake_csv"]),
            "output": str(kwargs["output"]),
            "rows": 2,
            "downloaded_rows": 1,
            "failed_rows": 1,
        }

    monkeypatch.setattr(outcome, "download_outcome_report_intake_assets", fake_download)
    base_argv = [
        "lifehack-datahub",
        "download-outcome-report-intake-assets",
        "--intake-csv",
        str(tmp_path / "intake.csv"),
        "--output",
        str(tmp_path / "downloaded.csv"),
    ]
    monkeypatch.setattr("sys.argv", base_argv)
    assert cli.main() == 1

    monkeypatch.setattr("sys.argv", [*base_argv, "--allow-failures"])
    assert cli.main() == 0


def test_merge_outcome_report_intake_results_requires_existing_file(tmp_path: Path):
    plan = tmp_path / "outcome_collection_plan.csv"
    rows = [
        _outcome_plan_row("school", "10140", "辽宁大学", "employment_rate", status="todo", priority_rank="1"),
        _outcome_plan_row("school", "10142", "沈阳工业大学", "employment_rate", status="todo", priority_rank="2"),
    ]
    rows[0]["metric_year"] = "2022"
    rows[1]["metric_year"] = "2024"
    _write_outcome_plan(plan, rows)
    source_result = build_outcome_report_source_plan(
        plan_csv=plan,
        output_dir=tmp_path / "report_sources",
        domains=["school"],
    )
    seeded_plan = tmp_path / "report_sources" / "outcome_report_source_plan_seeded.csv"
    apply_outcome_report_source_seeds(
        plan_csv=Path(source_result["csv"]),
        output=seeded_plan,
    )
    intake_result = build_outcome_report_intake_plan(
        report_source_csv=seeded_plan,
        output_dir=tmp_path / "report_intake",
    )
    with Path(intake_result["csv"]).open(encoding="utf-8", newline="") as f:
        intake_rows = list(csv.DictReader(f))
    local_pdf = tmp_path / "raw" / "lnu2022.pdf"
    local_pdf.parent.mkdir(parents=True)
    local_pdf.write_bytes(b"%PDF-1.4 fixture")
    intake_rows[0]["local_report_path"] = str(local_pdf)
    intake_rows[0]["intake_status"] = "downloaded"
    intake_rows[1]["local_report_path"] = str(tmp_path / "raw" / "missing.pdf")
    intake_rows[1]["intake_status"] = "downloaded"
    edited_intake = tmp_path / "report_intake" / "outcome_report_intake_plan_reviewed.csv"
    with edited_intake.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(intake_rows[0]))
        writer.writeheader()
        writer.writerows(intake_rows)

    output = tmp_path / "report_sources" / "outcome_report_source_plan_with_paths.csv"
    report = merge_outcome_report_intake_results(
        report_source_csv=seeded_plan,
        intake_csv=edited_intake,
        output=output,
    )

    assert report["approved_intake_rows"] == 2
    assert report["updated_rows"] == 1
    assert report["missing_file_rows"] == 1
    with output.open(encoding="utf-8", newline="") as f:
        merged_rows = list(csv.DictReader(f))
    lnu = next(row for row in merged_rows if row["entity_name"] == "辽宁大学")
    assert lnu["status"] == "ready"
    assert lnu["local_report_path"] == str(local_pdf)
    sut = next(
        row
        for row in merged_rows
        if row["entity_name"] == "沈阳工业大学"
        and row["report_scope"] == "undergraduate_teaching_quality_report"
    )
    assert sut["status"] == "candidate_found"


def test_build_outcome_report_extraction_plan_requires_local_file(tmp_path: Path):
    plan = tmp_path / "outcome_collection_plan.csv"
    rows = [
        _outcome_plan_row("school", "10140", "辽宁大学", "postgrad_rate", status="todo", priority_rank="1"),
    ]
    _write_outcome_plan(plan, rows)
    source_result = build_outcome_report_source_plan(
        plan_csv=plan,
        output_dir=tmp_path / "report_sources",
        domains=["school"],
    )
    with Path(source_result["csv"]).open(encoding="utf-8", newline="") as f:
        source_rows = list(csv.DictReader(f))
    local_pdf = tmp_path / "raw" / "lnu2025.pdf"
    local_pdf.parent.mkdir(parents=True)
    local_pdf.write_bytes(b"%PDF-1.4\n")
    local_ofd = tmp_path / "raw" / "lnu2025.ofd"
    local_ofd.write_bytes(b"ofd fixture")
    html_pdf = tmp_path / "raw" / "captcha.pdf"
    html_pdf.write_text("<!DOCTYPE html><html><body>验证码</body></html>", encoding="utf-8")
    source_rows[0].update({
        "status": "verified",
        "candidate_report_title": "辽宁大学2025届毕业生就业质量报告",
        "candidate_report_url": "https://example.edu/lnu2025.pdf",
        "candidate_source_date": "2025-12-31",
        "availability_date": "2026-01-05",
        "local_report_path": str(local_pdf),
    })
    source_rows[1].update({
        "status": "verified",
        "candidate_report_title": "辽宁大学2025年本科教学质量报告",
        "candidate_report_url": "https://example.edu/lnu_teaching2025.pdf",
        "candidate_source_date": "2025-12-31",
        "availability_date": "2026-01-05",
        "local_report_path": str(local_ofd),
    })
    html_row = {**source_rows[0]}
    html_row.update({
        "report_scope": "undergraduate_teaching_quality_report",
        "candidate_report_title": "辽宁大学2025年本科教学质量报告",
        "candidate_report_url": "https://example.edu/captcha.pdf",
        "local_report_path": str(html_pdf),
    })
    source_rows.append(html_row)
    report_source_csv = tmp_path / "outcome_report_source_plan_verified.csv"
    with report_source_csv.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=source_rows[0].keys())
        writer.writeheader()
        writer.writerows(source_rows)

    result = build_outcome_report_extraction_plan(
        report_source_csv=report_source_csv,
        output_dir=tmp_path / "extract",
    )

    assert result["rows"] == 3
    assert result["ready_rows"] == 2
    assert result["blocked_rows"] == 1
    with Path(result["csv"]).open(encoding="utf-8", newline="") as f:
        extraction_rows = list(csv.DictReader(f))
    assert extraction_rows[0]["extraction_status"] == "ready"
    assert extraction_rows[0]["input_path"] == str(local_pdf)
    assert extraction_rows[0]["output_path"].endswith("school_10140_2025_employment_quality_report_candidates.csv")
    assert extraction_rows[1]["extraction_status"] == "ready"
    assert extraction_rows[1]["block_reason"] == ""
    assert extraction_rows[2]["extraction_status"] == "blocked"
    assert extraction_rows[2]["block_reason"] == "local_report_path_is_html"


def test_build_outcome_report_extraction_plan_accepts_intake_status(tmp_path: Path):
    local_pdf = tmp_path / "raw" / "lnu2025.pdf"
    local_pdf.parent.mkdir(parents=True)
    local_pdf.write_bytes(b"%PDF-1.4\n")
    report_source_csv = tmp_path / "outcome_report_intake_plan_downloaded.csv"
    rows = [{
        "domain": "school",
        "entity_code": "10140",
        "entity_name": "辽宁大学",
        "metric_year": "2025",
        "report_scope": "employment_quality_report",
        "candidate_report_title": "辽宁大学2025届毕业生就业质量报告",
        "candidate_report_url": "https://example.edu/lnu2025.pdf",
        "candidate_source_date": "2025-12-31",
        "availability_date": "2026-01-05",
        "status": "",
        "intake_status": "downloaded",
        "local_report_path": str(local_pdf),
    }]
    with report_source_csv.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)

    result = build_outcome_report_extraction_plan(
        report_source_csv=report_source_csv,
        output_dir=tmp_path / "extract",
        statuses=["downloaded"],
    )

    assert result["rows"] == 1
    assert result["ready_rows"] == 1
    with Path(result["csv"]).open(encoding="utf-8", newline="") as f:
        extraction_rows = list(csv.DictReader(f))
    assert extraction_rows[0]["extraction_status"] == "ready"
    assert extraction_rows[0]["input_path"] == str(local_pdf)


def test_build_outcome_report_extraction_plan_defaults_to_downloaded_intake_status(tmp_path: Path):
    local_pdf = tmp_path / "raw" / "lnu2025.pdf"
    local_pdf.parent.mkdir(parents=True)
    local_pdf.write_bytes(b"%PDF-1.4\n")
    report_source_csv = tmp_path / "outcome_report_intake_results.csv"
    rows = [{
        "domain": "school",
        "entity_code": "10140",
        "entity_name": "辽宁大学",
        "metric_year": "2025",
        "report_scope": "employment_quality_report",
        "candidate_report_title": "辽宁大学2025届毕业生就业质量报告",
        "candidate_report_url": "https://example.edu/lnu2025.pdf",
        "candidate_source_date": "2025-12-31",
        "availability_date": "2026-01-05",
        "source_status": "candidate_found",
        "intake_status": "downloaded",
        "download_status": "downloaded",
        "local_report_path": str(local_pdf),
    }]
    with report_source_csv.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)

    result = build_outcome_report_extraction_plan(
        report_source_csv=report_source_csv,
        output_dir=tmp_path / "extract",
    )

    assert result["rows"] == 1
    assert result["ready_rows"] == 1


def test_build_outcome_report_extraction_plan_accepts_docx(tmp_path: Path):
    local_docx = tmp_path / "raw" / "fuxin.docx"
    local_docx.parent.mkdir(parents=True)
    with ZipFile(local_docx, "w") as zf:
        zf.writestr("word/document.xml", "<w:document xmlns:w=\"http://schemas.openxmlformats.org/wordprocessingml/2006/main\"/>")
    report_source_csv = tmp_path / "outcome_report_intake_results.csv"
    rows = [{
        "domain": "school",
        "entity_code": "1250",
        "entity_name": "阜新高等专科学校",
        "metric_year": "2024",
        "report_scope": "employment_quality_report",
        "candidate_report_title": "阜新高等专科学校2024年大专毕业生就业质量报告",
        "candidate_report_url": "https://www.fxgz.com.cn/showart/8804.html",
        "candidate_source_date": "2025-12-06",
        "availability_date": "2025-12-06",
        "intake_status": "downloaded",
        "local_report_path": str(local_docx),
    }]
    with report_source_csv.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)

    result = build_outcome_report_extraction_plan(
        report_source_csv=report_source_csv,
        output_dir=tmp_path / "extract",
    )

    assert result["rows"] == 1
    assert result["ready_rows"] == 1


def test_run_outcome_report_extraction_plan_writes_candidate_csv(tmp_path: Path, monkeypatch):
    input_pdf = tmp_path / "raw" / "lnu2025.pdf"
    input_pdf.parent.mkdir(parents=True)
    input_pdf.write_bytes(b"%PDF-1.4\n")
    output_csv = tmp_path / "candidates" / "lnu2025_candidates.csv"
    blocked_pdf = tmp_path / "raw" / "blocked.pdf"
    plan = tmp_path / "outcome_report_extraction_plan.csv"
    with plan.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "domain",
                "entity_code",
                "entity_name",
                "metric_year",
                "report_scope",
                "source_title",
                "source_url",
                "source_date",
                "availability_date",
                "input_path",
                "output_path",
                "planned_metric_keys",
                "extraction_status",
                "block_reason",
                "notes",
            ],
        )
        writer.writeheader()
        writer.writerow({
            "domain": "school",
            "entity_code": "10140",
            "entity_name": "辽宁大学",
            "metric_year": "2025",
            "report_scope": "employment_quality_report",
            "source_title": "辽宁大学2025届毕业生就业质量报告",
            "source_url": "https://example.edu/lnu2025.pdf",
            "source_date": "2025-12-31",
            "availability_date": "2026-01-05",
            "input_path": str(input_pdf),
            "output_path": str(output_csv),
            "planned_metric_keys": '["employment_rate"]',
            "extraction_status": "ready",
            "block_reason": "",
            "notes": "",
        })
        writer.writerow({
            "domain": "school",
            "entity_code": "10140",
            "entity_name": "辽宁大学",
            "metric_year": "2025",
            "report_scope": "undergraduate_teaching_quality_report",
            "source_title": "辽宁大学2025年本科教学质量报告",
            "source_url": "https://example.edu/lnu_teaching2025.pdf",
            "source_date": "2025-12-31",
            "availability_date": "2026-01-05",
            "input_path": str(blocked_pdf),
            "output_path": str(tmp_path / "candidates" / "blocked.csv"),
            "planned_metric_keys": '["postgrad_rate"]',
            "extraction_status": "blocked",
            "block_reason": "local_report_path_not_found",
            "notes": "",
        })

    def fake_extract(path: Path, **kwargs):
        assert path == input_pdf
        return [{
            "domain": kwargs["domain"],
            "entity_code": kwargs["entity_code"],
            "entity_name": kwargs["entity_name"],
            "metric_key": "employment_rate",
            "metric_label": "毕业去向落实率",
            "metric_unit": "ratio",
            "metric_year": kwargs["metric_year"],
            "candidate_value": "0.9236",
            "candidate_text_value": "92.36%",
            "source_title": kwargs["source_title"],
            "source_url": kwargs["source_url"],
            "evidence_quote": "毕业去向落实率为 92.36%",
            "metric_scope": "",
            "source_date": kwargs["source_date"],
            "availability_date": kwargs["availability_date"],
            "page_number": "1",
            "match_alias": "毕业去向落实率",
            "confidence": "medium",
            "review_status": "needs_review",
            "notes": "fixture",
        }]

    monkeypatch.setattr(
        "datahub.builders.outcome_report_extraction_runner.extract_outcome_metric_candidates_from_report",
        fake_extract,
    )
    report = run_outcome_report_extraction_plan(plan_csv=plan, report_path=tmp_path / "extract_report.json")

    assert report["errors"] == []
    assert report["ready_rows"] == 1
    assert report["skipped_rows"] == 1
    assert report["candidate_rows"] == 1
    assert output_csv.exists()
    with output_csv.open(encoding="utf-8", newline="") as f:
        candidates = list(csv.DictReader(f))
    assert candidates[0]["metric_key"] == "employment_rate"
    assert (tmp_path / "extract_report.json").exists()


def test_audit_outcome_collection_plan_reports_progress_and_errors(tmp_path: Path):
    plan = tmp_path / "outcome_collection_plan.csv"
    with plan.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "domain",
                "entity_code",
                "entity_name",
                "priority_rank",
                "plan_rows",
                "metric_key",
                "metric_label",
                "metric_unit",
                "metric_year",
                "search_queries",
                "status",
                "metric_value",
                "source_title",
                "source_url",
                "evidence_quote",
                "metric_scope",
                "denominator",
                "notes",
            ],
        )
        writer.writeheader()
        writer.writerow({
            "domain": "school",
            "entity_code": "10145",
            "entity_name": "东北大学",
            "priority_rank": "1",
            "plan_rows": "120",
            "metric_key": "postgrad_rate",
            "metric_label": "深造率",
            "metric_unit": "ratio",
            "metric_year": "2025",
            "search_queries": json.dumps(["东北大学 2025 就业质量报告"], ensure_ascii=False),
            "status": "verified",
            "metric_value": "46.2%",
            "source_title": "2025届毕业生就业质量报告",
            "source_url": "https://example.edu/report.pdf",
            "evidence_quote": "本科毕业生深造率为46.2%。",
            "metric_scope": "本科毕业生",
            "denominator": "",
            "notes": "",
        })
        writer.writerow({
            "domain": "school",
            "entity_code": "10145",
            "entity_name": "东北大学",
            "priority_rank": "1",
            "plan_rows": "120",
            "metric_key": "made_up_metric",
            "metric_label": "未知指标",
            "metric_unit": "ratio",
            "metric_year": "2025",
            "search_queries": "not-json",
            "status": "verified",
            "metric_value": "1",
            "source_title": "",
            "source_url": "",
            "evidence_quote": "",
            "metric_scope": "",
            "denominator": "",
            "notes": "",
        })
        writer.writerow({
            "domain": "school",
            "entity_code": "10145",
            "entity_name": "东北大学",
            "priority_rank": "1",
            "plan_rows": "120",
            "metric_key": "employment_rate",
            "metric_label": "毕业去向落实率",
            "metric_unit": "ratio",
            "metric_year": "2025",
            "search_queries": json.dumps(["东北大学 2025 就业质量报告"], ensure_ascii=False),
            "status": "typo_done",
            "metric_value": "",
            "source_title": "",
            "source_url": "",
            "evidence_quote": "",
            "metric_scope": "",
            "denominator": "",
            "notes": "",
        })

    report = audit_outcome_collection_plan(plan)

    assert report["rows"] == 3
    assert report["progress"]["complete_rows"] == 2
    assert report["evidence_counts"]["rows_with_source_url"] == 1
    assert any("unregistered outcome metric" in error for error in report["errors"])
    assert any("search_queries is not valid JSON" in error for error in report["errors"])
    assert any("complete status missing evidence" in error for error in report["errors"])
    assert any("unknown collection status" in error for error in report["errors"])


def test_build_outcome_collection_batch_limits_pending_rows(tmp_path: Path):
    plan = tmp_path / "outcome_collection_plan.csv"
    rows = [
        _outcome_plan_row("school", "10145", "东北大学", "postgrad_rate", status="todo", priority_rank="1"),
        _outcome_plan_row("school", "10145", "东北大学", "employment_rate", status="in_progress", priority_rank="1"),
        _outcome_plan_row("school", "10140", "辽宁大学", "postgrad_rate", status="verified", priority_rank="2"),
        _outcome_plan_row("major", "计算机类", "计算机类", "employment_rate", status="todo", priority_rank="1"),
        _outcome_plan_row("major", "自动化", "自动化", "employment_rate", status="blocked", priority_rank="2"),
    ]
    _write_outcome_plan(plan, rows)

    result = build_outcome_collection_batch(
        plan_csv=plan,
        output_dir=tmp_path / "batch",
        limit_per_domain=1,
    )

    assert result["rows"] == 2
    assert result["domain_counts"] == {"major": 1, "school": 1}
    with Path(result["csv"]).open(encoding="utf-8", newline="") as f:
        batch_rows = list(csv.DictReader(f))
    assert {row["domain"] for row in batch_rows} == {"school", "major"}
    assert {row["status"] for row in batch_rows} <= {"todo", "in_progress"}
    manifest = json.loads(Path(result["manifest"]).read_text(encoding="utf-8"))
    assert manifest["task_key_columns"] == ["domain", "entity_code", "metric_key", "metric_year"]
    assert "source_url" in manifest["editable_columns"]


def test_apply_outcome_collection_review_seeds_updates_matching_rows(tmp_path: Path):
    seed_config = load_outcome_collection_review_seeds()
    seed_count = len(seed_config["seeds"])
    verified_seed_count = sum(1 for seed in seed_config["seeds"] if seed["status"] == "verified")
    audit = audit_outcome_collection_review_seeds()
    assert audit["errors"] == []
    assert audit["seed_count"] == seed_count
    assert audit["status_counts"] == {"verified": verified_seed_count}

    plan = tmp_path / "outcome_collection_plan.csv"
    seeded = _outcome_plan_row("school", "0140", "辽宁大学", "employment_rate", status="todo", priority_rank="1")
    seeded["metric_year"] = "2024"
    dlpu = _outcome_plan_row("school", "0152", "大连工业大学", "postgrad_rate", status="todo", priority_rank="2")
    dlpu["metric_year"] = "2024"
    bohai = _outcome_plan_row("school", "0167", "渤海大学", "employment_rate", status="todo", priority_rank="3")
    bohai["metric_year"] = "2024"
    dufe = _outcome_plan_row("school", "0173", "东北财经大学", "employment_rate", status="todo", priority_rank="4")
    dufe["metric_year"] = "2024"
    dlu = _outcome_plan_row("school", "1258", "大连大学", "employment_rate", status="todo", priority_rank="5")
    dlu["metric_year"] = "2023"
    lnu_soe = _outcome_plan_row("school", "0140", "辽宁大学", "civil_service_rate", status="todo", priority_rank="6")
    lnu_soe["metric_year"] = "2022"
    lnnu_employment = _outcome_plan_row("school", "0165", "辽宁师范大学", "employment_rate", status="todo", priority_rank="7")
    lnnu_employment["metric_year"] = "2024"
    lnnu_postgrad = _outcome_plan_row("school", "0165", "辽宁师范大学", "postgrad_rate", status="todo", priority_rank="8")
    lnnu_postgrad["metric_year"] = "2024"
    jlu_employment = _outcome_plan_row("school", "0183", "吉林大学", "employment_rate", status="todo", priority_rank="9")
    jlu_employment["metric_year"] = "2024"
    jlu_postgrad = _outcome_plan_row("school", "0183", "吉林大学", "postgrad_rate", status="todo", priority_rank="10")
    jlu_postgrad["metric_year"] = "2024"
    lntu_employment = _outcome_plan_row("school", "0147", "辽宁工程技术大学", "employment_rate", status="todo", priority_rank="11")
    lntu_employment["metric_year"] = "2024"
    lntu_postgrad = _outcome_plan_row("school", "0147", "辽宁工程技术大学", "postgrad_rate", status="todo", priority_rank="12")
    lntu_postgrad["metric_year"] = "2024"
    sut_employment = _outcome_plan_row("school", "0142", "沈阳工业大学", "employment_rate", status="todo", priority_rank="13")
    sut_employment["metric_year"] = "2024"
    djtu_postgrad = _outcome_plan_row("school", "0150", "大连交通大学", "postgrad_rate", status="todo", priority_rank="14")
    djtu_postgrad["metric_year"] = "2024"
    pending = _outcome_plan_row("school", "0166", "沈阳师范大学", "employment_rate", status="todo", priority_rank="2")
    pending["metric_year"] = "2024"
    _write_outcome_plan(
        plan,
        [
            seeded,
            dlpu,
            bohai,
            dufe,
            dlu,
            lnu_soe,
            lnnu_employment,
            lnnu_postgrad,
            jlu_employment,
            jlu_postgrad,
            lntu_employment,
            lntu_postgrad,
            sut_employment,
            djtu_postgrad,
            pending,
        ],
    )

    output = tmp_path / "outcome_collection_plan_seeded.csv"
    report = apply_outcome_collection_review_seeds(plan_csv=plan, output=output)
    expected_matching_seeds = 14
    assert report["matched_rows"] == expected_matching_seeds
    assert report["updated_rows"] == expected_matching_seeds
    assert report["unmatched_seeds"] == seed_count - expected_matching_seeds

    with output.open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    by_entity = {(row["entity_code"], row["metric_key"]): row for row in rows}
    lnu = by_entity[("0140", "employment_rate")]
    assert lnu["status"] == "verified"
    assert lnu["metric_value"] == "0.8582"
    assert lnu["metric_scope"] == "2024届本科应届毕业生，毕业去向落实率"
    assert lnu["source_url"] == "https://xxgk.lnu.edu.cn/info/13534/68147.htm"
    assert "seed_review=" in lnu["notes"]
    dlpu_postgrad = by_entity[("0152", "postgrad_rate")]
    assert dlpu_postgrad["status"] == "verified"
    assert dlpu_postgrad["metric_value"] == "0.2547"
    assert "不是保研率" in dlpu_postgrad["metric_scope"]
    bohai_employment = by_entity[("0167", "employment_rate")]
    assert bohai_employment["status"] == "verified"
    assert bohai_employment["metric_value"] == "0.8754"
    dufe_employment = by_entity[("0173", "employment_rate")]
    assert dufe_employment["status"] == "verified"
    assert dufe_employment["metric_value"] == "0.8864"
    dlu_employment = by_entity[("1258", "employment_rate")]
    assert dlu_employment["status"] == "verified"
    assert dlu_employment["metric_value"] == "0.9031"
    assert dlu_employment["metric_year"] == "2023"
    lnu_soe = by_entity[("0140", "civil_service_rate")]
    assert lnu_soe["status"] == "verified"
    assert lnu_soe["metric_value"] == "0.2892"
    assert lnu_soe["metric_year"] == "2022"
    lnnu_employment = by_entity[("0165", "employment_rate")]
    assert lnnu_employment["status"] == "verified"
    assert lnnu_employment["metric_value"] == "0.9104"
    assert lnnu_employment["metric_year"] == "2024"
    lnnu_postgrad = by_entity[("0165", "postgrad_rate")]
    assert lnnu_postgrad["status"] == "verified"
    assert lnnu_postgrad["metric_value"] == "0.2557"
    assert lnnu_postgrad["metric_year"] == "2024"
    jlu_employment = by_entity[("0183", "employment_rate")]
    assert jlu_employment["status"] == "verified"
    assert jlu_employment["metric_value"] == "0.852"
    assert jlu_employment["metric_year"] == "2024"
    assert "8月27日" in jlu_employment["metric_scope"]
    jlu_postgrad = by_entity[("0183", "postgrad_rate")]
    assert jlu_postgrad["status"] == "verified"
    assert jlu_postgrad["metric_value"] == "0.4846"
    assert jlu_postgrad["metric_year"] == "2024"
    assert "42.96%" in jlu_postgrad["metric_scope"]
    lntu_employment = by_entity[("0147", "employment_rate")]
    assert lntu_employment["status"] == "verified"
    assert lntu_employment["metric_value"] == "0.8155"
    assert lntu_employment["metric_year"] == "2024"
    lntu_postgrad = by_entity[("0147", "postgrad_rate")]
    assert lntu_postgrad["status"] == "verified"
    assert lntu_postgrad["metric_value"] == "0.2419"
    assert lntu_postgrad["metric_year"] == "2024"
    sut_employment = by_entity[("0142", "employment_rate")]
    assert sut_employment["status"] == "verified"
    assert sut_employment["metric_value"] == "0.87"
    assert sut_employment["metric_year"] == "2024"
    djtu_postgrad = by_entity[("0150", "postgrad_rate")]
    assert djtu_postgrad["status"] == "verified"
    assert djtu_postgrad["metric_value"] == "0.2887"
    assert djtu_postgrad["metric_year"] == "2024"
    assert by_entity[("0166", "employment_rate")]["status"] == "todo"


def test_outcome_collection_source_evidence_policy_separates_metrics_from_recruitment_news():
    config = load_outcome_collection()
    policy = config["source_evidence_policy"]
    tiers = policy["tiers"]

    assert "直接给出本科毕业生统计口径" in policy["metric_publication_rule"]
    assert tiers["school_official_report"]["can_publish_outcome_metric"] is True
    assert "fa_fact_school_outcome" in tiers["school_official_report"]["allowed_evidence_targets"]
    assert tiers["school_official_news"]["can_publish_outcome_metric"] is True
    assert "普通招聘活动新闻只作为学校城市产业连接证据" in tiers["school_official_news"]["required_review"]
    assert tiers["government_talent_market"]["can_publish_outcome_metric"] is False
    assert "fa_mart_school_city_industry_fit" in tiers["government_talent_market"]["allowed_evidence_targets"]
    assert "不得把活动场次、参会企业数或岗位数当成毕业去向落实率" in tiers["government_talent_market"]["required_review"]
    assert tiers["local_news_recruitment_fair"]["can_publish_outcome_metric"] is False
    assert "必须回查学校、人社局或人才市场官方页面" in tiers["local_news_recruitment_fair"]["required_review"]
    assert tiers["self_media"]["can_publish_outcome_metric"] is False
    assert tiers["self_media"]["allowed_evidence_targets"] == ["research_candidate"]


def test_audit_outcome_collection_review_seeds_rejects_metric_value_out_of_range(monkeypatch: pytest.MonkeyPatch):
    seed = {
        "seed_id": "bad_outcome_metric_value",
        "domain": "school",
        "entity_code": "0001",
        "entity_name": "测试大学",
        "metric_key": "employment_rate",
        "metric_year": 2024,
        "status": "verified",
        "metric_value": 1.2,
        "source_title": "测试报告",
        "source_url": "https://example.edu/report.pdf",
        "evidence_quote": "毕业去向落实率为 120%。",
        "metric_scope": "测试口径",
        "source_date": "2024-12-31",
        "availability_date": "2024-12-31",
        "reviewer": "codex",
        "reviewed_at": "2026-05-14",
        "review_note": "测试越界值。",
    }
    monkeypatch.setattr(
        "datahub.builders.outcome_collection_seed_merge.load_outcome_collection_review_seeds",
        lambda: {"seeds": [seed]},
    )

    audit = audit_outcome_collection_review_seeds()

    assert audit["errors"] == ["seed 1 metric_value is above max_value 1: 1.2"]


def test_audit_outcome_collection_review_seeds_rejects_bad_year_and_dates(monkeypatch: pytest.MonkeyPatch):
    seed = {
        "seed_id": "bad_outcome_seed_dates",
        "domain": "school",
        "entity_code": "0001",
        "entity_name": "测试大学",
        "metric_key": "employment_rate",
        "metric_year": "2024.0",
        "status": "verified",
        "metric_value": 0.9,
        "source_title": "测试报告",
        "source_url": "https://example.edu/report.pdf",
        "evidence_quote": "毕业去向落实率为 90%。",
        "metric_scope": "测试口径",
        "source_date": "2024/12/31",
        "availability_date": "2024-12-31",
        "reviewer": "codex",
        "reviewed_at": "2026-14-05",
        "review_note": "测试年份和日期格式。",
    }
    monkeypatch.setattr(
        "datahub.builders.outcome_collection_seed_merge.load_outcome_collection_review_seeds",
        lambda: {"seeds": [seed]},
    )

    audit = audit_outcome_collection_review_seeds()

    assert audit["errors"] == [
        "seed 1 metric_year is not an integer",
        "seed 1 source_date must use YYYY-MM-DD",
        "seed 1 reviewed_at must use YYYY-MM-DD",
    ]


def test_audit_outcome_collection_review_seeds_rejects_non_http_source_url(monkeypatch: pytest.MonkeyPatch):
    seed = {
        "seed_id": "bad_outcome_source_url",
        "domain": "school",
        "entity_code": "0001",
        "entity_name": "测试大学",
        "metric_key": "employment_rate",
        "metric_year": 2024,
        "status": "verified",
        "metric_value": 0.9,
        "source_title": "测试报告",
        "source_url": "raw/outcome/report.pdf",
        "evidence_quote": "毕业去向落实率为 90%。",
        "metric_scope": "测试口径",
        "source_date": "2024-12-31",
        "availability_date": "2024-12-31",
        "reviewer": "codex",
        "reviewed_at": "2026-05-14",
        "review_note": "测试来源链接格式。",
    }
    monkeypatch.setattr(
        "datahub.builders.outcome_collection_seed_merge.load_outcome_collection_review_seeds",
        lambda: {"seeds": [seed]},
    )

    audit = audit_outcome_collection_review_seeds()

    assert audit["errors"] == ["seed 1 source_url must be an http(s) URL"]


def test_audit_outcome_collection_review_seeds_rejects_bad_date_order(monkeypatch: pytest.MonkeyPatch):
    seed = {
        "seed_id": "bad_outcome_date_order",
        "domain": "school",
        "entity_code": "0001",
        "entity_name": "测试大学",
        "metric_key": "employment_rate",
        "metric_year": 2024,
        "status": "verified",
        "metric_value": 0.9,
        "source_title": "测试报告",
        "source_url": "https://example.edu/report.pdf",
        "evidence_quote": "毕业去向落实率为 90%。",
        "metric_scope": "测试口径",
        "source_date": "2024-12-31",
        "availability_date": "2024-01-01",
        "reviewer": "codex",
        "reviewed_at": "2023-12-31",
        "review_note": "测试时间顺序。",
    }
    monkeypatch.setattr(
        "datahub.builders.outcome_collection_seed_merge.load_outcome_collection_review_seeds",
        lambda: {"seeds": [seed]},
    )

    audit = audit_outcome_collection_review_seeds()

    assert audit["errors"] == [
        "seed 1 source_date must not be after availability_date",
        "seed 1 reviewed_at must not be before availability_date",
    ]


def test_merge_outcome_collection_batch_updates_only_editable_columns(tmp_path: Path):
    plan = tmp_path / "outcome_collection_plan.csv"
    rows = [
        _outcome_plan_row("school", "10145", "东北大学", "postgrad_rate", status="todo", priority_rank="1"),
        _outcome_plan_row("major", "计算机类", "计算机类", "employment_rate", status="todo", priority_rank="2"),
    ]
    _write_outcome_plan(plan, rows)
    batch = tmp_path / "outcome_collection_batch.csv"
    edited_row = {**rows[0]}
    edited_row.update({
        "entity_name": "被篡改的学校名",
        "status": "verified",
        "metric_value": "46.2%",
        "source_title": "2025届毕业生就业质量报告",
        "source_url": "https://example.edu/report.pdf",
        "evidence_quote": "本科毕业生深造率为46.2%。",
        "metric_scope": "本科毕业生",
        "denominator": "1000",
        "source_date": "2025-12-31",
        "availability_date": "2026-01-05",
        "built_at": "2026-05-13T00:00:00",
        "notes": "人工核验",
    })
    _write_outcome_plan(batch, [edited_row])

    output = tmp_path / "outcome_collection_plan_merged.csv"
    report = merge_outcome_collection_batch(
        plan_csv=plan,
        batch_csv=batch,
        output=output,
    )

    assert report["updated_rows"] == 1
    assert report["status_counts"] == {"todo": 1, "verified": 1}
    with output.open(encoding="utf-8", newline="") as f:
        merged_rows = list(csv.DictReader(f))
    assert merged_rows[0]["entity_name"] == "东北大学"
    assert merged_rows[0]["status"] == "verified"
    assert merged_rows[0]["source_url"] == "https://example.edu/report.pdf"
    assert merged_rows[0]["notes"] == "人工核验"
    assert merged_rows[1]["entity_name"] == "计算机类"
    assert merged_rows[1]["status"] == "todo"


def test_merge_outcome_report_candidates_requires_approved_status(tmp_path: Path):
    plan = tmp_path / "outcome_collection_plan.csv"
    rows = [
        _outcome_plan_row("school", "10140", "辽宁大学", "civil_service_rate", status="todo", priority_rank="1"),
        _outcome_plan_row("school", "10142", "沈阳工业大学", "employment_rate", status="todo", priority_rank="2"),
    ]
    _write_outcome_plan(plan, rows)
    candidates = tmp_path / "outcome_candidates.csv"
    with candidates.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=CANDIDATE_COLUMNS)
        writer.writeheader()
        base = {column: "" for column in CANDIDATE_COLUMNS}
        writer.writerow({
            **base,
            "domain": "school",
            "entity_code": "10140",
            "entity_name": "被篡改的名称",
            "metric_key": "civil_service_rate",
            "metric_label": "体制内去向比例",
            "metric_unit": "ratio",
            "metric_year": "2025",
            "candidate_value": "0.2594",
            "candidate_text_value": "25.94%",
            "source_title": "辽宁大学2022届毕业生就业质量年度报告",
            "source_url": "https://www.lnu.edu.cn/info/15026/78891.htm",
            "evidence_quote": "国有企业占比为25.94%。",
            "metric_scope": "本科毕业生签约单位性质",
            "source_date": "2022-12-31",
            "availability_date": "2022-12-31",
            "page_number": "38",
            "match_alias": "国有企业",
            "confidence": "low",
            "review_status": "approved",
        })
        writer.writerow({
            **base,
            "domain": "school",
            "entity_code": "10142",
            "entity_name": "沈阳工业大学",
            "metric_key": "employment_rate",
            "metric_label": "毕业去向落实率",
            "metric_unit": "ratio",
            "metric_year": "2025",
            "candidate_value": "1",
            "source_title": "2023-2024年沈阳工业大学本科教学质量报告",
            "source_url": "https://www.sut.edu.cn/info/1584/67026.htm",
            "evidence_quote": "学位点就业率100%。",
            "source_date": "2024-12-31",
            "availability_date": "2024-12-31",
            "review_status": "needs_review",
        })

    output = tmp_path / "outcome_collection_plan_merged.csv"
    report = merge_outcome_report_candidates(
        plan_csv=plan,
        candidate_csv=candidates,
        output=output,
    )

    assert report["updated_rows"] == 1
    assert report["approved_candidate_rows"] == 1
    assert report["status_counts"] == {"todo": 1, "verified": 1}
    with output.open(encoding="utf-8", newline="") as f:
        merged_rows = list(csv.DictReader(f))
    assert merged_rows[0]["entity_name"] == "辽宁大学"
    assert merged_rows[0]["status"] == "verified"
    assert merged_rows[0]["metric_value"] == "0.2594"
    assert merged_rows[0]["source_url"] == "https://www.lnu.edu.cn/info/15026/78891.htm"
    assert "merged_from_report_candidate" in merged_rows[0]["notes"]
    assert merged_rows[1]["status"] == "todo"
    assert merged_rows[1]["metric_value"] == ""


def test_build_operational_gap_report_summarizes_existing_artifacts(tmp_path: Path):
    coverage = tmp_path / "coverage.json"
    coverage.write_text(json.dumps({
        "total_school_count": 2,
        "p0_blockers": [{"code": "SCHOOL_OUTCOME_NOT_OPERATIONAL"}],
        "coverage_areas": [{
            "key": "outcome",
            "label": "School outcome",
            "covered_school_count": 1,
            "total_school_count": 2,
            "coverage_rate": 0.5,
            "status": "below_threshold",
        }],
    }), encoding="utf-8")
    portfolio = tmp_path / "portfolio.json"
    portfolio.write_text(json.dumps({
        "category_counts": {"required_unavailable": 1},
        "p0_blockers": [{"code": "LN_SCORE_HISTORY_NOT_OPERATIONAL"}],
    }), encoding="utf-8")
    outcome = tmp_path / "outcome.json"
    outcome.write_text(json.dumps({
        "rows": 800,
        "progress": {"complete_rows": 12, "pending_rows": 788, "blocked_rows": 0},
        "errors": [],
        "warnings": [],
    }), encoding="utf-8")
    amap = tmp_path / "amap.json"
    amap.write_text(json.dumps({
        "ready_for_fetch": False,
        "key_present": False,
        "row_counts": {"input_rows": 1424, "requestable_rows": 1424},
        "errors": ["Amap Web API key missing; set AMAP_WEB_SERVICE_KEY"],
    }), encoding="utf-8")
    score = tmp_path / "score.json"
    score.write_text(json.dumps({
        "progress": {"ready_rows": 3391, "pending_rows": 21087},
        "status_counts": {"reviewed": 3391, "todo": 21087},
    }), encoding="utf-8")
    readiness = tmp_path / "campus_living.json"
    readiness.write_text(json.dumps({
        "ready_for_build": False,
        "location_rows": 0,
        "poi_rows": 0,
        "errors": ["location_input_missing", "poi_input_missing"],
    }), encoding="utf-8")

    report = build_operational_gap_report(
        coverage_report_path=coverage,
        portfolio_report_path=portfolio,
        outcome_audit_path=outcome,
        amap_readiness_path=amap,
        score_readiness_paths={"2023_2024": score},
        readiness_paths={"campus_living": readiness},
        report_path=tmp_path / "gap.json",
        markdown_path=tmp_path / "gap.md",
    )

    assert report["summary"]["ready_for_normal_operation"] is False
    assert report["summary"]["p0_blocker_signal_count"] == 6
    assert report["summary"]["unique_p0_blocker_count"] == 6
    assert report["readiness"]["campus_living"]["row_counts"] == {"location_rows": 0, "poi_rows": 0}
    assert (tmp_path / "gap.json").exists()
    markdown = (tmp_path / "gap.md").read_text(encoding="utf-8")
    assert "outcome_pending_rows" in markdown
    assert "score_reconciliation_pending_rows" in markdown
    assert "campus_living_not_ready_for_build" in markdown


def test_merge_outcome_report_candidates_requires_metric_scope(tmp_path: Path):
    plan = tmp_path / "outcome_collection_plan.csv"
    rows = [
        _outcome_plan_row("school", "10140", "辽宁大学", "employment_rate", status="todo", priority_rank="1"),
    ]
    _write_outcome_plan(plan, rows)
    candidates = tmp_path / "outcome_candidates.csv"
    with candidates.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=CANDIDATE_COLUMNS)
        writer.writeheader()
        base = {column: "" for column in CANDIDATE_COLUMNS}
        writer.writerow({
            **base,
            "domain": "school",
            "entity_code": "10140",
            "entity_name": "辽宁大学",
            "metric_key": "employment_rate",
            "metric_label": "毕业去向落实率",
            "metric_unit": "ratio",
            "metric_year": "2025",
            "candidate_value": "0.8582",
            "source_title": "辽宁大学2023-2024学年本科教学质量报告",
            "source_url": "https://example.edu/lnu.pdf",
            "evidence_quote": "毕业去向落实率为85.82%。",
            "source_date": "2024-12-06",
            "availability_date": "2024-12-06",
            "review_status": "approved",
        })

    output = tmp_path / "outcome_collection_plan_merged.csv"
    with pytest.raises(ValueError, match="incomplete approved candidate rows"):
        merge_outcome_report_candidates(
            plan_csv=plan,
            candidate_csv=candidates,
            output=output,
        )


def test_export_approved_scoped_stock_review_candidates_requires_manual_approval(tmp_path: Path):
    batch = tmp_path / "scoped_stock_review_batch.csv"
    with batch.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=[*CANDIDATE_COLUMNS, "scoped_review_class", "matched_scope_terms", "recommended_action"])
        writer.writeheader()
        base = {column: "" for column in writer.fieldnames}
        writer.writerow({
            **base,
            "domain": "school",
            "entity_code": "0162",
            "entity_name": "辽宁中医药大学",
            "metric_key": "employment_rate",
            "metric_label": "毕业去向落实率",
            "metric_unit": "ratio",
            "metric_year": "2024",
            "candidate_value": "0.85",
            "candidate_text_value": "85%以上",
            "source_title": "辽宁中医药大学2023-2024学年本科教学质量报告",
            "source_url": "https://example.edu/lnutcm.pdf",
            "evidence_quote": "辽宁省内毕业去向落实率85%以上。",
            "metric_scope": "liaoning_province_in_province_destination_rate_not_school_overall",
            "source_date": "2024-12-31",
            "availability_date": "2024-12-31",
            "review_status": "approved",
            "notes": "非学校总体就业率、非本科总体就业率。",
            "scoped_review_class": "scoped_official_candidate",
        })
        writer.writerow({
            **base,
            "domain": "school",
            "entity_code": "0163",
            "entity_name": "待复核学校",
            "metric_key": "employment_rate",
            "metric_year": "2024",
            "candidate_value": "0.90",
            "review_status": "needs_review",
            "scoped_review_class": "scoped_official_candidate",
        })

    output = tmp_path / "approved_candidates.csv"
    report = export_approved_scoped_stock_review_candidates(
        batch_csv=batch,
        output=output,
        report_path=tmp_path / "approved_candidates.json",
    )

    assert report["approved_rows"] == 1
    with output.open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    assert list(rows[0]) == CANDIDATE_COLUMNS
    assert len(rows) == 1
    assert rows[0]["entity_code"] == "0162"
    assert rows[0]["review_status"] == "approved"
    assert rows[0]["metric_scope"] == "liaoning_province_in_province_destination_rate_not_school_overall"


def test_build_scoped_outcome_stock_review_workspace_outputs_review_files(tmp_path: Path):
    batch = tmp_path / "scoped_stock_review_batch.csv"
    with batch.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=[*CANDIDATE_COLUMNS, "scoped_review_class", "matched_scope_terms", "recommended_action"])
        writer.writeheader()
        base = {column: "" for column in writer.fieldnames}
        writer.writerow({
            **base,
            "candidate_file": "staging/v1/extraction_merged/candidates/school.csv",
            "domain": "school",
            "entity_code": "0140",
            "entity_name": "辽宁大学",
            "metric_key": "employment_rate",
            "metric_year": "2024",
            "candidate_value": "0.8582",
            "candidate_text_value": "85.82%",
            "source_title": "辽宁大学本科教学质量报告",
            "source_url": "https://example.edu/lnu.htm",
            "evidence_quote": "本科应届毕业生毕业去向落实率为85.82%。",
            "review_status": "needs_review",
            "recommended_action": "review_scope_before_approval",
        })

    report = build_scoped_outcome_stock_review_workspace(
        batch_csv=batch,
        output_dir=tmp_path / "workspace",
    )

    assert report["rows"] == 1
    assert (tmp_path / "workspace" / "review.csv").exists()
    markdown = (tmp_path / "workspace" / "review.md").read_text(encoding="utf-8")
    assert "辽宁大学" in markdown
    assert "本科应届毕业生毕业去向落实率" in markdown
    assert "review_status" in markdown


def test_audit_scoped_outcome_stock_review_workspace_blocks_incomplete_approved_rows(tmp_path: Path):
    review = tmp_path / "review.csv"
    with review.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=[*CANDIDATE_COLUMNS, "scoped_review_class", "matched_scope_terms", "recommended_action"])
        writer.writeheader()
        base = {column: "" for column in writer.fieldnames}
        writer.writerow({
            **base,
            "domain": "school",
            "entity_code": "0140",
            "entity_name": "辽宁大学",
            "metric_key": "employment_rate",
            "metric_year": "2024",
            "candidate_value": "0.8582",
            "source_title": "辽宁大学本科教学质量报告",
            "source_url": "https://example.edu/lnu.htm",
            "evidence_quote": "本科应届毕业生毕业去向落实率为85.82%。",
            "source_date": "2024-12-31",
            "availability_date": "2024-12-31",
            "review_status": "approved",
        })

    report = audit_scoped_outcome_stock_review_workspace(
        review_csv=review,
        report_path=tmp_path / "audit.json",
    )

    assert report["ready_for_export"] is False
    assert report["approved_rows"] == 1
    assert "metric_scope" in report["errors"][0]
    assert (tmp_path / "audit.json").exists()


def test_audit_outcome_collection_core_coverage_detects_missing_admission_school(tmp_path: Path):
    db = tmp_path / "core.duckdb"
    con = duckdb.connect(str(db))
    con.execute("CREATE TABLE fa_dim_ln_admission_plan (school_code VARCHAR, school_name VARCHAR)")
    con.execute("INSERT INTO fa_dim_ln_admission_plan VALUES ('1001', '学校A'), ('1002', '学校B')")
    con.close()

    plan = tmp_path / "outcome_collection_plan.csv"
    rows = [
        _outcome_plan_row("school", "1001", "学校A", "employment_rate", status="todo", priority_rank="1"),
        _outcome_plan_row("school", "1001", "学校A", "postgrad_rate", status="todo", priority_rank="1"),
        _outcome_plan_row("school", "1001", "学校A", "keep_research_rate", status="todo", priority_rank="1"),
        _outcome_plan_row("school", "1001", "学校A", "civil_service_rate", status="todo", priority_rank="1"),
    ]
    _write_outcome_plan(plan, rows)

    report = audit_outcome_collection_core_coverage(
        plan_csv=plan,
        core_db=db,
        report_path=tmp_path / "coverage.json",
    )

    assert report["ready_for_full_universe_review"] is False
    assert report["core_school_count"] == 2
    assert report["missing_school_count"] == 1
    assert report["missing_school_sample"] == ["1002"]
    assert (tmp_path / "coverage.json").exists()


def test_build_scoped_outcome_stock_review_flags_official_scoped_candidates(tmp_path: Path):
    candidates = tmp_path / "school_candidates.csv"
    with candidates.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=CANDIDATE_COLUMNS)
        writer.writeheader()
        base = {column: "" for column in CANDIDATE_COLUMNS}
        writer.writerow({
            **base,
            "domain": "school",
            "entity_code": "0162",
            "entity_name": "辽宁中医药大学",
            "metric_key": "employment_rate",
            "metric_label": "毕业去向落实率",
            "metric_unit": "ratio",
            "metric_year": "2024",
            "candidate_value": "0.85",
            "source_title": "辽宁中医药大学2023-2024学年本科教学质量报告",
            "source_url": "https://xxgk.lnutcm.edu.cn/report.pdf",
            "evidence_quote": "辽宁省内毕业去向落实率 85%以上。",
            "review_status": "rejected",
            "notes": "非学校总体就业率。",
        })
        writer.writerow({
            **base,
            "domain": "school",
            "entity_code": "10140",
            "entity_name": "辽宁大学",
            "metric_key": "employment_rate",
            "metric_label": "毕业去向落实率",
            "metric_unit": "ratio",
            "metric_year": "2024",
            "candidate_value": "0.9",
            "source_title": "第三方汇总",
            "evidence_quote": "就业率 90%。",
            "review_status": "rejected",
            "notes": "第三方来源。",
        })

    report = build_scoped_outcome_stock_review(
        candidate_globs=[str(candidates)],
        output=tmp_path / "scoped_review.csv",
    )

    assert report["review_rows"] == 1
    with (tmp_path / "scoped_review.csv").open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    assert rows[0]["entity_name"] == "辽宁中医药大学"
    assert rows[0]["scoped_review_class"] == "scoped_official_candidate"
    assert "省内" in rows[0]["matched_scope_terms"]


def test_build_scoped_outcome_stock_review_batch_filters_and_prioritizes(tmp_path: Path):
    review_csv = tmp_path / "scoped_review.csv"
    with review_csv.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "candidate_file",
            *CANDIDATE_COLUMNS,
            "scoped_review_class",
            "matched_scope_terms",
            "recommended_action",
        ])
        writer.writeheader()
        base = {column: "" for column in writer.fieldnames or []}
        writer.writerow({
            **base,
            "entity_code": "2002",
            "entity_name": "后处理学校",
            "metric_key": "postgrad_rate",
            "scoped_review_class": "scoped_official_candidate",
        })
        writer.writerow({
            **base,
            "candidate_file": "staging/outcome_v46/extraction/candidates/priority.csv",
            "entity_code": "1001",
            "entity_name": "优先学校",
            "metric_key": "employment_rate",
            "metric_year": "2024",
            "candidate_value": "0.88",
            "source_title": "优先学校报告",
            "source_url": "https://example.edu/priority.pdf",
            "evidence_quote": "毕业去向落实率88%。",
            "scoped_review_class": "scoped_official_candidate",
        })
        writer.writerow({
            **base,
            "candidate_file": "staging/outcome_v48/extraction_merged/candidates/priority.csv",
            "entity_code": "1001",
            "entity_name": "优先学校更新版",
            "metric_key": "employment_rate",
            "metric_year": "2024",
            "candidate_value": "0.88",
            "source_title": "优先学校报告",
            "source_url": "https://example.edu/priority.pdf",
            "evidence_quote": "毕业去向落实率88%。",
            "scoped_review_class": "scoped_official_candidate",
        })
        writer.writerow({
            **base,
            "entity_code": "1000",
            "entity_name": "总体学校",
            "metric_key": "employment_rate",
            "scoped_review_class": "overall_approved_candidate",
        })

    report = build_scoped_outcome_stock_review_batch(
        review_csv=review_csv,
        output_dir=tmp_path / "batch",
        limit=2,
        review_class=["scoped_official_candidate"],
    )

    assert report["batch_rows"] == 2
    assert report["duplicate_filtered_rows"] == 1
    assert report["batch_metric_counts"] == {"employment_rate": 1, "postgrad_rate": 1}
    with (tmp_path / "batch" / "scoped_stock_review_batch.csv").open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    assert [row["entity_name"] for row in rows] == ["优先学校更新版", "后处理学校"]


def test_build_scoped_outcome_stock_review_batch_excludes_reviewed_rows(tmp_path: Path):
    review_csv = tmp_path / "scoped_review.csv"
    reviewed_csv = tmp_path / "reviewed.csv"
    fieldnames = [
        "candidate_file",
        *CANDIDATE_COLUMNS,
        "scoped_review_class",
        "matched_scope_terms",
        "recommended_action",
    ]
    base = {column: "" for column in fieldnames}
    with review_csv.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerow({
            **base,
            "entity_code": "1001",
            "entity_name": "已复核学校",
            "metric_key": "employment_rate",
            "metric_year": "2024",
            "candidate_value": "0.88",
            "source_title": "已复核学校报告",
            "source_url": "https://example.edu/reviewed.pdf",
            "evidence_quote": "毕业去向落实率88%。",
            "metric_scope": "manually_reviewed_scope",
            "scoped_review_class": "scoped_official_candidate",
        })
        writer.writerow({
            **base,
            "entity_code": "1002",
            "entity_name": "新学校",
            "metric_key": "employment_rate",
            "metric_year": "2024",
            "candidate_value": "0.86",
            "source_title": "新学校报告",
            "source_url": "https://example.edu/new.pdf",
            "evidence_quote": "毕业去向落实率86%。",
            "scoped_review_class": "scoped_official_candidate",
        })
    with reviewed_csv.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerow({
            **base,
            "entity_code": "1001",
            "entity_name": "已复核学校",
            "metric_key": "employment_rate",
            "metric_year": "2024",
            "candidate_value": "0.88",
            "source_title": "已复核学校报告",
            "source_url": "https://example.edu/reviewed.pdf",
            "evidence_quote": "毕业去向落实率88%。",
            "scoped_review_class": "scoped_official_candidate",
        })

    report = build_scoped_outcome_stock_review_batch(
        review_csv=review_csv,
        output_dir=tmp_path / "batch",
        limit=10,
        review_class=["scoped_official_candidate"],
        exclude_csv=[reviewed_csv],
    )

    assert report["excluded_row_keys"] == 1
    assert report["batch_rows"] == 1
    with (tmp_path / "batch" / "scoped_stock_review_batch.csv").open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    assert [row["entity_name"] for row in rows] == ["新学校"]


def test_build_outcome_packages_from_verified_collection_plan(tmp_path: Path):
    plan = tmp_path / "outcome_collection_plan.csv"
    fieldnames = [
        "domain",
        "entity_code",
        "entity_name",
        "priority_rank",
        "plan_rows",
        "metric_key",
        "metric_label",
        "metric_unit",
        "metric_year",
        "search_queries",
        "status",
        "metric_value",
        "source_title",
        "source_url",
        "evidence_quote",
        "metric_scope",
        "denominator",
        "source_date",
        "availability_date",
        "built_at",
        "notes",
    ]
    with plan.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerow({
            "domain": "school",
            "entity_code": "10145",
            "entity_name": "东北大学",
            "priority_rank": "1",
            "plan_rows": "120",
            "metric_key": "postgrad_rate",
            "metric_label": "深造率",
            "metric_unit": "ratio",
            "metric_year": "2025",
            "search_queries": json.dumps(["东北大学 2025 就业质量报告"], ensure_ascii=False),
            "status": "verified",
            "metric_value": "46.2%",
            "source_title": "2025届毕业生就业质量报告",
            "source_url": "https://example.edu/report.pdf",
            "evidence_quote": "本科毕业生深造率为46.2%。",
            "metric_scope": "本科毕业生",
            "denominator": "1000",
            "source_date": "2025-12-31",
            "availability_date": "2026-01-05",
            "built_at": "2026-05-13T00:00:00",
            "notes": "",
        })
        writer.writerow({
            "domain": "major",
            "entity_code": "080901",
            "entity_name": "计算机科学与技术",
            "priority_rank": "1",
            "plan_rows": "88",
            "metric_key": "exam_friendly_score",
            "metric_label": "考研友好度",
            "metric_unit": "score",
            "metric_year": "2025",
            "search_queries": json.dumps(["计算机科学与技术 考研友好度"], ensure_ascii=False),
            "status": "verified",
            "metric_value": "82",
            "source_title": "专业升学去向整理",
            "source_url": "https://example.edu/major.html",
            "evidence_quote": "计算机科学与技术专业升学方向延续性较强。",
            "metric_scope": "本科专业",
            "denominator": "",
            "source_date": "2025-12-31",
            "availability_date": "2026-01-05",
            "built_at": "2026-05-13T00:00:00",
            "notes": "",
        })
        writer.writerow({
            "domain": "school",
            "entity_code": "10145",
            "entity_name": "东北大学",
            "priority_rank": "1",
            "plan_rows": "120",
            "metric_key": "employment_rate",
            "metric_label": "毕业去向落实率",
            "metric_unit": "ratio",
            "metric_year": "2025",
            "search_queries": json.dumps(["东北大学 就业质量报告"], ensure_ascii=False),
            "status": "todo",
            "metric_value": "",
            "source_title": "",
            "source_url": "",
            "evidence_quote": "",
            "metric_scope": "",
            "denominator": "",
            "source_date": "",
            "availability_date": "",
            "built_at": "",
            "notes": "",
        })

    with pytest.raises(ValueError, match="pending_rows=1"):
        build_outcome_packages_from_collection_plan(
            plan_csv=plan,
            output_root=tmp_path / "exports_blocked",
            package_id="pkg-outcome-collection",
        )

    result = build_outcome_packages_from_collection_plan(
        plan_csv=plan,
        output_root=tmp_path / "exports",
        package_id="pkg-outcome-collection",
        allow_partial=True,
    )

    packages = {package["table"]: package for package in result["packages"]}
    assert set(packages) == {"fa_fact_school_outcome", "fa_fact_major_outcome"}
    assert packages["fa_fact_school_outcome"]["rows"] == 1
    assert packages["fa_fact_major_outcome"]["rows"] == 1
    school_package = Path(packages["fa_fact_school_outcome"]["package_dir"])
    manifest = json.loads((school_package / "manifest.json").read_text(encoding="utf-8"))
    assert manifest["source_lineage"]["source_kind"] == "verified_outcome_collection_plan"
    assert manifest["source_lineage"]["collection_plan"] == str(plan)
    assert manifest["source_lineage"]["allow_partial"] is True
    assert manifest["source_lineage"]["is_partial"] is True
    assert manifest["source_lineage"]["evidence_urls"] == ["https://example.edu/report.pdf"]
    assert packages["fa_fact_school_outcome"]["source_lineage"]["target_table"] == "fa_fact_school_outcome"
    partial_quality = json.loads((school_package / "quality_report.json").read_text(encoding="utf-8"))
    assert "partial_outcome_collection_package_not_for_core_import" in partial_quality["errors"]
    assert any("quality_report has errors" in err for err in validate_manifest(school_package / "manifest.json")["errors"])
    with (school_package / "fa_fact_school_outcome.csv").open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    assert rows[0]["school_code"] == "10145"
    assert rows[0]["metric_value"] == "0.462"


def test_cli_audit_outcome_collection_plan_returns_nonzero_on_errors(tmp_path: Path, monkeypatch):
    import datahub.cli as cli

    plan = tmp_path / "outcome_collection_plan.csv"
    rows = [_outcome_plan_row("school", "10140", "辽宁大学", "employment_rate", status="bad_status", priority_rank="1")]
    with plan.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)

    monkeypatch.setattr("sys.argv", [
        "lifehack-datahub",
        "audit-outcome-collection-plan",
        "--plan-csv",
        str(plan),
    ])

    assert cli.main() == 1


def test_build_major_outcome_from_civil_service_positions(tmp_path: Path):
    db = tmp_path / "core.duckdb"
    con = duckdb.connect(str(db))
    try:
        con.execute("""
            CREATE TABLE fa_dim_ln_admission_plan (
                major_short VARCHAR,
                major_full VARCHAR,
                batch VARCHAR
            )
        """)
        con.execute("""
            INSERT INTO fa_dim_ln_admission_plan VALUES
            ('120203K', '会计学', '本科批'),
            ('120204', '财务管理', '本科批'),
            ('030101K', '法学', '本科批'),
            ('050101', '汉语言文学', '本科批'),
            ('080901', '计算机科学与技术', '本科批')
        """)
    finally:
        con.close()

    positions = tmp_path / "positions.csv"
    fieldnames = [
        "source_key",
        "source_title",
        "source_url",
        "source_date",
        "availability_date",
        "sheet_name",
        "row_number",
        "department_name",
        "position_name",
        "position_code",
        "recruit_count",
        "major_requirement",
        "remarks",
    ]
    with positions.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerow({
            "source_key": "career_civil_service_posts",
            "source_title": "中央机关及其直属机构2026年度考试录用公务员招考简章",
            "source_url": "http://example.gov/positions.xls",
            "source_date": "2025-10-14",
            "availability_date": "2025-10-14",
            "sheet_name": "中央党群机关",
            "row_number": "4",
            "department_name": "中央办公厅",
            "position_name": "财务管理岗位",
            "position_code": "100110001002",
            "recruit_count": "2",
            "major_requirement": "本科：120203K会计学、120204财务管理；研究生：1253会计",
            "remarks": "研究生学历报考者须同时具有本科和研究生学历学位。",
        })
        writer.writerow({
            "source_key": "career_civil_service_posts",
            "source_title": "中央机关及其直属机构2026年度考试录用公务员招考简章",
            "source_url": "http://example.gov/positions.xls",
            "source_date": "2025-10-14",
            "availability_date": "2025-10-14",
            "sheet_name": "中央党群机关",
            "row_number": "6",
            "department_name": "中央办公厅",
            "position_name": "文秘岗位",
            "position_code": "100210002001",
            "recruit_count": "3",
            "major_requirement": "0301法学、0501中国语言文学",
            "remarks": "岗位要求具备较强的文稿写作能力。",
        })
        writer.writerow({
            "source_key": "career_civil_service_posts",
            "source_title": "中央机关及其直属机构2026年度考试录用公务员招考简章",
            "source_url": "http://example.gov/positions.xls",
            "source_date": "2025-10-14",
            "availability_date": "2025-10-14",
            "sheet_name": "中央党群机关",
            "row_number": "7",
            "department_name": "中央办公厅",
            "position_name": "综合岗位",
            "position_code": "100210002002",
            "recruit_count": "10",
            "major_requirement": "不限",
            "remarks": "",
        })

    result = build_major_outcome_from_civil_service_package(
        positions_csv=positions,
        core_db=db,
        output_root=tmp_path / "exports",
        package_id="major-civil-service-fit",
    )

    assert result["table"] == "fa_fact_major_outcome"
    assert result["rows"] == 4
    package_dir = Path(result["package_dir"])
    assert validate_manifest(package_dir / "manifest.json")["errors"] == []
    with (package_dir / "fa_fact_major_outcome.csv").open(encoding="utf-8", newline="") as f:
        rows = {row["major_code"]: row for row in csv.DictReader(f)}
    assert set(rows) == {"030101K", "050101", "120203K", "120204"}
    assert rows["120203K"]["metric_key"] == "civil_service_fit_score"
    assert rows["120203K"]["metric_unit"] == "score"
    assert rows["120203K"]["metric_scope"]
    assert "中央办公厅-财务管理岗位" in rows["120203K"]["evidence_quote"]
    assert "code_prefix_4" in rows["030101K"]["evidence_quote"]
    assert float(rows["050101"]["metric_value"]) > float(rows["120203K"]["metric_value"])


def test_audit_admission_plan_package_against_core_reports_scope_drift(tmp_path: Path):
    db = tmp_path / "core.duckdb"
    con = duckdb.connect(str(db))
    try:
        con.execute("""
            CREATE TABLE fa_dim_ln_admission_plan (
                school_code VARCHAR,
                school_name VARCHAR,
                major_code VARCHAR,
                major_full VARCHAR,
                major_short VARCHAR,
                batch VARCHAR,
                subject_cat VARCHAR,
                year INTEGER,
                school_tier VARCHAR,
                region VARCHAR,
                plan_count INTEGER,
                school_type VARCHAR,
                city_level_tag VARCHAR,
                postgrad_rate DOUBLE
            )
        """)
        con.execute("""
            INSERT INTO fa_dim_ln_admission_plan VALUES
                ('1001', '东北大学', '01', '计算机类', '计算机类', '本科批', '物理类', 2026, '985', '辽宁省沈阳市', 10, '公办', '新一线', 0.28),
                ('1002', '辽宁大学', '02', '法学', '法学', '本科批', '物理类', 2026, '211', '辽宁省沈阳市', 9, '公办', '新一线', 0.12),
                ('1003', '大连理工大学', '03', '软件工程', '软件工程', '本科批', '物理类', 2026, '985', '辽宁省大连市', 6, '公办', '新一线', 0.30),
                ('2001', '历史大学', '01', '汉语言文学', '汉语言文学', '本科批', '历史类', 2026, '普通本科', '辽宁省', 5, '公办', '其他', 0.05)
        """)
    finally:
        con.close()

    schema = get_table_schema("fa_dim_ln_admission_plan")
    package_dir = tmp_path / "exports" / "pkg-admission-plan-audit"
    package_dir.mkdir(parents=True)
    with (package_dir / "fa_dim_ln_admission_plan.csv").open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=schema["columns"], extrasaction="ignore")
        writer.writeheader()
        writer.writerows([
            {
                "school_code": "1001",
                "school_name": "东北大学",
                "major_code": "01",
                "major_full": "计算机类",
                "major_short": "计算机类",
                "batch": "本科批",
                "subject_cat": "物理类",
                "year": "2026",
                "school_tier": "985",
                "region": "辽宁省沈阳市",
                "plan_count": "10",
                "school_type": "公办",
                "city_level_tag": "新一线",
                "postgrad_rate": "0.28",
            },
            {
                "school_code": "1002",
                "school_name": "辽宁大学",
                "major_code": "02",
                "major_full": "法学",
                "major_short": "法学",
                "batch": "本科批",
                "subject_cat": "物理类",
                "year": "2026",
                "school_tier": "211",
                "region": "辽宁省沈阳市",
                "plan_count": "8",
                "school_type": "公办",
                "city_level_tag": "新一线",
                "postgrad_rate": "0.12",
            },
            {
                "school_code": "1004",
                "school_name": "沈阳工业大学",
                "major_code": "04",
                "major_full": "自动化",
                "major_short": "自动化",
                "batch": "本科批",
                "subject_cat": "物理类",
                "year": "2026",
                "school_tier": "普通本科",
                "region": "辽宁省沈阳市",
                "plan_count": "12",
                "school_type": "公办",
                "city_level_tag": "新一线",
                "postgrad_rate": "0.08",
            },
        ])
    (package_dir / "quality_report.json").write_text('{"errors":[]}\n', encoding="utf-8")
    (package_dir / "manifest.json").write_text(json.dumps({
        "package_id": "pkg-admission-plan-audit",
        "built_at": "2026-05-13T00:00:00",
        "source_version": "fixture",
        "tables": [{"name": "fa_dim_ln_admission_plan", "file": "fa_dim_ln_admission_plan.csv"}],
        "files": ["fa_dim_ln_admission_plan.csv"],
        "hashes": {},
        "quality_report": "quality_report.json",
    }, ensure_ascii=False), encoding="utf-8")

    report = audit_admission_plan_package_against_core(
        core_db=db,
        package_dirs=[package_dir],
        sample_limit=5,
    )

    assert report["errors"] == []
    assert report["counts"]["package_rows"] == 3
    assert report["counts"]["core_scoped_rows"] == 3
    assert report["counts"]["matched_rows"] == 2
    assert report["counts"]["different_rows"] == 1
    assert report["counts"]["package_only_rows"] == 1
    assert report["counts"]["core_only_rows"] == 1
    assert report["samples"]["different_rows"][0]["differences"] == [
        {"column": "plan_count", "package_value": 8, "core_value": 9}
    ]
    assert report["decision"]["reconciliation_required"] is True

    (package_dir / "quality_report.json").write_text('{"errors":["manual review is not complete"]}\n', encoding="utf-8")
    bad_quality_report = audit_admission_plan_package_against_core(
        core_db=db,
        package_dirs=[package_dir],
        sample_limit=5,
    )
    assert any("quality_report error" in error for error in bad_quality_report["errors"])
    assert bad_quality_report["decision"]["safe_to_import_without_reconciliation"] is False
    with pytest.raises(ValueError, match="quality_report error"):
        build_admission_plan_reconciliation_plan(
            core_db=db,
            package_dirs=[package_dir],
            output_dir=tmp_path / "blocked_admission_reconciliation",
        )


def test_build_admission_plan_reconciliation_plan_from_audit_inputs(tmp_path: Path):
    db = tmp_path / "core.duckdb"
    con = duckdb.connect(str(db))
    try:
        con.execute("""
            CREATE TABLE fa_dim_ln_admission_plan (
                school_code VARCHAR,
                school_name VARCHAR,
                major_code VARCHAR,
                major_full VARCHAR,
                major_short VARCHAR,
                batch VARCHAR,
                subject_cat VARCHAR,
                year INTEGER,
                region VARCHAR,
                plan_count INTEGER
            )
        """)
        con.execute("""
            INSERT INTO fa_dim_ln_admission_plan VALUES
                ('1001', '东北大学', '01', '计算机类', '计算机类', '本科批', '物理类', 2026, '辽宁省沈阳市', 10),
                ('1002', '辽宁大学', '02', '法学', '法学', '本科批', '物理类', 2026, '辽宁省沈阳市', 9),
                ('1003', '大连理工大学', '03', '软件工程', '软件工程', '本科批', '物理类', 2026, '辽宁省大连市', 6)
        """)
    finally:
        con.close()

    schema = get_table_schema("fa_dim_ln_admission_plan")
    package_dir = tmp_path / "exports" / "pkg-admission-plan-reconciliation"
    package_dir.mkdir(parents=True)
    with (package_dir / "fa_dim_ln_admission_plan.csv").open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=schema["columns"], extrasaction="ignore")
        writer.writeheader()
        writer.writerows([
            {
                "school_code": "1001",
                "school_name": "东北大学",
                "major_code": "01",
                "major_full": "计算机类",
                "major_short": "计算机类",
                "batch": "本科批",
                "subject_cat": "物理类",
                "year": "2026",
                "region": "辽宁省沈阳市",
                "plan_count": "10",
            },
            {
                "school_code": "1002",
                "school_name": "辽宁大学",
                "major_code": "02",
                "major_full": "法学",
                "major_short": "法学",
                "batch": "本科批",
                "subject_cat": "物理类",
                "year": "2026",
                "region": "辽宁省沈阳市",
                "plan_count": "8",
            },
            {
                "school_code": "1004",
                "school_name": "沈阳工业大学",
                "major_code": "04",
                "major_full": "自动化",
                "major_short": "自动化",
                "batch": "本科批",
                "subject_cat": "物理类",
                "year": "2026",
                "region": "辽宁省沈阳市",
                "plan_count": "12",
            },
        ])
    (package_dir / "quality_report.json").write_text('{"errors":[]}\n', encoding="utf-8")
    (package_dir / "manifest.json").write_text(json.dumps({
        "package_id": "pkg-admission-plan-reconciliation",
        "built_at": "2026-05-13T00:00:00",
        "source_version": "fixture",
        "tables": [{"name": "fa_dim_ln_admission_plan", "file": "fa_dim_ln_admission_plan.csv"}],
        "files": ["fa_dim_ln_admission_plan.csv"],
        "hashes": {},
        "quality_report": "quality_report.json",
    }, ensure_ascii=False), encoding="utf-8")

    result = build_admission_plan_reconciliation_plan(
        core_db=db,
        package_dirs=[package_dir],
        output_dir=tmp_path / "reconciliation",
    )

    assert result["rows"] == 3
    assert result["issue_counts"] == {
        "core_only_unmatched": 1,
        "package_only_unmatched": 1,
        "value_drift": 1,
    }
    with Path(result["csv"]).open(encoding="utf-8", newline="") as f:
        tasks = list(csv.DictReader(f))
    by_type = {task["issue_type"]: task for task in tasks}
    assert by_type["value_drift"]["status"] == "todo"
    assert by_type["value_drift"]["suggested_action"] == "review_admission_field_conflict"
    assert by_type["value_drift"]["package_plan_count"] == "8"
    assert by_type["value_drift"]["core_plan_count"] == "9"
    assert by_type["package_only_unmatched"]["package_major_code"] == "04"
    assert by_type["core_only_unmatched"]["core_major_code"] == "03"


def test_audit_admission_plan_reconciliation_plan_reports_progress(tmp_path: Path):
    plan = tmp_path / "admission_plan_reconciliation_plan.csv"
    rows = [
        _admission_reconciliation_row("t1", "value_drift", status="reviewed", review_decision="use_package_row"),
        _admission_reconciliation_row("t2", "core_only_unmatched", status="todo"),
    ]
    _write_admission_reconciliation_plan(plan, rows)

    report = audit_admission_plan_reconciliation_plan(plan)

    assert report["errors"] == []
    assert report["rows"] == 2
    assert report["status_counts"] == {"reviewed": 1, "todo": 1}
    assert report["decision_counts"] == {"use_package_row": 1}
    assert report["progress"]["ready_rows"] == 1
    assert report["progress"]["pending_rows"] == 1
    assert report["ready"]["review_complete"] is False
    assert report["ready"]["migration_ready"] is False


def test_audit_admission_plan_reconciliation_plan_rejects_side_mismatch(tmp_path: Path):
    plan = tmp_path / "admission_plan_reconciliation_plan.csv"
    package_decision_without_package = _admission_reconciliation_row(
        "bad-package",
        "core_only_unmatched",
        status="reviewed",
        review_decision="use_package_row",
    )
    package_decision_without_package.update({
        "package_major_code": "",
        "package_school_name": "",
        "package_major_full": "",
        "package_plan_count": "",
        "package_key_json": "{}",
    })
    core_decision_without_core = _admission_reconciliation_row(
        "bad-core",
        "package_only_unmatched",
        status="reviewed",
        review_decision="keep_core_row",
    )
    core_decision_without_core.update({
        "core_major_code": "",
        "core_school_name": "",
        "core_major_full": "",
        "core_plan_count": "",
        "core_key_json": "{}",
    })
    _write_admission_reconciliation_plan(plan, [package_decision_without_package, core_decision_without_core])

    report = audit_admission_plan_reconciliation_plan(plan)

    assert "row 2 use_package_row without package side" in report["errors"]
    assert "row 3 keep_core_row without core side" in report["errors"]
    assert report["ready"]["migration_ready"] is False


def test_build_and_merge_admission_plan_reconciliation_review_batch(tmp_path: Path):
    plan = tmp_path / "admission_plan_reconciliation_plan.csv"
    rows = [
        _admission_reconciliation_row("v1", "value_drift", priority="1", status="todo"),
        _admission_reconciliation_row("v2", "value_drift", priority="1", status="todo", school_code="1002"),
        _admission_reconciliation_row("p1", "package_only_unmatched", priority="2", status="todo", school_code="1003"),
        _admission_reconciliation_row("c1", "core_only_unmatched", priority="3", status="reviewed", review_decision="keep_core_row", school_code="1004"),
    ]
    _write_admission_reconciliation_plan(plan, rows)

    batch_result = build_admission_plan_reconciliation_review_batch(
        plan_csv=plan,
        output_dir=tmp_path / "batch",
        limit_per_issue=1,
    )

    assert batch_result["rows"] == 2
    assert batch_result["issue_counts"] == {"package_only_unmatched": 1, "value_drift": 1}
    with Path(batch_result["csv"]).open(encoding="utf-8", newline="") as f:
        batch_rows = list(csv.DictReader(f))
    batch_rows[0]["school_code"] = "tampered"
    batch_rows[0]["status"] = "reviewed"
    batch_rows[0]["review_decision"] = "use_package_row"
    batch_rows[0]["reviewer"] = "tester"
    batch_rows[0]["reviewed_at"] = "2026-05-13"
    batch_rows[0]["notes"] = "approved"
    _write_admission_reconciliation_plan(Path(batch_result["csv"]), batch_rows)

    output = tmp_path / "admission_plan_reconciliation_plan_merged.csv"
    merge_report = merge_admission_plan_reconciliation_review_batch(
        plan_csv=plan,
        batch_csv=Path(batch_result["csv"]),
        output=output,
    )

    assert merge_report["updated_rows"] == 1
    with output.open(encoding="utf-8", newline="") as f:
        merged_rows = list(csv.DictReader(f))
    by_id = {row["task_id"]: row for row in merged_rows}
    assert by_id["v1"]["school_code"] == "1001"
    assert by_id["v1"]["status"] == "reviewed"
    assert by_id["v1"]["review_decision"] == "use_package_row"
    assert by_id["p1"]["status"] == "todo"
    assert by_id["c1"]["status"] == "reviewed"


def test_build_admission_plan_delete_plan_rejects_unready(tmp_path: Path):
    plan = tmp_path / "admission_plan_reconciliation_plan.csv"
    rows = [
        _admission_reconciliation_row("c1", "core_only_unmatched", priority="3", status="todo", school_code="1004"),
    ]
    _write_admission_reconciliation_plan(plan, rows)

    try:
        build_admission_plan_delete_plan_from_reconciliation_plan(
            plan_csv=plan,
            output_dir=tmp_path / "delete_plan",
        )
        rejected = False
    except ValueError as exc:
        rejected = "not ready for delete planning" in str(exc)
    assert rejected


def test_build_admission_plan_delete_plan_from_reviewed_core_excludes(tmp_path: Path):
    plan = tmp_path / "admission_plan_reconciliation_plan.csv"
    core_exclude = _admission_reconciliation_row(
        "core-delete-1",
        "core_only_unmatched",
        priority="3",
        status="reviewed",
        review_decision="exclude_row",
        school_code="1004",
    )
    core_exclude.update({
        "package_major_code": "",
        "core_major_code": "04",
        "package_school_name": "",
        "core_school_name": "沈阳工业大学",
        "package_major_full": "",
        "core_major_full": "自动化",
        "package_plan_count": "",
        "core_plan_count": "12",
        "package_key_json": "{}",
        "core_key_json": json.dumps({
            "school_code": "1004",
            "major_code": "04",
            "batch": "本科批",
            "subject_cat": "物理类",
            "year": "2026",
        }, ensure_ascii=False),
    })
    package_exclude = _admission_reconciliation_row(
        "package-exclude-1",
        "package_only_unmatched",
        priority="2",
        status="reviewed",
        review_decision="exclude_row",
        school_code="1005",
    )
    package_exclude.update({
        "package_key_json": json.dumps({
            "school_code": "1005",
            "major_code": "05",
            "batch": "本科批",
            "subject_cat": "物理类",
            "year": "2026",
        }, ensure_ascii=False),
        "core_key_json": "{}",
    })
    _write_admission_reconciliation_plan(plan, [core_exclude, package_exclude])

    result = build_admission_plan_delete_plan_from_reconciliation_plan(
        plan_csv=plan,
        output_dir=tmp_path / "delete_plan",
    )

    assert result["rows"] == 1
    with Path(result["csv"]).open(encoding="utf-8", newline="") as f:
        delete_rows = list(csv.DictReader(f))
    assert delete_rows[0]["school_code"] == "1004"
    assert delete_rows[0]["major_code"] == "04"
    assert delete_rows[0]["year"] == "2026"
    assert delete_rows[0]["school_name"] == "沈阳工业大学"
    assert delete_rows[0]["plan_count"] == "12"
    manifest = json.loads(Path(result["manifest"]).read_text(encoding="utf-8"))
    assert manifest["csv"] == "admission_plan_delete_plan.csv"
    assert manifest["notes"].startswith("Delete migration plan only")


def _admission_reconciliation_row(
    task_id: str,
    issue_type: str,
    *,
    priority: str = "1",
    status: str = "todo",
    review_decision: str = "",
    school_code: str = "1001",
) -> dict[str, str]:
    return {
        "task_id": task_id,
        "issue_type": issue_type,
        "priority": priority,
        "status": status,
        "suggested_action": "review_admission_field_conflict",
        "match_confidence": "primary_key_match",
        "batch": "本科批",
        "subject_cat": "物理类",
        "year": "2026",
        "school_code": school_code,
        "package_major_code": "01",
        "core_major_code": "01",
        "package_school_name": "东北大学",
        "core_school_name": "东北大学",
        "package_major_full": "计算机类",
        "core_major_full": "计算机类",
        "package_plan_count": "8",
        "core_plan_count": "9",
        "package_key_json": json.dumps({
            "school_code": school_code,
            "major_code": "01",
            "batch": "本科批",
            "subject_cat": "物理类",
            "year": "2026",
        }, ensure_ascii=False),
        "core_key_json": json.dumps({
            "school_code": school_code,
            "major_code": "01",
            "batch": "本科批",
            "subject_cat": "物理类",
            "year": "2026",
        }, ensure_ascii=False),
        "differences_json": "[]",
        "review_decision": review_decision,
        "reviewer": "tester" if status == "reviewed" else "",
        "reviewed_at": "2026-05-13" if status == "reviewed" else "",
        "notes": "",
    }


def _write_admission_reconciliation_plan(path: Path, rows: list[dict[str, str]]) -> None:
    from datahub.builders.admission_plan_reconciliation_plan import PLAN_COLUMNS as ADMISSION_PLAN_COLUMNS

    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=ADMISSION_PLAN_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _outcome_plan_row(
    domain: str,
    entity_code: str,
    entity_name: str,
    metric_key: str,
    *,
    status: str,
    priority_rank: str,
) -> dict[str, str]:
    metric_label = {
        "postgrad_rate": "深造率",
        "employment_rate": "毕业去向落实率",
    }.get(metric_key, metric_key)
    return {
        "domain": domain,
        "entity_code": entity_code,
        "entity_name": entity_name,
        "priority_rank": priority_rank,
        "plan_rows": "20",
        "metric_key": metric_key,
        "metric_label": metric_label,
        "metric_unit": "ratio",
        "metric_year": "2025",
        "search_queries": json.dumps([f"{entity_name} 2025 {metric_label}"], ensure_ascii=False),
        "status": status,
        "metric_value": "",
        "source_title": "",
        "source_url": "",
        "evidence_quote": "",
        "metric_scope": "",
        "denominator": "",
        "source_date": "",
        "availability_date": "",
        "built_at": "",
        "notes": "",
    }


def _write_outcome_plan(path: Path, rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=OUTCOME_PLAN_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _career_plan_row(
    source_key: str,
    target_table: str,
    metric_key: str,
    *,
    status: str,
) -> dict[str, str]:
    source_name = {
        "career_recruitment_snapshot": "招聘需求与薪资快照",
        "career_civil_service_posts": "公考与事业编岗位目录",
    }.get(source_key, source_key)
    metric_label = {
        "salary_median": "月薪中位数",
        "salary_p75": "月薪75分位",
        "shortage_rank": "紧缺职业排行",
        "work_intensity_index": "工作强度指数",
        "civil_service_post_count": "公考岗位数",
    }.get(metric_key, metric_key)
    metric_unit = {
        "salary_median": "cny_month",
        "salary_p75": "cny_month",
        "shortage_rank": "rank",
        "work_intensity_index": "score",
        "civil_service_post_count": "count",
    }.get(metric_key, "")
    row = {column: "" for column in CAREER_PLAN_COLUMNS}
    row.update({
        "source_key": source_key,
        "source_name": source_name,
        "source_kind": "controlled_market_snapshot",
        "target_table": target_table,
        "occupation_code": "15-102",
        "occupation_name": "软件工程师",
        "tdx_l2": "T1205",
        "tdx_l2_name": "软件服务",
        "metric_key": metric_key,
        "metric_label": metric_label,
        "metric_unit": metric_unit,
        "metric_year": "2026",
        "city": "沈阳",
        "collection_methods": '["manual_platform_export"]',
        "official_distribution": "公开可复核来源",
        "evidence_urls": "[]",
        "search_queries": json.dumps([f"软件工程师 {metric_label} 沈阳 2026"], ensure_ascii=False),
        "status": status,
    })
    return row


def _write_career_plan(path: Path, rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=CAREER_PLAN_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def test_build_school_identity_package_matches_unique_school_names(tmp_path: Path):
    db = tmp_path / "core.duckdb"
    con = duckdb.connect(str(db))
    try:
        con.execute("""
            CREATE TABLE fa_dim_ln_admission_plan (
                school_code VARCHAR,
                school_name VARCHAR,
                major_code VARCHAR,
                major_full VARCHAR,
                batch VARCHAR,
                subject_cat VARCHAR
            )
        """)
        con.execute("""
            INSERT INTO fa_dim_ln_admission_plan VALUES
                ('0140', '辽宁大学', '01', '法学', '本科批', '历史类'),
                ('0140', '辽宁大学', '02', '汉语言文学', '本科批', '历史类'),
                ('0183', '吉林大学', '01', '计算机类', '本科批', '物理类'),
                ('1414', '中国石油大学(北京)', '01', '机械类', '本科批', '物理类'),
                ('6407', '香港中文大学', '01', '测试专业', '本科批', '物理类'),
                ('6407', '香港中文大学(深圳)', '02', '测试专业', '本科批', '物理类'),
                ('9999', '不存在大学', '01', '测试专业', '本科批', '物理类')
        """)
    finally:
        con.close()

    school_profile = tmp_path / "school_profile.csv"
    with school_profile.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "national_school_code",
                "school_name",
                "province",
                "city",
                "school_tier",
                "school_type",
                "ownership",
                "official_site",
                "competent_authority",
                "source_date",
                "availability_date",
                "built_at",
            ],
        )
        writer.writeheader()
        writer.writerow({
            "national_school_code": "4121010140",
            "school_name": "辽宁大学",
            "province": "辽宁省",
            "city": "沈阳市",
            "school_tier": "本科",
            "school_type": "",
            "ownership": "",
            "official_site": "",
            "competent_authority": "辽宁省",
            "source_date": "2025-06-20",
            "availability_date": "2025-06-27",
            "built_at": "2026-05-13T00:00:00",
        })
        writer.writerow({
            "national_school_code": "4122010183",
            "school_name": "吉林大学",
            "province": "吉林省",
            "city": "长春市",
            "school_tier": "本科",
            "school_type": "",
            "ownership": "",
            "official_site": "",
            "competent_authority": "教育部",
            "source_date": "2025-06-20",
            "availability_date": "2025-06-27",
            "built_at": "2026-05-13T00:00:00",
        })
        writer.writerow({
            "national_school_code": "4111011414",
            "school_name": "中国石油大学（北京）",
            "province": "北京市",
            "city": "北京市",
            "school_tier": "本科",
            "school_type": "",
            "ownership": "",
            "official_site": "",
            "competent_authority": "教育部",
            "source_date": "2025-06-20",
            "availability_date": "2025-06-27",
            "built_at": "2026-05-13T00:00:00",
        })
        writer.writerow({
            "national_school_code": "4144016407",
            "school_name": "香港中文大学（深圳）",
            "province": "广东省",
            "city": "深圳市",
            "school_tier": "本科",
            "school_type": "",
            "ownership": "",
            "official_site": "",
            "competent_authority": "广东省",
            "source_date": "2025-06-20",
            "availability_date": "2025-06-27",
            "built_at": "2026-05-13T00:00:00",
        })

    result = build_school_identity_package(
        core_db=db,
        school_profile_csv=school_profile,
        output_root=tmp_path / "exports",
        package_id="pkg-school-identity-test",
        source_version="fixture-school-identity",
    )
    package_dir = Path(result["package_dir"])
    assert validate_manifest(package_dir / "manifest.json")["errors"] == []
    assert result["rows"] == 4
    assert result["unmatched_rows"] == 1

    with (package_dir / "fa_bridge_school_identity.csv").open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    by_local_code = {row["local_school_code"]: row for row in rows}
    assert by_local_code["0140"]["national_school_code"] == "4121010140"
    assert by_local_code["0183"]["match_method"] == "unique_exact_school_name"
    assert by_local_code["1414"]["national_school_code"] == "4111011414"
    assert by_local_code["6407"]["local_school_name"] == "香港中文大学(深圳)"
    assert by_local_code["6407"]["national_school_code"] == "4144016407"


def test_build_merged_school_profile_package_preserves_base_rows_and_adds_reviewed_supplements(tmp_path: Path):
    fieldnames = [
        "national_school_code",
        "school_name",
        "province",
        "city",
        "school_tier",
        "school_type",
        "ownership",
        "official_site",
        "competent_authority",
        "source_date",
        "availability_date",
        "built_at",
    ]
    base_profile = tmp_path / "base_profile.csv"
    with base_profile.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerow({
            "national_school_code": "4115010127",
            "school_name": "内蒙古科技大学",
            "province": "内蒙古自治区",
            "city": "包头市",
            "school_tier": "本科",
            "school_type": "理工类",
            "ownership": "公办",
            "official_site": "https://www.imust.edu.cn/",
            "competent_authority": "内蒙古自治区",
            "source_date": "2025-06-20",
            "availability_date": "2025-06-27",
            "built_at": "2026-05-18T00:00:00",
        })

    supplemental_profile = tmp_path / "supplemental_profile.csv"
    with supplemental_profile.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerow({
            "national_school_code": "4115019127",
            "school_name": "内蒙古科技大学包头医学院",
            "province": "内蒙古自治区",
            "city": "包头市",
            "school_tier": "本科",
            "school_type": "医药类",
            "ownership": "公办",
            "official_site": "https://www.btmc.edu.cn/",
            "competent_authority": "内蒙古自治区",
            "source_date": "2026-05-18",
            "availability_date": "2026-05-18",
            "built_at": "2026-05-18T00:00:00",
        })

    result = build_merged_school_profile_package(
        base_profile_csv=base_profile,
        supplemental_profile_csv=supplemental_profile,
        output_root=tmp_path / "exports",
        package_id="pkg-school-profile-merged-test",
        source_version="fixture-school-profile-merged",
    )
    package_dir = Path(result["package_dir"])
    assert validate_manifest(package_dir / "manifest.json")["errors"] == []
    assert result["base_rows"] == 1
    assert result["supplemental_rows"] == 1
    assert result["rows"] == 2
    assert result["quality_report"]["merge_report"]["skipped_duplicate_supplemental_keys"] == []

    with (package_dir / "fa_dim_school_profile.csv").open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    by_code = {row["national_school_code"]: row for row in rows}
    assert by_code["4115010127"]["school_name"] == "内蒙古科技大学"
    assert by_code["4115019127"]["school_name"] == "内蒙古科技大学包头医学院"


def test_build_school_identity_review_plan_suggests_base_school(tmp_path: Path):
    db = tmp_path / "core.duckdb"
    con = duckdb.connect(str(db))
    try:
        con.execute("""
            CREATE TABLE fa_dim_ln_admission_plan (
                school_code VARCHAR,
                school_name VARCHAR
            )
        """)
        con.execute("""
            INSERT INTO fa_dim_ln_admission_plan VALUES
                ('9001', '北京大学医学部'),
                ('9999', '未知学院'),
                ('8888', '另一个未知学院')
        """)
    finally:
        con.close()

    school_profile = tmp_path / "school_profile.csv"
    with school_profile.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "national_school_code",
                "school_name",
                "province",
                "city",
                "school_tier",
                "school_type",
                "ownership",
                "official_site",
                "competent_authority",
                "source_date",
                "availability_date",
                "built_at",
            ],
        )
        writer.writeheader()
        writer.writerow({
            "national_school_code": "4111010001",
            "school_name": "北京大学",
            "province": "北京市",
            "city": "北京市",
            "school_tier": "本科",
            "school_type": "",
            "ownership": "",
            "official_site": "",
            "competent_authority": "教育部",
            "source_date": "2025-06-20",
            "availability_date": "2025-06-27",
            "built_at": "2026-05-13T00:00:00",
        })

    priority_missing = tmp_path / "identity_missing_schools.csv"
    with priority_missing.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "priority_rank",
                "priority_score",
                "school_code",
                "school_name",
                "plan_row_count",
                "major_count",
                "batches",
                "subject_cats",
                "coverage_area",
                "review_status",
                "notes",
            ],
        )
        writer.writeheader()
        writer.writerow({
            "priority_rank": "1",
            "priority_score": "109",
            "school_code": "9001",
            "school_name": "北京大学医学部",
            "plan_row_count": "10",
            "major_count": "9",
            "batches": "本科批",
            "subject_cats": "历史类|物理类",
            "coverage_area": "identity",
            "review_status": "todo",
            "notes": "",
        })
        writer.writerow({
            "priority_rank": "2",
            "priority_score": "88",
            "school_code": "8888",
            "school_name": "另一个未知学院",
            "plan_row_count": "8",
            "major_count": "8",
            "batches": "本科批",
            "subject_cats": "物理类",
            "coverage_area": "identity",
            "review_status": "todo",
            "notes": "",
        })

    result = build_school_identity_review_plan(
        core_db=db,
        school_profile_csv=school_profile,
        output_dir=tmp_path / "review",
        priority_missing_csv=priority_missing,
        source_date="2026-05-13",
    )

    assert result["rows"] == 3
    assert result["suggested_rows"] == 1
    with Path(result["csv"]).open(encoding="utf-8", newline="") as f:
        review_rows = list(csv.DictReader(f))
    assert [row["local_school_code"] for row in review_rows[:2]] == ["9001", "8888"]
    rows = {row["local_school_code"]: row for row in review_rows}
    assert rows["9001"]["priority_rank"] == "1"
    assert rows["9001"]["plan_row_count"] == "10"
    assert rows["9001"]["suggested_national_school_code"] == "4111010001"
    assert rows["9001"]["review_status"] == "todo"
    assert rows["9999"]["suggested_national_school_code"] == ""
    manifest = json.loads(Path(result["manifest"]).read_text(encoding="utf-8"))
    assert manifest["suggested_rows"] == 1
    assert manifest["priority_rows"] == 2

    reviewed_rows = list(rows.values())
    rows["9001"]["review_status"] = "approved"
    rows["9001"]["reviewed_national_school_code"] = rows["9001"]["suggested_national_school_code"]
    reviewed_plan = tmp_path / "reviewed_school_identity.csv"
    with reviewed_plan.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(reviewed_rows[0]))
        writer.writeheader()
        writer.writerows(reviewed_rows)
    package = build_school_identity_package(
        core_db=db,
        school_profile_csv=school_profile,
        output_root=tmp_path / "exports_reviewed",
        package_id="pkg-school-identity-reviewed-test",
        source_version="fixture-school-identity-reviewed",
        review_plan_csv=reviewed_plan,
    )
    assert package["rows"] == 1
    assert package["unmatched_rows"] == 2
    with (Path(package["package_dir"]) / "fa_bridge_school_identity.csv").open(encoding="utf-8", newline="") as f:
        bridge_rows = list(csv.DictReader(f))
    assert bridge_rows[0]["local_school_code"] == "9001"
    assert bridge_rows[0]["match_method"] == "reviewed_identity_mapping"


def test_audit_school_identity_review_plan_blocks_until_approved(tmp_path: Path):
    plan = tmp_path / "school_identity_review_plan.csv"
    rows = [
        {
            "priority_rank": "1",
            "priority_score": "109",
            "local_school_code": "9001",
            "local_school_name": "北京大学医学部",
            "plan_row_count": "10",
            "major_count": "9",
            "batches": "本科批",
            "subject_cats": "历史类|物理类",
            "reason": "unmatched",
            "candidate_count": "0",
            "suggested_national_school_code": "4111010001",
            "suggested_school_name": "北京大学",
            "suggested_province": "北京市",
            "suggested_city": "北京市",
            "suggestion_method": "base_name_contains_profile",
            "suggestion_count": "1",
            "review_status": "todo",
            "reviewed_national_school_code": "",
            "reviewer": "",
            "reviewed_at": "",
            "source_date": "2026-05-13",
            "availability_date": "2026-05-13",
            "built_at": "2026-05-13T00:00:00",
            "notes": "",
        }
    ]
    with plan.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)

    report = audit_school_identity_review_plan(plan_csv=plan)

    assert report["rows"] == 1
    assert report["blocking_rows"] == 1
    assert report["ready"]["ready_for_identity_package"] is False

    rows[0]["review_status"] = "approved"
    rows[0]["reviewed_national_school_code"] = "4111010001"
    with plan.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)
    approved_report = audit_school_identity_review_plan(plan_csv=plan, report_path=tmp_path / "audit.json")

    assert approved_report["approved_rows"] == 1
    assert approved_report["ready"]["ready_for_identity_package"] is True
    assert json.loads((tmp_path / "audit.json").read_text(encoding="utf-8"))["ready"]["ready_for_identity_package"] is True


def test_build_school_identity_review_batch_selects_priority_pending_rows(tmp_path: Path):
    plan = tmp_path / "school_identity_review_plan.csv"
    rows = [
        _school_identity_review_row("9002", "第二学校", priority_rank="2", review_status="todo"),
        _school_identity_review_row("9001", "第一学校", priority_rank="1", review_status="needs_review"),
        _school_identity_review_row("9003", "已批学校", priority_rank="3", review_status="approved"),
    ]
    _write_school_identity_review_plan(plan, rows)

    result = build_school_identity_review_batch(plan_csv=plan, output_dir=tmp_path / "batch", limit=2)

    assert result["rows"] == 2
    with Path(result["csv"]).open(encoding="utf-8", newline="") as f:
        batch_rows = list(csv.DictReader(f))
    assert [row["local_school_code"] for row in batch_rows] == ["9001", "9002"]
    manifest = json.loads(Path(result["manifest"]).read_text(encoding="utf-8"))
    assert manifest["editable_columns"] == ["notes", "review_status", "reviewed_at", "reviewed_national_school_code", "reviewer"]


def test_merge_school_identity_review_batch_updates_only_review_fields(tmp_path: Path):
    plan = tmp_path / "school_identity_review_plan.csv"
    rows = [
        _school_identity_review_row("9001", "北京大学医学部", priority_rank="1", review_status="todo"),
        _school_identity_review_row("9002", "第二学校", priority_rank="2", review_status="todo"),
    ]
    _write_school_identity_review_plan(plan, rows)

    batch = tmp_path / "school_identity_review_batch.csv"
    reviewed = [dict(rows[0])]
    reviewed[0]["local_school_name"] = "不应覆盖的名称"
    reviewed[0]["review_status"] = "approved"
    reviewed[0]["reviewed_national_school_code"] = "4111010001"
    reviewed[0]["reviewer"] = "reviewer-a"
    reviewed[0]["reviewed_at"] = "2026-05-18"
    reviewed[0]["notes"] = "matched by official MOE profile"
    _write_school_identity_review_plan(batch, reviewed)

    report = merge_school_identity_review_batch(
        plan_csv=plan,
        batch_csv=batch,
        output_csv=tmp_path / "merged.csv",
        report_path=tmp_path / "merge_report.json",
    )

    assert report["matched_rows"] == 1
    assert report["updated_rows"] == 1
    assert report["unknown_codes"] == []
    with (tmp_path / "merged.csv").open(encoding="utf-8", newline="") as f:
        merged_rows = {row["local_school_code"]: row for row in csv.DictReader(f)}
    assert merged_rows["9001"]["local_school_name"] == "北京大学医学部"
    assert merged_rows["9001"]["review_status"] == "approved"
    assert merged_rows["9001"]["reviewed_national_school_code"] == "4111010001"
    assert merged_rows["9001"]["reviewer"] == "reviewer-a"
    assert merged_rows["9002"]["review_status"] == "todo"
    assert json.loads((tmp_path / "merge_report.json").read_text(encoding="utf-8"))["updated_rows"] == 1


def test_school_identity_review_seeds_audit_and_apply(tmp_path: Path):
    plan = tmp_path / "school_identity_review_plan.csv"
    rows = [
        _school_identity_review_row("9001", "北京大学医学部", priority_rank="1", review_status="todo"),
        _school_identity_review_row("9002", "第二学校", priority_rank="2", review_status="todo"),
    ]
    _write_school_identity_review_plan(plan, rows)
    seeds = tmp_path / "school_identity_review_seeds.json"
    seeds.write_text(json.dumps({
        "version": "fixture",
        "seeds": [
            {
                "seed_id": "sid-9001",
                "local_school_code": "9001",
                "local_school_name": "北京大学医学部",
                "review_status": "approved",
                "reviewed_national_school_code": "4111010001",
                "reviewer": "reviewer-a",
                "reviewed_at": "2026-05-18",
                "review_note": "official MOE profile confirms identity",
            }
        ],
    }, ensure_ascii=False), encoding="utf-8")

    audit = audit_school_identity_review_seeds(seeds_path=seeds)
    assert audit["errors"] == []
    assert audit["status_counts"] == {"approved": 1}

    report = apply_school_identity_review_seeds(
        plan_csv=plan,
        output=tmp_path / "seeded_plan.csv",
        seeds_path=seeds,
        report_path=tmp_path / "seed_report.json",
    )

    assert report["matched_rows"] == 1
    assert report["updated_rows"] == 1
    with (tmp_path / "seeded_plan.csv").open(encoding="utf-8", newline="") as f:
        seeded_rows = {row["local_school_code"]: row for row in csv.DictReader(f)}
    assert seeded_rows["9001"]["review_status"] == "approved"
    assert seeded_rows["9001"]["reviewed_national_school_code"] == "4111010001"
    assert "seed_review=official MOE profile confirms identity" in seeded_rows["9001"]["notes"]
    assert seeded_rows["9002"]["review_status"] == "todo"
    assert json.loads((tmp_path / "seed_report.json").read_text(encoding="utf-8"))["updated_rows"] == 1


def test_school_identity_review_seeds_reject_approved_without_reviewed_code(tmp_path: Path):
    seeds = tmp_path / "bad_school_identity_review_seeds.json"
    seeds.write_text(json.dumps({
        "version": "fixture",
        "seeds": [
            {
                "seed_id": "sid-9001",
                "local_school_code": "9001",
                "local_school_name": "北京大学医学部",
                "review_status": "approved",
                "reviewer": "reviewer-a",
                "reviewed_at": "2026-05-18",
                "review_note": "missing reviewed code",
            }
        ],
    }, ensure_ascii=False), encoding="utf-8")

    audit = audit_school_identity_review_seeds(seeds_path=seeds)

    assert any("approved missing reviewed_national_school_code" in error for error in audit["errors"])


def _school_identity_review_row(
    local_school_code: str,
    local_school_name: str,
    *,
    priority_rank: str,
    review_status: str,
) -> dict[str, str]:
    return {
        "priority_rank": priority_rank,
        "priority_score": str(100 - int(priority_rank)),
        "local_school_code": local_school_code,
        "local_school_name": local_school_name,
        "plan_row_count": "10",
        "major_count": "8",
        "batches": "本科批",
        "subject_cats": "物理类",
        "reason": "unmatched",
        "candidate_count": "0",
        "suggested_national_school_code": "4111010001" if local_school_code == "9001" else "",
        "suggested_school_name": "北京大学" if local_school_code == "9001" else "",
        "suggested_province": "北京市" if local_school_code == "9001" else "",
        "suggested_city": "北京市" if local_school_code == "9001" else "",
        "suggestion_method": "base_name_contains_profile" if local_school_code == "9001" else "",
        "suggestion_count": "1" if local_school_code == "9001" else "0",
        "review_status": review_status,
        "reviewed_national_school_code": "",
        "reviewer": "",
        "reviewed_at": "",
        "source_date": "2026-05-13",
        "availability_date": "2026-05-13",
        "built_at": "2026-05-13T00:00:00",
        "notes": "",
    }


def _write_school_identity_review_plan(path: Path, rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)


def test_build_score_history_snapshot_filters_incomplete_rows(tmp_path: Path):
    db = tmp_path / "core.duckdb"
    con = duckdb.connect(str(db))
    try:
        con.execute("""
            CREATE TABLE fa_fact_ln_score_history (
                school_code VARCHAR,
                major_code VARCHAR,
                batch VARCHAR,
                subject_cat VARCHAR,
                score_year INTEGER,
                min_score DOUBLE,
                min_rank INTEGER,
                plan_count INTEGER,
                source_date VARCHAR
            )
        """)
        con.execute("""
            INSERT INTO fa_fact_ln_score_history VALUES
                ('0140', '01', '本科批', '历史类', 2025, 612, 12000, 12, '2025'),
                ('0140', '02', '本科批', '历史类', 2025, 580, NULL, 8, '2025')
        """)
    finally:
        con.close()

    result = build_score_history_snapshot_package(
        core_db=db,
        output_root=tmp_path / "exports",
        package_id="pkg-score-history-snapshot-test",
        source_version="fixture-score-history-snapshot",
    )
    package_dir = Path(result["package_dir"])
    assert validate_manifest(package_dir / "manifest.json")["errors"] == []
    assert result["rows"] == 1
    assert result["excluded_rows"] == 1
    assert result["quality_report"]["year_coverage"] == [2025]
    assert result["source_lineage"]["source_kind"] == "legacy_core_snapshot"

    with (package_dir / "fa_fact_ln_score_history.csv").open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    assert rows[0]["school_code"] == "0140"
    assert rows[0]["min_rank"] == "12000"


def test_build_admission_plan_snapshot_filters_incomplete_rows(tmp_path: Path):
    db = tmp_path / "core.duckdb"
    con = duckdb.connect(str(db))
    try:
        con.execute("""
            CREATE TABLE fa_dim_ln_admission_plan (
                school_code VARCHAR,
                school_name VARCHAR,
                major_code VARCHAR,
                major_full VARCHAR,
                major_short VARCHAR,
                batch VARCHAR,
                subject_cat VARCHAR,
                year INTEGER,
                school_tier VARCHAR,
                region VARCHAR,
                plan_count INTEGER,
                school_type VARCHAR,
                city VARCHAR,
                city_level_tag VARCHAR,
                school_rank VARCHAR,
                subject_eval VARCHAR,
                postgrad_rate DOUBLE,
                source_date VARCHAR
            )
        """)
        con.execute("""
            INSERT INTO fa_dim_ln_admission_plan VALUES
                ('0140', '辽宁大学', '01', '汉语言文学', '汉语言文学', '本科批', '历史类',
                 2026, '211', '沈阳', 12, '综合', '沈阳', '省会城市', '100', 'B+', 0.12, '2026-05-12'),
                ('0140', '辽宁大学', '02', NULL, '新闻学', '本科批', '历史类',
                 2026, '211', '沈阳', 8, '综合', '沈阳', '省会城市', '100', 'B+', 0.12, '2026-05-12')
        """)
    finally:
        con.close()

    result = build_admission_plan_snapshot_package(
        core_db=db,
        output_root=tmp_path / "exports",
        package_id="pkg-admission-plan-snapshot-test",
        source_version="fixture-admission-plan-snapshot",
    )
    package_dir = Path(result["package_dir"])
    assert validate_manifest(package_dir / "manifest.json")["errors"] == []
    assert result["rows"] == 1
    assert result["excluded_rows"] == 1
    assert result["quality_report"]["source_dates"] == ["2026-05-12"]
    assert result["source_lineage"]["source_kind"] == "legacy_core_snapshot"

    with (package_dir / "fa_dim_ln_admission_plan.csv").open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    assert rows[0]["school_code"] == "0140"
    assert rows[0]["major_full"] == "汉语言文学"


def test_download_remote_assets_from_config(tmp_path: Path, monkeypatch):
    remote = tmp_path / "remote.csv"
    remote.write_text("id,name\n1,alpha\n", encoding="utf-8")
    digest = hashlib.sha256(remote.read_bytes()).hexdigest()

    monkeypatch.setattr(
        "datahub.connectors.remote_files.load_sources",
        lambda: {
            "sources": {
                "demo_remote": {
                    "name": "demo",
                    "kind": "fixture_remote",
                    "target_tables": ["fa_demo"],
                    "acquisition": {
                        "official_distribution": "fixture distribution",
                        "evidence_urls": ["https://example.edu/source"],
                    },
                    "remote_files": [
                        {
                            "url": remote.as_uri(),
                            "file_name": "demo.csv",
                            "source_date": "2026-05-13",
                            "sha256": digest,
                        }
                    ],
                }
            }
        },
    )

    assets = download_remote_assets("demo_remote", tmp_path / "raw")
    assert len(assets) == 1
    assert assets[0].path.read_text(encoding="utf-8") == "id,name\n1,alpha\n"
    assert assets[0].path.name == "demo.csv"
    assert assets[0].source_date == "2026-05-13"
    manifest = json.loads((tmp_path / "raw/demo_remote/2026-05-13/_remote_manifest.json").read_text(encoding="utf-8"))
    assert manifest["source_key"] == "demo_remote"
    assert manifest["source_kind"] == "fixture_remote"
    assert manifest["target_tables"] == ["fa_demo"]
    assert manifest["evidence_urls"] == ["https://example.edu/source"]
    assert manifest["files"][0]["sha256"] == digest


def test_build_school_location_geocode_input_plan(tmp_path: Path):
    core_db = tmp_path / "core.duckdb"
    con = duckdb.connect(str(core_db))
    try:
        con.execute("""
            CREATE TABLE fa_dim_ln_admission_plan (
                school_code VARCHAR,
                school_name VARCHAR,
                region VARCHAR,
                batch VARCHAR
            )
        """)
        con.execute("""
            INSERT INTO fa_dim_ln_admission_plan VALUES
            ('10145', '东北大学', '辽宁沈阳', '本科批'),
            ('10145', '东北大学', '辽宁沈阳', '本科批'),
            ('10183', '吉林大学', '吉林省长春市', '本科批'),
            ('9145', '东北大学秦皇岛分校', '河北秦皇岛', '本科批'),
            ('99999', '测试学院', '辽宁大连', '本科批'),
            ('66666', '待审核学院', '辽宁沈阳', '本科批'),
            ('88888', '未匹配学院', '辽宁鞍山', '专科批'),
            ('6407', '香港中文大学(深圳)', '广东深圳市', '本科批'),
            ('6407', '香港中文大学', '广东深圳市', '艺术类本科批'),
            ('77777', '过滤学院', '辽宁沈阳', '艺术类')
        """)
    finally:
        con.close()

    profile_csv = tmp_path / "school_profile.csv"
    with profile_csv.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["national_school_code", "school_name", "province", "city"],
        )
        writer.writeheader()
        writer.writerow({
            "national_school_code": "4121010145",
            "school_name": "东北大学",
            "province": "辽宁省",
            "city": "沈阳市",
        })
        writer.writerow({
            "national_school_code": "4121099999",
            "school_name": "标准测试学院",
            "province": "辽宁省",
            "city": "大连市",
        })
        writer.writerow({
            "national_school_code": "4122010183",
            "school_name": "吉林大学",
            "province": "吉林省",
            "city": "长春市",
        })
        writer.writerow({
            "national_school_code": "4121066666",
            "school_name": "待审核学院",
            "province": "辽宁省",
            "city": "沈阳市",
        })
        writer.writerow({
            "national_school_code": "4144016407",
            "school_name": "香港中文大学(深圳)",
            "province": "广东省",
            "city": "深圳市",
        })

    identity_csv = tmp_path / "school_identity.csv"
    with identity_csv.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["local_school_code", "national_school_code", "review_status"])
        writer.writeheader()
        writer.writerow({"local_school_code": "99999", "national_school_code": "4121099999", "review_status": "approved"})
        writer.writerow({"local_school_code": "9145", "national_school_code": "4121010145", "review_status": "approved"})
        writer.writerow({"local_school_code": "6407", "national_school_code": "4144016407", "review_status": "approved"})
        writer.writerow({"local_school_code": "66666", "national_school_code": "4121066666", "review_status": "needs_review"})

    result = build_school_location_geocode_input_plan(
        core_db=core_db,
        output_dir=tmp_path / "staging",
        school_profile_csv=profile_csv,
        school_identity_csv=identity_csv,
        source_date="2026-05-13",
    )

    assert result["rows"] == 9
    assert result["distinct_local_school_count"] == 8
    assert result["duplicate_local_school_codes"] == [{
        "local_school_code": "6407",
        "request_rows": 2,
        "school_names": ["香港中文大学", "香港中文大学(深圳)"],
    }]
    assert result["ready_rows"] == 6
    assert result["blocked_rows"] == 3
    with Path(result["amap_input_csv"]).open(encoding="utf-8", newline="") as f:
        input_rows = list(csv.DictReader(f))
    assert {row["national_school_code"] for row in input_rows} == {"4121010145", "4121099999", "4122010183", "4144016407"}
    assert any(row["geocode_query"] == "沈阳市东北大学" for row in input_rows)
    assert any(row["city"] == "大连市" and row["local_school_code"] == "99999" for row in input_rows)
    branch = next(row for row in input_rows if row["local_school_code"] == "9145")
    assert branch["campus_key"] == "ln_9145"
    assert branch["city"] == "秦皇岛"
    assert branch["geocode_query"] == "秦皇岛东北大学秦皇岛分校"
    jlu = next(row for row in input_rows if row["local_school_code"] == "10183")
    assert jlu["city"] == "长春市"
    assert jlu["geocode_query"] == "长春市吉林大学"
    hk_rows = [row for row in input_rows if row["local_school_code"] == "6407"]
    assert len(hk_rows) == 2
    assert len({row["campus_key"] for row in hk_rows}) == 2
    assert all(row["campus_key"].startswith("ln_6407_") for row in hk_rows)
    with Path(result["plan_csv"]).open(encoding="utf-8", newline="") as f:
        plan_rows = list(csv.DictReader(f))
    blocked = [row for row in plan_rows if row["request_status"] == "blocked"]
    assert len(blocked) == 3
    assert {row["local_school_code"]: row["blocking_reason"] for row in blocked} == {
        "66666": "identity_not_approved;missing_national_school_code",
        "88888": "missing_national_school_code",
        "77777": "missing_national_school_code",
    }
    manifest = json.loads(Path(result["manifest"]).read_text(encoding="utf-8"))
    assert manifest["source_key"] == "school_location_geocode"
    assert manifest["distinct_local_school_count"] == 8
    assert "--address-column geocode_query" in manifest["fetch_command_hint"]

    audit = audit_school_location_geocode_input(
        plan_csv=Path(result["plan_csv"]),
        input_csv=Path(result["amap_input_csv"]),
        output=tmp_path / "staging" / "audit.json",
    )
    assert audit["errors"] == []
    assert audit["row_counts"]["ready_rows"] == 6
    assert audit["primary_key_checks"]["duplicate_count"] == 0
    assert audit["warnings"][0]["count"] == 3

    duplicate_input = tmp_path / "duplicate_input.csv"
    duplicate_input.write_text(
        Path(result["amap_input_csv"]).read_text(encoding="utf-8")
        + input_rows[0]["national_school_code"] + ",10145," + input_rows[0]["school_name"] + ","
        + input_rows[0]["campus_key"] + ",duplicate,admission_unit,辽宁,沈阳市,辽宁沈阳,沈阳市东北大学,2026-05-13,2026-05-13,2026-05-13T00:00:00\n",
        encoding="utf-8",
    )
    duplicate_audit = audit_school_location_geocode_input(
        plan_csv=Path(result["plan_csv"]),
        input_csv=duplicate_input,
    )
    assert any("duplicate input primary keys" in error for error in duplicate_audit["errors"])


def test_build_school_location_geocode_input_plan_uses_core_identity_profile(tmp_path: Path):
    core_db = tmp_path / "core.duckdb"
    con = duckdb.connect(str(core_db))
    try:
        con.execute("""
            CREATE TABLE fa_dim_ln_admission_plan (
                school_code VARCHAR,
                school_name VARCHAR,
                region VARCHAR,
                batch VARCHAR
            )
        """)
        con.execute("""
            INSERT INTO fa_dim_ln_admission_plan VALUES
            ('10145', '东北大学', '辽宁沈阳', '本科批'),
            ('10183', '吉林大学', '吉林省长春市', '本科批')
        """)
        con.execute("""
            CREATE TABLE fa_bridge_school_identity (
                local_school_code VARCHAR,
                national_school_code VARCHAR
            )
        """)
        con.execute("""
            INSERT INTO fa_bridge_school_identity VALUES
            ('10145', '4121010145'),
            ('10183', '4122010183')
        """)
        con.execute("""
            CREATE TABLE fa_dim_school_profile (
                national_school_code VARCHAR,
                school_name VARCHAR,
                province VARCHAR,
                city VARCHAR
            )
        """)
        con.execute("""
            INSERT INTO fa_dim_school_profile VALUES
            ('4121010145', '东北大学', '辽宁省', '沈阳市'),
            ('4122010183', '吉林大学', '吉林省', '长春市')
        """)
    finally:
        con.close()

    result = build_school_location_geocode_input_plan(
        core_db=core_db,
        output_dir=tmp_path / "staging",
        source_date="2026-05-18",
    )

    assert result["rows"] == 2
    assert result["distinct_local_school_count"] == 2
    assert result["ready_rows"] == 2
    assert result["blocked_rows"] == 0
    with Path(result["amap_input_csv"]).open(encoding="utf-8", newline="") as f:
        input_rows = list(csv.DictReader(f))
    assert {row["national_school_code"] for row in input_rows} == {"4121010145", "4122010183"}
    with Path(result["plan_csv"]).open(encoding="utf-8", newline="") as f:
        plan_rows = list(csv.DictReader(f))
    assert {row["match_method"] for row in plan_rows} == {"identity_bridge"}


def test_fetch_amap_web_api_geocode_writes_raw_manifest(tmp_path: Path, monkeypatch):
    source = tmp_path / "schools.csv"
    source.write_text("school_name,address,city\n东北大学,沈阳市和平区文化路3号巷11号,沈阳\n", encoding="utf-8")

    monkeypatch.setenv("AMAP_WEB_SERVICE_KEY", "fixture-key")
    monkeypatch.setattr(
        "datahub.connectors.amap_web_api.load_sources",
        lambda: {
            "sources": {
                "school_location_geocode": {
                    "name": "高校地理位置增强",
                    "kind": "verified_address_plus_amap_web_api_geocode",
                    "target_tables": ["fa_dim_school_location"],
                    "interfaces": {
                        "web_service": {
                            "provider": "amap_web_service",
                            "key_env": "AMAP_WEB_SERVICE_KEY",
                            "endpoints": {
                                "geocode": "https://restapi.amap.com/v3/geocode/geo",
                            },
                        },
                        "request_policy": {"timeout_seconds": 3, "rate_limit_per_second": 0},
                    },
                    "acquisition": {
                        "status": "web_api_configured_requires_connector",
                        "official_distribution": "fixture amap",
                        "evidence_urls": ["https://lbs.amap.com/api/webservice/guide/api/georegeo"],
                    },
                }
            }
        },
    )

    class FakeResponse:
        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, tb):
            return False

        def read(self):
            return b'{"status":"1","geocodes":[{"location":"123.1,41.8"}]}'

    requested = []

    def fake_urlopen(request, timeout):
        requested.append(request.full_url)
        return FakeResponse()

    monkeypatch.setattr("datahub.connectors.amap_web_api.urlopen", fake_urlopen)

    result = fetch_amap_web_api(
        source_key="school_location_geocode",
        operation="geocode",
        input_path=source,
        output_root=tmp_path / "raw",
        source_date="2026-05-13",
        address_column="address",
        city_column="city",
    )

    assert result["request_count"] == 1
    assert "fixture-key" in requested[0]
    manifest = json.loads(Path(result["manifest_path"]).read_text(encoding="utf-8"))
    assert manifest["source_key"] == "school_location_geocode"
    assert manifest["key_env"] == "AMAP_WEB_SERVICE_KEY"
    assert "key" not in manifest["request_params_without_key"][0]
    assert manifest["request_params_without_key"][0]["address"] == "沈阳市和平区文化路3号巷11号"
    lines = Path(result["jsonl_path"]).read_text(encoding="utf-8").splitlines()
    record = json.loads(lines[0])
    assert record["params"]["city"] == "沈阳"
    assert record["response"]["status"] == "1"
    assert "key" not in record["params"]


def test_audit_amap_web_api_readiness_blocks_missing_key(tmp_path: Path, monkeypatch):
    source = tmp_path / "schools.csv"
    source.write_text("school_name,address,city\n东北大学,沈阳市和平区文化路3号巷11号,沈阳\n", encoding="utf-8")
    monkeypatch.delenv("AMAP_WEB_SERVICE_KEY", raising=False)
    monkeypatch.setattr(
        "datahub.connectors.amap_web_api_readiness.load_sources",
        lambda: {
            "sources": {
                "school_location_geocode": {
                    "interfaces": {
                        "web_service": {
                            "provider": "amap_web_service",
                            "key_env": "AMAP_WEB_SERVICE_KEY",
                            "endpoints": {
                                "geocode": "https://restapi.amap.com/v3/geocode/geo",
                            },
                        }
                    },
                }
            }
        },
    )

    report = audit_amap_web_api_readiness(
        source_key="school_location_geocode",
        operation="geocode",
        input_path=source,
        address_column="address",
        output=tmp_path / "readiness.json",
    )

    assert report["ready_for_fetch"] is False
    assert report["key_present"] is False
    assert report["row_counts"]["requestable_rows"] == 1
    assert any("AMAP_WEB_SERVICE_KEY" in error for error in report["errors"])
    assert json.loads((tmp_path / "readiness.json").read_text(encoding="utf-8"))["ready_for_fetch"] is False


def test_audit_amap_web_api_readiness_passes_without_fetching(tmp_path: Path, monkeypatch):
    source = tmp_path / "schools.csv"
    source.write_text("school_name,address,city\n东北大学,沈阳市和平区文化路3号巷11号,沈阳\n", encoding="utf-8")
    monkeypatch.setenv("AMAP_WEB_SERVICE_KEY", "fixture-key")
    monkeypatch.setattr(
        "datahub.connectors.amap_web_api_readiness.load_sources",
        lambda: {
            "sources": {
                "school_location_geocode": {
                    "interfaces": {
                        "web_service": {
                            "provider": "amap_web_service",
                            "key_env": "AMAP_WEB_SERVICE_KEY",
                            "endpoints": {
                                "geocode": "https://restapi.amap.com/v3/geocode/geo",
                            },
                        }
                    },
                }
            }
        },
    )

    report = audit_amap_web_api_readiness(
        source_key="school_location_geocode",
        operation="geocode",
        input_path=source,
        address_column="address",
    )

    assert report["ready_for_fetch"] is True
    assert report["key_present"] is True
    assert report["errors"] == []


def test_fetch_amap_web_api_district_uses_config_scope(tmp_path: Path, monkeypatch):
    monkeypatch.setenv("AMAP_WEB_SERVICE_KEY", "fixture-key")
    monkeypatch.setattr(
        "datahub.connectors.amap_web_api.load_sources",
        lambda: {
            "sources": {
                "region_profile_geocode": {
                    "name": "城市与行政区基础信息",
                    "kind": "amap_web_api_district_profile",
                    "target_tables": ["fa_dim_region_profile"],
                    "interfaces": {
                        "web_service": {
                            "provider": "amap_web_service",
                            "key_env": "AMAP_WEB_SERVICE_KEY",
                            "endpoints": {
                                "district": "https://restapi.amap.com/v3/config/district",
                            },
                        },
                        "request_policy": {"timeout_seconds": 3, "rate_limit_per_second": 0},
                        "scope": {"province": "辽宁省"},
                    },
                    "acquisition": {
                        "status": "web_api_configured_requires_connector",
                        "official_distribution": "fixture district",
                        "evidence_urls": ["https://lbs.amap.com/api/webservice/guide/api/district"],
                    },
                }
            }
        },
    )

    class FakeResponse:
        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, tb):
            return False

        def read(self):
            return '{"status":"1","districts":[{"name":"辽宁省"}]}'.encode("utf-8")

    monkeypatch.setattr("datahub.connectors.amap_web_api.urlopen", lambda request, timeout: FakeResponse())

    result = fetch_amap_web_api(
        source_key="region_profile_geocode",
        operation="district",
        output_root=tmp_path / "raw",
        source_date="2026-05-13",
    )

    manifest = json.loads(Path(result["manifest_path"]).read_text(encoding="utf-8"))
    assert result["request_count"] == 1
    assert manifest["request_params_without_key"][0]["keywords"] == "辽宁省"
    assert manifest["request_params_without_key"][0]["subdistrict"] == "3"


def test_build_region_profile_package_from_amap_district(tmp_path: Path):
    raw_dir = tmp_path / "raw" / "region_profile_geocode" / "2026-05-13"
    raw_dir.mkdir(parents=True)
    raw_jsonl = raw_dir / "amap_web_api_district.jsonl"
    raw_manifest = raw_dir / "_amap_web_api_district.json"
    raw_jsonl.write_text(
        json.dumps({
            "request_index": 1,
            "operation": "district",
            "endpoint": "https://restapi.amap.com/v3/config/district",
            "params": {"keywords": "辽宁省", "subdistrict": "3"},
            "source_row": None,
            "raw_response_hash": "hash-region",
            "response": {
                "status": "1",
                "districts": [
                    {
                        "name": "辽宁省",
                        "adcode": "210000",
                        "citycode": [],
                        "center": "123.429096,41.796767",
                        "level": "province",
                        "districts": [
                            {
                                "name": "沈阳市",
                                "adcode": "210100",
                                "citycode": "024",
                                "center": "123.431474,41.805698",
                                "level": "city",
                                "districts": [
                                    {
                                        "name": "和平区",
                                        "adcode": "210102",
                                        "citycode": "024",
                                        "center": "123.395319,41.789766",
                                        "level": "district",
                                        "districts": [],
                                    }
                                ],
                            }
                        ],
                    }
                ],
            },
            "fetched_at": "2026-05-13T00:00:00",
        }, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    raw_manifest.write_text(json.dumps({
        "source_key": "region_profile_geocode",
        "source_name": "城市与行政区基础信息",
        "source_kind": "amap_web_api_district_profile",
        "source_date": "2026-05-13",
        "intake_at": "2026-05-13T00:00:00",
        "acquired_by": "datahub.fetch_amap_web_api",
        "official_distribution": "fixture district",
        "evidence_urls": ["https://lbs.amap.com/api/webservice/guide/api/district"],
        "target_tables": ["fa_dim_region_profile"],
        "operation": "district",
        "endpoint": "https://restapi.amap.com/v3/config/district",
        "key_env": "AMAP_WEB_SERVICE_KEY",
        "request_count": 1,
        "request_params_without_key": [{"keywords": "辽宁省", "subdistrict": "3"}],
        "files": [{"file_name": raw_jsonl.name, "path": str(raw_jsonl), "size_bytes": raw_jsonl.stat().st_size, "sha256": "fixture"}],
    }, ensure_ascii=False), encoding="utf-8")

    result = build_region_profile_package_from_amap_district(
        raw_jsonl=raw_jsonl,
        output_root=tmp_path / "exports",
        package_id="pkg-region-profile-test",
        source_version="fixture-region-profile",
    )

    package_dir = Path(result["package"]["package_dir"])
    assert validate_manifest(package_dir / "manifest.json")["errors"] == []
    assert result["rows"] == 3
    with (package_dir / "fa_dim_region_profile.csv").open(encoding="utf-8", newline="") as f:
        rows = {row["adcode"]: row for row in csv.DictReader(f)}
    assert rows["210000"]["region_level"] == "province"
    assert rows["210100"]["city"] == "沈阳市"
    assert rows["210102"]["district"] == "和平区"
    assert rows["210102"]["parent_adcode"] == "210100"
    assert rows["210102"]["coordinate_system"] == "GCJ-02"


def test_build_school_location_package_from_amap_geocode(tmp_path: Path, monkeypatch):
    raw_dir = tmp_path / "raw" / "school_location_geocode" / "2026-05-13"
    raw_dir.mkdir(parents=True)
    raw_jsonl = raw_dir / "amap_web_api_geocode.jsonl"
    raw_manifest = raw_dir / "_amap_web_api_geocode.json"
    raw_jsonl.write_text(
        json.dumps({
            "request_index": 1,
            "operation": "geocode",
            "endpoint": "https://restapi.amap.com/v3/geocode/geo",
            "params": {"address": "沈阳市和平区文化路3号巷11号", "city": "沈阳"},
            "source_row": {
                "national_school_code": "4121010145",
                "local_school_code": "10145",
                "school_name": "东北大学",
                "campus_key": "main",
                "campus_name": "南湖校区",
                "address": "沈阳市和平区文化路3号巷11号",
                "source_address_url": "https://example.edu/address",
            },
            "raw_response_hash": "abc123",
            "response": {
                "status": "1",
                "geocodes": [{
                    "formatted_address": "辽宁省沈阳市和平区文化路3号巷11号",
                    "province": "辽宁省",
                    "city": "沈阳市",
                    "district": "和平区",
                    "township": "南湖街道",
                    "street": "文化路",
                    "number": "3号巷11号",
                    "adcode": "210102",
                    "citycode": "024",
                    "location": "123.421,41.765",
                    "level": "门牌号",
                    "id": "B001",
                }],
            },
            "fetched_at": "2026-05-13T00:00:00",
        }, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    raw_manifest.write_text(json.dumps({
        "source_key": "school_location_geocode",
        "source_name": "高校地理位置增强",
        "source_kind": "verified_address_plus_amap_web_api_geocode",
        "source_date": "2026-05-13",
        "operation": "geocode",
        "intake_at": "2026-05-13T00:00:00",
        "acquired_by": "datahub.fetch_amap_web_api",
        "official_distribution": "fixture amap",
        "evidence_urls": ["https://lbs.amap.com/api/webservice/guide/api/georegeo"],
        "target_tables": ["fa_dim_school_location"],
    }, ensure_ascii=False), encoding="utf-8")

    monkeypatch.setattr(
        "datahub.builders.school_location_from_amap.load_sources",
        lambda: {
            "sources": {
                "school_location_geocode": {
                    "name": "高校地理位置增强",
                    "interfaces": {
                        "coordinate_system": "GCJ-02",
                        "default_campus_key": "main",
                        "geocode_confidence_by_level": {"门牌号": 0.95, "unknown": 0.5},
                    },
                }
            }
        },
    )

    result = build_school_location_package_from_amap_geocode(
        raw_jsonl=raw_jsonl,
        raw_manifest=raw_manifest,
        output_root=tmp_path / "exports",
        package_id="pkg-school-location-test",
        source_version="fixture-school-location",
    )

    package_dir = Path(result["package"]["package_dir"])
    assert validate_manifest(package_dir / "manifest.json")["errors"] == []
    assert result["rows"] == 1
    with (package_dir / "fa_dim_school_location.csv").open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    row = rows[0]
    assert row["national_school_code"] == "4121010145"
    assert row["campus_key"] == "main"
    assert row["coordinate_system"] == "GCJ-02"
    assert row["longitude"] == "123.421"
    assert row["latitude"] == "41.765"
    assert row["geocode_confidence"] == "0.95"
    assert row["geocode_raw_hash"] == "abc123"
    manifest = json.loads((package_dir / "manifest.json").read_text(encoding="utf-8"))
    assert manifest["source_lineage"]["raw_manifest"] == str(raw_manifest)
    assert manifest["source_lineage"]["source_kind"] == "parsed_amap_web_api_geocode"

    bad_jsonl = raw_dir / "amap_web_api_geocode_bad.jsonl"
    bad_jsonl.write_text(
        json.dumps({
            "request_index": 2,
            "operation": "geocode",
            "params": {"address": "沈阳市测试学院", "city": "沈阳"},
            "source_row": {
                "national_school_code": "4121000000",
                "local_school_code": "0000",
                "school_name": "测试学院",
                "campus_key": "main",
                "campus_name": "主校区",
                "city": "沈阳",
            },
            "raw_response_hash": "bad123",
            "response": {
                "status": "1",
                "geocodes": [{
                    "formatted_address": "辽宁省大连市测试路",
                    "province": "辽宁省",
                    "city": "大连市",
                    "district": "西岗区",
                    "adcode": "210203",
                    "citycode": "0411",
                    "location": "121.6,38.9",
                    "level": "市",
                }],
            },
        }, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    with pytest.raises(ValueError, match="confidence below minimum"):
        build_school_location_package_from_amap_geocode(
            raw_jsonl=bad_jsonl,
            raw_manifest=raw_manifest,
            output_root=tmp_path / "exports",
            package_id="pkg-school-location-bad-geocode-test",
            source_version="fixture-school-location",
        )

    duplicate_manifest = raw_dir / "_amap_web_api_geocode_duplicate.json"
    duplicate_manifest.write_text(
        (
            '{"source_key":"school_location_geocode",'
            '"source_key":"shadow_source",'
            '"source_date":"2026-05-13",'
            '"operation":"geocode"}\n'
        ),
        encoding="utf-8",
    )
    try:
        build_school_location_package_from_amap_geocode(
            raw_jsonl=raw_jsonl,
            raw_manifest=duplicate_manifest,
            output_root=tmp_path / "exports",
            package_id="pkg-school-location-duplicate-manifest-test",
            source_version="fixture-school-location",
        )
        duplicate_manifest_rejected = False
    except ValueError as exc:
        duplicate_manifest_rejected = "duplicate JSON key" in str(exc)
    assert duplicate_manifest_rejected


def test_download_page_images_from_config(tmp_path: Path, monkeypatch):
    image = tmp_path / "table.png"
    image.write_bytes(b"image-bytes")
    html = tmp_path / "page.html"
    html.write_text(f'<html><body><img src="{image.name}"></body></html>', encoding="utf-8")

    monkeypatch.setattr(
        "datahub.connectors.page_images.load_sources",
        lambda: {
            "sources": {
                "demo_images": {
                    "name": "demo images",
                    "kind": "official_page_images",
                    "target_tables": ["fa_fact_ln_score_distribution"],
                    "acquisition": {
                        "status": "remote_configured",
                        "official_distribution": "fixture page",
                        "evidence_urls": [html.as_uri()],
                    },
                    "page_image_sources": [
                        {
                            "page_url": html.as_uri(),
                            "kind": "mirror_page_images",
                            "source_date": "2024-06-25",
                            "file_prefix": "fixture",
                        }
                    ],
                }
            }
        },
    )

    result = download_page_images("demo_images", tmp_path / "raw")
    assert result["file_count"] == 1
    manifest = Path(result["pages"][0]["manifest_path"])
    data = json.loads(manifest.read_text(encoding="utf-8"))
    assert data["source_kind"] == "mirror_page_images"
    assert data["files"][0]["sha256"] == hashlib.sha256(b"image-bytes").hexdigest()
    assert Path(data["files"][0]["path"]).read_bytes() == b"image-bytes"


def test_ocr_page_images_writes_configured_manifest(tmp_path: Path, monkeypatch):
    image = tmp_path / "table.png"
    image.write_bytes(b"image-bytes")
    input_manifest = tmp_path / "raw" / "ln_score_distribution" / "2024-06-25" / "_page_images_index.json"
    input_manifest.parent.mkdir(parents=True)
    input_manifest.write_text(json.dumps({
        "source_key": "ln_score_distribution",
        "source_name": "辽宁普通高考成绩统计表",
        "source_kind": "official_page_images",
        "source_date": "2024-06-25",
        "evidence_urls": ["https://jyt.ln.gov.cn/example/index.shtml"],
        "target_tables": ["fa_fact_ln_score_distribution"],
        "files": [
            {
                "file_name": "table.png",
                "path": str(image),
                "sha256": "image-sha256-fixture",
            }
        ],
    }, ensure_ascii=False), encoding="utf-8")

    monkeypatch.setattr(
        "datahub.connectors.macos_vision_ocr.load_sources",
        lambda: {
            "sources": {
                "ln_score_distribution": {
                    "ocr": {
                        "engine": "macos_vision",
                        "recognition_languages": ["zh-Hans", "en-US"],
                        "recognition_level": "accurate",
                        "uses_language_correction": False,
                    }
                }
            }
        },
    )
    monkeypatch.setattr("datahub.connectors.macos_vision_ocr._compile_swift_ocr", lambda swiftc, binary_path: None)
    monkeypatch.setattr(
        "datahub.connectors.macos_vision_ocr._run_vision_ocr",
        lambda binary_path, image_paths, ocr_config: [
            {
                "image_path": str(image_paths[0]),
                "observations": [
                    {"text": "676及以上12", "confidence": 0.91, "x": 0.1, "y": 0.9, "width": 0.2, "height": 0.03}
                ],
            }
        ],
    )

    result = ocr_page_images(
        "ln_score_distribution",
        tmp_path / "raw",
        tmp_path / "ocr",
        manifest_paths=[input_manifest],
    )

    assert result["file_count"] == 1
    assert result["observation_count"] == 1
    ocr_manifest = json.loads(Path(result["pages"][0]["ocr_manifest"]).read_text(encoding="utf-8"))
    assert ocr_manifest["source_kind"] == "official_page_image_ocr"
    assert ocr_manifest["recognition_languages"] == ["zh-Hans", "en-US"]
    assert ocr_manifest["files"][0]["sha256"] == "image-sha256-fixture"
    jsonl = Path(result["pages"][0]["ocr_jsonl"]).read_text(encoding="utf-8")
    assert "676及以上12" in jsonl

    duplicate_manifest = tmp_path / "raw" / "ln_score_distribution" / "2024-06-25" / "_page_images_duplicate.json"
    duplicate_manifest.write_text(
        (
            '{"source_key":"ln_score_distribution",'
            '"source_key":"shadow_source",'
            '"source_date":"2024-06-25",'
            '"files":[{"file_name":"table.png","path":"'
            f'{image}'
            '"}]}\n'
        ),
        encoding="utf-8",
    )
    try:
        ocr_page_images(
            "ln_score_distribution",
            tmp_path / "raw",
            tmp_path / "ocr",
            manifest_paths=[duplicate_manifest],
        )
        duplicate_manifest_rejected = False
    except ValueError as exc:
        duplicate_manifest_rejected = "duplicate JSON key" in str(exc)
    assert duplicate_manifest_rejected


def test_parse_moe_major_catalog_lines():
    rows = parse_moe_major_catalog_lines([
        "01  学科门类：哲学",
        "0101 哲学类",
        "010101  哲学",
        "010103K  宗教学",
        "02",
        "学科门类：经济学",
        "0203 金融学类",
        "020306T  信用管理（注：可授经济学或管理学学士学位）",
        "0502 外国语言文学类",
        "0502100T 语言学",
    ])

    assert rows[0] == {
        "major_code": "010101",
        "major_name": "哲学",
        "major_category": "哲学",
        "major_class": "哲学类",
        "degree_type": "哲学学士",
        "study_years": None,
    }
    assert rows[2]["major_code"] == "020306T"
    assert rows[2]["major_name"] == "信用管理"
    assert rows[2]["degree_type"] == "可授经济学或管理学学士学位"
    assert rows[3]["major_code"] == "0502100T"
    assert rows[3]["major_name"] == "语言学"


def test_parse_ln_application_workbook_outputs_plan_and_score_history(tmp_path: Path):
    path = tmp_path / "application.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "物理类"
    ws.append([
        "院校代码",
        "专业代码",
        "专业简称",
        "院校名称",
        "院校水平",
        "专业全称",
        "保研率",
        "25年\n最低分",
        "25最低位次",
        "24年\n最低分",
        "24最低位次",
        "25年计划",
        "24年计划",
        "地域",
        "性质",
        "城市水平标签",
    ])
    ws.append([
        "0001",
        "56",
        "0002",
        "北京大学",
        "C9联盟/部委直属",
        "理科试验班类(理科基础类专业)",
        0.65,
        690,
        100,
        685,
        120,
        2,
        2,
        "北京市",
        "公办",
        "一线城市",
    ])
    ws.append([
        "0001",
        "56",
        "0002",
        "北京大学",
        "C9联盟/部委直属",
        "理科试验班类(理科基础类专业)",
        0.65,
        690,
        100,
        685,
        120,
        2,
        2,
        "北京市",
        "公办",
        "一线城市",
    ])
    ws2 = wb.create_sheet("物理类特殊")
    ws2.append(["院校代码", "专业代码", "院校名称", "专业全称", "25年最低分", "25年位次"])
    ws2.append(["0140", "AC", "辽宁大学", "法学类", 602, 12959])
    wb.save(path)

    result = parse_ln_application_workbooks([path])
    report = write_application_workbook_outputs(
        result,
        plan_output=tmp_path / "plan.csv",
        score_output=tmp_path / "score.csv",
        report_output=tmp_path / "report.json",
    )

    assert report["row_counts"]["fa_dim_ln_admission_plan"] == 2
    assert report["row_counts"]["fa_fact_ln_score_history"] == 3
    assert report["duplicate_counts"]["fa_dim_ln_admission_plan"] == 1
    assert any(item["sheet"] == "物理类特殊" for item in report["matched_sheets"])
    with (tmp_path / "plan.csv").open(encoding="utf-8", newline="") as f:
        plan_rows = list(csv.DictReader(f))
    assert plan_rows[0]["batch"] == "本科批"
    assert plan_rows[0]["subject_cat"] == "物理类"
    assert plan_rows[0]["year"] == "2026"
    assert plan_rows[0]["plan_count"] == "2"
    assert plan_rows[0]["keep_research_rate"] == "0.65"
    assert plan_rows[0]["school_nature"] == "公办"
    assert plan_rows[0]["source_date"] == "2025-08-27"
    assert plan_rows[0]["id"]
    assert plan_rows[1]["school_code"] == "0140"
    assert plan_rows[1]["major_code"] == "AC"
    assert plan_rows[1]["major_full"] == "法学类"
    with (tmp_path / "score.csv").open(encoding="utf-8", newline="") as f:
        score_rows = list(csv.DictReader(f))
    assert {row["score_year"] for row in score_rows} == {"2024", "2025"}
    assert score_rows[0]["min_rank"] == "100"
    assert score_rows[0]["score_type"] == "最低分"
    assert score_rows[0]["built_at"]


def test_parse_ln_application_workbook_rejects_duplicate_json_keys(tmp_path: Path):
    config = tmp_path / "ln_application_workbook.json"
    config.write_text('{"profiles":{},"profiles":{}}\n', encoding="utf-8")

    with pytest.raises(ValueError, match="duplicate JSON key"):
        parse_ln_application_workbooks([], config_path=config)


def test_parse_ln_projection_score_xlsx(tmp_path: Path):
    path = tmp_path / "projection.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "物理学科类"
    ws.append(["2025年辽宁省普通高等学校招生录取普通类本科批（物理学科类）投档最低分"])
    ws.append(["说明"])
    ws.append(["院校\n编号", "招生院校", "专业\n编号", "招生专业", "投档\n最低分"])
    ws.append([None, None, None, None, None, "（一）", "（二）", "（三）", "（四）", "（五）", "（六）", "（七）"])
    ws.append([None, None, None, None, None, "语数\n成绩", "语数\n最高\n成绩", "外语\n成绩", "首选\n科目\n成绩", "再选\n科目\n最高\n成绩", "再选\n科目\n次高\n成绩", "志\n愿\n号"])
    ws.append(["0378", "安徽财经大学", "13", "计算机类", "574", "229", "126", "130", "61", "82", "72", "10"])
    wb.save(path)

    rows = parse_ln_projection_score_file(
        path,
        score_year=2025,
        batch="本科批",
        source_date="2025-07-20",
        password_candidates=[],
    )

    assert len(rows) == 1
    assert rows[0]["subject_cat"] == "物理类"
    assert rows[0]["school_code"] == "0378"
    assert rows[0]["major_code"] == "13"
    assert rows[0]["min_score"] == 574
    assert "语数成绩" in rows[0]["tie_breaker_json"]


def test_parse_ln_projection_score_zip(tmp_path: Path):
    xlsx = tmp_path / "projection.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "历史学科类"
    ws.append(["2024年辽宁省普通高等学校招生录取普通类本科批（历史学科类）投档最低分"])
    ws.append(["说明"])
    ws.append(["院校\n编号", "招生院校", "专业\n编号", "招生专业", "投档\n最低分"])
    ws.append([None, None, None, None, None, "（一）", "（二）", "（三）", "（四）", "（五）", "（六）", "（七）"])
    ws.append([None, None, None, None, None, "语数\n成绩", "语数\n最高\n成绩", "外语\n成绩", "首选\n科目\n成绩", "再选\n科目\n最高\n成绩", "再选\n科目\n次高\n成绩", "志\n愿\n号"])
    ws.append(["0357", "安徽大学", "1A", "汉语言文学", "604", "236", "121", "121", "70", "89", "88", "7"])
    wb.save(xlsx)
    archive = tmp_path / "projection.zip"
    with ZipFile(archive, "w") as z:
        z.write(xlsx, arcname="projection.xlsx")

    rows = parse_ln_projection_score_file(
        archive,
        score_year=2024,
        batch="本科批",
        source_date="2024-07-20",
        password_candidates=[],
    )

    assert len(rows) == 1
    assert rows[0]["subject_cat"] == "历史类"
    assert rows[0]["score_year"] == 2024
    assert rows[0]["min_score"] == 604


def test_build_major_mapping_review_package_promotes_approved_rows(tmp_path: Path):
    db = tmp_path / "core.duckdb"
    con = duckdb.connect(str(db))
    try:
        con.execute("""
            CREATE TABLE fa_bridge_major_tdx (
                major_code VARCHAR,
                major_name VARCHAR,
                tdx_l2 VARCHAR,
                tdx_l2_name VARCHAR,
                tdx_l1_name VARCHAR,
                mapping_type VARCHAR,
                confidence VARCHAR,
                rationale VARCHAR,
                source_date DATE,
                availability_date DATE,
                built_at TIMESTAMP
            )
        """)
        con.execute("""
            INSERT INTO fa_bridge_major_tdx VALUES
                ('080901', '计算机科学与技术', 'T1205', '软件服务', '信息产业',
                 'primary', 'high', 'fixture', DATE '2026-05-12',
                 DATE '2026-05-12', TIMESTAMP '2026-05-12 00:00:00')
        """)
        con.execute("""
            CREATE TABLE fa_mart_major_mapping_review_queue (
                major_name VARCHAR,
                major_code_sample VARCHAR,
                plan_rows INTEGER,
                candidate_tdx_l2 VARCHAR,
                candidate_tdx_l2_name VARCHAR,
                candidate_tdx_l1_name VARCHAR,
                mapping_confidence VARCHAR,
                mapping_source VARCHAR,
                mapping_rationale VARCHAR,
                review_status VARCHAR,
                review_notes VARCHAR,
                source_date DATE,
                availability_date DATE,
                reason_codes_json VARCHAR,
                signal_contribution_json VARCHAR,
                pit_lineage_json VARCHAR,
                built_at TIMESTAMP
            )
        """)
        con.execute("""
            INSERT INTO fa_mart_major_mapping_review_queue VALUES
                ('会计学', '01', 235, 'T1001', '银行', '金融', 'medium',
                 'config:rules:accounting_finance', 'approved rationale',
                 'approved', 'manual ok', DATE '2026-05-13',
                 DATE '2026-05-13', '[]', '{}', '{}', TIMESTAMP '2026-05-13 00:00:00'),
                ('法学', '02', 214, 'T1301', '综合类', '综合类', 'low',
                 'config:entries', 'candidate rationale', 'candidate', NULL,
                 DATE '2026-05-13', DATE '2026-05-13', '[]', '{}', '{}',
                 TIMESTAMP '2026-05-13 00:00:00')
        """)
    finally:
        con.close()

    result = build_major_mapping_review_package(
        core_db=db,
        output_root=tmp_path / "exports",
        package_id="pkg-review-test",
        source_version="fixture-review",
    )
    package_dir = Path(result["package_dir"])
    report = validate_manifest(package_dir / "manifest.json")
    assert report["errors"] == []
    assert result["rows"] == 2
    assert result["promoted_rows"] == 1

    with (package_dir / "fa_bridge_major_tdx.csv").open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    promoted = next(row for row in rows if row["major_name"] == "会计学")
    assert promoted["major_code"].startswith("REVIEW_NAME_")
    assert promoted["major_code"] != "01"
    assert promoted["tdx_l2"] == "T1001"
    assert "manual ok" in promoted["rationale"]
